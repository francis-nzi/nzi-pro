"""Generate single-sheet Excel template from factor lookup data."""

import os
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.dataset_selector import (
    get_monthly_headers,
    get_reporting_period_display,
)
from services.fetch_existing_data import fetch_existing_scope_entries
from core.database import get_conn


def generate_single_sheet_template(
    job_id: int,
    client_name: str = "",
    site_name: str = "",
    job_number: str = "",
    reporting_year: str = "",
    report_from: str = "",
    report_to: str = "",
    include_custom_factors: bool = True,
    include_prev_year: bool = True,
) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Upload"

    ws["A1"] = "Client Name:"
    ws["B1"] = client_name
    ws["E1"] = "Job Number:"
    ws["F1"] = job_number

    ws["A2"] = "Site Name:"
    ws["B2"] = site_name
    ws["E2"] = "Reporting Period:"
    ws["F2"] = get_reporting_period_display(job_id)

    ws["A3"] = "Data Files:"
    ws["B3"] = "Standard UK Conversion Factors"

    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    for row in range(1, 4):
        for col in ["A", "C", "E"]:
            cell = ws[f"{col}{row}"]
            cell.font = Font(bold=True)
            cell.fill = header_fill

    monthly_headers = get_monthly_headers(job_id)
    headers = ["Scope", "Category", "Report Label", "ID", "UOM"] + monthly_headers + ["Qty", "Data Source", "Notes"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    factors = _load_reference_template_rows()
    existing_data = fetch_existing_scope_entries(job_id)

    current_row = 5
    for factor in factors:
        ws.cell(row=current_row, column=1, value=factor["scope"])
        ws.cell(row=current_row, column=2, value=factor["category"])
        ws.cell(row=current_row, column=3, value=factor["report_label"])
        ws.cell(row=current_row, column=4, value=factor["id"])
        ws.cell(row=current_row, column=5, value=factor["uom"])

        existing = existing_data.get(factor["id"], {})
        monthly_values = existing.get("monthly", [])
        for month_idx, month_col in enumerate(range(6, 18)):
            value = monthly_values[month_idx] if month_idx < len(monthly_values) else None
            ws.cell(row=current_row, column=month_col, value=value)

        month_start_col = get_column_letter(6)
        month_end_col = get_column_letter(17)
        ws.cell(row=current_row, column=18, value=f"=SUM({month_start_col}{current_row}:{month_end_col}{current_row})")
        ws.cell(row=current_row, column=19, value="Company Data")
        ws.cell(row=current_row, column=20, value=existing.get("notes"))
        current_row += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 12
    for col_idx in range(6, 18):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10
    ws.column_dimensions["R"].width = 12
    ws.column_dimensions["S"].width = 15
    ws.column_dimensions["T"].width = 30
    ws.column_dimensions["D"].hidden = True
    ws.freeze_panes = "A5"

    if include_prev_year:
        previous_year_sheets = _fetch_previous_year_rows(job_id=job_id, site_name=site_name)
        for year, rows in previous_year_sheets:
            if not rows:
                continue
            sheet_name = f"Previous {year}"
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws_prev = wb.create_sheet(title=sheet_name[:31])
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_prev.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            out_row = 2
            for row in rows:
                ws_prev.cell(row=out_row, column=1, value=row.get("scope"))
                ws_prev.cell(row=out_row, column=2, value=row.get("category"))
                ws_prev.cell(row=out_row, column=3, value=row.get("report_label"))
                ws_prev.cell(row=out_row, column=4, value=row.get("original_id"))
                ws_prev.cell(row=out_row, column=5, value=row.get("uom"))
                months = row.get("months") or []
                for month_idx, month_col in enumerate(range(6, 18)):
                    ws_prev.cell(row=out_row, column=month_col, value=(months[month_idx] if month_idx < len(months) else None))
                ws_prev.cell(row=out_row, column=18, value=row.get("qty"))
                ws_prev.cell(row=out_row, column=19, value="Previous Year")
                ws_prev.cell(row=out_row, column=20, value=row.get("notes"))
                out_row += 1

            ws_prev.column_dimensions["A"].width = 10
            ws_prev.column_dimensions["B"].width = 30
            ws_prev.column_dimensions["C"].width = 50
            ws_prev.column_dimensions["D"].width = 20
            ws_prev.column_dimensions["E"].width = 12
            for col_idx in range(6, 18):
                ws_prev.column_dimensions[get_column_letter(col_idx)].width = 10
            ws_prev.column_dimensions["R"].width = 12
            ws_prev.column_dimensions["S"].width = 15
            ws_prev.column_dimensions["T"].width = 30
            ws_prev.column_dimensions["D"].hidden = True
            ws_prev.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    def sanitize(s: str) -> str:
        if not s:
            return ""
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            s = s.replace(char, "_")
        return s.strip()

    job_no = sanitize(job_number) or f"job_{job_id}"
    client = sanitize(client_name) or "client"
    site = sanitize(site_name) or "site"
    year = sanitize(reporting_year) or "year"
    filename = f"{job_no} {client} {site} {year}.xlsx"

    return buffer.getvalue(), filename


def _load_reference_template_rows() -> list[dict]:
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(current_dir)
    template_path = os.path.join(project_root, "templates", "NZI Data Upload Template - Standard UK.xlsx")

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Reference template not found at: {template_path}")

    wb = load_workbook(template_path, data_only=True)
    ws = wb.active

    header_row = None
    for r in range(1, 20):
        vals = [ws.cell(r, c).value for c in range(1, 25)]
        norm = [str(x).strip().lower() if x else "" for x in vals]
        if "scope" in norm and "id" in norm:
            header_row = r
            break

    if not header_row:
        raise ValueError("Could not find header row in reference template")

    factors = []
    for r in range(header_row + 1, ws.max_row + 1):
        scope = ws.cell(r, 1).value
        category = ws.cell(r, 2).value
        report_label = ws.cell(r, 3).value
        factor_id = ws.cell(r, 4).value
        uom = ws.cell(r, 5).value

        if factor_id:
            factors.append(
                {
                    "scope": scope or "",
                    "category": category or "",
                    "report_label": report_label or "",
                    "id": str(factor_id).strip(),
                    "uom": uom or "",
                }
            )

    return factors


def _fetch_previous_year_rows(job_id: int, site_name: str) -> list[tuple[int, list[dict]]]:
    """Return prior-year rows grouped by reporting year for the same client/site."""
    with get_conn() as con:
        current = con.execute(
            """
            SELECT job_id, client_db_id, reporting_year, reporting_period_end
            FROM jobs
            WHERE job_id = ?
            """,
            [int(job_id)],
        ).fetchone()

        if not current:
            return []

        client_db_id = int(current[1]) if current[1] is not None else None
        current_year = int(current[2]) if current[2] is not None else None
        current_period_end = current[3]
        if client_db_id is None:
            return []

        prior_filter_sql = ""
        params: list[object] = [client_db_id]
        if current_period_end is not None:
            prior_filter_sql += " AND j.reporting_period_end IS NOT NULL AND j.reporting_period_end < ? "
            params.append(current_period_end)
        elif current_year is not None:
            prior_filter_sql += " AND j.reporting_year < ? "
            params.append(current_year)

        site_id = None
        normalized_site = (site_name or "").strip().lower()
        if normalized_site and normalized_site not in ("all", "all sites"):
            site_row = con.execute(
                """
                SELECT site_id
                FROM client_sites
                WHERE client_db_id = ? AND LOWER(TRIM(site_name)) = ?
                ORDER BY site_id
                LIMIT 1
                """,
                [client_db_id, normalized_site],
            ).fetchone()
            if site_row and site_row[0] is not None:
                site_id = int(site_row[0])

        prev_jobs = con.execute(
            f"""
            SELECT j.job_id, j.reporting_year, j.reporting_period_end
            FROM jobs j
            WHERE j.client_db_id = ?
            {prior_filter_sql}
            ORDER BY j.reporting_period_end DESC NULLS LAST, j.reporting_year DESC, j.job_id DESC
            """,
            params,
        ).fetchall()

        grouped: list[tuple[int, list[dict]]] = []
        for pj in prev_jobs or []:
            prev_job_id = int(pj[0])
            prev_year = int(pj[1]) if pj[1] is not None else 0
            if pj[2] is not None and hasattr(pj[2], "year"):
                prev_year = int(pj[2].year)

            if site_id is not None:
                rows = con.execute(
                    """
                    SELECT scope, category, COALESCE(NULLIF(report_label, ''), category) AS report_label,
                           original_id, uom, qty, notes,
                           month_1, month_2, month_3, month_4, month_5, month_6,
                           month_7, month_8, month_9, month_10, month_11, month_12
                    FROM job_scope_rows
                    WHERE job_id = ?
                      AND enabled = TRUE
                      AND site_id = ?
                    ORDER BY scope, category, report_label, original_id
                    """,
                    [prev_job_id, site_id],
                ).fetchall()
            else:
                rows = con.execute(
                    """
                    SELECT scope, category, COALESCE(NULLIF(report_label, ''), category) AS report_label,
                           original_id, uom, qty, notes,
                           month_1, month_2, month_3, month_4, month_5, month_6,
                           month_7, month_8, month_9, month_10, month_11, month_12
                    FROM job_scope_rows
                    WHERE job_id = ?
                      AND enabled = TRUE
                    ORDER BY scope, category, report_label, original_id
                    """,
                    [prev_job_id],
                ).fetchall()

            # If selected site has no rows in that prior job, fall back to all-site rows
            # so users still get prior-year context in the download.
            if site_id is not None and (not rows):
                rows = con.execute(
                    """
                    SELECT scope, category, COALESCE(NULLIF(report_label, ''), category) AS report_label,
                           original_id, uom, qty, notes,
                           month_1, month_2, month_3, month_4, month_5, month_6,
                           month_7, month_8, month_9, month_10, month_11, month_12
                    FROM job_scope_rows
                    WHERE job_id = ?
                      AND enabled = TRUE
                    ORDER BY scope, category, report_label, original_id
                    """,
                    [prev_job_id],
                ).fetchall()

            mapped: list[dict] = []
            for r in rows or []:
                mapped.append(
                    {
                        "scope": r[0],
                        "category": r[1],
                        "report_label": r[2],
                        "original_id": r[3],
                        "uom": r[4],
                        "qty": r[5],
                        "notes": r[6],
                        "months": [r[i] for i in range(7, 19)],
                    }
                )

            # Legacy fallback for older jobs that only have CRP scope entries.
            if not mapped:
                legacy_rows = con.execute(
                    """
                    SELECT scope, category, COALESCE(subcategory, description) AS report_label,
                           original_id, unit, amount, notes
                    FROM crp_scope_entries
                    WHERE job_id = ?
                      AND is_archived = FALSE
                    ORDER BY scope, category, subcategory, entry_id
                    """,
                    [prev_job_id],
                ).fetchall()
                for r in legacy_rows or []:
                    mapped.append(
                        {
                            "scope": r[0],
                            "category": r[1],
                            "report_label": r[2],
                            "original_id": r[3],
                            "uom": r[4],
                            "qty": r[5],
                            "notes": r[6],
                            "months": [],
                        }
                    )
            if mapped:
                grouped.append((prev_year, mapped))
        return grouped
