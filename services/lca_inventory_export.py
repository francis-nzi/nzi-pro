"""Downloadable Excel export of an LCA assessment's Inventory Breakdown --
the component x module pivot shown in JobLca.tsx's Inventory Items table.

Column span is every module in lca_modules_lookup (A1 through S15 --
LCA product modules and PCF/service Scope 3 categories share one lookup
table, sorted by sort_order), not just the modules a given assessment
happens to use -- this keeps the workbook's column layout stable
regardless of assessment type. Values are shown in kgCO2e (4dp), matching
the Inventory Breakdown table's current default unit.

File naming and job-data header-row conventions match the other job-scoped
downloads (services/download_filenames.py, api/spend_data_routes.py's
download_spend_template) rather than the simpler BOM upload template.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.database import get_conn
from services.download_filenames import build_download_filename
from services.lca_engine import resolve_line_emissions_tco2e, safe_float

TRANSPORT_MODULE_CODES = ("A2", "A4", "C2")


def _format_period_label(period_start: Any, period_end: Any) -> str:
    def _fmt(val: Any) -> str:
        if val is None:
            return "?"
        if hasattr(val, "strftime"):
            return val.strftime("%b %Y")
        try:
            return datetime.fromisoformat(str(val).strip()[:10]).strftime("%b %Y")
        except Exception:
            return str(val)
    return f"{_fmt(period_start)} - {_fmt(period_end)}"


def _job_and_assessment_context(con, job_id: int, assessment_id: int) -> dict[str, Any] | None:
    row = con.execute(
        """
        SELECT j.job_number, c.client_name, j.reporting_period_start, j.reporting_period_end,
               a.name, a.sku, a.lifecycle_boundary, a.total_tco2e
        FROM lca_assessments a
        JOIN jobs j ON j.job_id = a.job_id
        JOIN clients c ON c.db_id = j.client_db_id
        WHERE a.assessment_id = %s AND a.job_id = %s
        """,
        [int(assessment_id), int(job_id)],
    ).fetchone()
    if not row:
        return None
    return {
        "job_number": row[0],
        "client_name": row[1],
        "period_start": row[2],
        "period_end": row[3],
        "assessment_name": row[4],
        "sku": row[5],
        "lifecycle_boundary": row[6],
        "total_tco2e": safe_float(row[7]),
    }


def _all_modules(con) -> list[tuple[str, str]]:
    df = con.execute(
        "SELECT module_code, label FROM lca_modules_lookup WHERE is_active = TRUE ORDER BY sort_order"
    ).df()
    if df is None or df.empty:
        return []
    return [(str(r["module_code"]), str(r["label"])) for _, r in df.iterrows()]


def _line_items(con, assessment_id: int) -> list[dict[str, Any]]:
    df = con.execute(
        "SELECT * FROM lca_line_items WHERE assessment_id = %s ORDER BY module_code, line_item_id",
        [int(assessment_id)],
    ).df()
    if df is None or df.empty:
        return []
    df = df.astype(object).where(df.notna(), None)
    items = []
    for _, r in df.iterrows():
        row = dict(r)
        is_placeholder = bool(row.get("is_placeholder") or False)
        items.append(
            {
                "line_item_id": int(row["line_item_id"]),
                "component_id": row.get("component_id"),
                "line_label": row.get("line_label") or "",
                "module_code": row.get("module_code") or "",
                "quantity": safe_float(row.get("quantity")),
                "unit": row.get("unit") or "",
                "mapped_factor_source": row.get("mapped_factor_source"),
                "is_gap_filled": bool(row.get("is_gap_filled") or False),
                "is_placeholder": is_placeholder,
                "transport_emissions_tco2e": row.get("transport_emissions_tco2e"),
                "emissions_tco2e": 0.0 if is_placeholder else round(resolve_line_emissions_tco2e(row), 6),
            }
        )
    return items


def _line_item_status(item: dict[str, Any]) -> str:
    if item["is_placeholder"]:
        return "placeholder"
    needs_legs = item["module_code"] in TRANSPORT_MODULE_CODES and not item["transport_emissions_tco2e"]
    if needs_legs:
        return "needs_review"
    return "mapped" if (item["mapped_factor_source"] or item["is_gap_filled"]) else "needs_review"


def _build_breakdown_rows(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Component x module pivot -- same grouping as JobLca.tsx's
    inventoryBreakdown useMemo: keyed by component_id when present (so a
    component's several module lines merge into one row), else by label."""
    component_labels: dict[int, str] = {}
    for item in items:
        cid = item["component_id"]
        if cid is not None and cid not in component_labels:
            component_labels[int(cid)] = item["line_label"]

    rows: dict[str, dict[str, Any]] = {}
    for item in items:
        cid = item["component_id"]
        key = f"c-{int(cid)}" if cid is not None else f"l-{item['line_label']}"
        row = rows.get(key)
        if row is None:
            label = component_labels.get(int(cid)) if cid is not None else None
            row = {
                "key": key,
                "label": label or item["line_label"],
                "module_totals": {},
                "row_total": 0.0,
                "items": [],
            }
            rows[key] = row
        row["module_totals"][item["module_code"]] = row["module_totals"].get(item["module_code"], 0.0) + item["emissions_tco2e"]
        row["row_total"] += item["emissions_tco2e"]
        row["items"].append(item)

    breakdown_rows = []
    for row in rows.values():
        qty_item = next((i for i in row["items"] if i["module_code"] not in TRANSPORT_MODULE_CODES), row["items"][0])
        statuses = [_line_item_status(i) for i in row["items"]]
        if "needs_review" in statuses:
            status = "Needs review"
        elif "placeholder" in statuses:
            status = "Placeholder"
        else:
            status = "Mapped"
        breakdown_rows.append(
            {
                "label": row["label"],
                "quantity": qty_item["quantity"],
                "unit": qty_item["unit"],
                "module_totals": row["module_totals"],
                "row_total": row["row_total"],
                "status": status,
            }
        )
    breakdown_rows.sort(key=lambda r: r["label"].lower())
    return breakdown_rows


def generate_inventory_breakdown_export(job_id: int, assessment_id: int) -> tuple[bytes, str] | None:
    """Returns (xlsx_bytes, filename), or None if the job/assessment doesn't exist."""
    with get_conn() as con:
        context = _job_and_assessment_context(con, int(job_id), int(assessment_id))
        if context is None:
            return None
        modules = _all_modules(con)
        items = _line_items(con, int(assessment_id))

    job_number = context["job_number"] or f"job-{job_id}"
    client_name = context["client_name"] or "Client"
    breakdown_rows = _build_breakdown_rows(items)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Breakdown"

    ws["A1"] = "Client Name:"
    ws["B1"] = client_name
    ws["E1"] = "Job Number:"
    ws["F1"] = job_number
    ws["A2"] = "Assessment:"
    ws["B2"] = context["assessment_name"] or ""
    ws["E2"] = "Reporting Period:"
    ws["F2"] = _format_period_label(context["period_start"], context["period_end"])
    ws["A3"] = "Lifecycle Boundary:"
    ws["B3"] = str(context["lifecycle_boundary"] or "").replace("_", " ")
    ws["E3"] = "Total (kgCO2e):"
    ws["F3"] = round(context["total_tco2e"] * 1000, 4)

    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    for r in range(1, 4):
        for c in ("A", "C", "E"):
            cell = ws[f"{c}{r}"]
            cell.font = Font(bold=True)
            cell.fill = header_fill

    header_row = 5
    column_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    headers = ["Component", "Qty", "Unit"] + [code for code, _label in modules] + ["Total (kgCO2e)", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = column_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    module_col_start = 4
    total_col = module_col_start + len(modules)
    status_col = total_col + 1
    current_row = header_row + 1
    for row in breakdown_rows:
        ws.cell(row=current_row, column=1, value=row["label"])
        ws.cell(row=current_row, column=2, value=row["quantity"])
        ws.cell(row=current_row, column=3, value=row["unit"])
        for col_offset, (code, _label) in enumerate(modules):
            value = row["module_totals"].get(code)
            cell = ws.cell(row=current_row, column=module_col_start + col_offset)
            if value is not None:
                cell.value = round(value * 1000, 4)
                cell.number_format = "0.0000"
        total_cell = ws.cell(row=current_row, column=total_col, value=round(row["row_total"] * 1000, 4))
        total_cell.number_format = "0.0000"
        ws.cell(row=current_row, column=status_col, value=row["status"])
        current_row += 1

    grand_total_row = current_row
    ws.cell(row=grand_total_row, column=1, value="Grand Total").font = Font(bold=True)
    grand_total_tco2e = 0.0
    for col_offset, (code, _label) in enumerate(modules):
        column_total = sum(r["module_totals"].get(code, 0.0) for r in breakdown_rows)
        cell = ws.cell(row=grand_total_row, column=module_col_start + col_offset)
        if column_total:
            cell.value = round(column_total * 1000, 4)
            cell.number_format = "0.0000"
        cell.font = Font(bold=True)
    grand_total_tco2e = sum(r["row_total"] for r in breakdown_rows)
    gt_cell = ws.cell(row=grand_total_row, column=total_col, value=round(grand_total_tco2e * 1000, 4))
    gt_cell.number_format = "0.0000"
    gt_cell.font = Font(bold=True)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 8
    for col_offset in range(len(modules)):
        ws.column_dimensions[ws.cell(row=header_row, column=module_col_start + col_offset).column_letter].width = 10
    ws.column_dimensions[ws.cell(row=header_row, column=total_col).column_letter].width = 14
    ws.column_dimensions[ws.cell(row=header_row, column=status_col).column_letter].width = 14

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    file_name = build_download_filename(
        job_number=job_number,
        client_name=client_name,
        descriptor=f"LCA Inventory Breakdown - {context['assessment_name'] or 'Assessment'}",
        period_start=context["period_start"],
        period_end=context["period_end"],
    )
    return stream.getvalue(), file_name
