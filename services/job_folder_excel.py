import logging
import io
import os

from openpyxl import Workbook, load_workbook

from core.database import get_conn
from services.sites import list_sites

logger = logging.getLogger(__name__)


def _replace_sheet(wb, name: str):
    if name in wb.sheetnames:
        ws_old = wb[name]
        wb.remove(ws_old)
    return wb.create_sheet(title=name)


def _append_kv(ws, rows: list[tuple[str, object]]):
    for k, v in rows:
        ws.append([k, v])


def _init_scope_sheet(ws):
    ws["A1"].value = "Site Name:"
    ws["B1"].value = ""
    ws["C1"].value = "Report From:"
    ws["D1"].value = ""
    ws["E1"].value = "To:"
    ws["F1"].value = ""
    ws["A2"].value = "Data Files:"
    ws["B2"].value = ""
    ws["C2"].value = "Client Name:"
    ws["D2"].value = ""
    ws["E2"].value = "Job Number:"
    ws["F2"].value = ""
    ws.append([])
    ws.append(["ID", "Qty", "Apply"])


def _load_template_workbook() -> Workbook:
    try:
        base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
        override = (os.getenv("NZI_EXCEL_TEMPLATE_PATH") or "").strip()
        if override:
            template_path = override
            if not os.path.isabs(template_path):
                template_path = os.path.join(base_dir, template_path)
        else:
            template_path = os.path.join(base_dir, "templates", "NZI Data Upload Template - Basic UK.xlsx")
        if os.path.exists(template_path):
            return load_workbook(template_path)
    except Exception as exc:
        logger.warning("Falling back to generated Excel template workbook: %s", exc)

    wb = Workbook()
    try:
        wb.remove(wb.active)
    except Exception as exc:
        logger.warning("Unable to remove default workbook sheet; continuing with generated sheets: %s", exc)

    for name in ("Scope 1", "Scope 2", "Scope 3"):
        ws = wb.create_sheet(title=name)
        _init_scope_sheet(ws)
    return wb


def _job_core_data(job_id: int):
    try:
        with get_conn() as con:
            hdr = con.execute(
                """
                SELECT j.job_id, j.job_number, j.title, j.job_type, j.reporting_year, j.status,
                       j.start_date, j.due_date,
                       j.client_db_id, c.client_name
                FROM jobs j
                JOIN clients c ON c.db_id = j.client_db_id
                WHERE j.job_id=%s
                """,
                [int(job_id)],
            ).fetchone()

            crp = con.execute(
                """
                SELECT reporting_period_from, reporting_period_to,
                       client_order_number, client_contact_name, client_contact_email,
                       report_signee_name, report_signee_position,
                       num_employees, turnover_gbp, premises_size_m2,
                       vehicles_owned, vehicles_leased, premises_owned, premises_leased
                FROM crp_job_details
                WHERE job_id=%s
                """,
                [int(job_id)],
            ).fetchone()

            plan = con.execute(
                """
                SELECT data_collection_due, first_draft_due, final_report_due
                FROM job_plan
                WHERE job_id=%s
                """,
                [int(job_id)],
            ).fetchone()

        return hdr, crp, plan
    except Exception:
        return None, None, None


def build_excel_template_bytes(*, job_id: int, selected_site: str, include_prev_year: bool = True) -> tuple[bytes, str]:
    try:
        with get_conn() as con:
            row = con.execute(
                """
                SELECT j.job_id, j.job_number, j.reporting_year, j.client_db_id, c.client_name,
                       crp.reporting_period_from, crp.reporting_period_to
                FROM jobs j
                JOIN clients c ON c.db_id = j.client_db_id
                LEFT JOIN crp_job_details crp ON crp.job_id = j.job_id
                WHERE j.job_id=%s
                """,
                [int(job_id)],
            ).fetchone()
    except Exception:
        wb = _load_template_workbook()

        ws_core = _replace_sheet(wb, "Core Data")
        _append_kv(
            ws_core,
            [
                ("Job ID", int(job_id)),
                ("Template Site", selected_site),
                ("Include Previous Year", bool(include_prev_year)),
            ],
        )

        for ws in wb.worksheets:
            try:
                if ws["A1"].value and str(ws["A1"].value).strip().startswith("Site Name:"):
                    ws["B1"].value = selected_site
            except Exception as exc:
                logger.warning("Unable to update template site name on worksheet %s: %s", ws.title, exc)

        bio = io.BytesIO()
        wb.save(bio)
        data = bio.getvalue()
        filename = f"job_{int(job_id)}_data_upload_template.xlsx"
        return data, filename

    if not row:
        raise ValueError("Job not found")

    (_jid, job_number, reporting_year, client_db_id, client_name, rp_from, rp_to) = row

    sites_df = list_sites(int(client_db_id))

    data_files_ref = ""
    try:
        with get_conn() as con:
            df_ds = con.execute(
                """
                SELECT jsc.scope, d.name, d.year, d.analysis_type, d.country
                FROM job_scope_config jsc
                LEFT JOIN datasets d ON d.dataset_id = jsc.dataset_id
                WHERE jsc.job_id=%s
                ORDER BY jsc.scope
                """,
                [int(job_id)],
            ).df()

        parts = []
        if df_ds is not None and (not df_ds.empty):
            for _, r in df_ds.iterrows():
                if r.get("name") is None:
                    continue
                label = str(r.get("name") or "").strip()
                y = r.get("year")
                if y is not None and str(y) != "nan":
                    label = f"{label} {int(y)}"
                parts.append(label)
        data_files_ref = ", ".join(parts)
    except Exception as exc:
        logger.warning("Unable to derive Excel template data file references for job %s: %s", job_id, exc)
        data_files_ref = ""

    wb = _load_template_workbook()
    hdr, crp, plan = _job_core_data(int(job_id))

    if hdr is not None:
        (
            _jid2,
            _job_number,
            _title,
            _job_type,
            _year,
            _status,
            _start_date,
            _due_date,
            _client_db_id,
            _client_name,
        ) = hdr

        ws_core = _replace_sheet(wb, "Core Data")
        _append_kv(
            ws_core,
            [
                ("Client", _client_name),
                ("Job Number", _job_number),
                ("Job Title", _title),
                ("Job Type", _job_type),
                ("Job Status", _status),
                ("Reporting Year", _year),
                ("Start Date", _start_date),
                ("Due Date", _due_date),
                ("Reporting Period From", rp_from),
                ("Reporting Period To", rp_to),
                ("Template Site", selected_site),
            ],
        )

        if crp is not None:
            (
                crp_from,
                crp_to,
                client_order_number,
                client_contact_name,
                client_contact_email,
                report_signee_name,
                report_signee_position,
                num_employees,
                turnover_gbp,
                premises_size_m2,
                vehicles_owned,
                vehicles_leased,
                premises_owned,
                premises_leased,
            ) = crp

            ws_core.append([])
            _append_kv(
                ws_core,
                [
                    ("Client Order Number", client_order_number),
                    ("Client Contact Name", client_contact_name),
                    ("Client Contact Email", client_contact_email),
                    ("Report Signee Name", report_signee_name),
                    ("Report Signee Position", report_signee_position),
                    ("Employees", num_employees),
                    ("Turnover GBP", turnover_gbp),
                    ("Premises Size (m2)", premises_size_m2),
                    ("Vehicles Owned", vehicles_owned),
                    ("Vehicles Leased", vehicles_leased),
                    ("Premises Owned", premises_owned),
                    ("Premises Leased", premises_leased),
                ],
            )

            if crp_from or crp_to:
                ws_core.append([])
                _append_kv(
                    ws_core,
                    [
                        ("CRP Reporting Period From (stored)", crp_from),
                        ("CRP Reporting Period To (stored)", crp_to),
                    ],
                )

        if plan is not None:
            (data_collection_due, first_draft_due, final_report_due) = plan
            ws_core.append([])
            _append_kv(
                ws_core,
                [
                    ("Milestone: Data collection due", data_collection_due),
                    ("Milestone: First draft due", first_draft_due),
                    ("Milestone: Final report due", final_report_due),
                ],
            )

        try:
            ws_sites = _replace_sheet(wb, "Sites")
            if sites_df is not None and not sites_df.empty:
                ws_sites.append(list(sites_df.columns))
                for row_vals in sites_df.itertuples(index=False, name=None):
                    ws_sites.append(list(row_vals))
            else:
                ws_sites.append(["No sites found for this client."])
        except Exception as exc:
            logger.warning("Unable to append sites worksheet for job %s: %s", job_id, exc)

    for ws in wb.worksheets:
        if ws["A1"].value and str(ws["A1"].value).strip().startswith("Site Name:"):
            ws["B1"].value = selected_site
        if ws["C1"].value and str(ws["C1"].value).strip().startswith("Report From:"):
            ws["D1"].value = rp_from
        if ws["E1"].value and str(ws["E1"].value).strip().startswith("To:"):
            ws["F1"].value = rp_to
        if ws["A2"].value and str(ws["A2"].value).strip().startswith("Data Files:"):
            ws["B2"].value = data_files_ref or ws["B2"].value

        ws["C2"].value = "Client Name:"
        ws["D2"].value = client_name
        ws["E2"].value = "Job Number:"
        ws["F2"].value = job_number

    if include_prev_year:
        prev_year = int(reporting_year or 0) - 1 if reporting_year is not None else None
        prev_job_id = None
        if prev_year:
            try:
                with get_conn() as con:
                    r = con.execute(
                        """
                        SELECT job_id
                        FROM jobs
                        WHERE client_db_id=%s AND reporting_year=%s
                        ORDER BY job_id DESC
                        LIMIT 1
                        """,
                        [int(client_db_id), int(prev_year)],
                    ).fetchone()
                if r:
                    prev_job_id = int(r[0])
            except Exception as exc:
                logger.warning("Unable to resolve previous-year job for job %s: %s", job_id, exc)
                prev_job_id = None

        if prev_job_id is not None:
            try:
                with get_conn() as con:
                    prev_df = con.execute(
                        """
                        SELECT scope, category, subcategory, description, amount, unit, tco2e, method, notes, updated_at
                        FROM crp_scope_entries
                        WHERE job_id=%s AND is_archived=FALSE
                        ORDER BY scope, category, subcategory, entry_id
                        """,
                        [int(prev_job_id)],
                    ).df()

                if prev_df is not None and (not prev_df.empty):
                    name = f"Previous Year ({prev_year})"
                    if name in wb.sheetnames:
                        ws_prev = wb[name]
                        wb.remove(ws_prev)
                    ws_prev = wb.create_sheet(title=name)
                    ws_prev.append(["Client", client_name])
                    ws_prev.append(["Job", job_number])
                    ws_prev.append(["Site", selected_site])
                    ws_prev.append([])
                    ws_prev.append(list(prev_df.columns))
                    for row_vals in prev_df.itertuples(index=False, name=None):
                        ws_prev.append(list(row_vals))
            except Exception as exc:
                logger.warning("Unable to add previous-year worksheet for job %s: %s", job_id, exc)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn_site = "".join(ch if ch.isalnum() or ch in ("-", "_", " ") else "_" for ch in str(selected_site))
    filename = f"{job_number} - {fn_site} - NZI Data Upload Template.xlsx"
    return buf.getvalue(), filename
