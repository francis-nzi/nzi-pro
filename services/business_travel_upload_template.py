from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.database import get_conn
from services.dataset_selector import get_monthly_headers, get_reporting_period_display
from services.fetch_existing_data import fetch_existing_scope_entries


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _safe_filename_part(value: Any) -> str:
    text = _clean(value)
    if not text:
        return ""
    for char in '<>:"/\\|?*':
        text = text.replace(char, "_")
    text = text.replace(" ", "_")
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


def _job_context(con, job_id: int) -> dict[str, Any]:
    row = con.execute(
        """
        SELECT j.job_id, j.job_number, j.reporting_year, j.reporting_period_end,
               c.client_name, c.db_id
        FROM jobs j
        JOIN clients c ON c.db_id = j.client_db_id
        WHERE j.job_id=%s
        """,
        [int(job_id)],
    ).fetchone()
    if not row:
        return {}

    reporting_year = row[2]
    if reporting_year is None and row[3] is not None:
        try:
            reporting_year = int(getattr(row[3], "year", None) or str(row[3])[:4])
        except Exception:
            reporting_year = None

    return {
        "job_id": int(row[0]),
        "job_number": row[1],
        "reporting_year": reporting_year,
        "reporting_period_end": row[3],
        "client_name": row[4],
        "client_db_id": int(row[5]) if row[5] is not None else None,
    }


def _resolve_selected_site(con, client_db_id: int | None, site_id: int | None) -> dict[str, Any] | None:
    if client_db_id is None or site_id is None:
        return None
    row = con.execute(
        """
        SELECT site_id, site_name
        FROM client_sites
        WHERE client_db_id=%s
          AND site_id=%s
        LIMIT 1
        """,
        [int(client_db_id), int(site_id)],
    ).fetchone()
    if not row:
        return None
    return {"site_id": int(row[0]), "site_name": str(row[1]).strip() if row[1] is not None else None}

def _load_reference_rows_from_csv() -> list[dict[str, Any]]:
    current_dir = Path(__file__).resolve().parent
    template_path = current_dir.parent / "templates" / "NZI Data Upload Template - Standard UK.csv"
    if not template_path.exists():
        raise FileNotFoundError(f"Reference template not found at: {template_path}")

    rows: list[list[str]] = []
    with template_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows.extend(list(reader))

    header_row = None
    headers: list[str] = []
    for idx, row in enumerate(rows[:40]):
        norm = [_clean(value).lower() for value in row]
        if "scope" in norm and "id" in norm and "qty" in norm:
            header_row = idx
            headers = norm
            break

    if header_row is None:
        raise ValueError("Could not find header row in business travel reference template")

    def col(name: str) -> int | None:
        try:
            return headers.index(name.lower())
        except ValueError:
            return None

    scope_col = col("scope")
    category_col = col("category")
    report_label_col = col("report label")
    id_col = col("id")
    uom_col = col("uom")
    ghg_col = col("ghg unit")
    factor_col = col("factor")
    notes_col = col("notes")

    if scope_col is None or category_col is None or report_label_col is None or id_col is None or uom_col is None:
        raise ValueError("Business travel reference template is missing required columns")

    factors: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for row in rows[header_row + 1 :]:
        if len(row) <= max(scope_col, category_col, report_label_col, id_col, uom_col):
            continue
        scope = _clean(row[scope_col])
        category = _clean(row[category_col])
        report_label = _clean(row[report_label_col])
        original_id = _clean(row[id_col])
        if not original_id:
            continue
        if original_id in seen_ids:
            continue
        if scope != "Scope 3":
            continue
        if category.lower() != "business travel":
            continue

        factors.append(
            {
                "scope": scope or "Scope 3",
                "category": "Business Travel",
                "report_label": report_label,
                "original_id": original_id,
                "uom": _clean(row[uom_col]),
                "ghg_unit": _clean(row[ghg_col]) if ghg_col is not None and ghg_col < len(row) else "",
                "factor": _clean(row[factor_col]) if factor_col is not None and factor_col < len(row) else "",
                "notes": _clean(row[notes_col]) if notes_col is not None and notes_col < len(row) else "",
            }
        )
        seen_ids.add(original_id)

    factors.sort(key=lambda item: (item["report_label"].lower(), item["original_id"]))
    return factors


def _load_reference_rows_from_dataset(reporting_year: int | None = None) -> list[dict[str, Any]]:
    with get_conn() as con:
        params: list[object] = []
        where_parts = [
            "COALESCE(d.archived, FALSE) = FALSE",
            "LOWER(TRIM(COALESCE(fl.scope, ''))) = 'scope 3'",
            "LOWER(TRIM(COALESCE(fl.category, fl.level_1, ''))) = 'business travel'",
        ]
        if reporting_year is not None:
            where_parts.append("COALESCE(fl.year, d.year) = ?")
            params.append(int(reporting_year))

        df = con.execute(
            f"""
            SELECT
                fl.db_id,
                fl.dataset_id,
                d.name AS dataset,
                d.analysis_type,
                d.country,
                COALESCE(fl.year, d.year) AS year,
                fl.file_name,
                fl.original_id,
                fl.scope,
                fl.category,
                fl.level_1,
                fl.level_2,
                fl.level_3,
                fl.level_4,
                fl.column_text,
                fl.report_label,
                fl.uom,
                fl.ghg_unit,
                fl.factor,
                fl.source,
                fl.region,
                fl.currency,
                fl.method,
                fl.valid_from,
                fl.valid_to
            FROM v_factor_lookup fl
            LEFT JOIN datasets d ON d.dataset_id = fl.dataset_id
            WHERE {" AND ".join(where_parts)}
            ORDER BY COALESCE(fl.report_label, fl.column_text, fl.original_id), fl.original_id, fl.db_id
            """,
            params,
        ).df()

    factors: list[dict[str, Any]] = []
    if df is not None and not df.empty:
        safe_df = df.astype(object).where(df.notna(), None)
        for record in safe_df.to_dict(orient="records"):
            factor = _serialize_reference_row(record)
            if factor:
                factors.append(factor)
    return factors


def _serialize_reference_row(row: dict[str, Any]) -> dict[str, Any] | None:
    scope = _clean(row.get("scope"))
    category = _clean(row.get("category")) or _clean(row.get("level_1"))
    report_label = _clean(row.get("report_label")) or _clean(row.get("column_text")) or _clean(row.get("original_id"))
    original_id = _clean(row.get("original_id"))
    if not original_id:
        return None
    if scope != "Scope 3":
        return None
    if category.lower() != "business travel":
        return None
    return {
        "scope": scope or "Scope 3",
        "category": "Business Travel",
        "report_label": report_label,
        "original_id": original_id,
        "uom": _clean(row.get("uom")),
        "ghg_unit": _clean(row.get("ghg_unit")),
        "factor": _clean(row.get("factor")),
        "notes": _clean(row.get("notes")),
    }


def _load_reference_rows(reporting_year: int | None = None) -> list[dict[str, Any]]:
    factors = _load_reference_rows_from_dataset(reporting_year=reporting_year)
    if factors:
        return factors

    try:
        return _load_reference_rows_from_csv()
    except Exception:
        return []


def _fetch_previous_year_rows(job_id: int, site_id: int | None) -> list[tuple[int, list[dict[str, Any]]]]:
    with get_conn() as con:
        current = con.execute(
            """
            SELECT job_id, client_db_id, reporting_year, reporting_period_end
            FROM jobs
            WHERE job_id=%s
            """,
            [int(job_id)],
        ).fetchone()
        if not current:
            return []

        client_db_id = int(current[1]) if current[1] is not None else None
        current_year = int(current[2]) if current[2] is not None else None
        current_period_end = current[3]
        if client_db_id is None:
            return []

        prior_filter_sql = ""
        params: list[object] = [client_db_id]
        if current_period_end is not None:
            prior_filter_sql += " AND j.reporting_period_end IS NOT NULL AND j.reporting_period_end < ? "
            params.append(current_period_end)
        elif current_year is not None:
            prior_filter_sql += " AND j.reporting_year < ? "
            params.append(current_year)

        prev_jobs = con.execute(
            f"""
            SELECT j.job_id, j.reporting_year, j.reporting_period_end
            FROM jobs j
            WHERE j.client_db_id = ?
            {prior_filter_sql}
            ORDER BY j.reporting_period_end DESC NULLS LAST, j.reporting_year DESC, j.job_id DESC
            """,
            params,
        ).fetchall()

        grouped: list[tuple[int, list[dict[str, Any]]]] = []
        for pj in prev_jobs or []:
            prev_job_id = int(pj[0])
            prev_year = int(pj[1]) if pj[1] is not None else 0
            if pj[2] is not None and hasattr(pj[2], "year"):
                prev_year = int(pj[2].year)

            if site_id is not None:
                rows = con.execute(
                    """
                    SELECT row_id, scope, category, COALESCE(NULLIF(report_label, ''), category) AS report_label,
                           original_id, uom, qty, notes, data_source, data_confidence,
                           month_1, month_2, month_3, month_4, month_5, month_6,
                           month_7, month_8, month_9, month_10, month_11, month_12
                    FROM job_scope_rows
                    WHERE job_id = ?
                      AND enabled = TRUE
                      AND site_id = ?
                    ORDER BY scope, category, report_label, original_id
                    """,
                    [prev_job_id, site_id],
                ).fetchall()
            else:
                rows = con.execute(
                    """
                    SELECT row_id, scope, category, COALESCE(NULLIF(report_label, ''), category) AS report_label,
                           original_id, uom, qty, notes, data_source, data_confidence,
                           month_1, month_2, month_3, month_4, month_5, month_6,
                           month_7, month_8, month_9, month_10, month_11, month_12
                    FROM job_scope_rows
                    WHERE job_id = ?
                      AND enabled = TRUE
                    ORDER BY scope, category, report_label, original_id
                    """,
                    [prev_job_id],
                ).fetchall()

            if site_id is not None and not rows:
                rows = con.execute(
                    """
                    SELECT row_id, scope, category, COALESCE(NULLIF(report_label, ''), category) AS report_label,
                           original_id, uom, qty, notes, data_source, data_confidence,
                           month_1, month_2, month_3, month_4, month_5, month_6,
                           month_7, month_8, month_9, month_10, month_11, month_12
                    FROM job_scope_rows
                    WHERE job_id = ?
                      AND enabled = TRUE
                    ORDER BY scope, category, report_label, original_id
                    """,
                    [prev_job_id],
                ).fetchall()

            mapped: list[dict[str, Any]] = []
            for r in rows or []:
                scope = _clean(r[1])
                category = _clean(r[2]) or "Business Travel"
                report_label = _clean(r[3]) or category
                original_id = _clean(r[4])
                if not original_id:
                    continue
                if scope != "Scope 3":
                    continue
                if category.lower() != "business travel":
                    continue

                mapped.append(
                    {
                        "row_id": int(r[0]),
                        "scope": scope or "Scope 3",
                        "category": "Business Travel",
                        "report_label": report_label,
                        "original_id": original_id,
                        "uom": _clean(r[5]),
                        "qty": r[6],
                        "notes": r[7],
                        "data_source": _clean(r[8]) or "Company Data",
                        "data_confidence": _clean(r[9]) or "M",
                        "months": [r[10], r[11], r[12], r[13], r[14], r[15], r[16], r[17], r[18], r[19], r[20], r[21]],
                    }
                )

            if mapped:
                grouped.append((prev_year, mapped))

        return grouped


def generate_business_travel_upload_template(
    job_id: int,
    client_name: str = "",
    site_name: str = "",
    job_number: str = "",
    reporting_year: str = "",
    site_id: int | None = None,
    include_prev_year: bool = True,
) -> tuple[bytes, str]:
    with get_conn() as con:
        context = _job_context(con, int(job_id))
        client_name = client_name or str(context.get("client_name") or "Client")
        job_number = job_number or str(context.get("job_number") or f"job-{job_id}")
        reporting_year_val = context.get("reporting_year")
        if reporting_year_val is None:
            reporting_year_val = datetime.now().year
        reporting_year = reporting_year or str(reporting_year_val)
        selected_site = _resolve_selected_site(con, context.get("client_db_id"), site_id)
        selected_site_name = site_name or (selected_site.get("site_name") if selected_site else "")
        existing_data = fetch_existing_scope_entries(job_id)

    factors = _load_reference_rows(int(reporting_year) if str(reporting_year).strip() else None)
    monthly_headers = get_monthly_headers(job_id)
    headers = ["Scope", "Category", "Report Label", "ID", "UOM"] + monthly_headers + ["Qty", "Data Source", "Notes"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Upload"
    ws["A1"] = "Business Travel Data Upload"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        "Add business travel rows here, then import the workbook back into the job. "
        "Use the previous-year tabs to compare, reuse, and extend the same business travel factors."
    )
    ws["A3"] = "Job Number"
    ws["B3"] = job_number
    ws["E3"] = "Client"
    ws["F3"] = client_name
    ws["A4"] = "Site Name"
    ws["B4"] = selected_site_name or "All sites"
    ws["E4"] = "Reporting Period"
    ws["F4"] = get_reporting_period_display(job_id)
    ws["A5"] = "Data Files"
    ws["B5"] = "Business Travel Conversion Factors"

    header_row = 7
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = header_row + 1
    for factor in factors:
        ws.cell(row=current_row, column=1, value=factor["scope"])
        ws.cell(row=current_row, column=2, value="Business Travel")
        ws.cell(row=current_row, column=3, value=factor["report_label"])
        ws.cell(row=current_row, column=4, value=factor["original_id"])
        ws.cell(row=current_row, column=5, value=factor["uom"])

        existing = existing_data.get(factor["original_id"], {})
        monthly_values = existing.get("monthly", [])
        for month_idx, month_col in enumerate(range(6, 18)):
            value = monthly_values[month_idx] if month_idx < len(monthly_values) else None
            ws.cell(row=current_row, column=month_col, value=value)

        month_start_col = get_column_letter(6)
        month_end_col = get_column_letter(17)
        ws.cell(row=current_row, column=18, value=f"=SUM({month_start_col}{current_row}:{month_end_col}{current_row})")
        ws.cell(row=current_row, column=19, value="Business Travel Data")
        ws.cell(row=current_row, column=20, value=existing.get("notes") or factor.get("notes") or "")
        current_row += 1

    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 14,
        "B": 24,
        "C": 52,
        "D": 20,
        "E": 12,
        "R": 12,
        "S": 18,
        "T": 30,
    }
    for col_idx in range(6, 18):
        widths[get_column_letter(col_idx)] = 10
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.column_dimensions["D"].hidden = False

    if include_prev_year:
        previous_year_sheets = _fetch_previous_year_rows(int(job_id), site_id)
        for year, rows in previous_year_sheets:
            if not rows:
                continue
            sheet_name = f"Previous {year}"
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws_prev = wb.create_sheet(title=sheet_name[:31])
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_prev.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            out_row = 2
            for row in rows:
                ws_prev.cell(row=out_row, column=1, value=row.get("scope"))
                ws_prev.cell(row=out_row, column=2, value=row.get("category") or "Business Travel")
                ws_prev.cell(row=out_row, column=3, value=row.get("report_label"))
                ws_prev.cell(row=out_row, column=4, value=row.get("original_id"))
                ws_prev.cell(row=out_row, column=5, value=row.get("uom"))
                months = row.get("months") or []
                for month_idx, month_col in enumerate(range(6, 18)):
                    ws_prev.cell(row=out_row, column=month_col, value=(months[month_idx] if month_idx < len(months) else None))
                ws_prev.cell(row=out_row, column=18, value=row.get("qty"))
                ws_prev.cell(row=out_row, column=19, value=row.get("data_source") or "Previous Year")
                ws_prev.cell(row=out_row, column=20, value=row.get("notes"))
                out_row += 1

            ws_prev.freeze_panes = "A2"
            for col_idx in range(6, 18):
                ws_prev.column_dimensions[get_column_letter(col_idx)].width = 10
            for col, width in widths.items():
                ws_prev.column_dimensions[col].width = width

    file_name = "_".join(
        part
        for part in [
            _safe_filename_part(job_number) or f"job_{job_id}",
            _safe_filename_part(client_name) or "client",
            _safe_filename_part(selected_site_name) if selected_site_name else "",
            _safe_filename_part("business_travel_data_upload"),
            _safe_filename_part(reporting_year) or "year",
        ]
        if part
    ) + ".xlsx"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue(), file_name
