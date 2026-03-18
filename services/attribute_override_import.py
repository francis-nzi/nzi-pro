from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from core.database import get_conn

CLIENT_FIELDS: dict[str, dict[str, str]] = {
    "industry": {"type": "text", "column": "industry"},
    "crm_owner": {"type": "text", "column": "crm_owner"},
    "benchmark_period_start": {"type": "date", "column": "benchmark_period_start"},
    "benchmark_period_end": {"type": "date", "column": "benchmark_period_end"},
    "benchmark_year": {"type": "int", "column": "benchmark_year"},
    "company_reg": {"type": "text", "column": "company_reg"},
    "sic_code": {"type": "text", "column": "sic_code"},
    "year_end_month": {"type": "text", "column": "year_end_month"},
    "currency": {"type": "text", "column": "currency"},
    "description_long": {"type": "text", "column": "description_long"},
    "net_zero_year": {"type": "int", "column": "net_zero_year"},
}

JOB_FIELDS: dict[str, dict[str, str]] = {
    "crm_name": {"type": "text", "column": "crm_name"},
    "reporting_period_start": {"type": "date", "column": "reporting_period_start"},
    "reporting_period_end": {"type": "date", "column": "reporting_period_end"},
    "baseline_year": {"type": "int", "column": "baseline_year"},
    "title": {"type": "text", "column": "title"},
    "status": {"type": "text", "column": "status"},
    "start_date": {"type": "date", "column": "start_date"},
    "due_date": {"type": "date", "column": "due_date"},
}

CLIENT_MATCH_KEYS = ("client_db_id", "wfm_client_id", "client_name")
JOB_MATCH_KEYS = ("job_id", "wfm_job_id", "job_number")


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            return str(int(value))
    if isinstance(value, int):
        return str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    s = str(value).strip()
    if s.lower() in {"", "nan", "none", "null"}:
        return ""
    if s.endswith(".0"):
        try:
            whole = float(s)
            if whole.is_integer():
                return str(int(whole))
        except Exception:
            pass
    return s


def _parse_bool(value: Any) -> bool:
    s = _clean(value).lower()
    return s in {"1", "true", "yes", "y", "on"}


def _parse_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.date().isoformat()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = _clean(value)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            continue
    return None


def _parse_int(value: Any) -> int | None:
    s = _clean(value)
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _normalize_sheet_name(name: str) -> str:
    return _clean(name).strip().lower()


def _column_exists(con, table_name: str, column_name: str) -> bool:
    try:
        row = con.execute(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE lower(table_name) = lower(?)
              AND lower(column_name) = lower(?)
            LIMIT 1
            """,
            [str(table_name), str(column_name)],
        ).fetchone()
        return bool(row)
    except Exception:
        return True


def _ensure_audit_table(con) -> None:
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS bulk_attribute_override_audit (
          run_id VARCHAR,
          entity_type VARCHAR NOT NULL,
          record_id INTEGER NOT NULL,
          record_label VARCHAR,
          field_name VARCHAR NOT NULL,
          old_value TEXT,
          new_value TEXT,
          actor VARCHAR,
          applied_at TIMESTAMP NOT NULL DEFAULT NOW()
        )
        """
    )


def _sample_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return _clean(value)


def _empty_for_type(field_type: str) -> Any:
    return None if field_type in {"date", "int"} else None


def _build_reference_data(con) -> dict[str, Any]:
    client_columns = {field: _column_exists(con, "clients", meta["column"]) for field, meta in CLIENT_FIELDS.items()}
    job_columns = {field: _column_exists(con, "jobs", meta["column"]) for field, meta in JOB_FIELDS.items()}

    client_select_parts = ["db_id", "client_name"]
    for field, meta in CLIENT_FIELDS.items():
        col = meta["column"]
        if client_columns[field]:
            client_select_parts.append(col)
        else:
            client_select_parts.append(f"NULL AS {col}")
    client_rows = con.execute(f"SELECT {', '.join(client_select_parts)} FROM clients").fetchall()
    client_col_names = ["db_id", "client_name", *[meta["column"] for meta in CLIENT_FIELDS.values()]]

    clients_by_id: dict[int, dict[str, Any]] = {}
    clients_by_name: dict[str, list[dict[str, Any]]] = {}
    for row in client_rows:
        record = {client_col_names[idx]: row[idx] for idx in range(len(client_col_names))}
        record["label"] = _clean(record.get("client_name")) or f"Client {record.get('db_id')}"
        client_id = int(record["db_id"])
        clients_by_id[client_id] = record
        clients_by_name.setdefault(_clean(record.get("client_name")).lower(), []).append(record)

    job_select_parts = [
        "j.job_id",
        "j.job_number",
        "j.title",
        "c.client_name",
    ]
    for field, meta in JOB_FIELDS.items():
        col = meta["column"]
        if job_columns[field]:
            job_select_parts.append(f"j.{col}")
        else:
            job_select_parts.append(f"NULL AS {col}")
    job_rows = con.execute(
        f"""
        SELECT {', '.join(job_select_parts)}
        FROM jobs j
        LEFT JOIN clients c ON c.db_id = j.client_db_id
        """
    ).fetchall()
    job_col_names = ["job_id", "job_number", "title", "client_name", *[meta["column"] for meta in JOB_FIELDS.values()]]

    jobs_by_id: dict[int, dict[str, Any]] = {}
    jobs_by_number: dict[str, list[dict[str, Any]]] = {}
    for row in job_rows:
        record = {job_col_names[idx]: row[idx] for idx in range(len(job_col_names))}
        label = _clean(record.get("job_number")) or _clean(record.get("title")) or f"Job {record.get('job_id')}"
        client_name = _clean(record.get("client_name"))
        record["label"] = f"{label}{f' ({client_name})' if client_name else ''}"
        job_id = int(record["job_id"])
        jobs_by_id[job_id] = record
        jobs_by_number.setdefault(_clean(record.get("job_number")).lower(), []).append(record)

    wfm_rows = con.execute(
        """
        SELECT entity_type, wfm_id, nzi_id
        FROM wfm_import_map
        WHERE entity_type IN ('client', 'job')
        """
    ).fetchall()
    client_by_wfm: dict[str, dict[str, Any]] = {}
    job_by_wfm: dict[str, dict[str, Any]] = {}
    for entity_type, wfm_id, nzi_id in wfm_rows:
        key = _clean(wfm_id)
        entity = _clean(entity_type).lower()
        if entity == "client" and int(nzi_id) in clients_by_id:
            client_by_wfm[key] = clients_by_id[int(nzi_id)]
        if entity == "job" and int(nzi_id) in jobs_by_id:
            job_by_wfm[key] = jobs_by_id[int(nzi_id)]

    industry_names: set[str] = set()
    try:
        industry_rows = con.execute(
            """
            SELECT name
            FROM industries_lookup
            WHERE is_active = TRUE
            """
        ).fetchall()
        industry_names = {_clean(row[0]).lower() for row in industry_rows if _clean(row[0])}
    except Exception:
        industry_names = set()

    has_crp_job_details = False
    try:
        table_row = con.execute(
            """
            SELECT 1
            FROM information_schema.tables
            WHERE lower(table_name) = 'crp_job_details'
            LIMIT 1
            """
        ).fetchone()
        has_crp_job_details = bool(table_row)
    except Exception:
        has_crp_job_details = False

    return {
        "client_columns": client_columns,
        "job_columns": job_columns,
        "clients_by_id": clients_by_id,
        "clients_by_name": clients_by_name,
        "client_by_wfm": client_by_wfm,
        "jobs_by_id": jobs_by_id,
        "jobs_by_number": jobs_by_number,
        "job_by_wfm": job_by_wfm,
        "industry_names": industry_names,
        "has_crp_job_details": has_crp_job_details,
    }


def _match_record(entity: str, row: dict[str, Any], refs: dict[str, Any]) -> tuple[dict[str, Any] | None, str, str, list[str]]:
    warnings: list[str] = []
    match_by = _clean(row.get("match_by")).lower()
    keys = CLIENT_MATCH_KEYS if entity == "client" else JOB_MATCH_KEYS
    if match_by and match_by not in keys:
        return None, match_by, "", [f"Unsupported match_by '{match_by}' for {entity}s"]

    chosen_key = match_by
    if not chosen_key:
        for key in keys:
            if _clean(row.get(key)):
                chosen_key = key
                break
    if not chosen_key:
        return None, "", "", [f"No match column supplied for {entity} row"]

    match_value = _clean(row.get(chosen_key))
    if not match_value:
        return None, chosen_key, "", [f"Match value is blank for {chosen_key}"]

    if entity == "client":
        if chosen_key == "client_db_id":
            record = refs["clients_by_id"].get(_parse_int(match_value) or -1)
            return record, chosen_key, match_value, ([] if record else [f"Client ID {match_value} not found"])
        if chosen_key == "wfm_client_id":
            record = refs["client_by_wfm"].get(match_value)
            return record, chosen_key, match_value, ([] if record else [f"WFM client ID {match_value} not found"])
        matches = refs["clients_by_name"].get(match_value.lower(), [])
        if len(matches) == 1:
            return matches[0], chosen_key, match_value, warnings
        if len(matches) > 1:
            return None, chosen_key, match_value, [f"Client name '{match_value}' matched multiple records"]
        return None, chosen_key, match_value, [f"Client name '{match_value}' not found"]

    if chosen_key == "job_id":
        record = refs["jobs_by_id"].get(_parse_int(match_value) or -1)
        return record, chosen_key, match_value, ([] if record else [f"Job ID {match_value} not found"])
    if chosen_key == "wfm_job_id":
        record = refs["job_by_wfm"].get(match_value)
        return record, chosen_key, match_value, ([] if record else [f"WFM job ID {match_value} not found"])
    matches = refs["jobs_by_number"].get(match_value.lower(), [])
    if len(matches) == 1:
        return matches[0], chosen_key, match_value, warnings
    if len(matches) > 1:
        return None, chosen_key, match_value, [f"Job number '{match_value}' matched multiple records"]
    return None, chosen_key, match_value, [f"Job number '{match_value}' not found"]


def _prepare_updates(
    entity: str,
    row: dict[str, Any],
    record: dict[str, Any],
    refs: dict[str, Any],
) -> tuple[dict[str, Any], list[dict[str, Any]], list[str]]:
    config = CLIENT_FIELDS if entity == "client" else JOB_FIELDS
    available_columns = refs["client_columns"] if entity == "client" else refs["job_columns"]
    changes: list[dict[str, Any]] = []
    warnings: list[str] = []
    update_fields: dict[str, Any] = {}

    for field, meta in config.items():
        if not available_columns.get(field, True):
            if _clean(row.get(field)) or _parse_bool(row.get(f"clear_{field}")):
                warnings.append(f"Field '{field}' is not available in this environment")
            continue

        raw_value = row.get(field)
        clear_requested = _parse_bool(row.get(f"clear_{field}"))
        if clear_requested and _clean(raw_value):
            warnings.append(f"Row cannot set and clear '{field}' at the same time")
            continue
        if not clear_requested and _clean(raw_value) == "":
            continue

        field_type = meta["type"]
        if clear_requested:
            parsed_value = _empty_for_type(field_type)
            action = "clear"
        elif field_type == "date":
            parsed_value = _parse_date(raw_value)
            if parsed_value is None:
                warnings.append(f"Invalid date for '{field}': {_clean(raw_value)}")
                continue
            action = "set"
        elif field_type == "int":
            parsed_value = _parse_int(raw_value)
            if parsed_value is None:
                warnings.append(f"Invalid integer for '{field}': {_clean(raw_value)}")
                continue
            action = "set"
        else:
            parsed_value = _clean(raw_value) or None
            action = "set"

        if entity == "client" and field == "industry" and parsed_value:
            if refs["industry_names"] and _clean(parsed_value).lower() not in refs["industry_names"]:
                warnings.append(f"Industry '{parsed_value}' is not in industries lookup")
        if field == "year_end_month" and parsed_value:
            month_text = _clean(parsed_value)
            if month_text not in {f"{idx:02d}" for idx in range(1, 13)}:
                warnings.append("year_end_month must be 01-12")
                continue

        current_value = record.get(meta["column"])
        current_text = _sample_text(current_value)
        new_text = _sample_text(parsed_value)
        if current_text == new_text:
            continue

        update_fields[field] = parsed_value
        changes.append(
            {
                "field": field,
                "action": action,
                "from_value": current_text or None,
                "to_value": new_text or None,
            }
        )

    if entity == "client":
        start_value = update_fields.get("benchmark_period_start")
        end_value = update_fields.get("benchmark_period_end")
        start_current = _parse_date(record.get("benchmark_period_start"))
        end_current = _parse_date(record.get("benchmark_period_end"))
        start_date = start_value if "benchmark_period_start" in update_fields else start_current
        end_date = end_value if "benchmark_period_end" in update_fields else end_current
        if start_date and end_date and str(end_date) < str(start_date):
            warnings.append("benchmark_period_end cannot be earlier than benchmark_period_start")

    if entity == "job":
        start_value = update_fields.get("reporting_period_start")
        end_value = update_fields.get("reporting_period_end")
        start_current = _parse_date(record.get("reporting_period_start"))
        end_current = _parse_date(record.get("reporting_period_end"))
        start_date = start_value if "reporting_period_start" in update_fields else start_current
        end_date = end_value if "reporting_period_end" in update_fields else end_current
        if start_date and end_date and str(end_date) < str(start_date):
            warnings.append("reporting_period_end cannot be earlier than reporting_period_start")

    return update_fields, changes, warnings


def _sheet_rows(df: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    normalized_columns = [_normalize_sheet_name(str(col)) for col in df.columns]
    for idx, (_, series) in enumerate(df.iterrows(), start=2):
        row = {normalized_columns[col_idx]: series.iloc[col_idx] for col_idx in range(len(normalized_columns))}
        if not any(_clean(value) for value in row.values()):
            continue
        row["__row_number__"] = idx
        rows.append(row)
    return rows


def parse_override_workbook(raw: bytes, filename: str = "") -> dict[str, Any]:
    workbook = pd.read_excel(io.BytesIO(raw), sheet_name=None, dtype=object)
    sheets = {_normalize_sheet_name(name): df for name, df in (workbook or {}).items()}

    warnings: list[str] = []
    ready_rows: list[dict[str, Any]] = []
    blocked_rows: list[dict[str, Any]] = []
    skipped_rows: list[dict[str, Any]] = []
    totals = {"clients": 0, "jobs": 0}

    with get_conn() as con:
        refs = _build_reference_data(con)

    for entity, sheet_name in (("client", "clients"), ("job", "jobs")):
        df = sheets.get(sheet_name)
        if df is None:
            continue
        rows = _sheet_rows(df)
        totals[f"{entity}s"] = len(rows)
        for row in rows:
            row_number = int(row.get("__row_number__") or 0)
            record, match_key, match_value, match_errors = _match_record(entity, row, refs)
            base = {
                "entity": entity,
                "row_number": row_number,
                "match_key": match_key,
                "match_value": match_value,
                "status": "blocked",
                "warnings": list(match_errors),
                "changes": [],
            }
            if not record:
                blocked_rows.append(base)
                continue

            update_fields, changes, change_warnings = _prepare_updates(entity, row, record, refs)
            row_warnings = [*match_errors, *change_warnings]
            base.update(
                {
                    "target_id": int(record["db_id"] if entity == "client" else record["job_id"]),
                    "matched_label": record.get("label"),
                    "changes": changes,
                    "warnings": row_warnings,
                    "update_fields": update_fields,
                }
            )

            if change_warnings and not changes:
                blocked_rows.append(base)
                continue
            if not changes:
                base["status"] = "skipped"
                skipped_rows.append(base)
                continue

            if any("cannot be earlier" in warn for warn in row_warnings):
                blocked_rows.append(base)
                continue

            base["status"] = "ready"
            ready_rows.append(base)

    if "clients" not in sheets and "jobs" not in sheets:
        warnings.append("Workbook must include a 'clients' and/or 'jobs' sheet")

    return {
        "ok": True,
        "filename": _clean(filename),
        "summary": {
            "total_rows": int(totals["clients"] + totals["jobs"]),
            "client_rows": int(totals["clients"]),
            "job_rows": int(totals["jobs"]),
            "ready_rows": int(len(ready_rows)),
            "blocked_rows": int(len(blocked_rows)),
            "skipped_rows": int(len(skipped_rows)),
        },
        "warnings": warnings,
        "rows_ready": ready_rows,
        "rows_blocked": blocked_rows,
        "rows_skipped": skipped_rows,
    }


def commit_override_rows(rows: list[dict[str, Any]], actor: str = "") -> dict[str, Any]:
    if not isinstance(rows, list) or not rows:
        raise ValueError("rows_ready must be a non-empty list")

    run_id = f"bulk-override-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    applied_rows = 0
    applied_changes = 0
    applied_by_entity = {"client": 0, "job": 0}

    with get_conn() as con:
        refs = _build_reference_data(con)
        _ensure_audit_table(con)

        for row in rows:
            if not isinstance(row, dict):
                continue
            entity = _clean(row.get("entity")).lower()
            if entity not in {"client", "job"}:
                continue
            target_id = _parse_int(row.get("target_id"))
            update_fields = row.get("update_fields") if isinstance(row.get("update_fields"), dict) else {}
            if not target_id or not update_fields:
                continue

            config = CLIENT_FIELDS if entity == "client" else JOB_FIELDS
            record = (
                refs["clients_by_id"].get(int(target_id))
                if entity == "client"
                else refs["jobs_by_id"].get(int(target_id))
            )
            if not record:
                continue

            assignments: list[str] = []
            params: list[Any] = []
            field_count = 0

            for field, value in update_fields.items():
                meta = config.get(field)
                if not meta:
                    continue
                assignments.append(f"{meta['column']} = ?")
                params.append(value)
                old_value = record.get(meta["column"])
                con.execute(
                    """
                    INSERT INTO bulk_attribute_override_audit (
                      run_id, entity_type, record_id, record_label, field_name, old_value, new_value, actor
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    [
                        run_id,
                        entity,
                        int(target_id),
                        _clean(record.get("label")),
                        field,
                        _sample_text(old_value) or None,
                        _sample_text(value) or None,
                        _clean(actor) or None,
                    ],
                )
                field_count += 1

            if not assignments:
                continue

            if entity == "job":
                reporting_period_end = update_fields.get("reporting_period_end")
                if reporting_period_end:
                    try:
                        assignments.append("reporting_year = ?")
                        params.append(int(str(reporting_period_end)[:4]))
                    except Exception:
                        pass

            params.append(int(target_id))
            table_name = "clients" if entity == "client" else "jobs"
            pk_name = "db_id" if entity == "client" else "job_id"
            con.execute(f"UPDATE {table_name} SET {', '.join(assignments)} WHERE {pk_name} = ?", params)

            if entity == "job" and refs.get("has_crp_job_details"):
                crp_assignments: list[str] = []
                crp_params: list[Any] = []
                if "reporting_period_start" in update_fields:
                    crp_assignments.append("reporting_period_from = ?")
                    crp_params.append(update_fields.get("reporting_period_start"))
                if "reporting_period_end" in update_fields:
                    crp_assignments.append("reporting_period_to = ?")
                    crp_params.append(update_fields.get("reporting_period_end"))
                if "reporting_period_end" in update_fields and update_fields.get("reporting_period_end"):
                    try:
                        crp_assignments.append("reporting_year = ?")
                        crp_params.append(int(str(update_fields.get("reporting_period_end"))[:4]))
                    except Exception:
                        pass
                if crp_assignments:
                    crp_assignments.append("updated_at = NOW()")
                    crp_params.append(int(target_id))
                    con.execute(
                        f"UPDATE crp_job_details SET {', '.join(crp_assignments)} WHERE job_id = ?",
                        crp_params,
                    )

            applied_rows += 1
            applied_changes += field_count
            applied_by_entity[entity] += 1

    return {
        "ok": True,
        "run_id": run_id,
        "applied_rows": int(applied_rows),
        "applied_changes": int(applied_changes),
        "applied_by_entity": applied_by_entity,
    }


def build_override_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    instructions = [
        ["Bulk Attribute Override Import"],
        [""],
        ["Use optional sheets named 'clients' and/or 'jobs'."],
        ["Blank cells mean 'leave unchanged'."],
        ["To clear a field, leave the value blank and set clear_<field> to TRUE."],
        ["Match priority is taken from match_by, or the first populated ID column if match_by is blank."],
        ["Recommended date format: YYYY-MM-DD."],
    ]
    for row in instructions:
        ws.append(row)

    client_headers = [
        "match_by",
        "client_db_id",
        "wfm_client_id",
        "client_name",
        *CLIENT_FIELDS.keys(),
        *[f"clear_{field}" for field in CLIENT_FIELDS.keys()],
    ]
    client_ws = wb.create_sheet("clients")
    client_ws.append(client_headers)

    job_headers = [
        "match_by",
        "job_id",
        "wfm_job_id",
        "job_number",
        *JOB_FIELDS.keys(),
        *[f"clear_{field}" for field in JOB_FIELDS.keys()],
    ]
    job_ws = wb.create_sheet("jobs")
    job_ws.append(job_headers)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
