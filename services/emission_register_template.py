from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


SCOPE_OPTIONS = ["Scope 1", "Scope 2", "Scope 3"]
CATEGORY_OPTIONS = ["Fleet", "Travel", "Equipment"]
ROLLUP_METHOD_OPTIONS = ["sum", "weighted_sum", "headcount_scaled", "average"]


def _safe_filename_part(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    for char in '<>:"/\\|?*':
        text = text.replace(char, "_")
    text = text.replace(" ", "_")
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


def _register_token(source_type: str) -> str:
    return "business_travel_register" if str(source_type or "").strip() == "business_travel" else "asset_register"


def _job_context(con, job_id: int) -> dict[str, Any]:
    row = con.execute(
        """
        SELECT j.job_id, j.job_number, j.reporting_year, j.reporting_period_end,
               c.client_name, c.db_id
        FROM jobs j
        JOIN clients c ON c.db_id = j.client_db_id
        WHERE j.job_id=%s
        """,
        [int(job_id)],
    ).fetchone()
    if not row:
        return {}

    reporting_year = row[2]
    if reporting_year is None and row[3] is not None:
        try:
            reporting_year = int(getattr(row[3], "year", None) or str(row[3])[:4])
        except Exception:
            reporting_year = None

    return {
        "job_id": int(row[0]),
        "job_number": row[1],
        "reporting_year": reporting_year,
        "reporting_period_end": row[3],
        "client_name": row[4],
        "client_db_id": int(row[5]) if row[5] is not None else None,
    }


def _job_sites(con, client_db_id: int | None) -> list[str]:
    if client_db_id is None:
        return []
    rows = con.execute(
        """
        SELECT site_name
        FROM client_sites
        WHERE client_db_id=%s
          AND COALESCE(TRIM(site_name), '') <> ''
        ORDER BY site_name
        """,
        [int(client_db_id)],
    ).fetchall()
    return [str(row[0]).strip() for row in rows if str(row[0] or "").strip()]


def _add_column_validation(ws, cell_range: str, sheet_name: str, col_letter: str, start_row: int, end_row: int) -> None:
    if end_row < start_row:
        return
    dv = DataValidation(
        type="list",
        formula1=f"='{sheet_name}'!${col_letter}${start_row}:${col_letter}${end_row}",
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(cell_range)


def _style_header_row(ws, row_number: int, headers: list[str]) -> None:
    fill = PatternFill(fill_type="solid", fgColor="4472C4")
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_number, column=idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _set_column_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _populate_template_examples(source_type: str, site_name: str) -> tuple[list[list[Any]], list[list[Any]]]:
    site_label = site_name or "No site"
    if str(source_type or "").strip() == "business_travel":
        groups = [
            ["Business Travel - Grey Fleet - Example", "Scope 3", "Travel", site_label, "sum", "4_301_3070_9_1", "4_301_3070_9_1", 0.1987, "kgCO2e", "mile", "One group can hold many employee travel records."],
            ["Business Travel - Rail - Example", "Scope 3", "Travel", site_label, "sum", "4_301_3071_4_1", "4_301_3071_4_1", 0.0632, "kgCO2e", "mile", "Keep the group label stable from year to year."],
        ]
        sources = [
            ["Business Travel - Grey Fleet - Example", "Employee 001 - Grey Fleet", "", "Employee 001", "Grey fleet mileage", 1, 100, "Example Workbook", "H", "One of several employees in the same travel group."],
            ["Business Travel - Grey Fleet - Example", "Employee 002 - Grey Fleet", "", "Employee 002", "Grey fleet mileage", 1, 100, "Example Workbook", "H", "Same factor, separate source identity."],
            ["Business Travel - Rail - Example", "Employee 003 - Rail", "", "Employee 003", "Rail travel", 250, 100, "Example Workbook", "H", "A second group with a different travel mode."],
        ]
    else:
        groups = [
            ["Fleet - Medium Diesel - Example", "Scope 1", "Fleet", site_label, "sum", "4_301_3070_9_1", "4_301_3070_9_1", 0.1623, "kgCO2e", "vehicle", "Use one group for a natural roll-up bucket such as a vehicle fleet."],
            ["Equipment - Refrigeration - Example", "Scope 1", "Equipment", site_label, "sum", "4_301_3061_4_1", "4_301_3061_4_1", 0.2084, "kgCO2e", "unit", "Use the group to roll multiple similar assets into one report line."],
        ]
        sources = [
            ["Fleet - Medium Diesel - Example", "Vehicle 1", "REG-001", "", "Medium diesel vehicle", 1, 100, "Example Workbook", "H", "Two vehicles can share the same factor while keeping unique identities."],
            ["Fleet - Medium Diesel - Example", "Vehicle 2", "REG-002", "", "Medium diesel vehicle", 1, 100, "Example Workbook", "H", "This is the row-level asset record."],
            ["Equipment - Refrigeration - Example", "Refrigeration Unit", "REF-01", "", "Refrigeration", 1, 100, "Example Workbook", "H", "A different asset family in a separate group."],
        ]
    return groups, sources


def build_emission_register_workbook(
    con,
    *,
    job_id: int,
    source_type: str,
    kind: str = "template",
) -> tuple[bytes, str]:
    kind_value = "example" if str(kind or "").strip().lower() == "example" else "template"
    source_type_value = "business_travel" if str(source_type or "").strip() == "business_travel" else "asset"
    context = _job_context(con, int(job_id))
    client_name = str(context.get("client_name") or "Client")
    job_number = str(context.get("job_number") or f"job-{job_id}")
    reporting_year = context.get("reporting_year")
    if reporting_year is None:
        reporting_year = datetime.now().year
    register_name = _register_token(source_type_value)
    filename_parts = [
        _safe_filename_part(job_number),
        _safe_filename_part(client_name),
        _safe_filename_part(register_name),
        _safe_filename_part(reporting_year),
    ]
    base_name = "_".join(part for part in filename_parts if part)
    file_name = f"{base_name}{'_example' if kind_value == 'example' else ''}.xlsx"

    sites = _job_sites(con, context.get("client_db_id"))
    site_choices = ["No site", *sites] if sites else ["No site"]
    example_site = sites[0] if sites else "No site"

    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    groups_ws = wb.create_sheet("Groups")
    sources_ws = wb.create_sheet("Sources")
    lists_ws = wb.create_sheet("Lists")

    title_fill = PatternFill(fill_type="solid", fgColor="E8F0FE")

    instructions["A1"] = "Asset Register Workbook"
    instructions["A1"].font = Font(bold=True, size=14)
    instructions["A2"] = (
        "Add groups and asset rows here, then import the workbook back into the job. "
        "Groups own the scope, site, and factor family; asset rows inherit that configuration."
    )
    instructions["A4"] = "Job Number"
    instructions["B4"] = job_number
    instructions["A5"] = "Client"
    instructions["B5"] = client_name
    instructions["A6"] = "Data Register"
    instructions["B6"] = "Asset Register" if source_type_value == "asset" else "Business Travel Register"
    instructions["A7"] = "Reporting Year"
    instructions["B7"] = reporting_year
    instructions["A8"] = "Filename convention"
    instructions["B8"] = f"[job_no]_[client]_[data_register]_[reporting_year]{'_example' if kind_value == 'example' else ''}.xlsx"
    instructions["A10"] = "Naming guidance"
    instructions["A10"].font = Font(bold=True)
    for ref in ("A4", "A5", "A6", "A7", "A8"):
        instructions[ref].font = Font(bold=True)
        instructions[ref].fill = title_fill
    for row_idx, text in enumerate(
        [
            "Group name is the roll-up label and the place where Scope, Site, and factor metadata live.",
            "A group can contain many asset rows. Use it to roll similar assets together for reporting.",
            "Asset Name and Asset Identity describe the real-world item; Group rows carry Factor DB ID / Original ID / Factor / UOM.",
            "group_type is an internal family label. It usually stays as asset for the Asset Register or business_travel for Business Travel.",
            "Keep group names stable year to year so rollforward imports stay clean.",
        ],
        start=11,
    ):
        instructions[f"A{row_idx}"] = text
        instructions[f"A{row_idx}"].alignment = Alignment(wrap_text=True, vertical="top")
    instructions["A17"] = "Suggested naming pattern"
    instructions["A17"].font = Font(bold=True)
    instructions["B17"] = "[Category] - [Asset or Mode] - [Site or Team]"
    instructions["A18"] = "Example group names"
    instructions["A18"].font = Font(bold=True)
    instructions["B18"] = "Fleet - Medium Diesel - London Office"
    instructions["B19"] = "Equipment - Refrigeration - Head Office"
    instructions["B20"] = "Business Travel - Grey Fleet - Field Team"

    groups_headers = [
        "Group Name",
        "Scope",
        "Category",
        "Site Name",
        "Rollup Method",
        "Factor DB ID",
        "Original ID",
        "Factor",
        "GHG Unit",
        "UOM",
        "Notes",
    ]
    sources_headers = [
        "Group Name",
        "Asset Name",
        "Asset Identity",
        "Employee Name",
        "Asset Subtype",
        "Qty",
        "Apply %",
        "Data Source",
        "Data Confidence",
        "Notes",
    ]

    _style_header_row(groups_ws, 1, groups_headers)
    _style_header_row(sources_ws, 1, sources_headers)

    _set_column_widths(
        instructions,
        {
            "A": 28,
            "B": 80,
        },
    )
    _set_column_widths(
        groups_ws,
        {
            "A": 36,
            "B": 14,
            "C": 16,
            "D": 24,
            "E": 18,
            "F": 14,
            "G": 14,
            "H": 14,
            "I": 14,
            "J": 12,
            "K": 44,
        },
    )
    _set_column_widths(
        sources_ws,
        {
            "A": 36,
            "B": 28,
            "C": 20,
            "D": 20,
            "E": 24,
            "F": 12,
            "G": 12,
            "H": 18,
            "I": 14,
            "J": 36,
        },
    )

    # Validation lists
    lists_ws["A1"] = "Scope"
    for idx, value in enumerate(SCOPE_OPTIONS, start=2):
        lists_ws.cell(row=idx, column=1, value=value)
    lists_ws["B1"] = "Category"
    for idx, value in enumerate(CATEGORY_OPTIONS, start=2):
        lists_ws.cell(row=idx, column=2, value=value)
    lists_ws["C1"] = "Rollup Method"
    for idx, value in enumerate(ROLLUP_METHOD_OPTIONS, start=2):
        lists_ws.cell(row=idx, column=3, value=value)
    lists_ws["D1"] = "Site Name"
    for idx, value in enumerate(site_choices, start=2):
        lists_ws.cell(row=idx, column=4, value=value)
    lists_ws.sheet_state = "hidden"

    template_end_row = 250
    example_end_row = 20
    end_row = example_end_row if kind_value == "example" else template_end_row

    for row_idx in range(2, end_row + 1):
        groups_ws.cell(row=row_idx, column=1, value="" if kind_value == "template" else None)
        sources_ws.cell(row=row_idx, column=1, value="" if kind_value == "template" else None)

    _add_column_validation(groups_ws, f"B2:B{end_row}", "Lists", "A", 2, 1 + len(SCOPE_OPTIONS))
    _add_column_validation(groups_ws, f"C2:C{end_row}", "Lists", "B", 2, 1 + len(CATEGORY_OPTIONS))
    _add_column_validation(groups_ws, f"D2:D{end_row}", "Lists", "D", 2, 1 + len(site_choices))
    _add_column_validation(groups_ws, f"E2:E{end_row}", "Lists", "C", 2, 1 + len(ROLLUP_METHOD_OPTIONS))

    instructions.sheet_view.showGridLines = False
    groups_ws.freeze_panes = "A2"
    sources_ws.freeze_panes = "A2"

    if kind_value == "example":
        group_rows, source_rows = _populate_template_examples(source_type_value, example_site)
        for row_idx, row_values in enumerate(group_rows, start=2):
            for col_idx, value in enumerate(row_values, start=1):
                groups_ws.cell(row=row_idx, column=col_idx, value=value)
        for row_idx, row_values in enumerate(source_rows, start=2):
            for col_idx, value in enumerate(row_values, start=1):
                sources_ws.cell(row=row_idx, column=col_idx, value=value)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue(), file_name
