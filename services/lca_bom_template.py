"""Downloadable BOM upload template for LCA assessments.

Column names here match exactly what api/lca_routes.py:upload_bom_file's
column-alias resolution accepts (item_name/material/component/..., quantity/
qty/weight/..., unit/uom/..., etc.) -- see that function's `getv()` alias
lists if either side of this needs to change.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.database import get_conn
from services.download_filenames import safe_filename_part
from services.lca_material_categories import ensure_material_categories_deduped

BOM_HEADERS = [
    "item_name",
    "component_code",
    "module",
    "quantity",
    "unit",
    "origin_country",
    "material_category",
    "factor",
    "factor_unit",
]

EXAMPLE_ROWS = [
    ["Steel bracket", "BRK-001", "A1", 0.45, "kg", "China", "Metal", "", ""],
    ["Cardboard outer box", "PKG-014", "A1", 0.12, "kg", "UK", "Cardboard", "", ""],
    ["Assembly (grouping label only)", "", "A3", 0, "kg", "", "", "", ""],
]


def _job_and_assessment_context(con, job_id: int, assessment_id: int) -> dict[str, Any] | None:
    row = con.execute(
        """
        SELECT j.job_number, c.client_name, a.name, a.lifecycle_boundary
        FROM lca_assessments a
        JOIN jobs j ON j.job_id = a.job_id
        JOIN clients c ON c.db_id = j.client_db_id
        WHERE a.assessment_id = %s AND a.job_id = %s
        """,
        [int(assessment_id), int(job_id)],
    ).fetchone()
    if not row:
        return None
    return {
        "job_number": row[0],
        "client_name": row[1],
        "assessment_name": row[2],
        "lifecycle_boundary": row[3],
    }


def _module_reference_rows(con) -> list[tuple[str, str, str]]:
    df = con.execute(
        "SELECT module_code, label, module_group FROM lca_modules_lookup WHERE is_active = TRUE ORDER BY sort_order"
    ).df()
    if df is None or df.empty:
        return []
    return [(str(r["module_code"]), str(r["label"]), str(r["module_group"])) for _, r in df.iterrows()]


def _material_category_reference_rows(con) -> list[str]:
    ensure_material_categories_deduped(con)
    df = con.execute(
        "SELECT name FROM lca_material_categories_lookup WHERE is_active = TRUE ORDER BY name"
    ).df()
    if df is None or df.empty:
        return []
    return [str(v) for v in df["name"].tolist()]


def generate_lca_bom_template(job_id: int, assessment_id: int) -> tuple[bytes, str] | None:
    """Returns (xlsx_bytes, filename), or None if the job/assessment doesn't exist."""
    with get_conn() as con:
        context = _job_and_assessment_context(con, int(job_id), int(assessment_id))
        if context is None:
            return None
        modules = _module_reference_rows(con)
        material_categories = _material_category_reference_rows(con)

    job_number = context["job_number"] or f"job-{job_id}"
    client_name = context["client_name"] or "Client"
    assessment_name = context["assessment_name"] or "Assessment"

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws["A1"] = "LCA Bill of Materials Upload"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Fill in one row per material/component line, then import this workbook back into the assessment."
    ws["A3"] = "Job Number"
    ws["B3"] = job_number
    ws["D3"] = "Client"
    ws["E3"] = client_name
    ws["A4"] = "Assessment"
    ws["B4"] = assessment_name
    ws["D4"] = "Boundary"
    ws["E4"] = context["lifecycle_boundary"] or ""

    header_row = 6
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, header in enumerate(BOM_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = header_row + 1
    for example in EXAMPLE_ROWS:
        for col_idx, value in enumerate(example, start=1):
            ws.cell(row=current_row, column=col_idx, value=value if value != "" else None)
        current_row += 1

    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False
    widths = {"A": 32, "B": 16, "C": 10, "D": 12, "E": 10, "F": 16, "G": 16, "H": 10, "I": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    notes = wb.create_sheet(title="Column Notes")
    notes["A1"] = "Column"
    notes["B1"] = "Notes"
    for cell in (notes["A1"], notes["B1"]):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    notes_rows = [
        ("item_name", "Required. Also accepts: material, material_name, component, component_name, name, description."),
        ("component_code", "Optional. Links to an existing library component with this code, or creates one. Also accepts: part_code, part_number."),
        ("module", "Lifecycle module code -- see the Module Codes sheet. Defaults to A1 if blank or invalid. Also accepts: module_code, stage, stage_key."),
        ("quantity", "Also accepts: qty, weight, unit_weight, mass. A zero (or blank) quantity is kept as a placeholder/assembly-grouping row and excluded from the calculation -- use this for a sub-assembly label that has no mass of its own."),
        ("unit", "Defaults to kg. Also accepts: uom, units."),
        ("origin_country", "Optional, used for factor matching. Also accepts: country."),
        ("material_category", "Optional. Matched against the Material Categories sheet (case-insensitive) -- if it doesn't match an existing one, a new category is created automatically. Also accepts: category, category_name."),
        ("factor", "Optional. Leave blank to auto-match a factor; set this to skip auto-matching and use an explicit value. Also accepts: factor_value, emission_factor."),
        ("factor_unit", "Only used if factor is set. Defaults to kgCO2e/kg. Also accepts: emission_factor_unit."),
    ]
    for i, (col, note) in enumerate(notes_rows, start=2):
        notes.cell(row=i, column=1, value=col)
        notes.cell(row=i, column=2, value=note)
    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 100
    for row in notes.iter_rows(min_row=2, max_row=len(notes_rows) + 1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    if modules:
        mod_sheet = wb.create_sheet(title="Module Codes")
        mod_sheet["A1"] = "Code"
        mod_sheet["B1"] = "Label"
        mod_sheet["C1"] = "Group"
        for cell in (mod_sheet["A1"], mod_sheet["B1"], mod_sheet["C1"]):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
        for i, (code, label, group) in enumerate(modules, start=2):
            mod_sheet.cell(row=i, column=1, value=code)
            mod_sheet.cell(row=i, column=2, value=label)
            mod_sheet.cell(row=i, column=3, value=group)
        mod_sheet.column_dimensions["A"].width = 10
        mod_sheet.column_dimensions["B"].width = 34
        mod_sheet.column_dimensions["C"].width = 16

    if material_categories:
        cat_sheet = wb.create_sheet(title="Material Categories")
        cat_sheet["A1"] = "Existing categories (type a new one if yours isn't listed)"
        cat_sheet["A1"].font = Font(bold=True, color="FFFFFF")
        cat_sheet["A1"].fill = fill
        for i, name in enumerate(material_categories, start=2):
            cat_sheet.cell(row=i, column=1, value=name)
        cat_sheet.column_dimensions["A"].width = 34

    file_name = (
        f"{safe_filename_part(job_number)} {safe_filename_part(client_name)} "
        f"LCA BOM Template - {safe_filename_part(assessment_name)}.xlsx"
    )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue(), file_name
