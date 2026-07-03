from __future__ import annotations

import logging
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from functools import lru_cache
import hashlib
import io
from pathlib import Path
import json
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from core.database import db_backend, get_conn
from services.sites import add_site

logger = logging.getLogger(__name__)

ID_RE = re.compile(r"^\d+_\d+_\d+_\d+_\d+$")
SPEND_ID_RE = re.compile(r"^(?:[A-Z0-9]+-)?SPEND-[A-Z0-9.\-]+$", re.IGNORECASE)
NUMERIC_TEXT_RE = re.compile(r"^-?\d+(?:\.\d+)?$")
IGNORED_LEGACY_SECTIONS = {"company information", "company data"}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() == "nan":
        return ""
    return s


def _safe_float(value: Any) -> float | None:
    s = _clean(value)
    if not s:
        return None
    try:
        return float(s)
    except Exception as exc:
        logger.debug("Unable to parse float value %r: %s", s, exc)
        return None


def _norm_scope(value: Any) -> str | None:
    s = _clean(value).lower()
    if "scope 1" in s or s == "1":
        return "Scope 1"
    if "scope 2" in s or s == "2":
        return "Scope 2"
    if "scope 3" in s or s == "3":
        return "Scope 3"
    return None


def _is_ignored_legacy_section(section: Any) -> bool:
    return _clean(section).lower() in IGNORED_LEGACY_SECTIONS


def _is_factor_original_id(value: Any) -> bool:
    s = _clean(value)
    return bool(ID_RE.match(s) or SPEND_ID_RE.match(s))


def _legacy_storage_original_id(factor_original_id: str, lookup_key: str) -> str:
    oid = _clean(factor_original_id)
    lk = _clean(lookup_key).lower()
    if not oid:
        return ""
    digest = hashlib.sha1(f"{oid}|{lk}".encode("utf-8")).hexdigest()[:12]
    return f"LEGACY-{oid}-{digest}"


def _build_line_label(activity: Any, c2: Any, c3: Any, c4: Any, c5: Any, c6: Any, c7: Any, *, scope_col: int | None) -> str:
    activity_text = _clean(activity) or "Unspecified"
    values = [c2, c3, c4, c5, c6, c7]
    parts: list[str] = []
    seen: set[str] = {activity_text.lower()}
    for idx, raw in enumerate(values, start=3):
        if scope_col is not None and idx == scope_col:
            continue
        text = _clean(raw)
        normalized = text.lower()
        if not text or normalized in seen or NUMERIC_TEXT_RE.match(text):
            continue
        seen.add(normalized)
        parts.append(text)
    return activity_text if not parts else f"{activity_text} | {' | '.join(parts)}"


def _find_header_col(header_values: dict[int, str], candidates: set[str]) -> int | None:
    for col, value in header_values.items():
        if _clean(value).lower() in candidates:
            return int(col)
    return None


def _section_layout(ws, header_row: int, max_col: int) -> dict[str, Any]:
    header_values = {col: _clean(ws.cell(header_row, col).value) for col in range(1, max_col + 1)}
    month_cols: list[tuple[int, date, str]] = []
    for col, raw in header_values.items():
        md = _parse_month_header(raw)
        if md is not None:
            month_cols.append((int(col), md, _clean(raw)))
    return {
        "scope_col": _find_header_col(header_values, {"scope", "emissions scope", "ghg scope"}),
        "reporting_col": _find_header_col(header_values, {"is reporting"}),
        "month_cols": month_cols,
    }


def _row_text(ws, row_num: int, col_num: int | None) -> str:
    if col_num is None or int(col_num) < 1:
        return ""
    return _clean(ws.cell(int(row_num), int(col_num)).value)


def _factor_original_id_from_row(row: dict[str, Any]) -> str:
    return _clean(row.get("factor_original_id") or row.get("original_id"))


def _storage_original_id_from_row(row: dict[str, Any]) -> str:
    storage_original_id = _clean(row.get("storage_original_id"))
    if storage_original_id:
        return storage_original_id
    factor_original_id = _factor_original_id_from_row(row)
    lookup_key = _clean(row.get("lookup_key"))
    if factor_original_id and lookup_key:
        return _legacy_storage_original_id(factor_original_id, lookup_key)
    return _clean(row.get("original_id"))


def _build_staged_metadata(
    row: dict[str, Any],
    factor_rec: FactorRec | None,
    primary_dataset: int | None,
    *,
    factor_original_id: str,
    storage_original_id: str,
) -> dict[str, Any]:
    section = _clean(row.get("section")) or None
    activity = _clean(row.get("activity")) or None
    line_label = (
        _clean(row.get("line_label"))
        or activity
        or (factor_rec.report_label if factor_rec and factor_rec.report_label else None)
        or "Unspecified"
    )
    return {
        "scope": _clean(row.get("scope")),
        "original_id": storage_original_id,
        "storage_original_id": storage_original_id,
        "factor_original_id": factor_original_id,
        "lookup_key": _clean(row.get("lookup_key")) or None,
        "dataset_id": primary_dataset,
        "factor_db_id": (factor_rec.db_id if factor_rec else None),
        "level_1": (factor_rec.level_1 if factor_rec else None),
        "level_2": (factor_rec.level_2 if factor_rec else None),
        "level_3": (factor_rec.level_3 if factor_rec else None),
        "level_4": (factor_rec.level_4 if factor_rec else None),
        "column_text": line_label,
        "report_label": line_label,
        "category": section or activity or (factor_rec.level_2 if factor_rec else None),
        "section": section,
        "activity": activity,
        "match_source": _clean(row.get("match_source")) or None,
        "line_label": line_label,
    }


def _single_numeric(values: list[Any]) -> float | None:
    uniq: list[float] = []
    seen: set[str] = set()
    for value in values:
        if value is None:
            continue
        num = float(value)
        key = f"{num:.12g}"
        if key in seen:
            continue
        seen.add(key)
        uniq.append(num)
    if len(uniq) == 1:
        return float(uniq[0])
    return None


def _single_text(values: list[Any]) -> str | None:
    uniq: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = _clean(value)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        uniq.append(text)
    if len(uniq) == 1:
        return uniq[0]
    return None


def _single_int(values: list[Any]) -> int | None:
    uniq: list[int] = []
    seen: set[int] = set()
    for value in values:
        if value is None:
            continue
        num = int(value)
        if num in seen:
            continue
        seen.add(num)
        uniq.append(num)
    if len(uniq) == 1:
        return int(uniq[0])
    return None


def _quantity_storage_plan(
    *,
    month_quantities: dict[int, float],
    month_emissions: dict[int, float],
    month_datasets: dict[int, int],
    month_factor_db_ids: dict[int, int | None],
    month_factor_values: dict[int, float | None],
    month_uoms: dict[int, str | None],
    month_ghg_units: dict[int, str | None],
    factor_rec_primary: FactorRec | None,
    primary_dataset: int | None,
) -> dict[str, Any]:
    dataset_id = _single_int(list(month_datasets.values()))
    factor_db_id = _single_int([v for v in month_factor_db_ids.values() if v is not None])
    factor_value = _single_numeric([v for v in month_factor_values.values() if v is not None])
    uom = _single_text([v for v in month_uoms.values() if v is not None]) or (factor_rec_primary.uom if factor_rec_primary else None)
    ghg_unit = _single_text([v for v in month_ghg_units.values() if v is not None]) or (factor_rec_primary.ghg_unit if factor_rec_primary else None)
    source_qty = float(sum(month_quantities.values()))
    source_uom = uom

    raw_supported = (
        bool(month_quantities)
        and dataset_id is not None
        and factor_value is not None
        and bool(uom)
        and bool(ghg_unit)
        and len({int(v) for v in month_datasets.values()}) == 1
        and len({f"{float(v):.12g}" for v in month_factor_values.values() if v is not None}) == 1
        and len({_clean(v).lower() for v in month_uoms.values() if _clean(v)}) <= 1
        and len({_clean(v).lower() for v in month_ghg_units.values() if _clean(v)}) <= 1
    )

    if raw_supported:
        return {
            "value_mode": "quantity",
            "storage_reason": "single_dataset_factor",
            "dataset_id": dataset_id,
            "factor_db_id": factor_db_id if factor_db_id is not None else (factor_rec_primary.db_id if factor_rec_primary else None),
            "uom": uom,
            "ghg_unit": ghg_unit,
            "factor": float(factor_value),
            "qty": float(sum(month_quantities.values())),
            "calc_tco2e": float(sum(month_emissions.values())),
            "source_qty": source_qty,
            "source_uom": source_uom,
            "monthly_values": {i: float(month_quantities.get(i, 0.0)) for i in range(1, 13)},
        }

    reason = "multi_dataset_or_factor"
    if len({int(v) for v in month_datasets.values()}) <= 1 and len({f'{float(v):.12g}' for v in month_factor_values.values() if v is not None}) <= 1:
        reason = "unsupported_raw_metadata"
    return {
        "value_mode": "emissions",
        "storage_reason": reason,
        "dataset_id": primary_dataset,
        "factor_db_id": (factor_rec_primary.db_id if factor_rec_primary else None),
        "uom": "tCO2e",
        "ghg_unit": "tCO2e",
        "factor": 1.0,
        "qty": float(sum(month_emissions.values())),
        "calc_tco2e": float(sum(month_emissions.values())),
        "source_qty": source_qty,
        "source_uom": source_uom,
        "monthly_values": {i: float(month_emissions.get(i, 0.0)) for i in range(1, 13)},
    }


def _convert_entry_to_emissions(entry: dict[str, Any], reason: str) -> None:
    factor = float(entry.get("factor") or 0.0)
    ghg_unit = entry.get("ghg_unit")
    monthly_values = entry.get("monthly_values") or {}
    converted = {i: _to_tco2e(float(monthly_values.get(i, 0.0) or 0.0), factor, ghg_unit) for i in range(1, 13)}
    total = float(sum(converted.values()))
    entry["value_mode"] = "emissions"
    entry["storage_reason"] = reason
    entry["uom"] = "tCO2e"
    entry["ghg_unit"] = "tCO2e"
    entry["factor"] = 1.0
    entry["qty"] = total
    entry["calc_tco2e"] = total
    entry["monthly_values"] = converted


def _same_quantity_storage(entry: dict[str, Any], plan: dict[str, Any]) -> bool:
    if _clean(entry.get("value_mode")).lower() != "quantity" or _clean(plan.get("value_mode")).lower() != "quantity":
        return False
    same_dataset = _single_int([entry.get("dataset_id"), plan.get("dataset_id")]) is not None
    same_factor_db = _single_int([entry.get("factor_db_id"), plan.get("factor_db_id")]) is not None or (
        entry.get("factor_db_id") is None and plan.get("factor_db_id") is None
    )
    same_factor = _single_numeric([entry.get("factor"), plan.get("factor")]) is not None
    same_uom = _single_text([entry.get("uom"), plan.get("uom")]) is not None
    same_ghg_unit = _single_text([entry.get("ghg_unit"), plan.get("ghg_unit")]) is not None
    return bool(same_dataset and same_factor_db and same_factor and same_uom and same_ghg_unit)


def _coerce_row_payload_to_emissions(row: dict[str, Any], reason: str) -> dict[str, Any]:
    out = dict(row)
    if _clean(out.get("value_mode")).lower() == "emissions":
        out["storage_reason"] = _clean(out.get("storage_reason")) or reason
        return out
    factor = float(out.get("factor") or 0.0)
    ghg_unit = out.get("ghg_unit")
    converted = {
        i: _to_tco2e(float(out.get(f"month_{i}") or 0.0), factor, ghg_unit)
        for i in range(1, 13)
    }
    total = float(sum(converted.values()))
    for i in range(1, 13):
        out[f"month_{i}"] = float(converted[i])
    out["qty"] = total
    out["calc_tco2e"] = total
    out["uom"] = "tCO2e"
    out["ghg_unit"] = "tCO2e"
    out["factor"] = 1.0
    out["value_mode"] = "emissions"
    out["storage_reason"] = reason
    return out


def _template_lookup_path() -> Path:
    return Path(__file__).resolve().parents[1] / "wfm_import" / "analysis" / "wfm_template_id_lookup.json"


def _load_template_lookup() -> dict[str, str]:
    path = _template_lookup_path()
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        preferred = payload.get("preferred_items") or {}
        out: dict[str, str] = {}
        if isinstance(preferred, dict):
            for k, v in preferred.items():
                kk = _clean(k).lower()
                vv = _clean(v)
                if kk and vv:
                    out[kk] = vv
        return out
    except Exception as exc:
        logger.debug("Unable to load preferred lookup items from %s: %s", path, exc)
        return {}


def _lookup_key(section: str, activity: str, c2: str, c3: str, c4: str, c5: str, c6: str, c7: str) -> str:
    return "|".join(
        [
            _clean(section).lower(),
            _clean(activity).lower(),
            _clean(c2).lower(),
            _clean(c3).lower(),
            _clean(c4).lower(),
            _clean(c5).lower(),
            _clean(c6).lower(),
            _clean(c7).lower(),
        ]
    )


def _lookup_match(
    lookup: dict[str, str],
    *,
    section: str,
    activity: str,
    c2: str,
    c3: str,
    c4: str,
    c5: str,
    c6: str,
    c7: str,
    scope_col: int | None,
) -> dict[str, Any]:
    base_cols = [_clean(c2), _clean(c3), _clean(c4), _clean(c5), _clean(c6), _clean(c7)]
    base_key = _lookup_key(section, activity, *base_cols)
    exact_id = _clean(lookup.get(base_key))
    if exact_id:
        return {
            "lookup_key": base_key,
            "matched_lookup_key": base_key,
            "factor_original_id": exact_id,
            "match_source": "template_lookup",
            "scope_override": None,
            "match_note": "",
            "candidate_original_id": "",
        }

    scope_idx = int(scope_col) - 3 if scope_col is not None else -1
    if scope_idx < 0 or scope_idx >= len(base_cols):
        return {
            "lookup_key": base_key,
            "matched_lookup_key": "",
            "factor_original_id": "",
            "match_source": "unresolved",
            "scope_override": None,
            "match_note": "",
            "candidate_original_id": "",
        }

    original_scope_token = _clean(base_cols[scope_idx])
    candidates: dict[tuple[str, str], dict[str, Any]] = {}
    for alt_scope_token in ("1", "2", "3", "scope 1", "scope 2", "scope 3"):
        if _clean(alt_scope_token).lower() == original_scope_token.lower():
            continue
        alt_cols = list(base_cols)
        alt_cols[scope_idx] = alt_scope_token
        alt_key = _lookup_key(section, activity, *alt_cols)
        alt_id = _clean(lookup.get(alt_key))
        alt_scope = _norm_scope(alt_scope_token)
        if not alt_id or not alt_scope:
            continue
        candidates[(alt_id, alt_scope)] = {
            "lookup_key": base_key,
            "matched_lookup_key": alt_key,
            "factor_original_id": alt_id,
            "match_source": "template_lookup_scope_override",
            "scope_override": alt_scope,
            "match_note": f"Template lookup corrected scope from {original_scope_token or 'blank'} to {alt_scope_token}.",
        }

    if len(candidates) == 1:
        candidate = next(iter(candidates.values()))
        if original_scope_token:
            return {
                "lookup_key": base_key,
                "matched_lookup_key": candidate["matched_lookup_key"],
                "factor_original_id": "",
                "match_source": "scope_mismatch_candidate",
                "scope_override": None,
                "match_note": (
                    f"Workbook scope {original_scope_token} conflicts with template lookup "
                    f"{candidate['matched_lookup_key'].split('|')[4] or candidate['scope_override']}."
                ),
                "candidate_original_id": candidate["factor_original_id"],
            }
        return candidate

    return {
        "lookup_key": base_key,
        "matched_lookup_key": "",
        "factor_original_id": "",
        "match_source": "unresolved",
        "scope_override": None,
        "match_note": "",
        "candidate_original_id": "",
    }


def _parse_month_header(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    s = _clean(value)
    if not s:
        return None
    for fmt in ("%b-%Y", "%b %Y", "%B-%Y", "%B %Y", "%Y-%m"):
        try:
            d = datetime.strptime(s, fmt)
            return date(d.year, d.month, 1)
        except Exception as exc:
            logger.debug("Unable to parse month header %r with format %s: %s", s, fmt, exc)
    try:
        d = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.notna(d):
            return date(int(d.year), int(d.month), 1)
    except Exception as exc:
        logger.debug("Unable to parse month header %r via pandas date conversion: %s", s, exc)
    return None


def _is_reporting_yes(value: Any) -> bool:
    s = _clean(value).lower()
    if not s:
        return True
    return s in {"yes", "y", "true", "1"}


def _dataset_by_year() -> dict[int, int]:
    sql = """
        SELECT dataset_id, year, source, analysis_type, name
        FROM datasets
        ORDER BY year DESC NULLS LAST, dataset_id DESC
    """
    chosen: dict[int, int] = {}
    fallback: dict[int, int] = {}
    with get_conn() as con:
        df = con.execute(sql).df()
    if df is None or df.empty:
        return chosen
    for _, r in df.iterrows():
        y = r.get("year")
        dsid = r.get("dataset_id")
        if y is None or dsid is None:
            continue
        year = int(y)
        dsi = int(dsid)
        fallback.setdefault(year, dsi)
        source = _clean(r.get("source")).lower()
        analysis = _clean(r.get("analysis_type")).lower()
        name = _clean(r.get("name")).lower()
        if ("desnz-defra" in source or "activity-and-spend" in name) and ("activity" in analysis and "spend" in analysis):
            chosen.setdefault(year, dsi)
    for year, dsi in fallback.items():
        chosen.setdefault(year, dsi)
    return chosen


@dataclass
class FactorRec:
    db_id: int | None
    level_1: str | None
    level_2: str | None
    level_3: str | None
    level_4: str | None
    column_text: str | None
    report_label: str | None
    uom: str | None
    ghg_unit: str | None
    factor: float | None


def _factor_lookup_batch(req: dict[tuple[int, str], set[str]]) -> dict[tuple[int, str, str], FactorRec]:
    out: dict[tuple[int, str, str], FactorRec] = {}
    with get_conn() as con:
        for (dataset_id, scope), ids in req.items():
            id_list = sorted({_clean(x) for x in ids if _clean(x)})
            if not id_list:
                continue
            if db_backend() == "postgres":
                df = con.execute(
                    """
                    SELECT original_id, db_id, level_1, level_2, level_3, level_4,
                           column_text, report_label, uom, ghg_unit, factor
                    FROM v_factor_lookup
                    WHERE dataset_id=%s AND scope=%s AND original_id = ANY(%s)
                    """,
                    [int(dataset_id), str(scope), id_list],
                ).df()
            else:
                ph = ",".join(["?"] * len(id_list))
                df = con.execute(
                    f"""
                    SELECT original_id, db_id, level_1, level_2, level_3, level_4,
                           column_text, report_label, uom, ghg_unit, factor
                    FROM v_factor_lookup
                    WHERE dataset_id=? AND scope=? AND original_id IN ({ph})
                    """,
                    [int(dataset_id), str(scope), *id_list],
                ).df()
            if df is None or df.empty:
                continue
            for _, r in df.iterrows():
                oid = _clean(r.get("original_id"))
                if not oid:
                    continue
                out[(int(dataset_id), str(scope), oid)] = FactorRec(
                    db_id=int(r.get("db_id")) if r.get("db_id") is not None and str(r.get("db_id")) != "nan" else None,
                    level_1=_clean(r.get("level_1")) or None,
                    level_2=_clean(r.get("level_2")) or None,
                    level_3=_clean(r.get("level_3")) or None,
                    level_4=_clean(r.get("level_4")) or None,
                    column_text=_clean(r.get("column_text")) or None,
                    report_label=_clean(r.get("report_label")) or None,
                    uom=_clean(r.get("uom")) or None,
                    ghg_unit=_clean(r.get("ghg_unit")) or None,
                    factor=float(r.get("factor")) if r.get("factor") is not None and str(r.get("factor")) != "nan" else None,
                )
    return out


def _factor_rec_from_mapping(mapping: dict[str, Any]) -> FactorRec:
    return FactorRec(
        db_id=int(mapping.get("db_id")) if mapping.get("db_id") is not None and str(mapping.get("db_id")) != "nan" else None,
        level_1=_clean(mapping.get("level_1")) or None,
        level_2=_clean(mapping.get("level_2")) or None,
        level_3=_clean(mapping.get("level_3")) or None,
        level_4=_clean(mapping.get("level_4")) or None,
        column_text=_clean(mapping.get("column_text")) or None,
        report_label=_clean(mapping.get("report_label")) or None,
        uom=_clean(mapping.get("uom")) or None,
        ghg_unit=_clean(mapping.get("ghg_unit")) or None,
        factor=float(mapping.get("factor")) if mapping.get("factor") is not None and str(mapping.get("factor")) != "nan" else None,
    )


@lru_cache(maxsize=4096)
def _equivalent_factor_for_scope(dataset_id: int, desired_scope: str, source_original_id: str) -> tuple[str, FactorRec] | None:
    dataset_id = int(dataset_id)
    scope = _clean(desired_scope)
    original_id = _clean(source_original_id)
    if not scope or not original_id:
        return None

    with get_conn() as con:
        source = con.execute(
            """
            SELECT
                COALESCE(level_2, '') AS level_2,
                COALESCE(level_3, '') AS level_3,
                COALESCE(level_4, '') AS level_4,
                COALESCE(uom, '') AS uom,
                COALESCE(ghg_unit, '') AS ghg_unit
            FROM v_factor_lookup
            WHERE dataset_id=%s AND original_id=%s
            ORDER BY CASE WHEN scope=%s THEN 0 ELSE 1 END, db_id ASC
            LIMIT 1
            """,
            [dataset_id, original_id, scope],
        ).fetchone()
        if not source:
            return None

        level_2, level_3, level_4, uom, ghg_unit = [_clean(v) for v in source]
        df = con.execute(
            """
            SELECT original_id, db_id, level_1, level_2, level_3, level_4,
                   column_text, report_label, uom, ghg_unit, factor
            FROM v_factor_lookup
            WHERE dataset_id=%s
              AND scope=%s
              AND COALESCE(level_2, '')=%s
              AND COALESCE(level_3, '')=%s
              AND COALESCE(level_4, '')=%s
              AND COALESCE(uom, '')=%s
              AND COALESCE(ghg_unit, '')=%s
            ORDER BY db_id ASC
            """,
            [dataset_id, scope, level_2, level_3, level_4, uom, ghg_unit],
        ).df()

    if df is None or df.empty or len(df.index) != 1:
        return None

    record = df.iloc[0].to_dict()
    resolved_original_id = _clean(record.get("original_id"))
    if not resolved_original_id:
        return None
    return resolved_original_id, _factor_rec_from_mapping(record)


def _resolve_factor_for_scope(
    factors: dict[tuple[int, str, str], FactorRec],
    dataset_id: int,
    scope: str,
    original_id: str,
) -> tuple[str, FactorRec | None]:
    cleaned_scope = _clean(scope)
    cleaned_original_id = _clean(original_id)
    rec = factors.get((int(dataset_id), cleaned_scope, cleaned_original_id))
    if rec is not None and rec.factor is not None:
        return cleaned_original_id, rec
    equivalent = _equivalent_factor_for_scope(int(dataset_id), cleaned_scope, cleaned_original_id)
    if equivalent is None:
        return cleaned_original_id, None
    resolved_original_id, resolved_rec = equivalent
    if resolved_rec.factor is None:
        return cleaned_original_id, None
    return resolved_original_id, resolved_rec


def _to_tco2e(qty: float, factor: float, ghg_unit: str | None) -> float:
    emissions = float(qty) * float(factor)
    ghg = _clean(ghg_unit).replace(" ", "").lower()
    if ghg.startswith("kg"):
        return emissions / 1000.0
    return emissions


def parse_legacy_annual_workbook(raw_bytes: bytes) -> dict[str, Any]:
    wb = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    lookup = _load_template_lookup()
    dataset_by_year = _dataset_by_year()

    parsed_rows: list[dict[str, Any]] = []
    parse_warnings: list[str] = []
    ignored_rows = 0
    scope_override_rows = 0
    scope_mismatch_rows = 0
    scope_candidate_resolved_rows = 0

    for ws in wb.worksheets:
        current_section = ""
        month_cols: list[tuple[int, date, str]] = []
        scope_col: int | None = None
        reporting_col: int | None = None
        max_col = ws.max_column or 0
        for rnum in range(1, (ws.max_row or 0) + 1):
            c0 = _clean(ws.cell(rnum, 1).value)
            c1 = _clean(ws.cell(rnum, 2).value)
            c2 = _clean(ws.cell(rnum, 3).value)
            c3 = _clean(ws.cell(rnum, 4).value)
            c4 = _clean(ws.cell(rnum, 5).value)
            c5 = _clean(ws.cell(rnum, 6).value)
            c6 = _clean(ws.cell(rnum, 7).value)
            c7 = _clean(ws.cell(rnum, 8).value)

            if c0.lower().startswith("section"):
                current_section = c1 or c0
                layout = _section_layout(ws, rnum, max_col)
                month_cols = layout["month_cols"]
                scope_col = layout["scope_col"]
                reporting_col = layout["reporting_col"]
                continue

            if not current_section or not c1:
                continue
            if reporting_col is not None and not _is_reporting_yes(_row_text(ws, rnum, reporting_col)):
                continue
            if not month_cols:
                continue

            monthly_qty: dict[int, float] = {}
            has_non_zero = False
            for idx, (col, _d, _label) in enumerate(month_cols[:12], start=1):
                q = _safe_float(ws.cell(rnum, col).value)
                if q is None:
                    continue
                monthly_qty[idx] = float(q)
                if float(q) != 0:
                    has_non_zero = True
            if not has_non_zero:
                continue

            if _is_ignored_legacy_section(current_section):
                ignored_rows += 1
                continue

            scope = _norm_scope(_row_text(ws, rnum, scope_col) or c4)
            lookup_match = _lookup_match(
                lookup,
                section=current_section,
                activity=c1,
                c2=c2,
                c3=c3,
                c4=c4,
                c5=c5,
                c6=c6,
                c7=c7,
                scope_col=scope_col,
            )
            row_key = lookup_match.get("lookup_key") or _lookup_key(current_section, c1, c2, c3, c4, c5, c6, c7)
            factor_original_id = c0 if _is_factor_original_id(c0) else _clean(
                lookup_match.get("factor_original_id") or lookup_match.get("candidate_original_id")
            )
            if not _is_factor_original_id(c0) and lookup_match.get("scope_override"):
                scope = lookup_match.get("scope_override")
                scope_override_rows += 1
            storage_original_id = _legacy_storage_original_id(factor_original_id, row_key) if factor_original_id else ""
            match_source = "id_column" if _is_factor_original_id(c0) else (lookup_match.get("match_source") or "unresolved")
            line_label = _build_line_label(c1, c2, c3, c4, c5, c6, c7, scope_col=scope_col)

            parsed_rows.append(
                {
                    "sheet_name": ws.title,
                    "row_number": int(rnum),
                    "section": current_section,
                    "activity": c1,
                    "scope": scope,
                    "source_scope": _norm_scope(_row_text(ws, rnum, scope_col) or c4) or "",
                    "original_id": factor_original_id,
                    "factor_original_id": factor_original_id,
                    "storage_original_id": storage_original_id,
                    "match_source": match_source,
                    "lookup_key": row_key,
                    "matched_lookup_key": _clean(lookup_match.get("matched_lookup_key")),
                    "candidate_original_id": _clean(lookup_match.get("candidate_original_id")),
                    "match_note": _clean(lookup_match.get("match_note")),
                    "line_label": line_label,
                    "col_2": c2,
                    "col_3": c3,
                    "col_4": c4,
                    "col_5": c5,
                    "col_6": c6,
                    "col_7": c7,
                    "month_headers": [
                        {"position": i + 1, "label": label, "year": int(md.year), "month": int(md.month)}
                        for i, (_col, md, label) in enumerate(month_cols[:12])
                    ],
                    "monthly_qty": monthly_qty,
                }
            )

    factor_req: dict[tuple[int, str], set[str]] = defaultdict(set)
    for row in parsed_rows:
        scope = row.get("scope")
        oid = _factor_original_id_from_row(row)
        if not scope or not oid:
            continue
        for m in row.get("month_headers") or []:
            year = int(m.get("year"))
            dsid = dataset_by_year.get(year)
            if dsid is not None:
                factor_req[(int(dsid), str(scope))].add(oid)

    factors = _factor_lookup_batch(factor_req)

    staged_rows: list[dict[str, Any]] = []
    unresolved_rows: list[dict[str, Any]] = []
    collision_tracker: Counter[tuple[str, str]] = Counter()
    aggregate: dict[tuple[str, str], dict[str, Any]] = {}
    quantity_mode_rows = 0
    emissions_mode_rows = 0
    scope_equivalent_rows = 0

    for row in parsed_rows:
        scope = row.get("scope")
        oid = _factor_original_id_from_row(row)
        if not scope or not oid:
            reason = "missing scope or unresolved original_id"
            if row.get("match_source") == "scope_mismatch_candidate":
                reason = "workbook scope conflicts with template mapping"
            unresolved_rows.append({**row, "reason": reason})
            continue

        month_emissions: dict[int, float] = {}
        month_quantities: dict[int, float] = {}
        month_datasets: dict[int, int] = {}
        month_factor_db_ids: dict[int, int | None] = {}
        month_factor_values: dict[int, float | None] = {}
        month_uoms: dict[int, str | None] = {}
        month_ghg_units: dict[int, str | None] = {}
        month_missing: list[dict[str, Any]] = []
        factor_rec_primary: FactorRec | None = None
        resolved_oid = oid
        used_scope_equivalent = False

        for m in row.get("month_headers") or []:
            pos = int(m["position"])
            qty = _safe_float((row.get("monthly_qty") or {}).get(pos))
            if qty is None:
                continue
            year = int(m["year"])
            dsid = dataset_by_year.get(year)
            if dsid is None:
                month_missing.append({"position": pos, "reason": f"no dataset for year {year}"})
                continue
            month_oid, rec = _resolve_factor_for_scope(factors, int(dsid), str(scope), resolved_oid)
            if rec is None or rec.factor is None:
                month_missing.append({"position": pos, "reason": f"factor missing for dataset {dsid}"})
                continue
            if _clean(month_oid) and _clean(month_oid) != _clean(resolved_oid):
                resolved_oid = _clean(month_oid)
                used_scope_equivalent = True
            if factor_rec_primary is None:
                factor_rec_primary = rec
            month_quantities[pos] = float(qty)
            month_datasets[pos] = int(dsid)
            month_factor_db_ids[pos] = rec.db_id
            month_factor_values[pos] = float(rec.factor) if rec.factor is not None else None
            month_uoms[pos] = rec.uom
            month_ghg_units[pos] = rec.ghg_unit
            month_emissions[pos] = _to_tco2e(float(qty), float(rec.factor), rec.ghg_unit)

        if not month_emissions:
            reason = "no mappable monthly factor data"
            if row.get("match_source") == "scope_mismatch_candidate":
                reason = "workbook scope conflicts with template mapping and no factor matched that workbook scope"
                scope_mismatch_rows += 1
            unresolved_rows.append({**row, "reason": reason, "month_errors": month_missing})
            continue
        row = dict(row)
        if _clean(resolved_oid) and _clean(resolved_oid) != _clean(oid):
            row["original_id"] = resolved_oid
            row["factor_original_id"] = resolved_oid
            row["storage_original_id"] = _legacy_storage_original_id(resolved_oid, _clean(row.get("lookup_key")))
            row["match_source"] = "scope_equivalent_lookup"
            prior_note = _clean(row.get("match_note"))
            extra_note = f"Resolved to scope-matched factor {resolved_oid}."
            row["match_note"] = f"{prior_note} {extra_note}".strip()
        if used_scope_equivalent:
            scope_equivalent_rows += 1
        if row.get("match_source") == "scope_mismatch_candidate":
            scope_candidate_resolved_rows += 1

        dataset_counter = Counter(month_datasets.values())
        primary_dataset = int(dataset_counter.most_common(1)[0][0]) if dataset_counter else None
        total_emissions = float(sum(month_emissions.values()))
        storage_original_id = _storage_original_id_from_row(row)
        key = (str(scope), storage_original_id)
        collision_tracker[key] += 1
        storage_plan = _quantity_storage_plan(
            month_quantities=month_quantities,
            month_emissions=month_emissions,
            month_datasets=month_datasets,
            month_factor_db_ids=month_factor_db_ids,
            month_factor_values=month_factor_values,
            month_uoms=month_uoms,
            month_ghg_units=month_ghg_units,
            factor_rec_primary=factor_rec_primary,
            primary_dataset=primary_dataset,
        )

        entry = aggregate.get(key)
        if entry is None:
            entry = {
                **_build_staged_metadata(
                    row,
                    factor_rec_primary,
                    storage_plan["dataset_id"],
                    factor_original_id=oid,
                    storage_original_id=storage_original_id,
                ),
                "value_mode": storage_plan["value_mode"],
                "storage_reason": storage_plan["storage_reason"],
                "uom": storage_plan["uom"],
                "ghg_unit": storage_plan["ghg_unit"],
                "factor": storage_plan["factor"],
                "monthly_emissions": {i: 0.0 for i in range(1, 13)},
                "monthly_values": {i: 0.0 for i in range(1, 13)},
                "monthly_dataset_ids": {},
                "qty": 0.0,
                "calc_tco2e": 0.0,
                "source_qty": 0.0,
                "source_uom": storage_plan.get("source_uom"),
                "source_rows": [],
                "month_errors": [],
            }
            aggregate[key] = entry
        elif entry.get("value_mode") == "quantity" and (
            storage_plan["value_mode"] != "quantity" or not _same_quantity_storage(entry, storage_plan)
        ):
            _convert_entry_to_emissions(entry, "duplicate_rows_with_mixed_factor_context")

        if entry.get("value_mode") == "emissions":
            values_to_add = {i: float(month_emissions.get(i, 0.0)) for i in range(1, 13)}
            qty_to_add = total_emissions
            calc_to_add = total_emissions
        else:
            values_to_add = storage_plan["monthly_values"]
            qty_to_add = float(storage_plan["qty"])
            calc_to_add = float(storage_plan["calc_tco2e"])

        for pos, val in values_to_add.items():
            entry["monthly_values"][int(pos)] = float(entry["monthly_values"][int(pos)]) + float(val)
        for pos, val in month_emissions.items():
            entry["monthly_emissions"][int(pos)] = float(entry["monthly_emissions"][int(pos)]) + float(val)
            if pos in month_datasets:
                entry["monthly_dataset_ids"][int(pos)] = int(month_datasets[pos])
        entry["qty"] = float(entry["qty"]) + float(qty_to_add)
        entry["calc_tco2e"] = float(entry["calc_tco2e"]) + float(calc_to_add)
        entry["source_qty"] = float(entry.get("source_qty") or 0.0) + float(storage_plan.get("source_qty") or 0.0)
        incoming_source_uom = _clean(storage_plan.get("source_uom"))
        current_source_uom = _clean(entry.get("source_uom"))
        if incoming_source_uom:
            if not current_source_uom:
                entry["source_uom"] = incoming_source_uom
            elif current_source_uom.lower() != incoming_source_uom.lower():
                entry["source_uom"] = None
        entry["source_rows"].append({"sheet": row.get("sheet_name"), "row": row.get("row_number"), "activity": row.get("activity")})
        entry["month_errors"].extend(month_missing)

    for (_scope, _oid), entry in aggregate.items():
        if _clean(entry.get("value_mode")).lower() == "quantity":
            quantity_mode_rows += 1
        else:
            emissions_mode_rows += 1
        staged_rows.append(
            {
                **_build_staged_metadata(
                    entry,
                    None,
                    entry["dataset_id"],
                    factor_original_id=_clean(entry.get("factor_original_id")),
                    storage_original_id=_clean(entry.get("original_id")),
                ),
                "factor_db_id": entry["factor_db_id"],
                "level_1": entry["level_1"],
                "level_2": entry["level_2"],
                "level_3": entry["level_3"],
                "level_4": entry["level_4"],
                "value_mode": entry["value_mode"],
                "storage_reason": entry["storage_reason"],
                "qty": float(entry["qty"]),
                "uom": entry["uom"],
                "ghg_unit": entry["ghg_unit"],
                "factor": float(entry["factor"] or 0.0),
                "calc_tco2e": float(entry["calc_tco2e"]),
                "source_qty": float(entry.get("source_qty") or 0.0),
                "source_uom": entry.get("source_uom"),
                "month_1": float(entry["monthly_values"].get(1, 0.0)),
                "month_2": float(entry["monthly_values"].get(2, 0.0)),
                "month_3": float(entry["monthly_values"].get(3, 0.0)),
                "month_4": float(entry["monthly_values"].get(4, 0.0)),
                "month_5": float(entry["monthly_values"].get(5, 0.0)),
                "month_6": float(entry["monthly_values"].get(6, 0.0)),
                "month_7": float(entry["monthly_values"].get(7, 0.0)),
                "month_8": float(entry["monthly_values"].get(8, 0.0)),
                "month_9": float(entry["monthly_values"].get(9, 0.0)),
                "month_10": float(entry["monthly_values"].get(10, 0.0)),
                "month_11": float(entry["monthly_values"].get(11, 0.0)),
                "month_12": float(entry["monthly_values"].get(12, 0.0)),
                "monthly_dataset_ids": entry["monthly_dataset_ids"],
                "month_errors": entry["month_errors"],
                "source_rows": entry["source_rows"],
                "collision_count": int(collision_tracker[(entry["scope"], entry["original_id"])]),
            }
        )

    collision_rows = [x for x in staged_rows if int(x.get("collision_count") or 1) > 1]
    if collision_rows:
        parse_warnings.append(
            f"{len(collision_rows)} template lines were aggregated from duplicate source rows with the same scope and mapping key."
        )
    if emissions_mode_rows:
        parse_warnings.append(
            f"{emissions_mode_rows} rows were stored as emissions because their month-level factor context could not be represented as one raw quantity row."
        )
    if ignored_rows:
        parse_warnings.append(
            f"Skipped {ignored_rows} Company Information/Data rows because they are metadata and are not imported as emissions lines."
        )
    if scope_override_rows:
        parse_warnings.append(
            f"{scope_override_rows} rows were matched by correcting the workbook scope against the template mapping."
        )
    if scope_mismatch_rows:
        parse_warnings.append(
            f"{scope_mismatch_rows} rows were left unresolved because the workbook scope conflicts with the template mapping."
        )
    if scope_candidate_resolved_rows:
        parse_warnings.append(
            f"{scope_candidate_resolved_rows} rows were resolved by using a candidate template ID while keeping the workbook scope."
        )
    if scope_equivalent_rows:
        parse_warnings.append(
            f"{scope_equivalent_rows} rows were resolved by translating a cross-scope factor ID to the equivalent factor for the workbook scope."
        )

    parse_warnings.append(
        "Rows with one consistent dataset/factor are committed as raw quantities; mixed-factor rows fall back to monthly tCO2e storage."
    )

    return {
        "ok": True,
        "summary": {
            "parsed_rows": len(parsed_rows),
            "resolved_rows": len(staged_rows),
            "unresolved_rows": len(unresolved_rows),
            "collision_rows": len(collision_rows),
            "quantity_mode_rows": quantity_mode_rows,
            "emissions_mode_rows": emissions_mode_rows,
            "ignored_rows": ignored_rows,
            "scope_override_rows": scope_override_rows,
            "scope_mismatch_rows": scope_mismatch_rows,
            "scope_equivalent_rows": scope_equivalent_rows,
            "dataset_years_available": sorted(dataset_by_year.keys()),
        },
        "warnings": parse_warnings,
        "rows_ready": staged_rows,
        "rows_unresolved": unresolved_rows[:500],
    }


def _resolve_site_id(job_id: int, site_id: int | None) -> int:
    def _lookup_existing_site_id(client_db_id: int) -> int | None:
        with get_conn() as con:
            row = con.execute(
                """
                SELECT site_id
                FROM client_sites
                WHERE client_db_id=%s
                ORDER BY is_registered_office DESC, site_id ASC
                LIMIT 1
                """,
                [int(client_db_id)],
            ).fetchone()
        return int(row[0]) if row else None

    def _build_registered_office_location(client_row: tuple[Any, ...] | None) -> str:
        if not client_row:
            return "Registered Office"
        parts = [_clean(value) for value in client_row[1:] if _clean(value)]
        return ", ".join(parts) if parts else "Registered Office"

    with get_conn() as con:
        job = con.execute("SELECT client_db_id FROM jobs WHERE job_id=%s", [int(job_id)]).fetchone()
        if not job:
            raise ValueError("Job not found")
        client_db_id = int(job[0])
        if site_id is not None:
            ok = con.execute(
                "SELECT 1 FROM client_sites WHERE site_id=%s AND client_db_id=%s",
                [int(site_id), int(client_db_id)],
            ).fetchone()
            if not ok:
                raise ValueError("site_id does not belong to this job's client")
            return int(site_id)
        row = con.execute(
            """
            SELECT site_id
            FROM client_sites
            WHERE client_db_id=%s
            ORDER BY is_registered_office DESC, site_id ASC
            LIMIT 1
            """,
            [int(client_db_id)],
        ).fetchone()
        if not row:
            client_row = con.execute(
                """
                SELECT client_name, addr_line1, addr_line2, addr_city, addr_region, addr_postcode, addr_country
                FROM clients
                WHERE db_id=%s
                """,
                [int(client_db_id)],
            ).fetchone()
        else:
            client_row = None
    if row:
        return int(row[0])
    if not client_row:
        raise ValueError("No site found for this job/client")

    try:
        return int(add_site(int(client_db_id), "Registered Office", _build_registered_office_location(client_row), True))
    except Exception as exc:
        logger.warning("Unable to add registered office site for client %s: %s", client_db_id, exc)
        existing_site_id = _lookup_existing_site_id(int(client_db_id))
        if existing_site_id is not None:
            return int(existing_site_id)
        raise ValueError("No site found for this job/client")


def _ensure_job_scope_rows_schema(con) -> None:
    """Keep legacy commit writes tolerant of older production schemas."""
    ddl_statements = [
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS site_id INTEGER",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS category VARCHAR",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS level_4 VARCHAR",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS report_label VARCHAR",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS ghg_unit VARCHAR",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS enabled BOOLEAN DEFAULT TRUE",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS month_1 NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS month_2 NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS month_3 NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS month_4 NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS month_5 NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS month_6 NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS month_7 NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS month_8 NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS month_9 NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS month_10 NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS month_11 NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS month_12 NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS apply_pct NUMERIC DEFAULT 100",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS data_source VARCHAR DEFAULT 'Company Data'",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS data_confidence VARCHAR DEFAULT 'M'",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS is_custom_entry BOOLEAN DEFAULT FALSE",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS override_tco2e NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS override_reason VARCHAR",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS source_qty NUMERIC",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS source_uom VARCHAR",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT NOW()",
        "ALTER TABLE job_scope_rows ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT NOW()",
    ]
    for ddl in ddl_statements:
        try:
            con.execute(ddl)
        except Exception as exc:
            logger.warning("Ignoring legacy annual schema step %r: %s", ddl, exc)


def commit_legacy_rows(job_id: int, site_id: int | None, rows: list[dict[str, Any]]) -> dict[str, Any]:
    effective_site_id = _resolve_site_id(int(job_id), int(site_id) if site_id is not None else None)
    inserted = 0
    updated = 0
    disabled_existing = 0
    with get_conn() as con:
        _ensure_job_scope_rows_schema(con)
        disabled_row = con.execute(
            """
            SELECT COUNT(*)
            FROM job_scope_rows
            WHERE job_id=%s AND site_id=%s AND data_source='Legacy Annual Upload' AND enabled=TRUE
            """,
            [int(job_id), int(effective_site_id)],
        ).fetchone()
        disabled_existing = int(disabled_row[0] or 0) if disabled_row else 0
        con.execute(
            """
            UPDATE job_scope_rows
            SET enabled=FALSE, updated_at=NOW()
            WHERE job_id=%s AND site_id=%s AND data_source='Legacy Annual Upload'
            """,
            [int(job_id), int(effective_site_id)],
        )
        for r in rows:
            scope = _clean(r.get("scope"))
            original_id = _storage_original_id_from_row(r)
            factor_original_id = _factor_original_id_from_row(r)
            value_mode = _clean(r.get("value_mode")) or "emissions"
            if not scope or not original_id:
                continue

            if value_mode == "quantity":
                notes = "Legacy annual import; monthly values stored as raw quantities with factor metadata"
            else:
                notes = "Legacy annual import; monthly values stored as tCO2e emissions fallback"
            note_parts: list[str] = []
            if factor_original_id:
                note_parts.append(f"factor_original_id={factor_original_id}")
            if _clean(r.get("section")):
                note_parts.append(f"section={_clean(r.get('section'))}")
            if _clean(r.get("activity")):
                note_parts.append(f"activity={_clean(r.get('activity'))}")
            if _clean(r.get("storage_reason")):
                note_parts.append(f"storage_reason={_clean(r.get('storage_reason'))}")
            if note_parts:
                notes = f"{notes} ({'; '.join(note_parts)})"

            existing = con.execute(
                """
                SELECT row_id
                FROM job_scope_rows
                WHERE job_id=%s AND site_id=%s AND scope=%s AND original_id=%s
                LIMIT 1
                """,
                [int(job_id), int(effective_site_id), scope, original_id],
            ).fetchone()

            params_common = [
                r.get("dataset_id"),
                r.get("factor_db_id"),
                float(r.get("qty") or 0),
                r.get("uom") or "tCO2e",
                float(r.get("factor") or 1.0),
                r.get("ghg_unit") or "tCO2e",
                float(r.get("calc_tco2e") or 0),
                float(r.get("source_qty") or 0),
                r.get("source_uom") or r.get("uom"),
                r.get("level_1"),
                r.get("level_2"),
                r.get("level_3"),
                r.get("level_4"),
                r.get("column_text"),
                r.get("report_label"),
                r.get("category"),
                float(r.get("month_1") or 0),
                float(r.get("month_2") or 0),
                float(r.get("month_3") or 0),
                float(r.get("month_4") or 0),
                float(r.get("month_5") or 0),
                float(r.get("month_6") or 0),
                float(r.get("month_7") or 0),
                float(r.get("month_8") or 0),
                float(r.get("month_9") or 0),
                float(r.get("month_10") or 0),
                float(r.get("month_11") or 0),
                float(r.get("month_12") or 0),
            ]

            if existing:
                con.execute(
                    """
                    UPDATE job_scope_rows
                    SET enabled=TRUE,
                        dataset_id=%s,
                        factor_db_id=%s,
                        qty=%s,
                        uom=%s,
                        factor=%s,
                        ghg_unit=%s,
                        calc_tco2e=%s,
                        source_qty=%s,
                        source_uom=%s,
                        level_1=%s,
                        level_2=%s,
                        level_3=%s,
                        level_4=%s,
                        column_text=%s,
                        report_label=%s,
                        category=%s,
                        month_1=%s, month_2=%s, month_3=%s, month_4=%s, month_5=%s, month_6=%s,
                        month_7=%s, month_8=%s, month_9=%s, month_10=%s, month_11=%s, month_12=%s,
                        apply_pct=100,
                        data_source='Legacy Annual Upload',
                        data_confidence='M',
                        notes=%s,
                        updated_at=NOW()
                    WHERE row_id=%s
                    """,
                    [*params_common, notes, int(existing[0])],
                )
                updated += 1
            else:
                con.execute(
                    """
                    INSERT INTO job_scope_rows (
                        job_id, site_id, scope, category, dataset_id, factor_db_id, original_id,
                        level_1, level_2, level_3, level_4, column_text, report_label, notes, enabled,
                        qty, uom, factor, ghg_unit, calc_tco2e, source_qty, source_uom, override_tco2e, override_reason,
                        month_1, month_2, month_3, month_4, month_5, month_6,
                        month_7, month_8, month_9, month_10, month_11, month_12,
                        apply_pct, data_source, data_confidence, is_custom_entry, created_at, updated_at
                    )
                    VALUES (
                        %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, TRUE,
                        %s, %s, %s, %s, %s, %s, %s, NULL, NULL,
                        %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s,
                        100, 'Legacy Annual Upload', 'M', FALSE, NOW(), NOW()
                    )
                    """,
                    [
                        int(job_id),
                        int(effective_site_id),
                        scope,
                        r.get("category"),
                        r.get("dataset_id"),
                        r.get("factor_db_id"),
                        original_id,
                        r.get("level_1"),
                        r.get("level_2"),
                        r.get("level_3"),
                        r.get("level_4"),
                        r.get("column_text"),
                        r.get("report_label"),
                        notes,
                        float(r.get("qty") or 0),
                        r.get("uom") or "tCO2e",
                        float(r.get("factor") or 1.0),
                        r.get("ghg_unit") or "tCO2e",
                        float(r.get("calc_tco2e") or 0),
                        float(r.get("source_qty") or 0),
                        r.get("source_uom") or r.get("uom"),
                        float(r.get("month_1") or 0),
                        float(r.get("month_2") or 0),
                        float(r.get("month_3") or 0),
                        float(r.get("month_4") or 0),
                        float(r.get("month_5") or 0),
                        float(r.get("month_6") or 0),
                        float(r.get("month_7") or 0),
                        float(r.get("month_8") or 0),
                        float(r.get("month_9") or 0),
                        float(r.get("month_10") or 0),
                        float(r.get("month_11") or 0),
                        float(r.get("month_12") or 0),
                    ],
                )
                inserted += 1

    return {
        "ok": True,
        "job_id": int(job_id),
        "site_id": int(effective_site_id),
        "inserted": inserted,
        "updated": updated,
        "disabled_existing_legacy_rows": disabled_existing,
    }


def resolve_unresolved_rows(
    rows_ready: list[dict[str, Any]],
    rows_unresolved: list[dict[str, Any]],
    manual_lookup: dict[str, str] | None = None,
) -> dict[str, Any]:
    manual_lookup = manual_lookup or {}
    dataset_by_year = _dataset_by_year()

    seed_rows: list[dict[str, Any]] = list(rows_ready or [])
    unresolved_rows: list[dict[str, Any]] = []
    to_process: list[dict[str, Any]] = []

    for row in rows_unresolved or []:
        r = dict(row)
        if _is_ignored_legacy_section(r.get("section")):
            continue
        oid = _factor_original_id_from_row(r)
        lk = _clean(r.get("lookup_key"))
        if not oid and lk and _clean(manual_lookup.get(lk)):
            oid = _clean(manual_lookup.get(lk))
            r["original_id"] = oid
            r["factor_original_id"] = oid
            r["match_source"] = "manual_override"
        if not oid:
            r["reason"] = "missing original_id after manual override"
            unresolved_rows.append(r)
            continue
        if not _norm_scope(r.get("scope")) and not _clean(r.get("scope")):
            r["reason"] = "missing scope"
            unresolved_rows.append(r)
            continue
        to_process.append(r)

    factor_req: dict[tuple[int, str], set[str]] = defaultdict(set)
    for row in to_process:
        scope = _clean(row.get("scope"))
        oid = _factor_original_id_from_row(row)
        for m in row.get("month_headers") or []:
            year = int(m.get("year"))
            dsid = dataset_by_year.get(year)
            if dsid is not None:
                factor_req[(int(dsid), scope)].add(oid)
    factors = _factor_lookup_batch(factor_req)

    aggregate: dict[tuple[str, str], dict[str, Any]] = {}
    for row in to_process:
        scope = _clean(row.get("scope"))
        oid = _factor_original_id_from_row(row)
        month_emissions: dict[int, float] = {}
        month_quantities: dict[int, float] = {}
        month_datasets: dict[int, int] = {}
        month_factor_db_ids: dict[int, int | None] = {}
        month_factor_values: dict[int, float | None] = {}
        month_uoms: dict[int, str | None] = {}
        month_ghg_units: dict[int, str | None] = {}
        month_errors: list[dict[str, Any]] = []
        factor_primary: FactorRec | None = None
        resolved_oid = oid

        for m in row.get("month_headers") or []:
            pos = int(m.get("position"))
            qty = _safe_float((row.get("monthly_qty") or {}).get(pos))
            if qty is None:
                continue
            year = int(m.get("year"))
            dsid = dataset_by_year.get(year)
            if dsid is None:
                month_errors.append({"position": pos, "reason": f"no dataset for year {year}"})
                continue
            month_oid, rec = _resolve_factor_for_scope(factors, int(dsid), scope, resolved_oid)
            if rec is None or rec.factor is None:
                month_errors.append({"position": pos, "reason": f"factor missing for dataset {dsid}"})
                continue
            if _clean(month_oid) and _clean(month_oid) != _clean(resolved_oid):
                resolved_oid = _clean(month_oid)
            if factor_primary is None:
                factor_primary = rec
            month_quantities[pos] = float(qty)
            month_datasets[pos] = int(dsid)
            month_factor_db_ids[pos] = rec.db_id
            month_factor_values[pos] = float(rec.factor) if rec.factor is not None else None
            month_uoms[pos] = rec.uom
            month_ghg_units[pos] = rec.ghg_unit
            month_emissions[pos] = _to_tco2e(float(qty), float(rec.factor), rec.ghg_unit)

        if not month_emissions:
            row["reason"] = "manual original_id did not resolve to monthly factor data for the row scope"
            row["month_errors"] = month_errors
            unresolved_rows.append(row)
            continue
        row = dict(row)
        if _clean(resolved_oid) and _clean(resolved_oid) != _clean(oid):
            row["original_id"] = resolved_oid
            row["factor_original_id"] = resolved_oid
            row["storage_original_id"] = _legacy_storage_original_id(resolved_oid, _clean(row.get("lookup_key")))
            row["match_source"] = "scope_equivalent_lookup"
            prior_note = _clean(row.get("match_note"))
            extra_note = f"Resolved to scope-matched factor {resolved_oid}."
            row["match_note"] = f"{prior_note} {extra_note}".strip()

        storage_original_id = _storage_original_id_from_row(row)
        key = (scope, storage_original_id)
        ds_counter = Counter(month_datasets.values())
        primary_dataset = int(ds_counter.most_common(1)[0][0]) if ds_counter else None
        total = float(sum(month_emissions.values()))
        storage_plan = _quantity_storage_plan(
            month_quantities=month_quantities,
            month_emissions=month_emissions,
            month_datasets=month_datasets,
            month_factor_db_ids=month_factor_db_ids,
            month_factor_values=month_factor_values,
            month_uoms=month_uoms,
            month_ghg_units=month_ghg_units,
            factor_rec_primary=factor_primary,
            primary_dataset=primary_dataset,
        )
        if key not in aggregate:
            aggregate[key] = {
                **_build_staged_metadata(
                    row,
                    factor_primary,
                    storage_plan["dataset_id"],
                    factor_original_id=_factor_original_id_from_row(row),
                    storage_original_id=storage_original_id,
                ),
                "value_mode": storage_plan["value_mode"],
                "storage_reason": storage_plan["storage_reason"],
                "uom": storage_plan["uom"],
                "ghg_unit": storage_plan["ghg_unit"],
                "factor": storage_plan["factor"],
                "monthly_emissions": {i: 0.0 for i in range(1, 13)},
                "monthly_values": {i: 0.0 for i in range(1, 13)},
                "monthly_dataset_ids": {},
                "qty": 0.0,
                "calc_tco2e": 0.0,
                "source_qty": 0.0,
                "source_uom": storage_plan.get("source_uom"),
                "source_rows": [],
                "month_errors": [],
            }
        entry = aggregate[key]
        if entry.get("value_mode") == "quantity" and (
            storage_plan["value_mode"] != "quantity" or not _same_quantity_storage(entry, storage_plan)
        ):
            _convert_entry_to_emissions(entry, "duplicate_rows_with_mixed_factor_context")

        if entry.get("value_mode") == "emissions":
            values_to_add = {i: float(month_emissions.get(i, 0.0)) for i in range(1, 13)}
            qty_to_add = total
            calc_to_add = total
        else:
            values_to_add = storage_plan["monthly_values"]
            qty_to_add = float(storage_plan["qty"])
            calc_to_add = float(storage_plan["calc_tco2e"])

        for pos, val in values_to_add.items():
            entry["monthly_values"][int(pos)] = float(entry["monthly_values"][int(pos)]) + float(val)
        for pos, val in month_emissions.items():
            entry["monthly_emissions"][int(pos)] = float(entry["monthly_emissions"][int(pos)]) + float(val)
            if pos in month_datasets:
                entry["monthly_dataset_ids"][int(pos)] = int(month_datasets[pos])
        entry["qty"] = float(entry["qty"]) + float(qty_to_add)
        entry["calc_tco2e"] = float(entry["calc_tco2e"]) + float(calc_to_add)
        entry["source_qty"] = float(entry.get("source_qty") or 0.0) + float(storage_plan.get("source_qty") or 0.0)
        incoming_source_uom = _clean(storage_plan.get("source_uom"))
        current_source_uom = _clean(entry.get("source_uom"))
        if incoming_source_uom:
            if not current_source_uom:
                entry["source_uom"] = incoming_source_uom
            elif current_source_uom.lower() != incoming_source_uom.lower():
                entry["source_uom"] = None
        entry["source_rows"].append({"sheet": row.get("sheet_name"), "row": row.get("row_number"), "activity": row.get("activity")})
        entry["month_errors"].extend(month_errors)

    for _key, entry in aggregate.items():
        seed_rows.append(
            {
                **_build_staged_metadata(
                    entry,
                    None,
                    entry["dataset_id"],
                    factor_original_id=_clean(entry.get("factor_original_id")),
                    storage_original_id=_clean(entry.get("original_id")),
                ),
                "factor_db_id": entry["factor_db_id"],
                "level_1": entry["level_1"],
                "level_2": entry["level_2"],
                "level_3": entry["level_3"],
                "level_4": entry["level_4"],
                "value_mode": entry["value_mode"],
                "storage_reason": entry["storage_reason"],
                "qty": float(entry["qty"]),
                "uom": entry["uom"],
                "ghg_unit": entry["ghg_unit"],
                "factor": float(entry["factor"] or 0.0),
                "calc_tco2e": float(entry["calc_tco2e"]),
                "source_qty": float(entry.get("source_qty") or 0.0),
                "source_uom": entry.get("source_uom"),
                "month_1": float(entry["monthly_values"].get(1, 0.0)),
                "month_2": float(entry["monthly_values"].get(2, 0.0)),
                "month_3": float(entry["monthly_values"].get(3, 0.0)),
                "month_4": float(entry["monthly_values"].get(4, 0.0)),
                "month_5": float(entry["monthly_values"].get(5, 0.0)),
                "month_6": float(entry["monthly_values"].get(6, 0.0)),
                "month_7": float(entry["monthly_values"].get(7, 0.0)),
                "month_8": float(entry["monthly_values"].get(8, 0.0)),
                "month_9": float(entry["monthly_values"].get(9, 0.0)),
                "month_10": float(entry["monthly_values"].get(10, 0.0)),
                "month_11": float(entry["monthly_values"].get(11, 0.0)),
                "month_12": float(entry["monthly_values"].get(12, 0.0)),
                "monthly_dataset_ids": entry["monthly_dataset_ids"],
                "month_errors": entry["month_errors"],
                "source_rows": entry["source_rows"],
                "collision_count": 1,
            }
        )

    # Merge duplicates by scope+original_id across existing ready + newly resolved.
    merged: dict[tuple[str, str], dict[str, Any]] = {}
    for row in seed_rows:
        scope = _clean(row.get("scope"))
        oid = _storage_original_id_from_row(row)
        if not scope or not oid:
            continue
        row = {**row, "original_id": oid, "storage_original_id": oid}
        key = (scope, oid)
        if key not in merged:
            merged[key] = dict(row)
            continue
        cur = merged[key]
        row_to_merge = dict(row)
        if _clean(cur.get("value_mode")).lower() != _clean(row_to_merge.get("value_mode")).lower() or (
            _clean(cur.get("value_mode")).lower() == "quantity" and not _same_quantity_storage(cur, row_to_merge)
        ):
            cur = _coerce_row_payload_to_emissions(cur, "resolved_rows_with_mixed_factor_context")
            row_to_merge = _coerce_row_payload_to_emissions(row_to_merge, "resolved_rows_with_mixed_factor_context")
            merged[key] = cur
        for i in range(1, 13):
            cur[f"month_{i}"] = float(cur.get(f"month_{i}") or 0) + float(row_to_merge.get(f"month_{i}") or 0)
        cur["qty"] = float(cur.get("qty") or 0) + float(row_to_merge.get("qty") or 0)
        cur["calc_tco2e"] = float(cur.get("calc_tco2e") or 0) + float(row_to_merge.get("calc_tco2e") or 0)
        cur["source_qty"] = float(cur.get("source_qty") or 0) + float(row_to_merge.get("source_qty") or 0)
        cur_source_uom = _clean(cur.get("source_uom"))
        row_source_uom = _clean(row_to_merge.get("source_uom"))
        if row_source_uom:
            if not cur_source_uom:
                cur["source_uom"] = row_source_uom
            elif cur_source_uom.lower() != row_source_uom.lower():
                cur["source_uom"] = None
        cur["collision_count"] = int(cur.get("collision_count") or 1) + 1

    merged_rows = list(merged.values())
    return {
        "ok": True,
        "rows_ready": merged_rows,
        "rows_unresolved": unresolved_rows,
        "summary": {
            "resolved_rows": len(merged_rows),
            "unresolved_rows": len(unresolved_rows),
        },
    }
