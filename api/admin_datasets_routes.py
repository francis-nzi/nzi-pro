"""
Admin API routes for team, lookups, datasets, and system management.
"""

import logging
import hashlib

from fastapi import APIRouter, Depends, HTTPException, Body, UploadFile, File, Form, Query, Request
from fastapi.responses import StreamingResponse
from starlette.concurrency import run_in_threadpool
from api.auth import _current_user
from api.permissions import require_permission
from core.database import get_conn
from core.auth import set_user_password
from services.pdf_generation_queue import get_pdf_queue
from services.messaging_templates import build_email_content
from services.outbound_email import send_tracked_email
from services.tenancy import require_org, get_current_org_context, run_with_org_context
from ingest_conversion_factors import CONVERSION_FACTOR_COLUMNS
from pathlib import Path
import io
import zipfile
import tempfile
import secrets
import string
import re
import json
import os
from datetime import date, datetime, timedelta, timezone
import inspect
import pandas as pd
from threading import Lock
from decimal import Decimal, InvalidOperation
from services.legacy_annual_import import parse_legacy_annual_workbook, commit_legacy_rows, resolve_unresolved_rows
from services.attribute_override_import import (
    build_override_template_workbook,
    commit_override_rows,
    parse_override_workbook,
)
from services.audit_log import ensure_audit_log_table, parse_json_text, record_audit_event
from services.permissions import (
    ACCESS_SCOPES,
    ADMIN_ACCESS_PERMISSION,
    DEFAULT_INTERNAL_ACCESS_SCOPE,
    DEFAULT_PORTAL_ACCESS_SCOPE,
    PERMISSIONS,
    SUPERADMIN_ROLE,
    ensure_permission_schema,
    get_effective_permissions_for_user,
    invalidate_permission_cache,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from typing import Any

router = APIRouter(
    prefix="/admin",
    tags=["admin"],
    dependencies=[Depends(require_permission(ADMIN_ACCESS_PERMISSION))],
)
logger = logging.getLogger(__name__)

_LOOKUP_BOOTSTRAP_LOCK = Lock()
_LOOKUP_BOOTSTRAPPED: set[str] = set()
_ADMIN_USER_BOOTSTRAPPED = False
_ADMIN_USER_BOOTSTRAP_LOCK = Lock()
_ORG_SCOPED_LOOKUP_TABLES = {"job_types", "time_subjects", "portfolios_lookup"}
_ORG_ROLE_RANKS = {
    "owner": 40,
    "admin": 30,
    "billing": 20,
    "member": 10,
    "consultant": 10,
}
_ORG_ROLE_LABELS = {
    "owner": "Owner",
    "admin": "Admin",
    "billing": "Billing",
    "member": "Member",
    "consultant": "Consultant",
}
_ORG_ROLE_CAPABILITIES = {
    "owner": {
        "can_switch": True,
        "can_manage_members": True,
        "can_invite": True,
        "can_manage_organisation": True,
        "can_transfer_ownership": True,
        "can_manage_billing": True,
    },
    "admin": {
        "can_switch": True,
        "can_manage_members": True,
        "can_invite": True,
        "can_manage_organisation": True,
        "can_transfer_ownership": False,
        "can_manage_billing": False,
    },
    "billing": {
        "can_switch": True,
        "can_manage_members": False,
        "can_invite": False,
        "can_manage_organisation": False,
        "can_transfer_ownership": False,
        "can_manage_billing": True,
    },
    "member": {
        "can_switch": True,
        "can_manage_members": False,
        "can_invite": False,
        "can_manage_organisation": False,
        "can_transfer_ownership": False,
        "can_manage_billing": False,
    },
    "consultant": {
        "can_switch": True,
        "can_manage_members": False,
        "can_invite": False,
        "can_manage_organisation": False,
        "can_transfer_ownership": False,
        "can_manage_billing": False,
    },
}
_ORG_MANAGEMENT_ROLES = {"owner", "admin"}
_ORG_SWITCH_ROLES = {"owner", "admin", "billing", "member", "consultant"}
_ORG_BILLING_INVOICE_STATUSES = {"draft", "issued", "paid", "overdue", "void", "refunded"}
_ORG_BILLING_EVENT_TYPES = {
    "invoice_created",
    "invoice_issued",
    "payment_received",
    "payment_failed",
    "subscription_created",
    "subscription_updated",
    "subscription_canceled",
    "renewal",
    "reminder_sent",
    "note",
}



def _build_dataset_template_workbook(country: str, year: int) -> tuple[bytes, str]:
    country_norm = str(country or "").strip()
    if not country_norm:
        raise HTTPException(status_code=400, detail="country is required")

    with get_conn() as con:
        rows = con.execute(
            """
            SELECT
                d.name AS dataset_name,
                d.country AS dataset_country,
                d.year AS dataset_year,
                fl.scope,
                COALESCE(NULLIF(fl.category, ''), NULLIF(fl.level_1, ''), '') AS category,
                COALESCE(NULLIF(fl.report_label, ''), NULLIF(fl.column_text, ''), '') AS report_label,
                fl.original_id,
                COALESCE(NULLIF(fl.uom, ''), '') AS uom,
                COALESCE(NULLIF(fl.ghg_unit, ''), '') AS ghg_unit,
                fl.factor,
                fl.year AS factor_year,
                COALESCE(NULLIF(fl.level_1, ''), '') AS level_1,
                COALESCE(NULLIF(fl.level_2, ''), '') AS level_2,
                COALESCE(NULLIF(fl.level_3, ''), '') AS level_3,
                COALESCE(NULLIF(fl.level_4, ''), '') AS level_4,
                COALESCE(NULLIF(fl.column_text, ''), '') AS column_text,
                COALESCE(NULLIF(fl.source, ''), '') AS source,
                COALESCE(NULLIF(fl.region, ''), '') AS region,
                COALESCE(NULLIF(fl.method, ''), '') AS method
            FROM v_factor_lookup fl
            JOIN datasets d ON d.dataset_id = fl.dataset_id
            WHERE LOWER(TRIM(COALESCE(d.country, ''))) = LOWER(TRIM(%s))
              AND d.year = %s
              AND (d.archived IS NULL OR d.archived = FALSE)
            ORDER BY d.name, fl.scope, COALESCE(NULLIF(fl.category, ''), NULLIF(fl.level_1, ''), ''), COALESCE(NULLIF(fl.report_label, ''), NULLIF(fl.column_text, '')), fl.original_id
            """,
            [country_norm, int(year)],
        ).fetchall()

    if not rows:
        raise HTTPException(status_code=404, detail=f"No dataset rows found for {country_norm} {int(year)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Dataset Template"

    meta_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    even_fill = PatternFill(start_color="F8FBFF", end_color="F8FBFF", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    meta_rows = [
        ("Country", country_norm),
        ("Year", str(int(year))),
        ("Datasets", ", ".join(sorted({str(r[0] or "") for r in rows if r and r[0]}))),
    ]
    for idx, (label, value) in enumerate(meta_rows, start=1):
        ws.cell(row=idx, column=1, value=f"{label}:")
        ws.cell(row=idx, column=2, value=value)
        ws.cell(row=idx, column=1).font = Font(bold=True)
        ws.cell(row=idx, column=1).fill = meta_fill
        ws.cell(row=idx, column=2).fill = meta_fill

    ws["E1"] = "Template:"
    ws["F1"] = "Dataset rows workbook"
    ws["E2"] = "Generated:"
    ws["F2"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["E3"] = "Rows:"
    ws["F3"] = len(rows)
    for cell_ref in ("E1", "E2", "E3"):
        ws[cell_ref].font = Font(bold=True)
        ws[cell_ref].fill = meta_fill
    for cell_ref in ("F1", "F2", "F3"):
        ws[cell_ref].fill = meta_fill

    headers = [
        "Scope",
        "Category",
        "Report Label",
        "ID",
        "UOM",
        "GHG Unit",
        "Factor",
        "Year",
        "Source",
        "Region",
        "Method",
        "Dataset",
    ]
    header_row = 5
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=1):
        excel_row = header_row + row_idx
        row_fill = even_fill if row_idx % 2 == 0 else odd_fill
        values = [
            row[3],  # scope
            row[4],  # category
            row[5],  # report label
            row[6],  # original_id
            row[7],  # uom
            row[8],  # ghg_unit
            row[9],  # factor
            row[10],  # factor year
            row[16],  # source
            row[17],  # region
            row[18],  # method
            row[0],  # dataset_name
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.fill = row_fill
            if col_idx in (7, 8):
                cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A6"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["K"].width = 18
    ws.column_dimensions["L"].width = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    safe_country = re.sub(r"[^A-Za-z0-9._-]+", "_", country_norm).strip("_") or "country"
    filename = f"{safe_country}_{int(year)}_dataset_template.xlsx"
    return buffer.getvalue(), filename


@router.get("/templates/dataset-workbook")
def download_dataset_template_workbook(
    country: str = Query(..., min_length=1),
    year: int = Query(..., ge=1900, le=3000),
    _user: dict = Depends(_current_user),
):
    """Download a country/year workbook built from all dataset factor rows."""
    try:
        data, filename = _build_dataset_template_workbook(country, int(year))
        safe_filename = filename.replace('"', '\\"')
        return StreamingResponse(
            iter([data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to build dataset workbook: {e}")


@router.get("/datasets/blank-upload-template")
def download_blank_upload_template(_user: dict = Depends(_current_user)):
    """Download a blank XLSX template for uploading conversion factors via CSV upload."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Factors Upload"

    note_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    example_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    note_cell = ws.cell(row=1, column=1, value="Required columns: ID, Scope, Factor. Save as CSV (.csv) before uploading via Admin > Datasets.")
    note_cell.fill = note_fill
    ws.merge_cells("A1:R1")

    headers = [c.label for c in CONVERSION_FACTOR_COLUMNS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    example = [
        "1.A.1.a", "Scope 1", "Stationary combustion", "Fuels", "Gas", "", "",
        "Natural gas combustion", "Natural Gas", "kWh", "kgCO2e", 0.18289, 2025,
        "DESNZ", "United Kingdom", "Market-based", "", "",
    ]
    for col_idx, value in enumerate(example, start=1):
        ws.cell(row=3, column=col_idx, value=value).fill = example_fill

    widths = [14, 14, 24, 22, 22, 20, 20, 30, 30, 10, 12, 12, 8, 16, 16, 18, 12, 12]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A3"

    # Presentation-only prose, keyed by canonical field name. Kept separate from
    # CONVERSION_FACTOR_COLUMNS since descriptions/examples aren't parser-relevant
    # (unlike label/required/aliases, which the parser actually consumes).
    column_descriptions = {
        "original_id": "Unique identifier for this factor row",
        "scope": "Emission scope (e.g. Scope 1, Scope 2, Scope 3)",
        "factor": "Numeric emission factor value (kgCO2e per unit)",
        "category": "Top-level category label",
        "level_1": "Sub-category level 1",
        "level_2": "Sub-category level 2",
        "level_3": "Sub-category level 3",
        "level_4": "Sub-category level 4",
        "column_text": "Activity description shown in dropdowns",
        "report_label": "Label shown in reports",
        "uom": "Unit of measure",
        "ghg_unit": "Greenhouse gas unit",
        "year": "Reporting year this factor applies to",
        "source": "Data source name",
        "region": "Geographic region",
        "method": "Calculation method",
        "valid_from": "Date from which factor is valid (YYYY-MM-DD)",
        "valid_to": "Date to which factor is valid (YYYY-MM-DD)",
    }
    column_examples = {
        "original_id": "1.A.1.a",
        "scope": "Scope 1",
        "factor": "0.18289",
        "category": "Stationary combustion",
        "level_1": "Fuels",
        "level_2": "Gas",
        "level_3": "",
        "level_4": "",
        "column_text": "Natural gas combustion",
        "report_label": "Natural Gas",
        "uom": "kWh",
        "ghg_unit": "kgCO2e",
        "year": "2025",
        "source": "DESNZ",
        "region": "United Kingdom",
        "method": "Market-based",
        "valid_from": "2025-01-01",
        "valid_to": "2025-12-31",
    }

    ws2 = wb.create_sheet("Instructions")
    instr_rows = [("Column", "Required?", "Description", "Example")] + [
        (
            c.label,
            "Required" if c.required else "Optional",
            column_descriptions.get(c.field, ""),
            column_examples.get(c.field, ""),
        )
        for c in CONVERSION_FACTOR_COLUMNS
    ]
    req_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    for row_idx, row in enumerate(instr_rows, start=1):
        for col_idx, val in enumerate(row, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
            elif row[1] == "Required":
                cell.fill = req_fill
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 52
    ws2.column_dimensions["D"].width = 24

    buffer = io.BytesIO()
    wb.save(buffer)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="factors_upload_template.xlsx"'},
    )


def _ensure_factor_lookup_schema(con) -> None:
    for column, ddl_type in [
        ("dataset_id", "INTEGER"),
        ("file_name", "VARCHAR"),
        ("year", "INTEGER"),
        ("original_id", "VARCHAR"),
        ("scope", "VARCHAR"),
        ("category", "VARCHAR"),
        ("level_1", "VARCHAR"),
        ("level_2", "VARCHAR"),
        ("level_3", "VARCHAR"),
        ("level_4", "VARCHAR"),
        ("column_text", "VARCHAR"),
        ("report_label", "VARCHAR"),
        ("uom", "VARCHAR"),
        ("ghg_unit", "VARCHAR"),
        ("factor", "DOUBLE"),
        ("source", "VARCHAR"),
        ("region", "VARCHAR"),
        ("currency", "VARCHAR"),
        ("method", "VARCHAR"),
        ("valid_from", "DATE"),
        ("valid_to", "DATE"),
    ]:
        try:
            con.execute(f"ALTER TABLE factor_lookup ADD COLUMN IF NOT EXISTS {column} {ddl_type}")
        except Exception as exc:
            logger.warning("Ignoring factor_lookup schema step for %s: %s", column, exc)


def _serialize_factor_row(row: dict[str, Any]) -> dict[str, Any]:
    def _s(key: str) -> str:
        return str(row.get(key) or "")

    def _dt(key: str) -> str | None:
        value = row.get(key)
        if value is None:
            return None
        try:
            return value.isoformat()
        except Exception:
            text = str(value).strip()
            return text or None

    def _num(key: str) -> float | None:
        value = row.get(key)
        if value is None:
            return None
        try:
            return float(value)
        except Exception:
            return None

    def _int(key: str) -> int | None:
        value = row.get(key)
        if value is None:
            return None
        try:
            return int(value)
        except Exception:
            return None

    return {
        "db_id": _int("db_id"),
        "dataset_id": _int("dataset_id"),
        "dataset": _s("dataset"),
        "analysis_type": _s("analysis_type"),
        "country": _s("country"),
        "year": _int("year"),
        "file_name": _s("file_name"),
        "original_id": _s("original_id"),
        "scope": _s("scope"),
        "category": _s("category"),
        "level_1": _s("level_1"),
        "level_2": _s("level_2"),
        "level_3": _s("level_3"),
        "level_4": _s("level_4"),
        "column_text": _s("column_text"),
        "report_label": _s("report_label") or _s("column_text") or _s("original_id"),
        "uom": _s("uom"),
        "ghg_unit": _s("ghg_unit"),
        "factor": _num("factor"),
        "source": _s("source"),
        "region": _s("region"),
        "currency": _s("currency"),
        "method": _s("method"),
        "valid_from": _dt("valid_from"),
        "valid_to": _dt("valid_to"),
        "factor_definition_id": _int("factor_definition_id"),
        "factor_year_value_id": _int("factor_year_value_id"),
    }


def _load_factor_row(con, db_id: int) -> dict[str, Any] | None:
    row_df = con.execute(
        """
        SELECT
            fl.db_id,
            fl.dataset_id,
            d.name AS dataset,
            d.analysis_type,
            d.country,
            fl.year,
            fl.file_name,
            fl.original_id,
            fl.scope,
            fl.category,
            fl.level_1,
            fl.level_2,
            fl.level_3,
            fl.level_4,
            fl.column_text,
            fl.report_label,
            fl.uom,
            fl.ghg_unit,
            fl.factor,
            fl.source,
            fl.region,
            fl.currency,
            fl.method,
            fl.valid_from,
            fl.valid_to,
            fl.factor_definition_id,
            fl.factor_year_value_id
        FROM v_factor_lookup fl
        LEFT JOIN datasets d ON d.dataset_id = fl.dataset_id
        WHERE fl.db_id = %s
        LIMIT 1
        """,
        [int(db_id)],
    ).df()
    if row_df is None or row_df.empty:
        return None
    return _serialize_factor_row(row_df.iloc[0].to_dict())


def _build_factor_search_filters(
    *,
    q: str = "",
    db_id: int | None = None,
    dataset_id: int | None = None,
    country: str = "",
    year: int | None = None,
    scope: str = "",
    report_label: str = "",
    original_id: str = "",
    category: str = "",
    level_1: str = "",
    level_2: str = "",
    level_3: str = "",
    level_4: str = "",
    column_text: str = "",
    uom: str = "",
    ghg_unit: str = "",
    source: str = "",
    region: str = "",
    method: str = "",
) -> tuple[str, list[Any]]:
    params: list[Any] = []
    where_parts = ["COALESCE(d.archived, FALSE) = FALSE"]

    if q.strip():
        tokens = [part.strip() for part in re.split(r"\s+", q.strip()) if part.strip()]
        token_parts: list[str] = []
        token_params: list[Any] = []
        searchable_sql = """
            (
                CAST(fl.db_id AS VARCHAR) ILIKE %s
                OR CAST(fl.dataset_id AS VARCHAR) ILIKE %s
                OR CAST(d.year AS VARCHAR) ILIKE %s
                OR COALESCE(d.name, '') ILIKE %s
                OR COALESCE(d.country, '') ILIKE %s
                OR COALESCE(d.analysis_type, '') ILIKE %s
                OR COALESCE(fl.original_id, '') ILIKE %s
                OR COALESCE(fl.scope, '') ILIKE %s
                OR COALESCE(fl.category, '') ILIKE %s
                OR COALESCE(fl.level_1, '') ILIKE %s
                OR COALESCE(fl.level_2, '') ILIKE %s
                OR COALESCE(fl.level_3, '') ILIKE %s
                OR COALESCE(fl.level_4, '') ILIKE %s
                OR COALESCE(fl.column_text, '') ILIKE %s
                OR COALESCE(fl.report_label, '') ILIKE %s
                OR COALESCE(fl.uom, '') ILIKE %s
                OR COALESCE(fl.ghg_unit, '') ILIKE %s
                OR COALESCE(fl.source, '') ILIKE %s
                OR COALESCE(fl.region, '') ILIKE %s
                OR COALESCE(fl.method, '') ILIKE %s
            )
        """
        for token in tokens:
            needle = f"%{token}%"
            token_parts.append(searchable_sql)
            token_params.extend([needle] * 20)
        where_parts.append(" AND ".join(token_parts))
        params.extend(token_params)

    if db_id is not None:
        where_parts.append("fl.db_id = %s")
        params.append(int(db_id))
    if dataset_id is not None:
        where_parts.append("fl.dataset_id = %s")
        params.append(int(dataset_id))
    if country.strip():
        where_parts.append("LOWER(TRIM(COALESCE(d.country, ''))) = LOWER(TRIM(%s))")
        params.append(country.strip())
    if year is not None:
        where_parts.append("d.year = %s")
        params.append(int(year))
    if scope.strip():
        where_parts.append("LOWER(TRIM(COALESCE(fl.scope, ''))) = LOWER(TRIM(%s))")
        params.append(scope.strip())
    if report_label.strip():
        where_parts.append("COALESCE(fl.report_label, fl.column_text, '') ILIKE %s")
        params.append(f"%{report_label.strip()}%")
    if original_id.strip():
        where_parts.append("COALESCE(fl.original_id, '') ILIKE %s")
        params.append(f"%{original_id.strip()}%")
    if category.strip():
        where_parts.append("COALESCE(fl.category, fl.level_1, '') ILIKE %s")
        params.append(f"%{category.strip()}%")
    if level_1.strip():
        where_parts.append("COALESCE(fl.level_1, '') ILIKE %s")
        params.append(f"%{level_1.strip()}%")
    if level_2.strip():
        where_parts.append("COALESCE(fl.level_2, '') ILIKE %s")
        params.append(f"%{level_2.strip()}%")
    if level_3.strip():
        where_parts.append("COALESCE(fl.level_3, '') ILIKE %s")
        params.append(f"%{level_3.strip()}%")
    if level_4.strip():
        where_parts.append("COALESCE(fl.level_4, '') ILIKE %s")
        params.append(f"%{level_4.strip()}%")
    if column_text.strip():
        where_parts.append("COALESCE(fl.column_text, '') ILIKE %s")
        params.append(f"%{column_text.strip()}%")
    if uom.strip():
        where_parts.append("COALESCE(fl.uom, '') ILIKE %s")
        params.append(f"%{uom.strip()}%")
    if ghg_unit.strip():
        where_parts.append("COALESCE(fl.ghg_unit, '') ILIKE %s")
        params.append(f"%{ghg_unit.strip()}%")
    if source.strip():
        where_parts.append("COALESCE(fl.source, '') ILIKE %s")
        params.append(f"%{source.strip()}%")
    if region.strip():
        where_parts.append("COALESCE(fl.region, '') ILIKE %s")
        params.append(f"%{region.strip()}%")
    if method.strip():
        where_parts.append("COALESCE(fl.method, '') ILIKE %s")
        params.append(f"%{method.strip()}%")

    return " AND ".join(where_parts), params


def _parse_bulk_terms(raw_terms: Any) -> list[str]:
    if raw_terms is None:
        return []
    if isinstance(raw_terms, str):
        parts = re.split(r"[\n,;|]+", raw_terms)
    elif isinstance(raw_terms, (list, tuple, set)):
        parts = [str(part) for part in raw_terms]
    else:
        parts = [str(raw_terms)]

    terms: list[str] = []
    seen: set[str] = set()
    for part in parts:
        text = str(part or "").strip()
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        terms.append(text)
    return terms


def _normalize_report_label_text(value: str | None, remove_terms: list[str]) -> str:
    text = str(value or "").strip()
    if not text:
        return ""

    for term in remove_terms:
        if not term:
            continue
        text = re.sub(re.escape(term), "", text, flags=re.IGNORECASE)

    text = re.sub(r"\s*-\s*-\s*", " - ", text)
    text = re.sub(r"\s{2,}", " ", text)
    text = re.sub(r"\s*([,;:])\s*", r"\1 ", text)
    text = re.sub(r"\s{2,}", " ", text)
    return text.strip(" -|")


@router.get("/factors")
def search_factors(
    q: str = "",
    db_id: int | None = None,
    dataset_id: int | None = None,
    country: str = "",
    year: int | None = None,
    scope: str = "",
    report_label: str = "",
    original_id: str = "",
    category: str = "",
    level_1: str = "",
    level_2: str = "",
    level_3: str = "",
    level_4: str = "",
    column_text: str = "",
    uom: str = "",
    ghg_unit: str = "",
    source: str = "",
    region: str = "",
    method: str = "",
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=1000),
    _user: dict = Depends(_current_user)
):
    """Search conversion factors (excludes archived datasets)."""
    try:
        where_sql, params = _build_factor_search_filters(
            q=q,
            db_id=db_id,
            dataset_id=dataset_id,
            country=country,
            year=year,
            scope=scope,
            report_label=report_label,
            original_id=original_id,
            category=category,
            level_1=level_1,
            level_2=level_2,
            level_3=level_3,
            level_4=level_4,
            column_text=column_text,
            uom=uom,
            ghg_unit=ghg_unit,
            source=source,
            region=region,
            method=method,
        )

        with get_conn() as con:
            _ensure_factor_lookup_schema(con)

            total_row = con.execute(
                f"""
                SELECT COUNT(*)
                FROM v_factor_lookup fl
                LEFT JOIN datasets d ON d.dataset_id = fl.dataset_id
                WHERE {where_sql}
                """,
                params,
            ).fetchone()
            total = int(total_row[0]) if total_row else 0

            df = con.execute(
                f"""
                SELECT
                    fl.db_id,
                    fl.dataset_id,
                    d.name AS dataset,
                    d.analysis_type,
                    d.country,
                    fl.year,
                    fl.file_name,
                    fl.original_id,
                    fl.scope,
                    fl.category,
                    fl.level_1,
                    fl.level_2,
                    fl.level_3,
                    fl.level_4,
                    fl.column_text,
                    fl.report_label,
                    fl.uom,
                    fl.ghg_unit,
                    fl.factor,
                    fl.source,
                    fl.region,
                    fl.currency,
                    fl.method,
                    fl.valid_from,
                    fl.valid_to,
                    fl.factor_definition_id,
                    fl.factor_year_value_id
                FROM v_factor_lookup fl
                LEFT JOIN datasets d ON d.dataset_id = fl.dataset_id
                WHERE {where_sql}
                ORDER BY d.year DESC NULLS LAST, d.name, COALESCE(fl.category, fl.level_1), fl.column_text, fl.original_id
                LIMIT %s OFFSET %s
                """,
                params + [int(per_page), int((page - 1) * per_page)],
            ).df()

        items: list[dict] = []
        if df is not None and not df.empty:
            # pandas turns SQL NULL numerics into NaN, which is not JSON-serialisable
            # and would otherwise escape the try/except during response encoding.
            safe_df = df.astype(object).where(df.notna(), None)
            for record in safe_df.to_dict(orient="records"):
                items.append(_serialize_factor_row(record))

        pages = (total + per_page - 1) // per_page if per_page else 0
        return {
            "items": items,
            "total": total,
            "page": page,
            "per_page": per_page,
            "pages": pages,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to search factors: {e}")


@router.post("/datasets/{dataset_id}/factors")
def create_factor_row(
    dataset_id: int,
    body: dict = Body(...),
    _user: dict = Depends(_current_user),
):
    try:
        with get_conn() as con:
            _ensure_factor_lookup_schema(con)
            dataset_row = con.execute(
                "SELECT dataset_id, name, analysis_type, country, year FROM datasets WHERE dataset_id = %s",
                [int(dataset_id)],
            ).fetchone()
            if not dataset_row:
                raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")

            original_id = str(body.get("original_id") or "").strip()
            scope = str(body.get("scope") or "").strip()
            factor_val = body.get("factor")
            if not original_id:
                raise HTTPException(status_code=400, detail="original_id is required")
            if not scope:
                raise HTTPException(status_code=400, detail="scope is required")
            if factor_val is None or str(factor_val).strip() == "":
                raise HTTPException(status_code=400, detail="factor is required")

            row = con.execute(
                """
                INSERT INTO factor_lookup (
                    dataset_id, file_name, year, original_id, scope,
                    category, level_1, level_2, level_3, level_4,
                    column_text, report_label, uom, ghg_unit, factor,
                    source, region, currency, method, valid_from, valid_to
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING db_id
                """,
                [
                    int(dataset_id),
                    str(body.get("file_name") or "").strip() or None,
                    int(body.get("year")) if body.get("year") is not None and str(body.get("year")).strip() != "" else int(dataset_row[4]) if dataset_row[4] is not None else None,
                    original_id,
                    scope,
                    str(body.get("category") or "").strip() or None,
                    str(body.get("level_1") or "").strip() or None,
                    str(body.get("level_2") or "").strip() or None,
                    str(body.get("level_3") or "").strip() or None,
                    str(body.get("level_4") or "").strip() or None,
                    str(body.get("column_text") or "").strip() or None,
                    str(body.get("report_label") or "").strip() or None,
                    str(body.get("uom") or "").strip() or None,
                    str(body.get("ghg_unit") or "").strip() or None,
                    float(factor_val),
                    str(body.get("source") or "").strip() or None,
                    str(body.get("region") or "").strip() or None,
                    str(body.get("currency") or "").strip() or None,
                    str(body.get("method") or "").strip() or None,
                    str(body.get("valid_from") or "").strip() or None,
                    str(body.get("valid_to") or "").strip() or None,
                ],
            ).fetchone()
            if not row:
                raise HTTPException(status_code=500, detail="Failed to create factor row")
            factor_row = _load_factor_row(con, int(row[0]))
            if not factor_row:
                raise HTTPException(status_code=500, detail="Failed to load created factor row")
            return {"ok": True, "item": factor_row}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create factor row: {e}")


# Fields stored in emission_factor_definitions (shared across all years for this factor)
_DEFINITION_FIELDS = frozenset({
    "scope", "category", "level_1", "level_2", "level_3", "level_4",
    "report_label", "uom", "ghg_unit", "source", "region", "method", "column_text",
})
# Fields stored in emission_factor_year_values (one value per dataset/year)
_YEAR_VALUE_FIELDS = frozenset({"factor", "year", "original_id", "currency", "valid_from", "valid_to"})
# Fields that always live only in factor_lookup
_LEGACY_ONLY_FIELDS = frozenset({"dataset_id", "file_name"})


@router.patch("/factors/{db_id}")
def update_factor_row(
    db_id: int,
    body: dict = Body(...),
    _user: dict = Depends(_current_user),
):
    try:
        with get_conn() as con:
            _ensure_factor_lookup_schema(con)
            # Fetch the row with its canonical definition and year-value IDs
            existing = con.execute(
                """
                SELECT fl.db_id, efa.factor_id AS factor_definition_id, efyv.factor_year_value_id
                FROM factor_lookup fl
                LEFT JOIN emission_factor_aliases efa
                    ON efa.dataset_id = fl.dataset_id
                    AND TRIM(efa.original_id) = TRIM(fl.original_id)
                LEFT JOIN emission_factor_year_values efyv
                    ON efyv.factor_id = efa.factor_id
                    AND efyv.dataset_id = fl.dataset_id
                    AND efyv.superseded_by IS NULL
                    AND (efyv.year = fl.year OR (efyv.year IS NULL AND fl.year IS NULL))
                WHERE fl.db_id = %s
                """,
                [int(db_id)],
            ).fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail=f"Factor row {db_id} not found")

            factor_definition_id: int | None = existing[1]
            factor_year_value_id: int | None = existing[2]

            # Parse and coerce all submitted fields
            all_keys = list(_DEFINITION_FIELDS) + list(_YEAR_VALUE_FIELDS) + list(_LEGACY_ONLY_FIELDS)
            parsed: dict[str, Any] = {}
            for key in all_keys:
                if key not in body:
                    continue
                value = body[key]
                if key in ("dataset_id", "year") and value is not None and str(value).strip() != "":
                    value = int(value)
                elif key == "factor" and value is not None and str(value).strip() != "":
                    value = float(value)
                else:
                    value = str(value).strip() if value is not None else None
                    if value == "":
                        value = None
                parsed[key] = value

            if not parsed:
                return {"ok": True, "message": "No fields to update"}

            # 1. Write definition-layer fields to emission_factor_definitions (affects ALL years)
            if factor_definition_id:
                def_payload = {k: v for k, v in parsed.items() if k in _DEFINITION_FIELDS}
                if def_payload:
                    set_clause = ", ".join(f"{k} = %s" for k in def_payload)
                    con.execute(
                        f"UPDATE emission_factor_definitions SET {set_clause} WHERE factor_id = %s",
                        list(def_payload.values()) + [factor_definition_id],
                    )

            # 2. Write year-specific numeric fields to emission_factor_year_values
            if factor_year_value_id:
                yv_payload = {k: v for k, v in parsed.items() if k in _YEAR_VALUE_FIELDS}
                if yv_payload:
                    set_clause = ", ".join(f"{k} = %s" for k in yv_payload)
                    con.execute(
                        f"UPDATE emission_factor_year_values SET {set_clause} WHERE factor_year_value_id = %s",
                        list(yv_payload.values()) + [factor_year_value_id],
                    )

            # 3. Write to factor_lookup: always for legacy-only fields; also for any
            #    field whose canonical table doesn't exist yet for this row (unlinked).
            fl_payload: dict[str, Any] = {k: v for k, v in parsed.items() if k in _LEGACY_ONLY_FIELDS}
            if not factor_definition_id:
                fl_payload.update({k: v for k, v in parsed.items() if k in _DEFINITION_FIELDS})
            if not factor_year_value_id:
                fl_payload.update({k: v for k, v in parsed.items() if k in _YEAR_VALUE_FIELDS})
            if fl_payload:
                set_clause = ", ".join(f"{k} = %s" for k in fl_payload)
                con.execute(
                    f"UPDATE factor_lookup SET {set_clause} WHERE db_id = %s",
                    list(fl_payload.values()) + [int(db_id)],
                )

            factor_row = _load_factor_row(con, int(db_id))
            if not factor_row:
                raise HTTPException(status_code=500, detail="Failed to reload updated factor row")
            return {"ok": True, "item": factor_row}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update factor row: {e}")


@router.post("/factors/bulk-normalize-report-labels")
def bulk_normalize_factor_report_labels(
    body: dict = Body(...),
    _user: dict = Depends(_current_user),
):
    try:
        def _optional_int(value: Any) -> int | None:
            if value in (None, ""):
                return None
            try:
                return int(value)
            except Exception as exc:
                raise HTTPException(status_code=400, detail=f"Invalid integer value: {value}") from exc

        remove_terms = _parse_bulk_terms(body.get("remove_terms") or body.get("remove_text"))
        if not remove_terms:
            raise HTTPException(status_code=400, detail="At least one remove term is required")

        dry_run = bool(body.get("dry_run", True))
        preview_limit = int(body.get("preview_limit", 50) or 50)
        if preview_limit < 0:
            preview_limit = 50

        where_sql, params = _build_factor_search_filters(
            q=str(body.get("q") or ""),
            db_id=_optional_int(body.get("db_id")),
            dataset_id=_optional_int(body.get("dataset_id")),
            country=str(body.get("country") or ""),
            year=_optional_int(body.get("year")),
            scope=str(body.get("scope") or ""),
            report_label=str(body.get("report_label") or ""),
            original_id=str(body.get("original_id") or ""),
            category=str(body.get("category") or ""),
            level_1=str(body.get("level_1") or ""),
            level_2=str(body.get("level_2") or ""),
            level_3=str(body.get("level_3") or ""),
            level_4=str(body.get("level_4") or ""),
            column_text=str(body.get("column_text") or ""),
            uom=str(body.get("uom") or ""),
            ghg_unit=str(body.get("ghg_unit") or ""),
            source=str(body.get("source") or ""),
            region=str(body.get("region") or ""),
            method=str(body.get("method") or ""),
        )

        with get_conn() as con:
            _ensure_factor_lookup_schema(con)
            df = con.execute(
                f"""
                SELECT
                    fl.db_id,
                    fl.dataset_id,
                    d.name AS dataset,
                    d.analysis_type,
                    d.country,
                    fl.year,
                    fl.file_name,
                    fl.original_id,
                    fl.scope,
                    fl.category,
                    fl.level_1,
                    fl.level_2,
                    fl.level_3,
                    fl.level_4,
                    fl.column_text,
                    COALESCE(fl.report_label, fl.column_text, fl.original_id, '') AS current_label,
                    fl.report_label,
                    fl.uom,
                    fl.ghg_unit,
                    fl.factor,
                    fl.source,
                    fl.region,
                    fl.currency,
                    fl.method,
                    fl.valid_from,
                    fl.valid_to,
                    fl.factor_definition_id
                FROM v_factor_lookup fl
                LEFT JOIN datasets d ON d.dataset_id = fl.dataset_id
                WHERE {where_sql}
                ORDER BY d.year DESC NULLS LAST, d.name, COALESCE(fl.category, fl.level_1), COALESCE(fl.report_label, fl.column_text, fl.original_id), fl.original_id
                """,
                params,
            ).df()

            preview_items: list[dict[str, Any]] = []
            # (db_id, new_label, factor_definition_id) -- factor_definition_id
            # is None for rows with no emission_factor_definitions match.
            changed_rows: list[tuple[int, str, int | None]] = []
            matched_count = 0
            if df is not None and not df.empty:
                matched_count = int(len(df.index))
                safe_df = df.astype(object).where(df.notna(), None)
                for record in safe_df.to_dict(orient="records"):
                    current_label = str(record.get("current_label") or "").strip()
                    new_label = _normalize_report_label_text(current_label, remove_terms)
                    if not current_label or not new_label or new_label == current_label:
                        continue
                    factor_definition_id = (
                        int(record["factor_definition_id"]) if record.get("factor_definition_id") is not None else None
                    )
                    changed_rows.append((int(record["db_id"]), new_label, factor_definition_id))
                    if len(preview_items) < preview_limit:
                        preview_items.append(
                            {
                                "db_id": int(record["db_id"]),
                                "dataset_id": int(record["dataset_id"]) if record.get("dataset_id") is not None else None,
                                "dataset": str(record.get("dataset") or ""),
                                "country": str(record.get("country") or ""),
                                "year": int(record["year"]) if record.get("year") is not None else None,
                                "scope": str(record.get("scope") or ""),
                                "category": str(record.get("category") or ""),
                                "original_id": str(record.get("original_id") or ""),
                                "current_label": current_label,
                                "new_label": new_label,
                            }
                        )

            updated_count = 0
            if not dry_run and changed_rows:
                # v_factor_lookup.report_label prefers emission_factor_definitions.
                # report_label over factor_lookup.report_label whenever the former
                # is set (sql_migrations/0050_phase2_dual_read_view.sql) -- for the
                # ~91% of rows now backed by a definitions row (confirmed live,
                # 2026-08-03), writing factor_lookup alone left the displayed label
                # completely unchanged. Both tables get the new label so it takes
                # effect regardless of which one the view is currently preferring.
                for db_id_value, new_label, factor_definition_id in changed_rows:
                    con.execute(
                        "UPDATE factor_lookup SET report_label = %s WHERE db_id = %s",
                        [new_label, db_id_value],
                    )
                    if factor_definition_id is not None:
                        con.execute(
                            "UPDATE emission_factor_definitions SET report_label = %s WHERE factor_id = %s",
                            [new_label, factor_definition_id],
                        )
                updated_count = len(changed_rows)

        return {
            "ok": True,
            "dry_run": dry_run,
            "matched_count": matched_count,
            "changed_count": len(changed_rows),
            "updated_count": updated_count,
            "remove_terms": remove_terms,
            "preview": preview_items,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to bulk normalize report labels: {e}")


# =========================
# PORTAL DATA ENTRY BUCKETS (Client Portal Data Entry, Phase 1)
# =========================

@router.get("/portal-data-entry-buckets")
def list_portal_data_entry_buckets(_user: dict = Depends(_current_user)):
    from services.portal_data_entry import BUCKET_KEYS, BUCKET_LABELS, ensure_portal_data_entry_schema

    with get_conn() as con:
        ensure_portal_data_entry_schema(con)
        df = con.execute(
            "SELECT bucket_id, bucket_key, match_category, match_level_1 FROM portal_data_entry_buckets ORDER BY bucket_key, match_category"
        ).df()

    rows = [] if df is None or df.empty else df.where(df.notna(), None).to_dict("records")
    return {
        "buckets": [{"bucket_key": k, "label": BUCKET_LABELS[k]} for k in BUCKET_KEYS],
        "mappings": rows,
    }


@router.post("/portal-data-entry-buckets")
def create_portal_data_entry_bucket_mapping(
    body: dict = Body(...),
    _user: dict = Depends(_current_user),
):
    from services.portal_data_entry import BUCKET_KEYS, ensure_portal_data_entry_schema

    bucket_key = str(body.get("bucket_key") or "").strip()
    match_category = str(body.get("match_category") or "").strip() or None
    match_level_1 = str(body.get("match_level_1") or "").strip() or None
    if bucket_key not in BUCKET_KEYS:
        raise HTTPException(status_code=400, detail=f"bucket_key must be one of {BUCKET_KEYS}")
    if not match_category and not match_level_1:
        raise HTTPException(status_code=400, detail="match_category or match_level_1 is required")

    with get_conn() as con:
        ensure_portal_data_entry_schema(con)
        try:
            row = con.execute(
                """
                INSERT INTO portal_data_entry_buckets (bucket_key, match_category, match_level_1)
                VALUES (%s, %s, %s)
                RETURNING bucket_id
                """,
                [bucket_key, match_category, match_level_1],
            ).fetchone()
        except Exception as e:
            if "ux_portal_data_entry_buckets_category" in str(e):
                raise HTTPException(status_code=409, detail="That category is already mapped to a bucket")
            raise HTTPException(status_code=500, detail=f"Failed to create mapping: {e}")

    return {"ok": True, "bucket_id": int(row[0])}


@router.delete("/portal-data-entry-buckets/{bucket_id}")
def delete_portal_data_entry_bucket_mapping(
    bucket_id: int,
    _user: dict = Depends(_current_user),
):
    with get_conn() as con:
        con.execute("DELETE FROM portal_data_entry_buckets WHERE bucket_id = %s", [int(bucket_id)])
    return {"ok": True}


# =========================
# LOOKUPS MANAGEMENT
# =========================

@router.get("/datasets")
def list_datasets(
    q: str = Query(""),
    country: str = Query(""),
    year: int | None = Query(None),
    archived: str = Query("false"),  # "false" | "true" | "all"
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=1000),
    _user: dict = Depends(_current_user),
):
    """List datasets with factor counts, paginated and filterable."""
    try:
        with get_conn() as con:
            def _table_exists(name: str) -> bool:
                try:
                    row = con.execute(
                        """
                        SELECT 1
                        FROM information_schema.tables
                        WHERE table_schema = 'public'
                          AND table_name = %s
                        LIMIT 1
                        """,
                        [name],
                    ).fetchone()
                    return bool(row)
                except Exception:
                    return False

            def _column_exists(table_name: str, column_name: str) -> bool:
                try:
                    row = con.execute(
                        """
                        SELECT 1
                        FROM information_schema.columns
                        WHERE table_schema = 'public'
                          AND table_name = %s
                          AND column_name = %s
                        LIMIT 1
                        """,
                        [table_name, column_name],
                    ).fetchone()
                    return bool(row)
                except Exception:
                    return False

            dataset_columns = [
                "dataset_id",
                "name",
                "source",
                "analysis_type",
                "country",
                "region",
                "currency",
                "year",
                "version",
                "license",
                "notes",
                "valid_from",
                "valid_to",
                "archived",
                "archived_at",
                "archived_by",
            ]

            select_parts: list[str] = []
            group_parts: list[str] = []
            for column in dataset_columns:
                if _column_exists("datasets", column):
                    select_parts.append(f"d.{column} AS {column}")
                    group_parts.append(f"d.{column}")
                else:
                    select_parts.append(f"NULL AS {column}")

            if _table_exists("factor_lookup"):
                select_parts.append("COUNT(f.db_id) AS factor_count")
                factor_join = "LEFT JOIN v_factor_lookup f ON f.dataset_id = d.dataset_id"
            else:
                select_parts.append("0 AS factor_count")
                factor_join = ""

            group_by_sql = ", ".join(group_parts) if group_parts else "d.dataset_id"
            select_sql = ",\n                       ".join(select_parts)

            where_clauses: list[str] = []
            where_params: list[Any] = []
            if q:
                like = f"%{q.strip().lower()}%"
                where_clauses.append("(LOWER(COALESCE(d.name, '')) LIKE %s OR LOWER(COALESCE(d.source, '')) LIKE %s)")
                where_params.extend([like, like])
            if country:
                where_clauses.append("d.country = %s")
                where_params.append(country)
            if year is not None:
                where_clauses.append("d.year = %s")
                where_params.append(int(year))
            if archived == "true":
                where_clauses.append("d.archived = TRUE")
            elif archived != "all":
                where_clauses.append("(d.archived = FALSE OR d.archived IS NULL)")
            where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

            total_row = con.execute(
                f"SELECT COUNT(*) FROM datasets d {where_sql}",
                where_params,
            ).fetchone()
            total = int(total_row[0] or 0)

            offset = (page - 1) * per_page
            df = con.execute(
                f"""
                SELECT {select_sql}
                FROM datasets d
                {factor_join}
                {where_sql}
                GROUP BY {group_by_sql}
                ORDER BY year DESC NULLS LAST, name
                LIMIT %s OFFSET %s
                """,
                where_params + [per_page, offset],
            ).df()

        items = []
        if df is not None and not df.empty:
            for _, r in df.iterrows():
                items.append({
                    "dataset_id": int(r.get("dataset_id")),
                    "name": str(r.get("name") or ""),
                    "source": str(r.get("source") or ""),
                    "analysis_type": str(r.get("analysis_type") or ""),
                    "country": str(r.get("country") or ""),
                    "region": str(r.get("region") or ""),
                    "currency": str(r.get("currency") or ""),
                    "year": int(r.get("year")) if r.get("year") else None,
                    "version": str(r.get("version") or ""),
                    "valid_from": str(r.get("valid_from")) if r.get("valid_from") else None,
                    "valid_to": str(r.get("valid_to")) if r.get("valid_to") else None,
                    "archived": bool(r.get("archived")) if r.get("archived") is not None else False,
                    "archived_at": str(r.get("archived_at")) if r.get("archived_at") else None,
                    "archived_by": str(r.get("archived_by")) if r.get("archived_by") else None,
                    "factor_count": int(r.get("factor_count") or 0),
                })

        with get_conn() as con:
            countries = [
                r[0] for r in con.execute(
                    "SELECT DISTINCT country FROM datasets WHERE country IS NOT NULL AND TRIM(country) != '' ORDER BY country"
                ).fetchall()
            ]

        return {
            "items": items,
            "total": total,
            "page": page,
            "per_page": per_page,
            "pages": max(1, (total + per_page - 1) // per_page),
            "countries": countries,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list datasets: {e}")


@router.post("/datasets")
def create_dataset(body: dict = Body(...), _user: dict = Depends(_current_user)):
    """Create a new dataset."""
    try:
        name = body.get("name", "").strip()
        source = body.get("source", "").strip()
        
        if not name or not source:
            raise HTTPException(status_code=400, detail="Name and source are required")
        
        with get_conn() as con:
            row = con.execute(
                """
                INSERT INTO datasets
                (name, source, analysis_type, country, region, currency, year, version, license, notes, valid_from, valid_to)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING dataset_id
                """,
                [
                    name,
                    source,
                    body.get("analysis_type", "Activity"),
                    body.get("country", "UK"),
                    body.get("region"),
                    body.get("currency", "GBP"),
                    body.get("year"),
                    body.get("version"),
                    body.get("license"),
                    body.get("notes"),
                    body.get("valid_from"),
                    body.get("valid_to"),
                ],
            ).fetchone()
            
            dataset_id = int(row[0])
        
        return {"ok": True, "dataset_id": dataset_id}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create dataset: {e}")


@router.put("/datasets/{dataset_id}")
def update_dataset(
    dataset_id: int,
    body: dict = Body(...),
    _user: dict = Depends(_current_user)
):
    """Update an existing dataset."""
    try:
        with get_conn() as con:
            # Check if dataset exists
            existing = con.execute(
                "SELECT dataset_id FROM datasets WHERE dataset_id = %s",
                [dataset_id]
            ).fetchone()
            
            if not existing:
                raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
            
            # Update dataset
            con.execute(
                """
                UPDATE datasets
                SET name = %s, source = %s, analysis_type = %s, country = %s,
                    region = %s, currency = %s, year = %s, version = %s,
                    license = %s, notes = %s, valid_from = %s, valid_to = %s
                WHERE dataset_id = %s
                """,
                [
                    body.get("name"),
                    body.get("source"),
                    body.get("analysis_type"),
                    body.get("country"),
                    body.get("region"),
                    body.get("currency"),
                    body.get("year"),
                    body.get("version"),
                    body.get("license"),
                    body.get("notes"),
                    body.get("valid_from"),
                    body.get("valid_to"),
                    dataset_id,
                ]
            )
        
        return {"ok": True, "message": "Dataset updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update dataset: {e}")


@router.patch("/datasets/{dataset_id}/archive")
def archive_dataset(
    dataset_id: int,
    body: dict = Body(...),
    _user: dict = Depends(_current_user)
):
    """Archive or unarchive a dataset."""
    try:
        archived = body.get("archived", True)
        user_email = _user.get("user", "admin")
        
        with get_conn() as con:
            # Check if dataset exists
            existing = con.execute(
                "SELECT dataset_id FROM datasets WHERE dataset_id = %s",
                [dataset_id]
            ).fetchone()
            
            if not existing:
                raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
            
            if archived:
                # Archive the dataset
                con.execute(
                    """
                    UPDATE datasets
                    SET archived = TRUE, archived_at = NOW(), archived_by = %s
                    WHERE dataset_id = %s
                    """,
                    [user_email, dataset_id]
                )
                message = "Dataset archived successfully"
            else:
                # Unarchive the dataset
                con.execute(
                    """
                    UPDATE datasets
                    SET archived = FALSE, archived_at = NULL, archived_by = NULL
                    WHERE dataset_id = %s
                    """,
                    [dataset_id]
                )
                message = "Dataset unarchived successfully"
        
        return {"ok": True, "message": message}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to archive dataset: {e}")


@router.delete("/datasets/{dataset_id}")
def delete_dataset(
    dataset_id: int,
    _user: dict = Depends(_current_user)
):
    """Permanently delete a dataset. Admin only. Should only be used on archived datasets."""
    try:
        with get_conn() as con:
            # Check if dataset exists and is archived
            existing = con.execute(
                "SELECT archived FROM datasets WHERE dataset_id = %s",
                [dataset_id]
            ).fetchone()
            
            if not existing:
                raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
            
            # Delete associated factors first
            con.execute(
                "DELETE FROM factor_lookup WHERE dataset_id = %s",
                [dataset_id]
            )
            
            # Delete the dataset
            con.execute(
                "DELETE FROM datasets WHERE dataset_id = %s",
                [dataset_id]
            )
        
        return {"ok": True, "message": "Dataset permanently deleted"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete dataset: {e}")


@router.get("/datasets/{dataset_id}/export")
def export_dataset(
    dataset_id: int,
    _user: dict = Depends(_current_user)
):
    """Export all factors from a dataset as CSV."""
    try:
        import io
        import csv
        from fastapi.responses import StreamingResponse
        
        with get_conn() as con:
            # Get dataset info
            dataset_row = con.execute(
                "SELECT name, source, analysis_type, year FROM datasets WHERE dataset_id = %s",
                [dataset_id]
            ).fetchone()
            
            if not dataset_row:
                raise HTTPException(status_code=404, detail=f"Dataset {dataset_id} not found")
            
            dataset_name = dataset_row[0]
            
            # Get all factors for this dataset
            df = con.execute(
                """
                SELECT original_id, scope, category, level_1, level_2, level_3, level_4,
                       column_text, report_label, uom, ghg_unit, factor, year,
                       valid_from, valid_to, source, region, method
                FROM v_factor_lookup
                WHERE dataset_id = %s
                ORDER BY scope, COALESCE(category, level_1), level_2, level_3, original_id
                """,
                [dataset_id]
            ).df()
        
        if df is None or df.empty:
            raise HTTPException(status_code=404, detail=f"No factors found for dataset {dataset_id}")
        
        # Convert to CSV
        output = io.StringIO()
        df.to_csv(output, index=False)
        output.seek(0)
        
        # Create filename
        safe_name = dataset_name.replace(" ", "_").replace("/", "-")
        filename = f"{safe_name}_dataset_{dataset_id}.csv"
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export dataset: {e}")


@router.get("/datasets/cross-country-audit")
def cross_country_dataset_audit(_user: dict = Depends(_current_user)):
    """
    Find all job_scope_rows where the row's dataset country doesn't match the
    job's client country — indicates legacy-import factor mis-resolution.
    Also reports rows where factor_db_id IS NULL but dataset_id IS NOT NULL
    (orphaned reference left after a dataset upload deleted stale factors).
    """
    try:
        with get_conn() as con:
            # Orphaned: factor_db_id NULL but dataset_id still set
            orphaned = con.execute(
                """
                SELECT
                    jsr.row_id,
                    jsr.job_id,
                    j.job_number,
                    j.title AS job_title,
                    c.client_name,
                    c.addr_country AS client_country,
                    jsr.dataset_id,
                    d.name AS dataset_name,
                    d.country AS dataset_country,
                    jsr.scope,
                    jsr.original_id,
                    jsr.report_label
                FROM job_scope_rows jsr
                JOIN jobs j ON j.job_id = jsr.job_id
                JOIN clients c ON c.db_id = j.client_db_id
                LEFT JOIN datasets d ON d.dataset_id = jsr.dataset_id
                WHERE jsr.factor_db_id IS NULL
                  AND jsr.dataset_id IS NOT NULL
                  AND COALESCE(jsr.enabled, TRUE) = TRUE
                ORDER BY j.job_number, jsr.row_id
                """
            ).fetchall()

            # Cross-country: dataset country doesn't match client country
            cross_country = con.execute(
                """
                SELECT
                    jsr.row_id,
                    jsr.job_id,
                    j.job_number,
                    j.title AS job_title,
                    c.client_name,
                    c.addr_country AS client_country,
                    jsr.dataset_id,
                    d.name AS dataset_name,
                    d.country AS dataset_country,
                    jsr.scope,
                    jsr.original_id,
                    jsr.report_label
                FROM job_scope_rows jsr
                JOIN jobs j ON j.job_id = jsr.job_id
                JOIN clients c ON c.db_id = j.client_db_id
                JOIN datasets d ON d.dataset_id = jsr.dataset_id
                WHERE COALESCE(jsr.enabled, TRUE) = TRUE
                  AND d.country IS NOT NULL
                  AND c.addr_country IS NOT NULL
                  AND LOWER(TRIM(d.country)) <> LOWER(TRIM(c.addr_country))
                ORDER BY j.job_number, jsr.row_id
                """
            ).fetchall()

            def _row_dict(r):
                return {
                    "row_id": r[0], "job_id": r[1], "job_number": r[2],
                    "job_title": r[3], "client_name": r[4], "client_country": r[5],
                    "dataset_id": r[6], "dataset_name": r[7], "dataset_country": r[8],
                    "scope": r[9], "original_id": r[10], "report_label": r[11],
                }

            orphaned_rows = [_row_dict(r) for r in orphaned]
            cross_rows = [_row_dict(r) for r in cross_country]

            # Summarise by job
            orphaned_jobs: dict = {}
            for r in orphaned_rows:
                key = r["job_number"] or str(r["job_id"])
                orphaned_jobs.setdefault(key, {"job_number": r["job_number"], "job_title": r["job_title"],
                                               "client_name": r["client_name"], "row_count": 0})
                orphaned_jobs[key]["row_count"] += 1

            cross_jobs: dict = {}
            for r in cross_rows:
                key = r["job_number"] or str(r["job_id"])
                cross_jobs.setdefault(key, {"job_number": r["job_number"], "job_title": r["job_title"],
                                            "client_name": r["client_name"], "client_country": r["client_country"],
                                            "datasets": set(), "row_count": 0})
                cross_jobs[key]["datasets"].add(r["dataset_name"] or str(r["dataset_id"]))
                cross_jobs[key]["row_count"] += 1
            for v in cross_jobs.values():
                v["datasets"] = sorted(v["datasets"])

            return {
                "orphaned_dataset_id": {
                    "description": "Rows where factor_db_id is NULL but dataset_id is still set (stale reference)",
                    "total_rows": len(orphaned_rows),
                    "jobs": list(orphaned_jobs.values()),
                },
                "cross_country_mismatch": {
                    "description": "Rows where the row's dataset country does not match the client's country",
                    "total_rows": len(cross_rows),
                    "jobs": [dict(j) for j in cross_jobs.values()],
                },
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Audit failed: {e}")


@router.post("/datasets/fix-cross-country")
def fix_cross_country_mismatches(_user: dict = Depends(_current_user)):
    """
    Re-resolve cross-country mismatched job_scope_rows to the correct country's
    dataset. For each affected row, looks for a factor with the same original_id
    in a dataset matching the client's country and the same year. If found,
    updates factor_db_id and dataset_id. If no match, NULLs both fields.
    """
    try:
        with get_conn() as con:
            rows = con.execute(
                """
                SELECT
                    jsr.row_id,
                    jsr.original_id,
                    jsr.dataset_id AS current_dataset_id,
                    d_current.year AS current_year,
                    c.addr_country AS client_country
                FROM job_scope_rows jsr
                JOIN jobs j ON j.job_id = jsr.job_id
                JOIN clients c ON c.db_id = j.client_db_id
                JOIN datasets d_current ON d_current.dataset_id = jsr.dataset_id
                WHERE COALESCE(jsr.enabled, TRUE) = TRUE
                  AND d_current.country IS NOT NULL
                  AND c.addr_country IS NOT NULL
                  AND LOWER(TRIM(d_current.country)) <> LOWER(TRIM(c.addr_country))
                """
            ).fetchall()

            resolved = 0
            nulled = 0
            for row_id, original_id, _, current_year, client_country in rows:
                # Try to find a matching factor in the correct country / same year
                match = con.execute(
                    """
                    SELECT fl.db_id, fl.dataset_id
                    FROM v_factor_lookup fl
                    JOIN datasets d ON d.dataset_id = fl.dataset_id
                    WHERE fl.original_id = %s
                      AND d.year = %s
                      AND LOWER(TRIM(d.country)) = LOWER(TRIM(%s))
                    ORDER BY fl.db_id DESC
                    LIMIT 1
                    """,
                    [original_id, current_year, client_country],
                ).fetchone()

                if match:
                    con.execute(
                        "UPDATE job_scope_rows SET factor_db_id = %s, dataset_id = %s, updated_at = NOW() WHERE row_id = %s",
                        [int(match[0]), int(match[1]), int(row_id)],
                    )
                    resolved += 1
                else:
                    con.execute(
                        "UPDATE job_scope_rows SET factor_db_id = NULL, dataset_id = NULL, updated_at = NOW() WHERE row_id = %s",
                        [int(row_id)],
                    )
                    nulled += 1

            return {
                "ok": True,
                "total": len(rows),
                "re_resolved": resolved,
                "nulled_no_match": nulled,
                "message": f"Fixed {len(rows)} rows: {resolved} re-resolved to correct country dataset, {nulled} NULLed (no matching factor found).",
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Fix failed: {e}")



