"""Client Portal Data Entry Phase 2 — Purchased Goods & Services.

Clients submit raw ledger lines (GL/Nominal Code, Description, Net Value,
VAT%, optionally Currency/Conversion Rate) manually or via a simple upload,
get a suggested spend category (services.spend reuse), confirm/override it,
and a CRM reviews/approves before the line becomes eligible for
sync_spend_to_scope() -- see CLIENT_PORTAL_DATA_ENTRY_SCOPE.md Phase 2.
"""
from __future__ import annotations

import io
import logging
from typing import Any

import pandas as pd
from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from api.portal_auth_routes import portal_user_dep
from api.spend_data_routes import (
    _ensure_spend_tables,
    _factor_by_id,
    _factor_category_expr,
    _factor_label_expr,
    _gross_from_net,
    _parse_upload,
    _persist_spend_row,
    _safe_float,
    _safe_optional_int,
)
from core.database import get_conn
from services.portal import PORTAL_ROLE_CAN_MANAGE_ACTIONS
from services.portal_data_entry import (
    PORTAL_DATA_ENTRY_EXPIRED_MESSAGE,
    get_job_summary,
    get_portal_data_entry_status,
    get_top_spend_categories,
    resolve_current_job_for_client,
)
from services.spend_line_matching import suggest_spend_lines

logger = logging.getLogger(__name__)
router = APIRouter(tags=["portal-spend"])

_PGS_CATEGORY = "Purchased Goods and Services"

MAX_GL_CODE_LENGTH = 15
MAX_VAT_PCT = 100
MAX_NET_VALUE = 999_999_999


def _validate_spend_line_fields(reference_code: str | None, amount_net: float | None, vat_pct: float | None) -> None:
    if reference_code is not None and len(reference_code) > MAX_GL_CODE_LENGTH:
        raise HTTPException(status_code=400, detail=f"GL / Nominal Code must be {MAX_GL_CODE_LENGTH} characters or fewer")
    if amount_net is not None and not (0 <= amount_net <= MAX_NET_VALUE):
        raise HTTPException(status_code=400, detail=f"Net Value must be between 0 and {MAX_NET_VALUE:,}")
    if vat_pct is not None and not (0 <= vat_pct <= MAX_VAT_PCT):
        raise HTTPException(status_code=400, detail=f"VAT % must be between 0 and {MAX_VAT_PCT}")


def _assert_can_manage(current_user: dict) -> None:
    if current_user.get("role", "ClientAdmin") not in PORTAL_ROLE_CAN_MANAGE_ACTIONS:
        raise HTTPException(status_code=403, detail="Your portal role doesn't allow this action")


def _resolve_job_or_404(con, client_db_id: int) -> int:
    job_id = resolve_current_job_for_client(con, client_db_id)
    if job_id is None:
        raise HTTPException(status_code=404, detail="No open job found for this account yet — contact your NZI consultant")
    return job_id


def _assert_data_entry_open(con, job_id: int) -> None:
    if get_portal_data_entry_status(con, job_id)["portal_data_entry_expired"]:
        raise HTTPException(status_code=403, detail=PORTAL_DATA_ENTRY_EXPIRED_MESSAGE)


@router.get("/portal/spend/rows")
def portal_spend_list_rows(current_user: dict = Depends(portal_user_dep)):
    client_db_id = int(current_user["client_db_id"])
    with get_conn() as con:
        _ensure_spend_tables(con)
        job_id = _resolve_job_or_404(con, client_db_id)
        job_summary = get_job_summary(con, job_id)
        df = con.execute(
            """
            SELECT entry_id, reference_code, spend_description, currency, conversion_rate,
                   amount_net, amount_gross, vat_pct,
                   mapped_scope, mapped_category, mapped_report_label,
                   mapping_status, review_status, review_note, created_at,
                   month_1, month_2, month_3, month_4, month_5, month_6,
                   month_7, month_8, month_9, month_10, month_11, month_12
            FROM job_spend_entries
            WHERE job_id = %s AND COALESCE(is_deleted, FALSE) = FALSE AND submitted_by_portal = TRUE
            ORDER BY entry_id DESC
            """,
            [int(job_id)],
        ).df()
    if df is None or df.empty:
        return {"job_id": job_id, "rows": [], **job_summary}
    # astype(object) first -- see api/portal_data_entry_routes.py for why the
    # plain df.where(df.notna(), None) is a no-op on float64 columns and
    # breaks JSON serialization for rows with a null numeric field.
    df = df.astype(object).where(df.notna(), None)
    rows = [{k: row.get(k) for k in row.index} for _, row in df.iterrows()]
    for r in rows:
        if r.get("created_at") is not None:
            r["created_at"] = str(r["created_at"])
    return {"job_id": job_id, "rows": rows, **job_summary}


@router.get("/portal/spend/template")
def portal_spend_template(current_user: dict = Depends(portal_user_dep)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Spend Data"
    headers = ["GL / Nominal Code", "Description", "Net Value (excl VAT)", "VAT %", "Currency", "Conversion Rate"]
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    ws.cell(row=2, column=1, value="1234")
    ws.cell(row=2, column=2, value="Example: IT consultancy services")
    ws.cell(row=2, column=3, value=1000)
    ws.cell(row=2, column=4, value=20)
    ws.cell(row=2, column=5, value="GBP")
    ws.cell(row=2, column=6, value=1)
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="spend-data-template.xlsx"'},
    )


@router.post("/portal/spend/upload-preview")
async def portal_spend_upload_preview(
    file: UploadFile = File(...),
    current_user: dict = Depends(portal_user_dep),
):
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty upload file")
    with get_conn() as con:
        _ensure_spend_tables(con)
        df = _parse_upload(data, file.filename or "upload.csv")
    return {"count": len(df), "preview": df.head(20).to_dict("records")}


@router.post("/portal/spend/upload-commit")
async def portal_spend_upload_commit(
    file: UploadFile = File(...),
    current_user: dict = Depends(portal_user_dep),
):
    _assert_can_manage(current_user)
    client_db_id = int(current_user["client_db_id"])
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty upload file")

    with get_conn() as con:
        _ensure_spend_tables(con)
        job_id = _resolve_job_or_404(con, client_db_id)
        _assert_data_entry_open(con, job_id)
        df = _parse_upload(data, file.filename or "upload.csv")

        inserted = 0
        skipped: list[dict[str, Any]] = []
        for idx, r in df.iterrows():
            reference_code = str(r.get("reference_code") or "").strip()
            amount_net = _safe_float(r.get("amount_net"), 0.0)
            vat_pct = _safe_float(r.get("vat_pct"), 0.0)
            description = str(r.get("spend_description") or "").strip()
            try:
                _validate_spend_line_fields(reference_code, amount_net, vat_pct)
            except HTTPException as exc:
                skipped.append({"row": int(idx) + 1, "spend_description": description, "reason": exc.detail})
                continue

            _persist_spend_row(
                con=con,
                job_id=int(job_id),
                client_db_id=int(client_db_id),
                site_id=None,
                source_type="portal_upload",
                code_type="nominal_code",
                reference_code=reference_code,
                spend_description=description,
                currency=str(r.get("currency") or "GBP").strip().upper(),
                conversion_currency=str(r.get("conversion_currency") or "GBP").strip().upper(),
                conversion_rate=_safe_float(r.get("conversion_rate"), 1.0),
                amount_net=amount_net,
                vat_pct=vat_pct,
                notes=None,
                actor=str(current_user.get("email") or "portal"),
                submitted_by_portal=True,
            )
            inserted += 1

    return {"ok": True, "job_id": job_id, "inserted": inserted, "skipped": skipped}


@router.post("/portal/spend/rows")
def portal_spend_create_row(
    payload: dict = Body(...),
    current_user: dict = Depends(portal_user_dep),
):
    _assert_can_manage(current_user)
    client_db_id = int(current_user["client_db_id"])

    description = str(payload.get("spend_description") or "").strip()
    if not description:
        raise HTTPException(status_code=400, detail="spend_description is required")
    try:
        amount_net = float(payload.get("amount_net"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="amount_net is required and must be a number")
    reference_code = str(payload.get("reference_code") or "").strip()
    vat_pct = _safe_float(payload.get("vat_pct"), 0.0)
    _validate_spend_line_fields(reference_code, amount_net, vat_pct)

    with get_conn() as con:
        _ensure_spend_tables(con)
        job_id = _resolve_job_or_404(con, client_db_id)
        _assert_data_entry_open(con, job_id)
        saved = _persist_spend_row(
            con=con,
            job_id=int(job_id),
            client_db_id=client_db_id,
            site_id=None,
            source_type="portal_manual",
            code_type="nominal_code",
            reference_code=reference_code,
            spend_description=description,
            currency=str(payload.get("currency") or "GBP").strip().upper(),
            conversion_currency=str(payload.get("currency") or "GBP").strip().upper(),
            conversion_rate=_safe_float(payload.get("conversion_rate"), 1.0),
            amount_net=amount_net,
            vat_pct=vat_pct,
            notes=None,
            actor=str(current_user.get("email") or "portal"),
            submitted_by_portal=True,
        )

    return {"ok": True, "job_id": job_id, "entry_id": saved["entry_id"], "auto_mapped": saved["auto_mapped"]}


@router.get("/portal/spend/categories/top")
def portal_spend_top_categories(current_user: dict = Depends(portal_user_dep)):
    client_db_id = int(current_user["client_db_id"])
    with get_conn() as con:
        _ensure_spend_tables(con)
        items = get_top_spend_categories(con, client_db_id)
    return {"items": items}


@router.get("/portal/spend/rows/{entry_id}/suggest-category")
def portal_spend_suggest_category(entry_id: int, current_user: dict = Depends(portal_user_dep)):
    """Keyword-scores the row's own GL code + description against the
    Admin Centre's curated Spend Lines list -- see
    services/spend_line_matching.py. Returns an empty list (never an
    error) when nothing scores above zero, so the UI can fall back to the
    normal search."""
    client_db_id = int(current_user["client_db_id"])
    with get_conn() as con:
        _ensure_spend_tables(con)
        row = con.execute(
            """
            SELECT e.reference_code, e.spend_description
            FROM job_spend_entries e JOIN jobs j ON j.job_id = e.job_id
            WHERE e.entry_id = %s AND j.client_db_id = %s AND COALESCE(e.is_deleted, FALSE) = FALSE
            """,
            [int(entry_id), client_db_id],
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Spend row not found")
        text = f"{row[1] or ''} {row[0] or ''}"
        items = suggest_spend_lines(con, text)
    return {"entry_id": entry_id, "items": items}


@router.get("/portal/spend/categories/search")
def portal_spend_search_categories(
    q: str = Query("", min_length=0),
    limit: int = Query(25, ge=1, le=100),
    current_user: dict = Depends(portal_user_dep),
):
    client_db_id = int(current_user["client_db_id"])
    with get_conn() as con:
        _ensure_spend_tables(con)
        job_id = _resolve_job_or_404(con, client_db_id)
        category_expr = _factor_category_expr(con, "f")
        label_expr = _factor_label_expr(con, "f")
        df = con.execute(
            f"""
            SELECT f.db_id, f.dataset_id, f.original_id, f.scope, {category_expr} AS category, {label_expr} AS report_label
            FROM v_factor_lookup f
            LEFT JOIN datasets d ON d.dataset_id = f.dataset_id
            WHERE (d.archived IS NULL OR d.archived = FALSE)
              AND {category_expr} = %s
              AND (
                   %s = ''
                   OR {label_expr} ILIKE %s
                   OR f.column_text ILIKE %s
                   OR f.original_id ILIKE %s
              )
            ORDER BY {label_expr}
            LIMIT %s
            """,
            [_PGS_CATEGORY, q, f"%{q}%", f"%{q}%", f"%{q}%", int(limit)],
        ).df()

    items: list[dict[str, Any]] = []
    if df is not None and not df.empty:
        for _, r in df.iterrows():
            items.append(
                {
                    "db_id": int(r["db_id"]),
                    "original_id": r.get("original_id"),
                    "scope": r.get("scope"),
                    "category": r.get("category"),
                    "report_label": r.get("report_label"),
                }
            )
    return {"job_id": job_id, "items": items}


@router.post("/portal/spend/rows/{entry_id}/confirm-category")
def portal_spend_confirm_category(
    entry_id: int,
    payload: dict = Body(...),
    current_user: dict = Depends(portal_user_dep),
):
    """The client's own confirm/override action. Deliberately does NOT call
    _upsert_client_mapping -- that memory-write is deferred to CRM approval
    (PATCH /jobs/{job_id}/spend-data/{entry_id}/review) so a client's mistake
    can't poison next year's auto-suggestions before anyone has checked it."""
    _assert_can_manage(current_user)
    client_db_id = int(current_user["client_db_id"])
    factor_db_id = _safe_optional_int(payload.get("factor_db_id"))
    if not factor_db_id:
        raise HTTPException(status_code=400, detail="factor_db_id is required")

    with get_conn() as con:
        _ensure_spend_tables(con)
        row = con.execute(
            """
            SELECT e.entry_id, e.amount_net, e.vat_pct, e.job_id
            FROM job_spend_entries e
            JOIN jobs j ON j.job_id = e.job_id
            WHERE e.entry_id = %s AND j.client_db_id = %s AND COALESCE(e.is_deleted, FALSE) = FALSE
            """,
            [int(entry_id), client_db_id],
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Spend row not found")
        _assert_data_entry_open(con, int(row[3]))

        factor = _factor_by_id(con, factor_db_id)
        if not factor:
            raise HTTPException(status_code=404, detail="Category not found")
        if str(factor.get("category") or "") != _PGS_CATEGORY:
            raise HTTPException(status_code=400, detail="That isn't a Purchased Goods and Services category")

        amount = _gross_from_net(_safe_float(row[1], 0.0), _safe_float(row[2], 0.0))
        emissions = amount * _safe_float(factor.get("factor"), 0.0)
        con.execute(
            """
            UPDATE job_spend_entries
            SET dataset_id=%s, factor_db_id=%s, factor_original_id=%s,
                mapped_scope=%s, mapped_category=%s, mapped_report_label=%s,
                mapping_status='mapped', mapping_confidence='High',
                mapped_by=%s, mapped_at=NOW(),
                estimated_emissions_tco2e=%s,
                review_status='pending_review', review_note=NULL, updated_at=NOW()
            WHERE entry_id=%s
            """,
            [
                factor.get("dataset_id"), factor.get("db_id"), factor.get("original_id"),
                factor.get("scope"), factor.get("category"), factor.get("report_label"),
                str(current_user.get("email") or "portal"),
                emissions,
                int(entry_id),
            ],
        )

    return {"ok": True, "entry_id": int(entry_id), "review_status": "pending_review"}


@router.patch("/portal/spend/rows/{entry_id}")
def portal_spend_update_row(
    entry_id: int,
    payload: dict = Body(...),
    current_user: dict = Depends(portal_user_dep),
):
    _assert_can_manage(current_user)
    client_db_id = int(current_user["client_db_id"])

    with get_conn() as con:
        _ensure_spend_tables(con)
        existing = con.execute(
            """
            SELECT e.entry_id, e.review_status, e.job_id
            FROM job_spend_entries e JOIN jobs j ON j.job_id = e.job_id
            WHERE e.entry_id = %s AND j.client_db_id = %s
              AND e.submitted_by_portal = TRUE AND COALESCE(e.is_deleted, FALSE) = FALSE
            """,
            [int(entry_id), client_db_id],
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Spend row not found")
        if existing[1] == "approved":
            raise HTTPException(status_code=409, detail="This row has already been approved and can no longer be edited here")
        _assert_data_entry_open(con, int(existing[2]))

        if "reference_code" in payload:
            _validate_spend_line_fields(str(payload.get("reference_code") or "").strip(), None, None)
        if "amount_net" in payload:
            _validate_spend_line_fields(None, _safe_float(payload.get("amount_net"), 0.0), None)
        if "vat_pct" in payload:
            _validate_spend_line_fields(None, None, _safe_float(payload.get("vat_pct"), 0.0))

        set_clauses: list[str] = []
        params: list[Any] = []
        for field in ["reference_code", "spend_description", "currency", "notes"]:
            if field in payload:
                set_clauses.append(f"{field} = %s")
                params.append(str(payload.get(field) or "").strip() or None)
        if "amount_net" in payload:
            set_clauses.append("amount_net = %s")
            params.append(_safe_float(payload.get("amount_net"), 0.0))
        if "vat_pct" in payload:
            set_clauses.append("vat_pct = %s")
            params.append(_safe_float(payload.get("vat_pct"), 0.0))
        if "conversion_rate" in payload:
            set_clauses.append("conversion_rate = %s")
            params.append(_safe_float(payload.get("conversion_rate"), 1.0))
        for i in range(1, 13):
            field = f"month_{i}"
            if field in payload:
                set_clauses.append(f"{field} = %s")
                value = payload.get(field)
                params.append(_safe_float(value, 0.0) if value not in (None, "") else None)

        if not set_clauses:
            return {"ok": True, "entry_id": entry_id, "review_status": existing[1]}

        set_clauses.append("review_status = 'pending_review'")
        set_clauses.append("review_note = NULL")
        set_clauses.append("updated_at = NOW()")
        params.append(int(entry_id))
        con.execute(f"UPDATE job_spend_entries SET {', '.join(set_clauses)} WHERE entry_id = %s", params)

        if "amount_net" in payload or "vat_pct" in payload or "conversion_rate" in payload:
            row = con.execute(
                "SELECT amount_net, vat_pct, factor_db_id FROM job_spend_entries WHERE entry_id = %s",
                [int(entry_id)],
            ).fetchone()
            net = _safe_float(row[0], 0.0)
            vat = _safe_float(row[1], 0.0)
            gross = _gross_from_net(net, vat)
            emissions = None
            if row[2]:
                factor = _factor_by_id(con, int(row[2]))
                if factor:
                    emissions = gross * _safe_float(factor.get("factor"), 0.0)
            con.execute(
                "UPDATE job_spend_entries SET amount_gross = %s, estimated_emissions_tco2e = %s, updated_at = NOW() WHERE entry_id = %s",
                [gross, emissions, int(entry_id)],
            )

    return {"ok": True, "entry_id": entry_id, "review_status": "pending_review"}


@router.delete("/portal/spend/rows/{entry_id}")
def portal_spend_delete_row(entry_id: int, current_user: dict = Depends(portal_user_dep)):
    _assert_can_manage(current_user)
    client_db_id = int(current_user["client_db_id"])

    with get_conn() as con:
        _ensure_spend_tables(con)
        existing = con.execute(
            """
            SELECT e.entry_id, e.review_status, e.job_id
            FROM job_spend_entries e JOIN jobs j ON j.job_id = e.job_id
            WHERE e.entry_id = %s AND j.client_db_id = %s
              AND e.submitted_by_portal = TRUE AND COALESCE(e.is_deleted, FALSE) = FALSE
            """,
            [int(entry_id), client_db_id],
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Spend row not found")
        if existing[1] == "approved":
            raise HTTPException(status_code=409, detail="This row has already been approved and can no longer be deleted here")
        _assert_data_entry_open(con, int(existing[2]))

        con.execute(
            "UPDATE job_spend_entries SET is_deleted = TRUE, updated_at = NOW() WHERE entry_id = %s",
            [int(entry_id)],
        )

    return {"ok": True, "entry_id": entry_id}
