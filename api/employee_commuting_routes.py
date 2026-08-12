"""
Employee commuting and working-from-home template workflow.

This route set provides:
- a downloadable workbook template
- upload preview/validation
- commit into job_scope_rows for workbook uploads and into the source register for direct entries
- a lightweight summary for the Jobs -> Data screen
"""

from __future__ import annotations

import io
import json
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from api.auth import _current_user
from api.job_scope_data_routes import (
    _additional_dataset_ids,
    _ensure_job_scope_rows_schema,
    _legacy_scope_dataset_map,
)
from api.job_emission_register_routes import _ensure_schema as _ensure_emission_register_schema
from core.database import get_conn
from services.audit_log import record_audit_event
from services.dataset_selector import get_applicable_datasets, get_scope_primary_datasets
from services.download_filenames import build_download_filename
from services.vehicle_categorization import categorize_vehicle
from services.vehicle_lookup import lookup_vehicle_by_registration
from services.virus_scan import VirusScanError, scan_bytes

router = APIRouter()


COMMUTING_DATA_SOURCE = "Employee Commuting Template"
DIRECT_COMMUTING_DATA_SOURCE = "Employee Commuting Direct Entry"
TEMPLATE_VERSION = "Employee Commuting Template v2"
WEEKS_PER_YEAR_DEFAULT = 48
COMMUTING_SHEET = "Employee Commuting"
WFH_SHEET = "Working From Home"
LOOKUPS_SHEET = "Lookups"
HEADER_ROW = 9
DATA_START_ROW = HEADER_ROW + 1
BLANK_TEMPLATE_ROWS = 120
MILES_TO_KM = 1.609344

COMMUTE_HEADERS = [
    "Employee / Team",
    "Commute Mode",
    "Vehicle / Service Type",
    "Distance Unit",
    "One-Way Distance",
    "Office Days / Week",
    "Weeks / Year",
    "Annual Distance",
    "Notes",
]

WFH_HEADERS = [
    "Employee / Team",
    "Annual WFH Days",
    "Hours Per Day",
    "Annual WFH Hours",
    "Notes",
]

COMMUTE_MODE_OPTIONS = [
    "Car - Petrol",
    "Car - Diesel",
    "Car - Hybrid",
    "Car - Electric",
    "Motorbike",
    "Taxi",
    "Bus",
    "Rail",
    "Ferry",
    "Walking",
    "Cycling",
]

SERVICE_TYPE_OPTIONS = [
    "Average",
    "Small",
    "Medium",
    "Large",
    "Regular",
    "Black",
    "Local Bus",
    "Local Bus (London)",
    "National",
    "International",
    "Light / Tram",
    "London Underground",
    "Foot Passenger",
    "Car Passenger",
]

UNIT_OPTIONS = ["miles", "km"]

MODE_ALIASES = {
    "car petrol": "car - petrol",
    "car - petrol": "car - petrol",
    "petrol car": "car - petrol",
    "car diesel": "car - diesel",
    "car - diesel": "car - diesel",
    "diesel car": "car - diesel",
    "car hybrid": "car - hybrid",
    "car - hybrid": "car - hybrid",
    "hybrid car": "car - hybrid",
    "car electric": "car - electric",
    "car - electric": "car - electric",
    "electric car": "car - electric",
    "motorbike": "motorbike",
    "motor bike": "motorbike",
    "motorcycle": "motorbike",
    "taxi": "taxis",
    "taxis": "taxis",
    "cab": "taxis",
    "bus": "bus",
    "rail": "rail",
    "train": "rail",
    "ferry": "ferry",
    "walking": "walking",
    "walk": "walking",
    "cycling": "cycling",
    "cycle": "cycling",
    "bike": "cycling",
    "bicycle": "cycling",
}

VARIANT_ALIASES = {
    "": "",
    "average": "average",
    "small": "small",
    "medium": "medium",
    "large": "large",
    "regular": "regular",
    "black": "black",
    "local bus": "local bus",
    "local bus london": "local bus (london)",
    "local bus (london)": "local bus (london)",
    "national": "national",
    "international": "international",
    "light tram": "light/tram",
    "light / tram": "light/tram",
    "light/tram": "light/tram",
    "tram": "light/tram",
    "london underground": "london underground",
    "tube": "london underground",
    "foot passenger": "foot passenger",
    "car passenger": "car passenger",
}

UNIT_ALIASES = {
    "mi": "miles",
    "mile": "miles",
    "miles": "miles",
    "km": "km",
    "kilometre": "km",
    "kilometres": "km",
    "kilometer": "km",
    "kilometers": "km",
    "passenger km": "passenger.km",
    "passenger kilometer": "passenger.km",
    "passenger kilometers": "passenger.km",
    "passenger kilometre": "passenger.km",
    "passenger kilometres": "passenger.km",
    "passenger.km": "passenger.km",
    "pkm": "passenger.km",
}

LOOKUP_PATH = (
    Path(__file__).resolve().parents[1]
    / "wfm_import"
    / "analysis"
    / "wfm_template_id_lookup.json"
)


def _norm_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9()]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _safe_float(value: Any, default: float | None = None) -> float | None:
    if value is None:
        return default
    if isinstance(value, bool):
        return float(value)
    try:
        text = str(value).replace(",", "").strip()
        if not text:
            return default
        out = float(text)
        if not math.isfinite(out):
            return default
        return out
    except Exception:
        return default


def _safe_int(value: Any) -> int | None:
    out = _safe_float(value)
    if out is None:
        return None
    try:
        return int(out)
    except Exception:
        return None


def _safe_str(value: Any) -> str:
    return str(value or "").strip()


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _fmt_period_part(value: Any) -> str:
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d-%b-%Y")
    text = str(value).strip()
    try:
        return datetime.fromisoformat(text[:10]).strftime("%d-%b-%Y")
    except Exception:
        return text


def _load_lookup_payload() -> tuple[dict[tuple[str, str, str], str], str | None]:
    payload = json.loads(LOOKUP_PATH.read_text(encoding="utf-8"))
    preferred = payload.get("preferred_items") or {}
    commute_map: dict[tuple[str, str, str], str] = {}
    wfh_original_id: str | None = None

    for raw_key, raw_original_id in preferred.items():
        if not isinstance(raw_key, str) or not isinstance(raw_original_id, str):
            continue

        parts = raw_key.split("|")
        if len(parts) < 8:
            continue

        bucket = _norm_text(parts[0])
        activity = _norm_text(parts[1])

        if bucket == "employee commuting":
            activity = MODE_ALIASES.get(_norm_text(parts[1]), _norm_text(parts[1]))
            variant = VARIANT_ALIASES.get(_norm_text(parts[6]), _norm_text(parts[6]))
            unit = UNIT_ALIASES.get(_norm_text(parts[7]), _norm_text(parts[7]))
            commute_map[(activity, variant, unit)] = raw_original_id
            continue

        if bucket == "detailed carbon" and activity == "wfh electricity":
            wfh_original_id = raw_original_id

    return commute_map, wfh_original_id


COMMUTE_FACTOR_MAP, WFH_ORIGINAL_ID = _load_lookup_payload()


def _canonical_mode(value: Any) -> str:
    return MODE_ALIASES.get(_norm_text(value), "")


def _canonical_variant(value: Any) -> str:
    return VARIANT_ALIASES.get(_norm_text(value), "")


def _canonical_unit(value: Any) -> str:
    return UNIT_ALIASES.get(_norm_text(value), "")


def _distance_unit_base(value: Any) -> str:
    unit = _canonical_unit(value) or _norm_text(value)
    if unit in {"passenger.km", "passenger km", "pkm"}:
        return "km"
    return unit


def _hours_equivalent_unit(value: Any) -> bool:
    """True if `value` describes an hours-denominated rate -- e.g. the WFH
    factor's uom "per FTE Working Hour", which is numerically 1:1 with plain
    "hours", just phrased descriptively in the source data. _convert_quantity
    only understands distance units (miles/km/passenger.km), so calling it
    for the WFH "hours" -> "per FTE Working Hour" pair always returned None
    and made every WFH submission unresolvable -- confirmed live 2026-08."""
    return "hour" in _norm_text(value)


def _default_variant_for_mode(mode: str) -> str | None:
    if mode.startswith("car - "):
        return "average"
    if mode == "bus":
        return "average"
    if mode == "taxis":
        return "regular"
    if mode in {"motorbike", "walking", "cycling"}:
        return ""
    return None


def _default_unit_for_mode(mode: str) -> str | None:
    if mode.startswith("car - ") or mode == "motorbike":
        return "miles"
    if mode in {"taxis", "bus", "rail", "ferry", "walking", "cycling"}:
        return "km"
    return None


def _table_columns(con, table_name: str) -> set[str]:
    try:
        rows = con.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = current_schema()
              AND table_name = %s
            """,
            [str(table_name)],
        ).fetchall()
        return {str(row[0]).strip().lower() for row in (rows or [])}
    except Exception:
        return set()


def _factor_select_parts(con) -> dict[str, str]:
    cols = _table_columns(con, "factor_lookup")
    has = lambda col: col in cols
    return {
        "category": (
            "COALESCE(category, level_2, level_1, 'Employee Commuting')"
            if has("category")
            else "COALESCE(level_2, level_1, 'Employee Commuting')"
        ),
        "level_4": "level_4" if has("level_4") else "NULL::text",
        "report_label": "report_label" if has("report_label") else "NULL::text",
        "ghg_unit": "ghg_unit" if has("ghg_unit") else "NULL::text",
    }


def _job_meta(con, job_id: int) -> dict[str, Any]:
    row = con.execute(
        """
        SELECT
          j.job_id,
          j.job_number,
          j.reporting_year,
          j.reporting_period_start,
          j.reporting_period_end,
          j.client_db_id,
          c.client_name
        FROM jobs j
        LEFT JOIN clients c ON c.db_id = j.client_db_id
        WHERE j.job_id = %s
        LIMIT 1
        """,
        [int(job_id)],
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Job not found")
    return {
        "job_id": int(row[0]),
        "job_number": str(row[1] or f"Job-{job_id}"),
        "reporting_year": row[2],
        "reporting_period_start": row[3],
        "reporting_period_end": row[4],
        "client_db_id": int(row[5]) if row[5] is not None else None,
        "client_name": str(row[6] or "Client"),
    }


def _job_employee_count(con, job_id: int) -> int | None:
    """
    Resolve the job headcount from intensity metrics first, then legacy CRP details.
    """
    try:
        row = con.execute(
            "SELECT intensity_metrics FROM jobs WHERE job_id = %s",
            [int(job_id)],
        ).fetchone()
        if row and isinstance(row[0], dict):
            employees_metric = row[0].get("employees", {})
            if isinstance(employees_metric, dict):
                employee_value = _safe_int(employees_metric.get("value"))
                if employee_value is not None and employee_value > 0:
                    return int(employee_value)
    except Exception:
        pass

    try:
        row = con.execute(
            "SELECT num_employees FROM crp_job_details WHERE job_id = %s",
            [int(job_id)],
        ).fetchone()
        employee_value = _safe_int(row[0]) if row else None
        if employee_value is not None and employee_value > 0:
            return int(employee_value)
    except Exception:
        pass

    return None


def _effective_employee_count(
    con,
    job_id: int,
    override_employee_count: int | None,
) -> int | None:
    if override_employee_count is not None and override_employee_count > 0:
        return int(override_employee_count)
    return _job_employee_count(con, int(job_id))


def _job_site_label(con, job_id: int, site_id: int | None) -> tuple[int | None, str]:
    if site_id is None:
        return None, "All_Staff"

    row = con.execute(
        """
        SELECT s.site_id, s.site_name
        FROM jobs j
        JOIN client_sites s ON s.client_db_id = j.client_db_id
        WHERE j.job_id = %s
          AND s.site_id = %s
        LIMIT 1
        """,
        [int(job_id), int(site_id)],
    ).fetchone()
    if not row:
        raise HTTPException(status_code=400, detail="Selected site is not linked to this job")
    return int(row[0]), str(row[1] or f"Site-{site_id}")


def _template_filename(meta: dict[str, Any], site_label: str) -> str:
    return build_download_filename(
        job_number=meta.get("job_number") or f"Job-{meta.get('job_id')}",
        client_name=meta.get("client_name") or "Client",
        descriptor="Employee Commuting",
        period_start=meta.get("reporting_period_start"),
        period_end=meta.get("reporting_period_end"),
        reporting_year=meta.get("reporting_year"),
    )


def _apply_header_row(ws, row_number: int, headers: list[str]) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_number, column=idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill


def _set_column_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _add_list_validation(ws, cell_range: str, formula: str) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def _build_template_workbook(meta: dict[str, Any], site_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = COMMUTING_SHEET
    wfh = wb.create_sheet(WFH_SHEET)
    lookups = wb.create_sheet(LOOKUPS_SHEET)

    title_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    for target, title, intro in [
        (
            ws,
            "Employee Commuting",
            "Add one row per employee or team. Leave blank rows empty. Annual Distance is calculated automatically.",
        ),
        (
            wfh,
            "Working From Home",
            "Add one row per employee or team. Enter annual WFH days and the sheet will calculate annual hours.",
        ),
    ]:
        target["A1"] = title
        target["A1"].font = Font(bold=True, size=14)
        target["A2"] = intro
        target["A4"] = "Client"
        target["B4"] = meta.get("client_name")
        target["A5"] = "Job Number"
        target["B5"] = meta.get("job_number")
        target["A6"] = "Site"
        target["B6"] = site_label
        target["A7"] = "Reporting Period"
        target["B7"] = (
            f"{_fmt_period_part(meta.get('reporting_period_start'))}"
            f" to "
            f"{_fmt_period_part(meta.get('reporting_period_end'))}"
        ).strip()
        target["D4"] = "Template"
        target["E4"] = TEMPLATE_VERSION
        for ref_cell in ("A4", "A5", "A6", "A7", "D4"):
            target[ref_cell].font = Font(bold=True)
            target[ref_cell].fill = title_fill

    _apply_header_row(ws, HEADER_ROW, COMMUTE_HEADERS)
    _apply_header_row(wfh, HEADER_ROW, WFH_HEADERS)

    _set_column_widths(
        ws,
        {
            "A": 24,
            "B": 20,
            "C": 24,
            "D": 16,
            "E": 18,
            "F": 18,
            "G": 16,
            "H": 18,
            "I": 36,
        },
    )
    _set_column_widths(
        wfh,
        {
            "A": 24,
            "B": 18,
            "C": 16,
            "D": 18,
            "E": 36,
        },
    )

    lookups["A1"] = "Commute Modes"
    for row_num, value in enumerate(COMMUTE_MODE_OPTIONS, start=2):
        lookups.cell(row=row_num, column=1, value=value)

    lookups["B1"] = "Service Types"
    for row_num, value in enumerate(SERVICE_TYPE_OPTIONS, start=2):
        lookups.cell(row=row_num, column=2, value=value)

    lookups["C1"] = "Distance Units"
    for row_num, value in enumerate(UNIT_OPTIONS, start=2):
        lookups.cell(row=row_num, column=3, value=value)

    lookups.sheet_state = "hidden"

    commute_end_row = DATA_START_ROW + BLANK_TEMPLATE_ROWS - 1
    wfh_end_row = DATA_START_ROW + BLANK_TEMPLATE_ROWS - 1

    _add_list_validation(ws, f"B{DATA_START_ROW}:B{commute_end_row}", f"={LOOKUPS_SHEET}!$A$2:$A${len(COMMUTE_MODE_OPTIONS) + 1}")
    _add_list_validation(ws, f"C{DATA_START_ROW}:C{commute_end_row}", f"={LOOKUPS_SHEET}!$B$2:$B${len(SERVICE_TYPE_OPTIONS) + 1}")
    _add_list_validation(ws, f"D{DATA_START_ROW}:D{commute_end_row}", f"={LOOKUPS_SHEET}!$C$2:$C${len(UNIT_OPTIONS) + 1}")

    for row_num in range(DATA_START_ROW, commute_end_row + 1):
        ws.cell(row=row_num, column=7, value=WEEKS_PER_YEAR_DEFAULT)
        ws.cell(
            row=row_num,
            column=8,
            value=f'=IF(OR(E{row_num}="",F{row_num}="",G{row_num}=""),"",E{row_num}*2*F{row_num}*G{row_num})',
        )

    for row_num in range(DATA_START_ROW, wfh_end_row + 1):
        wfh.cell(row=row_num, column=3, value=7.5)
        wfh.cell(
            row=row_num,
            column=4,
            value=f'=IF(OR(B{row_num}="",C{row_num}=""),"",B{row_num}*C{row_num})',
        )

    ws.freeze_panes = f"A{DATA_START_ROW}"
    wfh.freeze_panes = f"A{DATA_START_ROW}"

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _sheet_row_empty(values: list[Any]) -> bool:
    return all(_is_blank(value) for value in values)


def _commuting_row_has_user_input(values: list[Any]) -> bool:
    # Ignore helper/default columns (weeks/year and calculated annual distance)
    # so unused template rows remain blank to the parser.
    user_columns = [0, 1, 2, 3, 4, 5, 8]
    return any(not _is_blank(values[idx]) for idx in user_columns if idx < len(values))


def _wfh_row_has_user_input(values: list[Any]) -> bool:
    # Ignore helper/default columns (hours/day and calculated annual hours)
    # so unused template rows remain blank to the parser.
    user_columns = [0, 1, 4]
    return any(not _is_blank(values[idx]) for idx in user_columns if idx < len(values))


def _parse_commuting_sheet(ws) -> list[dict[str, Any]]:
    parsed: list[dict[str, Any]] = []
    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        values = [ws.cell(row=row_num, column=col_idx).value for col_idx in range(1, 10)]
        if not _commuting_row_has_user_input(values):
            continue

        employee_name = _safe_str(values[0])
        mode_value = _safe_str(values[1])
        service_value = _safe_str(values[2])
        unit_value = _safe_str(values[3])
        one_way_distance = _safe_float(values[4])
        office_days = _safe_float(values[5])
        weeks_per_year = _safe_float(values[6])
        annual_distance = _safe_float(values[7])
        notes = _safe_str(values[8])

        if annual_distance is None and None not in (one_way_distance, office_days, weeks_per_year):
            annual_distance = float(one_way_distance) * 2.0 * float(office_days) * float(weeks_per_year)

        parsed.append(
            {
                "sheet": COMMUTING_SHEET,
                "row_number": row_num,
                "row_type": "commuting",
                "employee_name": employee_name,
                "mode_value": mode_value,
                "service_value": service_value,
                "unit_value": unit_value,
                "annual_quantity": annual_distance,
                "notes": notes,
                "one_way_distance": one_way_distance,
                "office_days": office_days,
                "weeks_per_year": weeks_per_year,
            }
        )
    return parsed


def _parse_wfh_sheet(ws) -> list[dict[str, Any]]:
    parsed: list[dict[str, Any]] = []
    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        values = [ws.cell(row=row_num, column=col_idx).value for col_idx in range(1, 6)]
        if not _wfh_row_has_user_input(values):
            continue

        employee_name = _safe_str(values[0])
        annual_days = _safe_float(values[1])
        hours_per_day = _safe_float(values[2])
        annual_hours = _safe_float(values[3])
        notes = _safe_str(values[4])

        if annual_hours is None and annual_days is not None and hours_per_day is not None:
            annual_hours = float(annual_days) * float(hours_per_day)

        parsed.append(
            {
                "sheet": WFH_SHEET,
                "row_number": row_num,
                "row_type": "wfh",
                "employee_name": employee_name,
                "annual_quantity": annual_hours,
                "annual_days": annual_days,
                "hours_per_day": hours_per_day,
                "notes": notes,
            }
        )
    return parsed


def _parse_template(raw_bytes: bytes) -> list[dict[str, Any]]:
    try:
        wb = load_workbook(io.BytesIO(raw_bytes), data_only=False)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read workbook: {exc}")

    if COMMUTING_SHEET not in wb.sheetnames or WFH_SHEET not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail="This workbook is not a valid employee commuting template. Please download a fresh template from the app.",
        )

    parsed = []
    parsed.extend(_parse_commuting_sheet(wb[COMMUTING_SHEET]))
    parsed.extend(_parse_wfh_sheet(wb[WFH_SHEET]))
    return parsed


def _candidate_dataset_ids(con, job_id: int, scope: str) -> list[int]:
    ordered: list[int] = []

    def _add(value: Any) -> None:
        dataset_id = _safe_int(value)
        if dataset_id is None:
            return
        if dataset_id not in ordered:
            ordered.append(dataset_id)

    primary = get_scope_primary_datasets(int(job_id))
    _add(primary.get(scope))

    applicable = get_applicable_datasets(int(job_id))
    for dataset_id in applicable.get(scope, []) or []:
        _add(dataset_id)

    legacy = _legacy_scope_dataset_map(int(job_id))
    _add(legacy.get(scope))

    for dataset_id in _additional_dataset_ids(con, int(job_id)):
        _add(dataset_id)

    return ordered


def _resolve_factor_record(con, job_id: int, original_id: str, scope: str) -> dict[str, Any] | None:
    parts = _factor_select_parts(con)
    rows = con.execute(
        f"""
        SELECT
          db_id,
          dataset_id,
          original_id,
          scope,
          {parts["category"]} AS category,
          level_1,
          level_2,
          level_3,
          {parts["level_4"]} AS level_4,
          column_text,
          COALESCE({parts["report_label"]}, column_text) AS report_label,
          uom,
          factor,
          {parts["ghg_unit"]} AS ghg_unit
        FROM v_factor_lookup
        WHERE original_id = %s
        """,
        [str(original_id)],
    ).fetchall()

    if not rows:
        return None

    preferred_ids = _candidate_dataset_ids(con, int(job_id), scope)
    preferred_rank = {dataset_id: idx for idx, dataset_id in enumerate(preferred_ids)}

    def _row_rank(row: Any) -> tuple[int, int]:
        dataset_id = _safe_int(row[1])
        if dataset_id in preferred_rank:
            return (0, preferred_rank[dataset_id])
        return (1, dataset_id or 999999)

    row = sorted(rows, key=_row_rank)[0]
    return {
        "factor_db_id": int(row[0]),
        "dataset_id": _safe_int(row[1]),
        "original_id": str(row[2]),
        "scope": str(row[3] or scope),
        "category": _safe_str(row[4]) or "Employee Commuting",
        "level_1": _safe_str(row[5]) or "Employee Commuting",
        "level_2": _safe_str(row[6]) or "Employee Commuting",
        "level_3": _safe_str(row[7]) or None,
        "level_4": _safe_str(row[8]) or None,
        "column_text": _safe_str(row[9]) or None,
        "report_label": _safe_str(row[10]) or _safe_str(row[9]) or "Employee Commuting",
        "uom": _safe_str(row[11]) or None,
        "factor": _safe_float(row[12], 0.0) or 0.0,
        "ghg_unit": _safe_str(row[13]) or "tCO2e",
    }


def _resolve_commuting_factor_record(con, job_id: int, original_id: str) -> dict[str, Any] | None:
    """
    Resolve a commuting factor record, preferring the dedicated employee commuting
    identifier suffix when it exists.
    """
    candidates = []
    base_id = str(original_id or "").strip()
    if not base_id:
        return None

    if not base_id.endswith("-c"):
        candidates.append(f"{base_id}-c")
    candidates.append(base_id)

    for candidate in candidates:
        factor_record = _resolve_factor_record(con, int(job_id), candidate, "Scope 3")
        if factor_record:
            return factor_record
    return None


def _dataset_id_for_insert(con, dataset_id: Any) -> int | None:
    dsid = _safe_int(dataset_id)
    if dsid is None:
        return None
    exists = con.execute(
        "SELECT 1 FROM datasets WHERE dataset_id = %s",
        [dsid],
    ).fetchone()
    return dsid if exists else None


def _convert_quantity(quantity: float, from_unit: str, to_unit: str) -> float | None:
    src = _distance_unit_base(from_unit)
    dst = _distance_unit_base(to_unit)
    if not src or not dst:
        return None
    if src == dst:
        return float(quantity)
    if src == "miles" and dst == "km":
        return float(quantity) * MILES_TO_KM
    if src == "km" and dst == "miles":
        return float(quantity) / MILES_TO_KM
    return None


def _resolve_commuting_original_id(
    mode_value: str,
    service_value: str,
    unit_value: str,
) -> tuple[str | None, str | None, str | None, str | None, str | None]:
    mode = _canonical_mode(mode_value)
    if not mode:
        return None, None, None, "Commute Mode is required or not recognised", None

    variant = _canonical_variant(service_value)
    if service_value and not variant:
        return None, None, None, "Vehicle / Service Type is not recognised", None

    if not service_value:
        variant = _default_variant_for_mode(mode) or ""

    unit = _canonical_unit(unit_value)
    if unit_value and not unit:
        return None, None, None, "Distance Unit is not recognised", None
    if not unit:
        unit = _default_unit_for_mode(mode) or ""

    candidate_keys = [(mode, variant, unit)]
    matched_unit = unit
    if mode in {"taxis", "bus", "rail", "ferry", "walking", "cycling"} and unit in {"miles", "km"}:
        alternate_unit = "km" if unit == "miles" else "miles"
        candidate_keys.append((mode, variant, alternate_unit))

    original_id = None
    for candidate_key in candidate_keys:
        original_id = COMMUTE_FACTOR_MAP.get(candidate_key)
        if original_id:
            matched_unit = candidate_key[2]
            break

    if not original_id:
        return None, mode, variant, "This commute mode/service combination does not match an emissions factor", None

    return original_id, mode, variant, None, matched_unit


def _build_commuting_notes(parsed_row: dict[str, Any], mode: str, variant: str) -> str:
    details = [
        f"Employee/Team: {parsed_row.get('employee_name') or 'Unspecified'}",
        f"Mode: {mode}",
    ]
    if variant:
        details.append(f"Service Type: {variant}")
    if parsed_row.get("one_way_distance") is not None:
        details.append(f"One Way Distance: {parsed_row['one_way_distance']}")
    if parsed_row.get("office_days") is not None:
        details.append(f"Office Days/Week: {parsed_row['office_days']}")
    if parsed_row.get("weeks_per_year") is not None:
        details.append(f"Weeks/Year: {parsed_row['weeks_per_year']}")
    if parsed_row.get("notes"):
        details.append(f"Notes: {parsed_row['notes']}")
    return " | ".join(details)


def _build_wfh_notes(parsed_row: dict[str, Any]) -> str:
    details = [f"Employee/Team: {parsed_row.get('employee_name') or 'Unspecified'}"]
    if parsed_row.get("annual_days") is not None:
        details.append(f"Annual WFH Days: {parsed_row['annual_days']}")
    if parsed_row.get("hours_per_day") is not None:
        details.append(f"Hours/Day: {parsed_row['hours_per_day']}")
    if parsed_row.get("notes"):
        details.append(f"Notes: {parsed_row['notes']}")
    return " | ".join(details)


def _calc_commuting_tco2e(quantity: Any, factor: Any, apply_pct: Any, ghg_unit: Any) -> float:
    value = float(quantity or 0.0) * float(factor or 0.0) * (float(apply_pct or 100.0) / 100.0)
    if "kg" in str(ghg_unit or "kgCO2e").lower():
        value /= 1000.0
    return float(value)


def _resolve_preview_rows(
    con,
    job_id: int,
    site_id: int | None,
    parsed_rows: list[dict[str, Any]],
    employee_count_override: int | None = None,
    disable_scaling: bool = False,
) -> dict[str, Any]:
    ready_rows: list[dict[str, Any]] = []
    unresolved_rows: list[dict[str, Any]] = []
    unit_fallback_count = 0
    job_employee_headcount = _job_employee_count(con, int(job_id))
    employee_headcount = _effective_employee_count(con, int(job_id), employee_count_override)
    responding_employees = {
        _safe_str(row.get("employee_name")).strip().lower()
        for row in parsed_rows
        if (_safe_float(row.get("annual_quantity")) or 0.0) > 0 and _safe_str(row.get("employee_name")).strip()
    }
    commuting_response_count = len(responding_employees)
    commuting_scale_factor = 1.0
    if not disable_scaling and employee_headcount and commuting_response_count and commuting_response_count < employee_headcount:
        commuting_scale_factor = float(employee_headcount) / float(commuting_response_count)

    for parsed_row in parsed_rows:
        row_number = int(parsed_row["row_number"])
        employee_name = _safe_str(parsed_row.get("employee_name"))

        quantity = _safe_float(parsed_row.get("annual_quantity"))
        if quantity is None or quantity <= 0:
            unresolved_rows.append(
                {
                    "sheet": parsed_row["sheet"],
                    "row_number": row_number,
                    "employee_name": employee_name,
                    "reason": "Annual quantity is missing or zero",
                }
            )
            continue

        if parsed_row["row_type"] == "commuting":
            original_id, mode, variant, error, matched_unit = _resolve_commuting_original_id(
                parsed_row.get("mode_value"),
                parsed_row.get("service_value"),
                parsed_row.get("unit_value"),
            )
            if error or not original_id or not mode:
                unresolved_rows.append(
                    {
                        "sheet": parsed_row["sheet"],
                        "row_number": row_number,
                        "employee_name": employee_name,
                        "reason": error or "Could not resolve commute factor",
                    }
                )
                continue

            factor_record = _resolve_commuting_factor_record(con, int(job_id), original_id)
            if not factor_record:
                unresolved_rows.append(
                    {
                        "sheet": parsed_row["sheet"],
                        "row_number": row_number,
                        "employee_name": employee_name,
                        "reason": f"Original ID {original_id} was not found in available emission factors",
                        "original_id": original_id,
                    }
                )
                continue

            resolved_original_id = str(factor_record.get("original_id") or original_id)
            input_unit = _canonical_unit(parsed_row.get("unit_value")) or (_default_unit_for_mode(mode) or "")
            factor_uom = _safe_str(factor_record.get("uom")) or input_unit
            converted_qty = _convert_quantity(float(quantity), input_unit, factor_uom)
            if converted_qty is None:
                unresolved_rows.append(
                    {
                        "sheet": parsed_row["sheet"],
                        "row_number": row_number,
                        "employee_name": employee_name,
                        "reason": f"Could not convert quantity from {input_unit or 'blank'} to {factor_uom or 'blank'}",
                        "original_id": original_id,
                    }
                )
                continue

            notes = _build_commuting_notes(parsed_row, mode, variant or "")
            if matched_unit and input_unit and matched_unit != input_unit:
                unit_fallback_count += 1
                notes = f"{notes} | Unit fallback applied: resolved as {matched_unit}"
            scaled_qty = float(converted_qty) * float(commuting_scale_factor)
            calc_tco2e = _calc_commuting_tco2e(
                scaled_qty,
                factor_record.get("factor"),
                100,
                factor_record.get("ghg_unit"),
            )
            if commuting_scale_factor > 1.0:
                notes = (
                    f"{notes} | Scaled to full workforce: "
                    f"{commuting_response_count}/{employee_headcount} distinct employees "
                    f"({commuting_scale_factor:.2f}x)"
                )
            ready_rows.append(
                {
                    "sheet": parsed_row["sheet"],
                    "row_number": row_number,
                    "row_type": "commuting",
                    "employee_name": employee_name,
                    "site_id": site_id,
                    "scope": "Scope 3",
                    "original_id": resolved_original_id,
                    "dataset_id": factor_record.get("dataset_id"),
                    "factor_db_id": factor_record.get("factor_db_id"),
                    "category": "Employee Commuting",
                    "level_1": "Employee Commuting",
                    "level_2": mode.title(),
                    "level_3": variant.title() if variant else None,
                    "level_4": None,
                    "column_text": factor_record.get("column_text"),
                    "report_label": (
                        f"Employee Commuting - {mode.title()}"
                        + (f" - {variant.title()}" if variant else "")
                    ),
                    "qty": scaled_qty,
                    "uom": factor_uom,
                    "factor": factor_record.get("factor"),
                    "ghg_unit": factor_record.get("ghg_unit"),
                    "calc_tco2e": calc_tco2e,
                    "apply_pct": 100,
                    "data_source": COMMUTING_DATA_SOURCE,
                    "data_confidence": "M",
                    "notes": notes,
                    "is_custom_entry": False,
                }
            )
            continue

        original_id = WFH_ORIGINAL_ID
        if not original_id:
            unresolved_rows.append(
                {
                    "sheet": parsed_row["sheet"],
                    "row_number": row_number,
                    "employee_name": employee_name,
                    "reason": "Working from home factor ID is not configured",
                }
            )
            continue

        factor_record = _resolve_factor_record(con, int(job_id), original_id, "Scope 3")
        if not factor_record:
            unresolved_rows.append(
                {
                    "sheet": parsed_row["sheet"],
                    "row_number": row_number,
                    "employee_name": employee_name,
                    "reason": f"Original ID {original_id} was not found in available emission factors",
                    "original_id": original_id,
                }
            )
            continue

        factor_uom = _safe_str(factor_record.get("uom")) or "hours"
        converted_qty = (
            float(quantity)
            if _hours_equivalent_unit(factor_uom)
            else _convert_quantity(float(quantity), "hours", factor_uom)
        )
        if converted_qty is None:
            unresolved_rows.append(
                {
                    "sheet": parsed_row["sheet"],
                    "row_number": row_number,
                    "employee_name": employee_name,
                    "reason": f"Could not convert WFH hours to {factor_uom}",
                    "original_id": original_id,
                }
            )
            continue

        notes = _build_wfh_notes(parsed_row)
        calc_tco2e = _calc_commuting_tco2e(
            converted_qty,
            factor_record.get("factor"),
            100,
            factor_record.get("ghg_unit"),
        )
        ready_rows.append(
            {
                "sheet": parsed_row["sheet"],
                "row_number": row_number,
                "row_type": "wfh",
                "employee_name": employee_name,
                "site_id": site_id,
                "scope": "Scope 3",
                "original_id": original_id,
                "dataset_id": factor_record.get("dataset_id"),
                "factor_db_id": factor_record.get("factor_db_id"),
                "category": "Employee Commuting",
                "level_1": "Employee Commuting",
                "level_2": "Working From Home",
                "level_3": None,
                "level_4": None,
                "column_text": factor_record.get("column_text"),
                "report_label": "Employee Commuting - Working From Home",
                "qty": converted_qty,
                "uom": factor_uom,
                "factor": factor_record.get("factor"),
                "ghg_unit": factor_record.get("ghg_unit"),
                "calc_tco2e": calc_tco2e,
                "apply_pct": 100,
                "data_source": COMMUTING_DATA_SOURCE,
                "data_confidence": "M",
                "notes": notes,
                "is_custom_entry": False,
            }
        )

    return {
        "parsed_count": len(parsed_rows),
        "ready_count": len(ready_rows),
        "unresolved_count": len(unresolved_rows),
        "total_tco2e": round(sum(float(row["calc_tco2e"] or 0.0) for row in ready_rows), 6),
        "employee_headcount": employee_headcount,
        "job_employee_headcount": job_employee_headcount,
        "commuting_response_count": commuting_response_count,
        "commuting_scale_factor": round(float(commuting_scale_factor), 6),
        "unit_fallback_count": unit_fallback_count,
        "ready_rows": ready_rows,
        "unresolved_rows": unresolved_rows,
    }


def _merge_commuting_ready_rows(ready_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, Any, Any, Any, Any], dict[str, Any]] = {}
    order: list[tuple[Any, Any, Any, Any, Any]] = []

    for row in ready_rows:
        key = (
            row.get("site_id"),
            row.get("scope"),
            row.get("original_id"),
            row.get("data_source"),
            row.get("row_type"),
        )
        if key not in grouped:
            grouped[key] = dict(row)
            grouped[key]["_source_row_count"] = 1
            grouped[key]["_employee_names"] = []
            grouped[key]["_notes"] = []
            employee_name = _safe_str(row.get("employee_name"))
            note_text = _safe_str(row.get("notes"))
            if employee_name:
                grouped[key]["_employee_names"].append(employee_name)
            if note_text:
                grouped[key]["_notes"].append(note_text)
            order.append(key)
            continue

        bucket = grouped[key]
        bucket["_source_row_count"] = int(bucket.get("_source_row_count") or 0) + 1

        employee_name = _safe_str(row.get("employee_name"))
        if employee_name and employee_name not in bucket["_employee_names"]:
            bucket["_employee_names"].append(employee_name)

        note_text = _safe_str(row.get("notes"))
        if note_text and note_text not in bucket["_notes"]:
            bucket["_notes"].append(note_text)

        bucket["qty"] = float(bucket.get("qty") or 0.0) + float(row.get("qty") or 0.0)
        bucket["calc_tco2e"] = float(bucket.get("calc_tco2e") or 0.0) + float(row.get("calc_tco2e") or 0.0)

    merged_rows: list[dict[str, Any]] = []
    for key in order:
        row = grouped[key]
        row_count = int(row.pop("_source_row_count", 1) or 1)
        employee_names = [name for name in row.pop("_employee_names", []) if name]
        notes = [note for note in row.pop("_notes", []) if note]

        if row_count > 1:
            summary_bits = [f"Aggregated {row_count} workbook rows"]
            if employee_names:
                sample_names = ", ".join(employee_names[:5])
                if len(employee_names) > 5:
                    sample_names = f"{sample_names}, ..."
                summary_bits.append(f"employees: {sample_names}")
            if notes:
                summary_bits.append(f"notes: {' | '.join(notes[:3])}")

            aggregation_note = "; ".join(summary_bits)
            existing_note = _safe_str(row.get("notes"))
            row["notes"] = f"{aggregation_note} | {existing_note}" if existing_note else aggregation_note

        merged_rows.append(row)

    return merged_rows


def _is_commuting_unique_violation(exc: Exception) -> bool:
    message = str(exc)
    return (
        "job_scope_rows_job_site_scope_active_uidx" in message
        or "duplicate key value violates unique constraint" in message
        or "UNIQUE constraint failed" in message
    )


def _disable_existing_commuting_rows(con, job_id: int, site_id: int | None) -> int:
    # Build separate queries per site_id nullability — psycopg3 cannot infer
    # the type of a bare `%s IS NULL` expression (IndeterminateDatatype).
    if site_id is None:
        rows = con.execute(
            """
            UPDATE job_scope_rows
            SET enabled = FALSE, updated_at = NOW()
            WHERE job_id = %s
              AND COALESCE(data_source, '') = %s
              AND site_id IS NULL
              AND enabled = TRUE
            RETURNING row_id
            """,
            [int(job_id), COMMUTING_DATA_SOURCE],
        ).fetchall()
    else:
        rows = con.execute(
            """
            UPDATE job_scope_rows
            SET enabled = FALSE, updated_at = NOW()
            WHERE job_id = %s
              AND COALESCE(data_source, '') = %s
              AND site_id = %s
              AND enabled = TRUE
            RETURNING row_id
            """,
            [int(job_id), COMMUTING_DATA_SOURCE, int(site_id)],
        ).fetchall()
    return len(rows or [])


def _insert_ready_rows(con, job_id: int, ready_rows: list[dict[str, Any]]) -> int:
    inserted = 0
    for row in ready_rows:
        dataset_id = _dataset_id_for_insert(con, row.get("dataset_id"))
        con.execute(
            """
            INSERT INTO job_scope_rows (
              job_id, scope, site_id, dataset_id, factor_db_id, original_id,
              category, level_1, level_2, level_3, level_4, column_text, report_label,
              qty, uom, factor, ghg_unit, calc_tco2e, apply_pct, data_source,
              data_confidence, notes, is_custom_entry
            )
            VALUES (
              %s, %s, %s, %s, %s, %s,
              %s, %s, %s, %s, %s, %s, %s,
              %s, %s, %s, %s, %s, %s, %s,
              %s, %s, %s
            )
            """,
            [
                int(job_id),
                row.get("scope"),
                row.get("site_id"),
                dataset_id,
                row.get("factor_db_id"),
                row.get("original_id"),
                row.get("category"),
                row.get("level_1"),
                row.get("level_2"),
                row.get("level_3"),
                row.get("level_4"),
                row.get("column_text"),
                row.get("report_label"),
                row.get("qty"),
                row.get("uom"),
                row.get("factor"),
                row.get("ghg_unit"),
                row.get("calc_tco2e"),
                row.get("apply_pct", 100),
                row.get("data_source", COMMUTING_DATA_SOURCE),
                row.get("data_confidence", "M"),
                row.get("notes"),
                row.get("is_custom_entry", False),
            ],
        )
        inserted += 1
    return inserted


def _parse_months(entry: dict[str, Any]) -> list[float | None] | None:
    """Optional 12-value monthly breakdown from the portal's monthly entry
    grid -- months[0]=Jan .. months[11]=Dec, same calendar-indexed convention
    used everywhere else (see services/monthly_emissions.py). Returns None
    if the entry has no `months` field at all (e.g. legacy Excel upload)."""
    raw = entry.get("months")
    if not isinstance(raw, list):
        return None
    months: list[float | None] = [None] * 12
    for i in range(min(12, len(raw))):
        months[i] = _safe_float(raw[i])
    return months


def _months_sum(months: list[float | None] | None) -> float | None:
    if not months:
        return None
    values = [m for m in months if m is not None]
    return float(sum(values)) if values else None


def _manual_entry_to_parsed_row(entry: dict[str, Any]) -> dict[str, Any] | None:
    row_type = _safe_str(entry.get("row_type")).lower()
    employee_name = _safe_str(entry.get("employee_name"))
    notes = _safe_str(entry.get("notes"))

    if row_type == "commuting":
        one_way_distance = _safe_float(entry.get("one_way_distance"))
        office_days = _safe_float(entry.get("office_days"))
        weeks_per_year = _safe_float(entry.get("weeks_per_year"))
        months = _parse_months(entry)
        annual_quantity = _safe_float(entry.get("annual_quantity"))
        if annual_quantity is None:
            annual_quantity = _months_sum(months)
        if annual_quantity is None and None not in (one_way_distance, office_days, weeks_per_year):
            annual_quantity = float(one_way_distance) * 2.0 * float(office_days) * float(weeks_per_year)
        return {
            "sheet": "Direct Entry",
            "row_number": _safe_int(entry.get("row_number")) or 0,
            "row_type": "commuting",
            "employee_name": employee_name,
            "mode_value": _safe_str(entry.get("mode_value")),
            "service_value": _safe_str(entry.get("service_value")),
            "unit_value": _safe_str(entry.get("unit_value")),
            "annual_quantity": annual_quantity,
            "months": months,
            "notes": notes,
            "one_way_distance": one_way_distance,
            "office_days": office_days,
            "weeks_per_year": weeks_per_year,
        }

    if row_type == "wfh":
        annual_days = _safe_float(entry.get("annual_days"))
        hours_per_day = _safe_float(entry.get("hours_per_day"))
        annual_quantity = _safe_float(entry.get("annual_quantity"))
        if annual_quantity is None and annual_days is not None and hours_per_day is not None:
            annual_quantity = float(annual_days) * float(hours_per_day)
        return {
            "sheet": "Direct Entry",
            "row_number": _safe_int(entry.get("row_number")) or 0,
            "row_type": "wfh",
            "employee_name": employee_name,
            "annual_quantity": annual_quantity,
            "annual_days": annual_days,
            "hours_per_day": hours_per_day,
            "notes": notes,
        }

    return None


def _resolve_manual_commuting_rows(
    con,
    job_id: int,
    site_id: int | None,
    entries: list[dict[str, Any]],
) -> dict[str, Any]:
    ready_rows: list[dict[str, Any]] = []
    unresolved_rows: list[dict[str, Any]] = []

    for idx, entry in enumerate(entries, start=1):
        parsed_row = _manual_entry_to_parsed_row({**entry, "row_number": idx})
        if not parsed_row:
            unresolved_rows.append(
                {
                    "sheet": "Direct Entry",
                    "row_number": idx,
                    "employee_name": _safe_str(entry.get("employee_name")),
                    "reason": "Row type must be commuting or wfh",
                }
            )
            continue

        row_type = str(parsed_row.get("row_type") or "").lower()
        employee_name = _safe_str(parsed_row.get("employee_name"))
        if not employee_name:
            unresolved_rows.append(
                {
                    "sheet": "Direct Entry",
                    "row_number": idx,
                    "employee_name": "",
                    "reason": "Employee / team name is required",
                }
            )
            continue
        quantity = _safe_float(parsed_row.get("annual_quantity"))
        if quantity is None or quantity <= 0:
            unresolved_rows.append(
                {
                    "sheet": "Direct Entry",
                    "row_number": idx,
                    "employee_name": employee_name,
                    "reason": "Annual quantity is missing or zero",
                }
            )
            continue

        if row_type == "commuting":
            original_id, mode, variant, error, _matched_unit = _resolve_commuting_original_id(
                parsed_row.get("mode_value"),
                parsed_row.get("service_value"),
                parsed_row.get("unit_value"),
            )
            if error or not original_id or not mode:
                unresolved_rows.append(
                    {
                        "sheet": "Direct Entry",
                        "row_number": idx,
                        "employee_name": employee_name,
                        "reason": error or "Could not resolve commute factor",
                    }
                )
                continue

            factor_record = _resolve_commuting_factor_record(con, int(job_id), original_id)
            if not factor_record:
                unresolved_rows.append(
                    {
                        "sheet": "Direct Entry",
                        "row_number": idx,
                        "employee_name": employee_name,
                        "reason": f"Original ID {original_id} was not found in available emission factors",
                        "original_id": original_id,
                    }
                )
                continue

            resolved_original_id = str(factor_record.get("original_id") or original_id)
            input_unit = _canonical_unit(parsed_row.get("unit_value")) or (_default_unit_for_mode(mode) or "")
            factor_uom = _safe_str(factor_record.get("uom")) or input_unit
            converted_qty = _convert_quantity(float(quantity), input_unit, factor_uom)
            if converted_qty is None:
                unresolved_rows.append(
                    {
                        "sheet": "Direct Entry",
                        "row_number": idx,
                        "employee_name": employee_name,
                        "reason": f"Could not convert quantity from {input_unit or 'blank'} to {factor_uom or 'blank'}",
                        "original_id": original_id,
                    }
                )
                continue

            notes = _build_commuting_notes(parsed_row, mode, variant or "")
            calc_tco2e = _calc_commuting_tco2e(
                converted_qty,
                factor_record.get("factor"),
                100,
                factor_record.get("ghg_unit"),
            )
            # Convert each month individually (same unit conversion as the
            # annual total) so month_1..12 stay in factor_uom, matching qty.
            months_in = parsed_row.get("months")
            months_converted: list[float | None] = [None] * 12
            if months_in:
                for i, month_val in enumerate(months_in):
                    if month_val is None:
                        continue
                    months_converted[i] = _convert_quantity(float(month_val), input_unit, factor_uom)

            ready_row = {
                "scope": "Scope 3",
                "site_id": site_id,
                "source_type": "employee_commuting",
                "source_subtype": "commuting",
                "source_name": (
                    f"{employee_name} - Employee Commuting - {mode.title()}"
                    + (f" - {variant.title()}" if variant else "")
                ).strip(" -"),
                "asset_identifier": None,
                "employee_name": employee_name or None,
                "dataset_id": factor_record.get("dataset_id"),
                "factor_db_id": factor_record.get("factor_db_id"),
                "original_id": resolved_original_id,
                "category": "Employee Commuting",
                "qty": float(converted_qty),
                "uom": factor_uom,
                "factor": factor_record.get("factor"),
                "ghg_unit": factor_record.get("ghg_unit"),
                "calc_tco2e": calc_tco2e,
                "apply_pct": 100,
                "data_source": DIRECT_COMMUTING_DATA_SOURCE,
                "data_confidence": "M",
                "notes": notes,
                "detail_json": {
                    "entry_type": "commuting",
                    "mode_value": parsed_row.get("mode_value"),
                    "service_value": parsed_row.get("service_value"),
                    "unit_value": parsed_row.get("unit_value"),
                    "one_way_distance": parsed_row.get("one_way_distance"),
                    "office_days": parsed_row.get("office_days"),
                    "weeks_per_year": parsed_row.get("weeks_per_year"),
                    "manual_entry": True,
                },
            }
            for i in range(12):
                ready_row[f"month_{i + 1}"] = months_converted[i]
            ready_rows.append(ready_row)
            continue

        original_id = WFH_ORIGINAL_ID
        if not original_id:
            unresolved_rows.append(
                {
                    "sheet": "Direct Entry",
                    "row_number": idx,
                    "employee_name": employee_name,
                    "reason": "Working from home factor ID is not configured",
                }
            )
            continue

        factor_record = _resolve_factor_record(con, int(job_id), original_id, "Scope 3")
        if not factor_record:
            unresolved_rows.append(
                {
                    "sheet": "Direct Entry",
                    "row_number": idx,
                    "employee_name": employee_name,
                    "reason": f"Original ID {original_id} was not found in available emission factors",
                    "original_id": original_id,
                }
            )
            continue

        factor_uom = _safe_str(factor_record.get("uom")) or "hours"
        converted_qty = (
            float(quantity)
            if _hours_equivalent_unit(factor_uom)
            else _convert_quantity(float(quantity), "hours", factor_uom)
        )
        if converted_qty is None:
            unresolved_rows.append(
                {
                    "sheet": "Direct Entry",
                    "row_number": idx,
                    "employee_name": employee_name,
                    "reason": f"Could not convert WFH hours to {factor_uom}",
                    "original_id": original_id,
                }
            )
            continue

        notes = _build_wfh_notes(parsed_row)
        calc_tco2e = _calc_commuting_tco2e(
            converted_qty,
            factor_record.get("factor"),
            100,
            factor_record.get("ghg_unit"),
        )
        ready_rows.append(
            {
                "scope": "Scope 3",
                "site_id": site_id,
                "source_type": "employee_commuting",
                "source_subtype": "wfh",
                "source_name": f"{employee_name} - Employee Commuting - Working From Home".strip(" -"),
                "asset_identifier": None,
                "employee_name": employee_name or None,
                "dataset_id": factor_record.get("dataset_id"),
                "factor_db_id": factor_record.get("factor_db_id"),
                "original_id": original_id,
                "category": "Employee Commuting",
                "qty": float(converted_qty),
                "uom": factor_uom,
                "factor": factor_record.get("factor"),
                "ghg_unit": factor_record.get("ghg_unit"),
                "calc_tco2e": calc_tco2e,
                "apply_pct": 100,
                "data_source": DIRECT_COMMUTING_DATA_SOURCE,
                "data_confidence": "M",
                "notes": notes,
                "detail_json": {
                    "entry_type": "wfh",
                    "annual_days": parsed_row.get("annual_days"),
                    "hours_per_day": parsed_row.get("hours_per_day"),
                    "manual_entry": True,
                },
            }
        )

    return {
        "parsed_count": len(entries),
        "ready_count": len(ready_rows),
        "unresolved_count": len(unresolved_rows),
        "total_tco2e": round(sum(float(row["calc_tco2e"] or 0.0) for row in ready_rows), 6),
        "ready_rows": ready_rows,
        "unresolved_rows": unresolved_rows,
    }


def _insert_manual_commuting_rows(
    con, job_id: int, ready_rows: list[dict[str, Any]], submitted_by_portal: bool = False
) -> tuple[int, list[int]]:
    inserted = 0
    inserted_ids: list[int] = []
    for row in ready_rows:
        dataset_id = _dataset_id_for_insert(con, row.get("dataset_id"))
        month_cols = ", ".join(f"month_{i}" for i in range(1, 13))
        month_placeholders = ", ".join(["%s"] * 12)
        result = con.execute(
            f"""
            INSERT INTO job_emission_sources (
              job_id, group_id, scope, category, source_type, source_subtype, site_id,
              source_name, asset_identifier, employee_name, dataset_id, factor_db_id, original_id,
              qty, uom, factor, ghg_unit, apply_pct, data_source, data_confidence, notes,
              detail_json, calc_tco2e, enabled, submitted_by_portal, review_status, {month_cols}
            )
            VALUES (
              %s, %s, %s, %s, %s, %s, %s,
              %s, %s, %s, %s, %s, %s,
              %s, %s, %s, %s, %s, %s, %s, %s,
              %s, %s, %s, %s, %s, {month_placeholders}
            )
            RETURNING source_id
            """,
            [
                int(job_id),
                None,
                row.get("scope"),
                row.get("category"),
                row.get("source_type"),
                row.get("source_subtype"),
                row.get("site_id"),
                row.get("source_name"),
                row.get("asset_identifier"),
                row.get("employee_name"),
                dataset_id,
                row.get("factor_db_id"),
                row.get("original_id"),
                row.get("qty"),
                row.get("uom"),
                row.get("factor"),
                row.get("ghg_unit"),
                row.get("apply_pct", 100),
                row.get("data_source", DIRECT_COMMUTING_DATA_SOURCE),
                row.get("data_confidence", "M"),
                row.get("notes"),
                row.get("detail_json") or {},
                row.get("calc_tco2e"),
                not submitted_by_portal,
                bool(submitted_by_portal),
                "pending_review" if submitted_by_portal else None,
                *[row.get(f"month_{i}") for i in range(1, 13)],
            ],
        ).fetchone()
        inserted += 1
        if result:
            inserted_ids.append(int(result[0]))
    return inserted, inserted_ids


def _update_manual_commuting_row(con, job_id: int, source_id: int, row: dict[str, Any]) -> None:
    dataset_id = _dataset_id_for_insert(con, row.get("dataset_id"))
    con.execute(
        """
        UPDATE job_emission_sources
        SET
          group_id = %s,
          scope = %s,
          category = %s,
          source_type = %s,
          source_subtype = %s,
          site_id = %s,
          source_name = %s,
          asset_identifier = %s,
          employee_name = %s,
          dataset_id = %s,
          factor_db_id = %s,
          original_id = %s,
          qty = %s,
          uom = %s,
          factor = %s,
          ghg_unit = %s,
          apply_pct = %s,
          data_source = %s,
          data_confidence = %s,
          notes = %s,
          detail_json = %s,
          calc_tco2e = %s,
          enabled = TRUE,
          updated_at = NOW()
        WHERE source_id = %s
          AND job_id = %s
        """,
        [
            None,
            row.get("scope"),
            row.get("category"),
            row.get("source_type"),
            row.get("source_subtype"),
            row.get("site_id"),
            row.get("source_name"),
            row.get("asset_identifier"),
            row.get("employee_name"),
            dataset_id,
            row.get("factor_db_id"),
            row.get("original_id"),
            row.get("qty"),
            row.get("uom"),
            row.get("factor"),
            row.get("ghg_unit"),
            row.get("apply_pct", 100),
            row.get("data_source", DIRECT_COMMUTING_DATA_SOURCE),
            row.get("data_confidence", "M"),
            row.get("notes"),
            row.get("detail_json") or {},
            row.get("calc_tco2e"),
            int(source_id),
            int(job_id),
        ],
    )


def _disable_existing_manual_commuting_rows(con, job_id: int, site_id: int | None) -> int:
    if site_id is None:
        rows = con.execute(
            """
            UPDATE job_emission_sources
            SET enabled = FALSE, updated_at = NOW()
            WHERE job_id = %s
              AND source_type = %s
              AND site_id IS NULL
              AND enabled = TRUE
            RETURNING source_id
            """,
            [int(job_id), "employee_commuting"],
        ).fetchall()
    else:
        rows = con.execute(
            """
            UPDATE job_emission_sources
            SET enabled = FALSE, updated_at = NOW()
            WHERE job_id = %s
              AND source_type = %s
              AND site_id = %s
              AND enabled = TRUE
            RETURNING source_id
            """,
            [int(job_id), "employee_commuting", int(site_id)],
        ).fetchall()
    return len(rows or [])


@router.get("/jobs/{job_id}/employee-commuting/template")
def download_employee_commuting_template(
    job_id: int,
    site_id: int | None = Query(None),
    _user: dict[str, str] = Depends(_current_user),
):
    with get_conn(autocommit=False) as con:
        meta = _job_meta(con, int(job_id))
        _, site_label = _job_site_label(con, int(job_id), site_id)
        workbook_bytes = _build_template_workbook(meta, site_label)
        filename = _template_filename(meta, site_label)

    safe_filename = filename.replace('"', '\\"')
    return Response(
        content=workbook_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_filename}"',
            "X-Filename": filename,
        },
    )


@router.get("/jobs/{job_id}/employee-commuting/summary")
def employee_commuting_summary(
    job_id: int,
    _user: dict[str, str] = Depends(_current_user),
):
    with get_conn() as con:
        _ensure_job_scope_rows_schema(con)
        _ensure_emission_register_schema(con)
        _job_meta(con, int(job_id))
        employee_headcount = _job_employee_count(con, int(job_id))
        row = con.execute(
            """
            SELECT
              COALESCE(SUM(row_count), 0) AS row_count,
              COALESCE(SUM(total_tco2e), 0) AS total_tco2e,
              COUNT(DISTINCT site_key) AS site_count
            FROM (
              SELECT
                COUNT(*) AS row_count,
                COALESCE(SUM(calc_tco2e), 0) AS total_tco2e,
                COALESCE(site_id, -1) AS site_key
              FROM job_scope_rows
              WHERE job_id = %s
                AND COALESCE(data_source, '') IN (%s, %s)
                AND enabled = TRUE
              GROUP BY COALESCE(site_id, -1)
              UNION ALL
              SELECT
                COUNT(*) AS row_count,
                COALESCE(SUM(calc_tco2e), 0) AS total_tco2e,
                COALESCE(site_id, -1) AS site_key
              FROM job_emission_sources
              WHERE job_id = %s
                AND source_type = %s
                AND enabled = TRUE
              GROUP BY COALESCE(site_id, -1)
            ) x
            """,
            [int(job_id), COMMUTING_DATA_SOURCE, DIRECT_COMMUTING_DATA_SOURCE, int(job_id), "employee_commuting"],
        ).fetchone()

    return {
        "job_id": int(job_id),
        "row_count": int(row[0] or 0) if row else 0,
        "total_tco2e": float(row[1] or 0.0) if row else 0.0,
        "site_count": int(row[2] or 0) if row else 0,
        "employee_headcount": employee_headcount,
        "data_source": COMMUTING_DATA_SOURCE,
    }


@router.post("/jobs/{job_id}/employee-commuting/upload-preview")
async def preview_employee_commuting_upload(
    job_id: int,
    site_id: int | None = Query(None),
    employee_count: int | None = Query(None, ge=1),
    disable_scaling: bool = Query(False),
    file: UploadFile = File(...),
    _user: dict[str, str] = Depends(_current_user),
):
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty upload")
    try:
        scan_bytes(raw, filename=file.filename)
    except VirusScanError as e:
        raise HTTPException(status_code=400, detail=str(e))

    with get_conn() as con:
        _ensure_job_scope_rows_schema(con)
        _job_meta(con, int(job_id))
        validated_site_id, site_label = _job_site_label(con, int(job_id), site_id)
        parsed_rows = _parse_template(raw)
        preview = _resolve_preview_rows(con, int(job_id), validated_site_id, parsed_rows, employee_count, disable_scaling)

    return {
        "job_id": int(job_id),
        "site_id": validated_site_id,
        "site_label": site_label,
        "template_version": TEMPLATE_VERSION,
        **preview,
    }


@router.post("/jobs/{job_id}/employee-commuting/upload-commit")
async def commit_employee_commuting_upload(
    request: Request,
    job_id: int,
    site_id: int | None = Query(None),
    replace_existing: bool = Query(True),
    employee_count: int | None = Query(None, ge=1),
    disable_scaling: bool = Query(False),
    file: UploadFile = File(...),
    _user: dict[str, str] = Depends(_current_user),
):
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty upload")
    try:
        scan_bytes(raw, filename=file.filename)
    except VirusScanError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Run DDL migrations in a separate autocommit connection so they cannot
    # abort the transactional block below (PostgreSQL marks a transaction as
    # "aborted" when any statement inside it fails, even if the exception is
    # silently caught by try/except).
    with get_conn() as _schema_con:
        _ensure_job_scope_rows_schema(_schema_con)

    try:
        with get_conn(autocommit=False) as con:
            meta = _job_meta(con, int(job_id))
            validated_site_id, site_label = _job_site_label(con, int(job_id), site_id)
            parsed_rows = _parse_template(raw)
            preview = _resolve_preview_rows(con, int(job_id), validated_site_id, parsed_rows, employee_count, disable_scaling)

            if preview["unresolved_count"] > 0:
                raise HTTPException(
                    status_code=400,
                    detail={
                        "message": "Upload contains unresolved rows. Please correct the workbook and preview again before import.",
                        "preview": {
                            "job_id": int(job_id),
                            "site_id": validated_site_id,
                            "site_label": site_label,
                            "template_version": TEMPLATE_VERSION,
                            **preview,
                        },
                    },
                )

            if preview["ready_count"] <= 0:
                raise HTTPException(status_code=400, detail="No importable employee commuting rows were found")

            grouped_rows = _merge_commuting_ready_rows(preview["ready_rows"])
            disabled = 0
            if replace_existing:
                disabled = _disable_existing_commuting_rows(con, int(job_id), validated_site_id)

            inserted = _insert_ready_rows(con, int(job_id), grouped_rows)

            record_audit_event(
                con,
                request=request,
                actor=_user,
                action="import",
                entity_type="employee_commuting_import",
                entity_id=f"{int(job_id)}:{validated_site_id if validated_site_id is not None else 'all'}",
                client_id=meta.get("client_db_id"),
                job_id=int(job_id),
                metadata={
                    "site_id": validated_site_id,
                    "site_label": site_label,
                    "replace_existing": bool(replace_existing),
                    "inserted": int(inserted),
                    "disabled": int(disabled),
                    "parsed_count": int(preview.get("parsed_count") or 0),
                    "ready_count": int(preview.get("ready_count") or 0),
                    "grouped_count": int(len(grouped_rows)),
                    "unresolved_count": int(preview.get("unresolved_count") or 0),
                    "total_tco2e": float(preview.get("total_tco2e") or 0.0),
                    "employee_headcount": preview.get("employee_headcount"),
                    "job_employee_headcount": preview.get("job_employee_headcount"),
                    "commuting_response_count": preview.get("commuting_response_count"),
                    "commuting_scale_factor": preview.get("commuting_scale_factor"),
                    "employee_count": employee_count,
                    "data_source": COMMUTING_DATA_SOURCE,
                },
            )
    except HTTPException:
        raise
    except Exception as exc:
        import traceback
        err_detail = f"{type(exc).__name__}: {exc}\n{traceback.format_exc()}"
        if _is_commuting_unique_violation(exc):
            raise HTTPException(
                status_code=409,
                detail=(
                    "This commuting import conflicts with an existing active row for the same commute factor. "
                    "Use Replace previous commuting import, or clear the existing commuting rows for this site and try again."
                ),
            ) from exc
        raise HTTPException(status_code=500, detail=err_detail) from exc

    return {
        "ok": True,
        "job_id": int(job_id),
        "site_id": validated_site_id,
        "site_label": site_label,
        "inserted": int(inserted),
        "source_row_count": int(preview.get("ready_count") or 0),
        "disabled": int(disabled),
        "total_tco2e": float(preview["total_tco2e"]),
        "employee_headcount": preview.get("employee_headcount"),
        "job_employee_headcount": preview.get("job_employee_headcount"),
        "commuting_response_count": preview.get("commuting_response_count"),
        "commuting_scale_factor": preview.get("commuting_scale_factor"),
        "employee_count": employee_count,
        "data_source": COMMUTING_DATA_SOURCE,
    }


@router.post("/jobs/{job_id}/employee-commuting/direct-entry-by-vehicle")
def create_employee_commuting_entry_by_vehicle(
    request: Request,
    job_id: int,
    payload: dict = Body(...),
    site_id: int | None = Query(None),
    _user: dict[str, str] = Depends(_current_user),
):
    """CRM equivalent of the portal's "I drive my own car" path -- resolves
    a factor directly from a DVLA registration lookup instead of the
    mode/service dropdowns. Staff-entered, so the row lands enabled=TRUE
    immediately (no portal review gate) -- see
    api/portal_commuting_routes.py's portal_commuting_create_row_by_vehicle
    for the client-facing equivalent. The registration number is used
    transiently here and never persisted."""
    employee_name = str(payload.get("employee_name") or "").strip()
    registration = str(payload.get("registration_number") or "").strip()
    if not employee_name:
        raise HTTPException(status_code=400, detail="employee_name is required")
    if not registration:
        raise HTTPException(status_code=400, detail="registration_number is required")
    try:
        annual_quantity = float(payload.get("annual_quantity"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="annual_quantity is required and must be a number")
    if annual_quantity <= 0:
        raise HTTPException(status_code=400, detail="annual_quantity must be greater than zero")

    vehicle_data, lookup_error = lookup_vehicle_by_registration(registration)
    if lookup_error:
        status = 503 if "not configured" in lookup_error else 404
        raise HTTPException(status_code=status, detail=lookup_error)

    with get_conn() as con:
        _ensure_job_scope_rows_schema(con)
        _ensure_emission_register_schema(con)
        meta = _job_meta(con, int(job_id))
        validated_site_id, site_label = _job_site_label(con, int(job_id), site_id)

        factor, category_error = categorize_vehicle(con, int(job_id), vehicle_data)
        if category_error:
            raise HTTPException(status_code=422, detail=category_error)

        calc_tco2e = _calc_commuting_tco2e(annual_quantity, factor.get("factor"), 100, factor.get("ghg_unit"))
        ready_row = {
            "scope": "Scope 3",
            "site_id": validated_site_id,
            "source_type": "employee_commuting",
            "source_subtype": "commuting",
            "source_name": f"{employee_name} - Employee Commuting - {factor.get('report_label')}".strip(" -"),
            "asset_identifier": None,
            "employee_name": employee_name,
            "dataset_id": factor.get("dataset_id"),
            "factor_db_id": factor.get("factor_db_id"),
            "original_id": factor.get("original_id"),
            "category": "Employee Commuting",
            "qty": float(annual_quantity),
            "uom": factor.get("uom"),
            "factor": factor.get("factor"),
            "ghg_unit": factor.get("ghg_unit"),
            "calc_tco2e": calc_tco2e,
            "apply_pct": 100,
            "data_source": DIRECT_COMMUTING_DATA_SOURCE,
            "data_confidence": "M",
            "notes": f"Employee/Team: {employee_name} — matched via registration lookup to {factor.get('report_label')}",
            "detail_json": {"entry_type": "commuting", "manual_entry": True, "via": "registration_lookup"},
        }

        _inserted, inserted_ids = _insert_manual_commuting_rows(con, int(job_id), [ready_row])

        record_audit_event(
            con,
            request=request,
            actor=_user,
            action="create",
            entity_type="employee_commuting_direct_entry",
            entity_id=inserted_ids[0] if inserted_ids else None,
            client_id=meta.get("client_db_id"),
            job_id=int(job_id),
            metadata={
                "site_id": validated_site_id,
                "site_label": site_label,
                "matched_category": factor.get("report_label"),
                "data_source": DIRECT_COMMUTING_DATA_SOURCE,
                "via": "registration_lookup",
            },
        )

    return {
        "ok": True,
        "job_id": int(job_id),
        "source_id": inserted_ids[0] if inserted_ids else None,
        "matched_category": factor.get("report_label"),
        "make": vehicle_data.get("make"),
        "fuel_type": vehicle_data.get("fuel_type"),
    }


@router.post("/jobs/{job_id}/employee-commuting/direct-entry-commit")
def commit_employee_commuting_direct_entries(
    request: Request,
    job_id: int,
    payload: dict = Body(...),
    site_id: int | None = Query(None),
    replace_existing: bool = Query(False),
    _user: dict[str, str] = Depends(_current_user),
):
    entries = payload.get("entries") or []
    if not isinstance(entries, list) or not entries:
        raise HTTPException(status_code=400, detail="No direct employee commuting entries were provided")

    with get_conn() as con:
        _ensure_job_scope_rows_schema(con)
        _ensure_emission_register_schema(con)
        meta = _job_meta(con, int(job_id))
        validated_site_id, site_label = _job_site_label(con, int(job_id), site_id)
        preview = _resolve_manual_commuting_rows(con, int(job_id), validated_site_id, entries)

        if preview["unresolved_count"] > 0:
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "Direct entries contain unresolved rows. Please correct them before saving.",
                    "preview": {
                        "job_id": int(job_id),
                        "site_id": validated_site_id,
                        "site_label": site_label,
                        "template_version": "Direct Entry",
                        **preview,
                    },
                },
            )

        if preview["ready_count"] <= 0:
            raise HTTPException(status_code=400, detail="No importable direct employee commuting rows were found")

        disabled = 0
        if replace_existing:
            disabled = _disable_existing_manual_commuting_rows(con, int(job_id), validated_site_id)

        inserted, _inserted_ids = _insert_manual_commuting_rows(con, int(job_id), preview["ready_rows"])

        record_audit_event(
            con,
            request=request,
            actor=_user,
            action="import",
            entity_type="employee_commuting_direct_entry",
            entity_id=f"{int(job_id)}:{validated_site_id if validated_site_id is not None else 'all'}",
            client_id=meta.get("client_db_id"),
            job_id=int(job_id),
            metadata={
                "site_id": validated_site_id,
                "site_label": site_label,
                "replace_existing": bool(replace_existing),
                "inserted": int(inserted),
                "disabled": int(disabled),
                "parsed_count": int(preview.get("parsed_count") or 0),
                "ready_count": int(preview.get("ready_count") or 0),
                "unresolved_count": int(preview.get("unresolved_count") or 0),
                "total_tco2e": float(preview.get("total_tco2e") or 0.0),
                "data_source": DIRECT_COMMUTING_DATA_SOURCE,
            },
        )

    return {
        "ok": True,
        "job_id": int(job_id),
        "site_id": validated_site_id,
        "site_label": site_label,
        "inserted": int(inserted),
        "disabled": int(disabled),
        "total_tco2e": float(preview["total_tco2e"]),
        "data_source": DIRECT_COMMUTING_DATA_SOURCE,
    }


@router.patch("/jobs/{job_id}/employee-commuting/direct-entry/{source_id}")
def update_employee_commuting_direct_entry(
    request: Request,
    job_id: int,
    source_id: int,
    payload: dict = Body(...),
    site_id: int | None = Query(None),
    _user: dict[str, str] = Depends(_current_user),
):
    entry = payload.get("entry") if isinstance(payload, dict) else None
    if not isinstance(entry, dict):
        raise HTTPException(status_code=400, detail="A direct entry payload is required")

    with get_conn() as con:
        _ensure_job_scope_rows_schema(con)
        _ensure_emission_register_schema(con)
        meta = _job_meta(con, int(job_id))
        validated_site_id, site_label = _job_site_label(con, int(job_id), site_id)
        existing = con.execute(
            """
            SELECT source_id
            FROM job_emission_sources
            WHERE source_id = %s
              AND job_id = %s
              AND source_type = %s
              AND COALESCE(enabled, TRUE) = TRUE
            """,
            [int(source_id), int(job_id), "employee_commuting"],
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Saved direct entry not found")

        preview = _resolve_manual_commuting_rows(con, int(job_id), validated_site_id, [entry])

        if preview["unresolved_count"] > 0:
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "Updated direct entry contains unresolved rows. Please correct it before saving.",
                    "preview": {
                        "job_id": int(job_id),
                        "site_id": validated_site_id,
                        "site_label": site_label,
                        "template_version": "Direct Entry",
                        **preview,
                    },
                },
            )

        if preview["ready_count"] <= 0:
            raise HTTPException(status_code=400, detail="No importable direct employee commuting rows were found")

        row = preview["ready_rows"][0]
        _update_manual_commuting_row(con, int(job_id), int(source_id), row)

        record_audit_event(
            con,
            request=request,
            actor=_user,
            action="update",
            entity_type="employee_commuting_direct_entry",
            entity_id=f"{int(job_id)}:{int(source_id)}",
            client_id=meta.get("client_db_id"),
            job_id=int(job_id),
            metadata={
                "source_id": int(source_id),
                "site_id": validated_site_id,
                "site_label": site_label,
                "parsed_count": int(preview.get("parsed_count") or 0),
                "ready_count": int(preview.get("ready_count") or 0),
                "unresolved_count": int(preview.get("unresolved_count") or 0),
                "total_tco2e": float(preview.get("total_tco2e") or 0.0),
                "data_source": DIRECT_COMMUTING_DATA_SOURCE,
            },
        )

    return {
        "ok": True,
        "job_id": int(job_id),
        "source_id": int(source_id),
        "site_id": validated_site_id,
        "site_label": site_label,
        "total_tco2e": float(preview["total_tco2e"]),
        "data_source": DIRECT_COMMUTING_DATA_SOURCE,
    }


@router.get("/jobs/{job_id}/employee-commuting/pending-review")
def list_pending_review_commuting_rows(job_id: int, _user: dict[str, str] = Depends(_current_user)):
    """Client-portal-submitted commuting rows awaiting CRM approval."""
    with get_conn() as con:
        _ensure_emission_register_schema(con)
        df = con.execute(
            """
            SELECT source_id, site_id, employee_name, source_subtype, qty, uom,
                   factor, calc_tco2e, notes, review_status, review_note, created_at
            FROM job_emission_sources
            WHERE job_id = %s AND source_type = 'employee_commuting'
              AND submitted_by_portal = TRUE AND review_status IN ('pending_review', 'rejected')
            ORDER BY created_at DESC
            """,
            [int(job_id)],
        ).df()
        if df is None or df.empty:
            return {"job_id": int(job_id), "rows": []}
        # astype(object) first -- see api/portal_data_entry_routes.py for why the
        # plain df.where(df.notna(), None) is a no-op on float64 columns and
        # breaks JSON serialization for rows with a null qty/factor/calc_tco2e.
        df = df.astype(object).where(df.notna(), None)
        rows = []
        for _, r in df.iterrows():
            row = {k: r.get(k) for k in r.index}
            if row.get("created_at") is not None:
                row["created_at"] = str(row["created_at"])
            rows.append(row)
        return {"job_id": int(job_id), "rows": rows}


@router.patch("/jobs/{job_id}/employee-commuting/{source_id}/review")
def review_commuting_row(
    request: Request,
    job_id: int,
    source_id: int,
    body: dict = Body(...),
    _user: dict[str, str] = Depends(_current_user),
):
    """Approve or reject a client-portal-submitted employee-commuting row.
    Approving is the only thing that flips enabled to TRUE -- until then the
    row is invisible to load_combined_reporting_rows() and the commuting
    summary, both of which already filter enabled=TRUE."""
    decision = str(body.get("decision") or "").strip().lower()
    if decision not in ("approve", "reject"):
        raise HTTPException(status_code=400, detail="decision must be 'approve' or 'reject'")
    note = body.get("note")
    reviewer = str(_user.get("email") or _user.get("full_name") or "crm")

    with get_conn() as con:
        _ensure_emission_register_schema(con)
        row = con.execute(
            """
            SELECT submitted_by_portal, review_status
            FROM job_emission_sources
            WHERE source_id = %s AND job_id = %s AND source_type = 'employee_commuting'
            """,
            [int(source_id), int(job_id)],
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Commuting row not found")
        if not row[0]:
            raise HTTPException(status_code=400, detail="This row wasn't submitted via the client portal")
        if row[1] == "approved":
            raise HTTPException(status_code=409, detail="Already approved")

        if decision == "approve":
            con.execute(
                """
                UPDATE job_emission_sources
                SET enabled = TRUE, review_status = 'approved', review_note = %s,
                    reviewed_by = %s, reviewed_at = NOW()
                WHERE source_id = %s
                """,
                [note, reviewer, int(source_id)],
            )
        else:
            con.execute(
                """
                UPDATE job_emission_sources
                SET review_status = 'rejected', review_note = %s, reviewed_by = %s, reviewed_at = NOW()
                WHERE source_id = %s
                """,
                [note, reviewer, int(source_id)],
            )

        record_audit_event(
            con,
            request=request,
            actor=_user,
            action=f"commuting_row_{decision}",
            entity_type="job_emission_source",
            entity_id=int(source_id),
            job_id=int(job_id),
            metadata={"decision": decision, "note": note},
        )

    return {"ok": True, "source_id": int(source_id), "review_status": "approved" if decision == "approve" else "rejected"}
