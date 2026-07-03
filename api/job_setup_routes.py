from __future__ import annotations

import io
import os
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, Response
from openpyxl import load_workbook

from api.auth import _current_user
from api.job_setup_helpers import (
    _col_exists,
    _ensure_job_additional_datasets_table,
    _ensure_job_original_portfolio_column,
    _job_scope_config_audit_snapshot,
    _resolve_scope_dataset_map,
    _table_exists,
)
from api.job_template_helpers import _job_template_paths
from api.permissions import assert_client_access, assert_job_access, assert_permission
from core.database import db_backend, get_conn
from api.client_management_helpers import _list_sites
from services.audit_log import record_audit_event
from services.client_benchmark import ensure_client_benchmark_columns
from services.tenancy import require_org
from api.org_admin_helpers import _require_org_plan_active

router = APIRouter()


@router.get("/lookups/portfolios")
def list_portfolios_lookup(_user: dict = Depends(_current_user)):
    """Return active portfolio options from portfolios_lookup.

    Intentionally outside the /admin router so any authenticated user
    (including CRMs without admin.access) can fetch options for the
    Original Portfolio field in Job Setup.
    """
    assert_permission(_user, "jobs.view")
    org_id = require_org(_user)
    with get_conn() as con:
        rows = con.execute(
            """
            SELECT name FROM portfolios_lookup
            WHERE COALESCE(is_active, TRUE) = TRUE
            ORDER BY name ASC
            """,
        ).fetchall()
    names = [str(r[0] or "").strip() for r in rows if r[0]]
    if not names:
        names = ["NZI", "NZN"]
    return {"items": names}


@router.get("/jobs/{job_id}/sites")
def job_sites(job_id: int, _user: dict[str, str] = Depends(_current_user)):
    with get_conn() as con:
        row = con.execute(
            "SELECT client_db_id FROM jobs WHERE job_id=?",
            [int(job_id)],
        ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Job not found")

    client_db_id = int(row[0])
    df = _list_sites(client_db_id)

    sites: list[dict[str, object]] = []
    if df is not None and (not df.empty):
        for _, r in df.iterrows():
            sites.append(
                {
                    "site_id": int(r.get("site_id")) if r.get("site_id") is not None else None,
                    "site_name": r.get("site_name"),
                    "location": r.get("location"),
                    "is_registered_office": bool(r.get("is_registered_office")) if r.get("is_registered_office") is not None else False,
                }
            )

    return {"job_id": int(job_id), "client_db_id": client_db_id, "sites": sites}


@router.get("/datasets")
def list_datasets(_user: dict[str, str] = Depends(_current_user)):
    try:
        with get_conn() as con:
            df = con.execute(
                """
                SELECT dataset_id, name, source, analysis_type, country, region,
                       currency, year, version, archived, archived_at, archived_by
                FROM datasets
                ORDER BY dataset_id DESC
                """
            ).df()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"/datasets failed: {e}")

    items: list[dict[str, object]] = []
    if df is not None and (not df.empty):
        for _, r in df.iterrows():
            items.append(
                {
                    "dataset_id": int(r.get("dataset_id")),
                    "name": r.get("name"),
                    "source": r.get("source"),
                    "analysis_type": r.get("analysis_type"),
                    "country": r.get("country"),
                    "region": r.get("region"),
                    "currency": r.get("currency"),
                    "year": (int(r.get("year")) if r.get("year") is not None else None),
                    "version": r.get("version"),
                    "archived": bool(r.get("archived")) if r.get("archived") is not None else False,
                    "archived_at": r.get("archived_at"),
                    "archived_by": r.get("archived_by"),
                }
            )

    return {"items": items}


@router.get("/jobs/{job_id}/scope-config")
def get_job_scope_config(job_id: int, _user: dict[str, str] = Depends(_current_user)):
    try:
        _ensure_job_additional_datasets_table()
        with get_conn() as con:
            exists_job = con.execute("SELECT 1 FROM jobs WHERE job_id=?", [int(job_id)]).fetchone()
            if not exists_job:
                raise HTTPException(status_code=404, detail="Job not found")

            df = con.execute(
                """
                SELECT scope, include_scope, dataset_id, factor_method
                FROM job_scope_config
                WHERE job_id=?
                ORDER BY scope
                """,
                [int(job_id)],
            ).df()

            add_df = con.execute(
                """
                SELECT dataset_id
                FROM job_additional_datasets
                WHERE job_id=?
                ORDER BY dataset_id
                """,
                [int(job_id)],
            ).df()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load scope config: {e}")

    allowed_scopes = ["Scope 1", "Scope 2", "Scope 3"]
    legacy_by_scope: dict[str, dict[str, object]] = {
        scope: {
            "scope": scope,
            "include_scope": True,
            "dataset_id": None,
            "factor_method": None,
        }
        for scope in allowed_scopes
    }

    if df is not None and (not df.empty):
        for _, r in df.iterrows():
            scope = str(r.get("scope") or "").strip()
            if scope not in legacy_by_scope:
                continue
            dsid_raw = r.get("dataset_id")
            dsid = (
                int(dsid_raw)
                if dsid_raw is not None and str(dsid_raw) != "nan"
                else None
            )
            legacy_by_scope[scope] = {
                "scope": scope,
                "include_scope": bool(r.get("include_scope")) if r.get("include_scope") is not None else True,
                "dataset_id": dsid,
                "factor_method": r.get("factor_method"),
            }

    effective_ds_map, auto_resolution, auto_warnings = _resolve_scope_dataset_map(int(job_id))

    items: list[dict[str, object]] = []
    legacy_items: list[dict[str, object]] = []
    for scope in allowed_scopes:
        legacy_item = dict(legacy_by_scope.get(scope) or {})
        legacy_items.append(legacy_item)

        effective_item = dict(legacy_item)
        if scope in effective_ds_map:
            effective_item["dataset_id"] = int(effective_ds_map[scope])
        items.append(effective_item)

    auto_payload = None
    if auto_resolution:
        auto_payload = {
            "country": auto_resolution.get("country"),
            "reporting_period_start": auto_resolution.get("reporting_period_start"),
            "reporting_period_end": auto_resolution.get("reporting_period_end"),
            "uses_legacy_fallback": bool(auto_resolution.get("uses_legacy_fallback")),
            "scope_summaries": auto_resolution.get("scope_summaries") or [],
            "datasets_for_report": auto_resolution.get("datasets_for_report") or [],
            "unresolved_scopes": auto_resolution.get("unresolved_scopes") or [],
        }

    additional_dataset_ids: list[int] = []
    if add_df is not None and (not add_df.empty):
        for _, row in add_df.iterrows():
            raw = row.get("dataset_id")
            if raw is None or str(raw) == "nan":
                continue
            try:
                additional_dataset_ids.append(int(raw))
            except Exception:
                continue

    return {
        "job_id": int(job_id),
        "mode": "automatic" if auto_resolution else "legacy",
        "items": items,
        "legacy_items": legacy_items,
        "additional_dataset_ids": additional_dataset_ids,
        "warnings": auto_warnings,
        "auto_resolution": auto_payload,
    }


@router.put("/jobs/{job_id}/scope-config")
def update_job_scope_config(
    request: Request,
    job_id: int,
    payload: dict[str, object] = Body(...),
    _user: dict[str, str] = Depends(_current_user),
):
    raw_items = payload.get("items")
    if not isinstance(raw_items, list):
        raise HTTPException(status_code=400, detail="items must be a list")
    raw_additional = payload.get("additional_dataset_ids")
    if raw_additional is None:
        raw_additional = []
    if not isinstance(raw_additional, list):
        raise HTTPException(status_code=400, detail="additional_dataset_ids must be a list")

    allowed_scopes = {"Scope 1", "Scope 2", "Scope 3"}
    updates: list[tuple[str, int | None, bool | None, str | None]] = []
    for it in raw_items:
        if not isinstance(it, dict):
            continue
        scope = str(it.get("scope") or "").strip()
        if scope not in allowed_scopes:
            continue

        ds_raw = it.get("dataset_id")
        dsid: int | None
        if ds_raw is None or str(ds_raw).strip() == "" or str(ds_raw).strip().lower() in ("none", "null"):
            dsid = None
        else:
            try:
                dsid = int(ds_raw)
            except Exception:
                raise HTTPException(status_code=400, detail=f"Invalid dataset_id for {scope}")

        inc_raw = it.get("include_scope")
        include_scope: bool | None
        if inc_raw is None:
            include_scope = None
        else:
            include_scope = bool(inc_raw)

        fm_raw = it.get("factor_method")
        factor_method: str | None = str(fm_raw).strip() if fm_raw is not None and str(fm_raw).strip() else None
        updates.append((scope, dsid, include_scope, factor_method))

    if not updates:
        raise HTTPException(status_code=400, detail="No valid scope config items")

    additional_dataset_ids: list[int] = []
    for raw in raw_additional:
        if raw is None or str(raw).strip() in ("", "none", "null"):
            continue
        try:
            additional_dataset_ids.append(int(raw))
        except Exception:
            raise HTTPException(status_code=400, detail=f"Invalid additional dataset_id: {raw}")
    additional_dataset_ids = sorted(set(additional_dataset_ids))

    try:
        _ensure_job_additional_datasets_table()
        with get_conn() as con:
            before = _job_scope_config_audit_snapshot(con, int(job_id))
            exists_job = con.execute("SELECT 1 FROM jobs WHERE job_id=?", [int(job_id)]).fetchone()
            if not exists_job:
                raise HTTPException(status_code=404, detail="Job not found")

            # Ensure rows exist
            for scope in allowed_scopes:
                con.execute(
                    """
                    INSERT INTO job_scope_config (job_id, scope, include_scope, dataset_id, factor_method)
                    VALUES (?, ?, TRUE, NULL, NULL)
                    ON CONFLICT (job_id, scope) DO NOTHING
                    """,
                    [int(job_id), scope],
                )

            for scope, dsid, include_scope, factor_method in updates:
                if include_scope is None and factor_method is None:
                    con.execute(
                        "UPDATE job_scope_config SET dataset_id=? WHERE job_id=? AND scope=?",
                        [dsid, int(job_id), scope],
                    )
                else:
                    con.execute(
                        """
                        UPDATE job_scope_config
                        SET dataset_id=?,
                            include_scope=COALESCE(?, include_scope),
                            factor_method=COALESCE(?, factor_method)
                        WHERE job_id=? AND scope=?
                        """,
                        [dsid, include_scope, factor_method, int(job_id), scope],
                    )

            con.execute("DELETE FROM job_additional_datasets WHERE job_id=?", [int(job_id)])
            for dsid in additional_dataset_ids:
                con.execute(
                    """
                    INSERT INTO job_additional_datasets (job_id, dataset_id)
                    VALUES (?, ?)
                    ON CONFLICT (job_id, dataset_id) DO NOTHING
                    """,
                    [int(job_id), int(dsid)],
                )
            after = _job_scope_config_audit_snapshot(con, int(job_id))
            record_audit_event(
                con,
                request=request,
                actor=_user,
                action="update",
                entity_type="job_scope_config",
                entity_id=int(job_id),
                job_id=int(job_id),
                before=before,
                after=after,
                metadata={
                    "updated_scopes": sorted([scope for scope, _, _, _ in updates]),
                    "additional_dataset_ids": additional_dataset_ids,
                },
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update scope config: {e}")

    return {"ok": True, "job_id": int(job_id)}


@router.post("/jobs/{job_id}/excel-import-preflight")
def job_excel_import_preflight(
    job_id: int,
    payload: dict[str, object] = Body(...),
    _user: dict[str, str] = Depends(_current_user),
):
    """
    Returns rows that would overwrite an existing scope row that already has
    a non-zero qty. The frontend uses this to warn the user before importing.
    """
    site_id_raw = payload.get("site_id")
    rows_ready = payload.get("rows_ready")
    if site_id_raw is None or not isinstance(rows_ready, list):
        raise HTTPException(status_code=400, detail="site_id and rows_ready are required")
    try:
        site_id = int(site_id_raw)
    except Exception:
        raise HTTPException(status_code=400, detail="site_id must be an integer")

    conflicts = []
    try:
        with get_conn() as con:
            for r in rows_ready:
                if not isinstance(r, dict):
                    continue
                scope = str(r.get("scope") or "").strip()
                original_id = str(r.get("original_id") or "").strip()
                if not scope or not original_id:
                    continue
                existing = con.execute(
                    """
                    SELECT row_id, qty, report_label
                    FROM job_scope_rows
                    WHERE job_id=%s AND site_id=%s AND scope=%s AND original_id=%s
                      AND COALESCE(enabled, TRUE) = TRUE
                    LIMIT 1
                    """,
                    [int(job_id), site_id, scope, original_id],
                ).fetchone()
                if existing and existing[1] is not None and float(existing[1]) != 0:
                    conflicts.append({
                        "scope": scope,
                        "original_id": original_id,
                        "report_label": r.get("report_label") or (str(existing[2]) if existing[2] else original_id),
                        "existing_qty": float(existing[1]),
                        "upload_qty": float(r["qty"]) if r.get("qty") is not None else None,
                    })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Preflight check failed: {e}")

    return {"conflicts": conflicts}


@router.post("/jobs/{job_id}/excel-import")
def job_excel_import(
    job_id: int,
    payload: dict[str, object] = Body(...),
    _user: dict[str, str] = Depends(_current_user),
):
    site_id_raw = payload.get("site_id")
    rows_ready = payload.get("rows_ready")

    if site_id_raw is None:
        raise HTTPException(status_code=400, detail="site_id is required")
    try:
        site_id = int(site_id_raw)
    except Exception:
        raise HTTPException(status_code=400, detail="site_id must be an integer")

    if not isinstance(rows_ready, list):
        raise HTTPException(status_code=400, detail="rows_ready must be a list")
    if not rows_ready:
        raise HTTPException(status_code=400, detail="rows_ready is empty")

    inserted = 0
    updated = 0
    month_fields = [f"month_{idx}" for idx in range(1, 13)]
    month_placeholders = ", ".join(["?"] * len(month_fields))

    try:
        with get_conn() as con:
            job_row = con.execute("SELECT client_db_id FROM jobs WHERE job_id=?", [int(job_id)]).fetchone()
            if not job_row:
                raise HTTPException(status_code=404, detail="Job not found")
            client_db_id = int(job_row[0])

            site_ok = con.execute(
                "SELECT 1 FROM client_sites WHERE site_id=? AND client_db_id=?",
                [int(site_id), int(client_db_id)],
            ).fetchone()
            if not site_ok:
                raise HTTPException(status_code=400, detail="site_id does not belong to this job's client")

            for r in rows_ready:
                if not isinstance(r, dict):
                    continue

                scope = str(r.get("scope") or "").strip()
                original_id = str(r.get("original_id") or "").strip()
                if not scope or not original_id:
                    continue

                dataset_id = r.get("dataset_id")
                factor_db_id = r.get("db_id")
                if dataset_id is None or factor_db_id is None:
                    continue

                qty = r.get("qty")
                uom = r.get("uom")
                factor = r.get("factor")
                ghg_unit = r.get("ghg_unit")
                calc_tco2e = r.get("calc_tco2e")
                notes = r.get("notes")
                data_source = r.get("data_source")
                data_confidence = r.get("data_confidence")
                apply_pct = r.get("apply_pct")
                report_label = r.get("report_label")

                level_1 = r.get("level_1")
                level_2 = r.get("level_2")
                level_3 = r.get("level_3")
                level_4 = r.get("level_4")
                column_text = r.get("column_text")
                month_values = [r.get(field) for field in month_fields]

                exists = con.execute(
                    """
                    SELECT row_id
                    FROM job_scope_rows
                    WHERE job_id=? AND site_id=? AND scope=? AND original_id=?
                    LIMIT 1
                    """,
                    [int(job_id), int(site_id), str(scope), str(original_id)],
                ).fetchone()

                if exists:
                    con.execute(
                        """
                        UPDATE job_scope_rows
                        SET enabled=TRUE,
                            dataset_id=?,
                            factor_db_id=?,
                            qty=?,
                            uom=?,
                            factor=?,
                            ghg_unit=?,
                            calc_tco2e=?,
                            apply_pct=?,
                            data_source=?,
                            data_confidence=?,
                            notes=?,
                            month_1=?,
                            month_2=?,
                            month_3=?,
                            month_4=?,
                            month_5=?,
                            month_6=?,
                            month_7=?,
                            month_8=?,
                            month_9=?,
                            month_10=?,
                            month_11=?,
                            month_12=?,
                            level_1=?,
                            level_2=?,
                            level_3=?,
                            level_4=?,
                            report_label=?,
                            column_text=?,
                            updated_at=NOW()
                        WHERE row_id=?
                        """,
                        [
                            int(dataset_id),
                            int(factor_db_id),
                            float(qty) if qty is not None else None,
                            (str(uom).strip() if uom is not None else None),
                            float(factor) if factor is not None else None,
                            (str(ghg_unit).strip() if ghg_unit is not None else None),
                            float(calc_tco2e) if calc_tco2e is not None else None,
                            float(apply_pct) if apply_pct is not None else None,
                            (str(data_source).strip() if data_source is not None else None),
                            (str(data_confidence).strip() if data_confidence is not None else None),
                            (str(notes).strip() if notes is not None else None),
                            float(month_values[0]) if month_values[0] is not None else None,
                            float(month_values[1]) if month_values[1] is not None else None,
                            float(month_values[2]) if month_values[2] is not None else None,
                            float(month_values[3]) if month_values[3] is not None else None,
                            float(month_values[4]) if month_values[4] is not None else None,
                            float(month_values[5]) if month_values[5] is not None else None,
                            float(month_values[6]) if month_values[6] is not None else None,
                            float(month_values[7]) if month_values[7] is not None else None,
                            float(month_values[8]) if month_values[8] is not None else None,
                            float(month_values[9]) if month_values[9] is not None else None,
                            float(month_values[10]) if month_values[10] is not None else None,
                            float(month_values[11]) if month_values[11] is not None else None,
                            (str(level_1).strip() if level_1 is not None else None),
                            (str(level_2).strip() if level_2 is not None else None),
                            (str(level_3).strip() if level_3 is not None else None),
                            (str(level_4).strip() if level_4 is not None else None),
                            (str(report_label).strip() if report_label is not None else None),
                            (str(column_text).strip() if column_text is not None else None),
                            int(exists[0]),
                        ],
                    )
                    updated += 1
                else:
                    con.execute(
                        f"""
                        INSERT INTO job_scope_rows
                          (job_id, site_id, scope, category, dataset_id, factor_db_id, original_id,
                           level_1, level_2, level_3, level_4, column_text,
                           report_label, notes, enabled,
                           qty, uom, factor, ghg_unit, apply_pct, data_source, data_confidence,
                           month_1, month_2, month_3, month_4, month_5, month_6,
                           month_7, month_8, month_9, month_10, month_11, month_12,
                           calc_tco2e, override_tco2e, override_reason,
                           created_at, updated_at)
                        VALUES
                          (?, ?, ?, ?, ?, ?, ?,
                           ?, ?, ?, ?, ?,
                           ?, ?, TRUE,
                           ?, ?, ?, ?, ?, ?, ?,
                           {month_placeholders},
                           ?, NULL, NULL,
                           NOW(), NOW())
                        """,
                        [
                            int(job_id),
                            int(site_id),
                            str(scope),
                            (str(level_1).strip() if level_1 is not None else None),  # category = top-level dataset category
                            int(dataset_id),
                            int(factor_db_id),
                            str(original_id),
                            (str(level_1).strip() if level_1 is not None else None),
                            (str(level_2).strip() if level_2 is not None else None),
                            (str(level_3).strip() if level_3 is not None else None),
                            (str(level_4).strip() if level_4 is not None else None),
                            (str(column_text).strip() if column_text is not None else None),
                            (str(report_label).strip() if report_label is not None else None),
                            (str(notes).strip() if notes is not None else None),
                            float(qty) if qty is not None else None,
                            (str(uom).strip() if uom is not None else None),
                            float(factor) if factor is not None else None,
                            (str(ghg_unit).strip() if ghg_unit is not None else None),
                            float(apply_pct) if apply_pct is not None else None,
                            (str(data_source).strip() if data_source is not None else None),
                            (str(data_confidence).strip() if data_confidence is not None else None),
                            float(month_values[0]) if month_values[0] is not None else None,
                            float(month_values[1]) if month_values[1] is not None else None,
                            float(month_values[2]) if month_values[2] is not None else None,
                            float(month_values[3]) if month_values[3] is not None else None,
                            float(month_values[4]) if month_values[4] is not None else None,
                            float(month_values[5]) if month_values[5] is not None else None,
                            float(month_values[6]) if month_values[6] is not None else None,
                            float(month_values[7]) if month_values[7] is not None else None,
                            float(month_values[8]) if month_values[8] is not None else None,
                            float(month_values[9]) if month_values[9] is not None else None,
                            float(month_values[10]) if month_values[10] is not None else None,
                            float(month_values[11]) if month_values[11] is not None else None,
                            float(calc_tco2e) if calc_tco2e is not None else None,
                        ],
                    )
                    inserted += 1
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to import rows: {e}")

    return {"ok": True, "job_id": int(job_id), "site_id": int(site_id), "inserted": int(inserted), "updated": int(updated)}


@router.get("/jobs/{job_id}/excel-template")
def job_excel_template(
    job_id: int,
    site: str = Query(..., min_length=1),
    include_prev_year: bool = Query(True),
    template_format: str = Query("single", regex="^(single|multi)$"),
    template_id: int | None = Query(None),
    _user: dict[str, str] = Depends(_current_user),
):
    _tmp_template_file = None
    try:
        # Get job/client metadata for filename convention.
        with get_conn() as con:
            job_row = con.execute(
                """
                SELECT
                    j.job_number,
                    j.reporting_year,
                    j.reporting_period_start,
                    j.reporting_period_end,
                    c.client_name,
                    j.job_template_id
                FROM jobs j
                LEFT JOIN clients c ON c.db_id = j.client_db_id
                WHERE j.job_id = ?
                """,
                [job_id],
            ).fetchone()
        job_number = str(job_row[0] or f"Job-{job_id}") if job_row else f"Job-{job_id}"
        reporting_year = job_row[1] if job_row and len(job_row) > 1 else ""
        reporting_period_start = job_row[2] if job_row and len(job_row) > 2 else None
        reporting_period_end = job_row[3] if job_row and len(job_row) > 3 else None
        client_name = str(job_row[4] or "Client") if job_row and len(job_row) > 4 else "Client"
        job_template_id = int(job_row[5]) if job_row and len(job_row) > 5 and job_row[5] is not None else None
        resolved_template_id = int(template_id) if template_id is not None else job_template_id

        if resolved_template_id is None:
            raise HTTPException(status_code=400, detail="Please select a Job Template before downloading.")

        with get_conn() as con:
            template_row = con.execute(
                """
                SELECT
                    jt.excel_template_path,
                    jt.file_content,
                    jt.template_name,
                    jt.template_key
                FROM job_templates jt
                WHERE jt.job_template_id = ?
                  AND COALESCE(jt.is_active, TRUE) = TRUE
                LIMIT 1
                """,
                [int(resolved_template_id)],
            ).fetchone()
        if not template_row:
            raise HTTPException(status_code=400, detail="Please select a valid Job Template before downloading.")

        template_excel_path = template_row[0] if len(template_row) > 0 else None
        template_file_content = template_row[1] if len(template_row) > 1 else None

        def _fmt_period_part(val):
            if not val:
                return ""
            if hasattr(val, "strftime"):
                return val.strftime("%d-%b-%Y")
            txt = str(val).strip()
            try:
                return datetime.fromisoformat(txt[:10]).strftime("%d-%b-%Y")
            except Exception:
                return txt

        def _safe_name_part(value: str) -> str:
            s = str(value or "").strip()
            s = re.sub(r'[<>:"/\\|?*]+', "", s)
            s = re.sub(r"\s+", "-", s)
            return s.strip("-") or "Unknown"

        period_part = f"{_fmt_period_part(reporting_period_start)}-to-{_fmt_period_part(reporting_period_end)}"
        if period_part == "-to-":
            period_part = str(reporting_year or datetime.now().year)

        paths = _job_template_paths(int(job_id))
        reference_template_path = str(template_excel_path or "").strip() or paths.get("excel_template_path")

        # If the assigned template has DB content (uploaded via Admin), write to a temp
        # file so the generator uses the correct file instead of a stale disk copy.
        if template_file_content:
            import tempfile
            file_content_bytes = bytes(template_file_content) if not isinstance(template_file_content, bytes) else template_file_content
            _tmp_template_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            _tmp_template_file.write(file_content_bytes)
            _tmp_template_file.close()
            reference_template_path = _tmp_template_file.name

        if template_format == "single":
            from services.generate_single_sheet_template import generate_single_sheet_template

            data, filename = generate_single_sheet_template(
                job_id=int(job_id),
                client_name=client_name,
                site_name=site,
                job_number=job_number,
                reporting_year=str(reporting_year) if reporting_year else "",
                report_from="",
                report_to="",
                include_custom_factors=True,
                include_prev_year=bool(include_prev_year),
                reference_template_path=str(reference_template_path) if reference_template_path else None,
            )
        else:
            # Legacy multi-sheet template
            os.environ["NZI_EXCEL_TEMPLATE_PATH"] = str(paths.get("excel_template_path") or "")
            data, filename = build_excel_template_bytes(
                job_id=int(job_id),
                selected_site=str(site),
                include_prev_year=bool(include_prev_year),
            )
        filename = "_".join(
            [
                _safe_name_part(job_number),
                _safe_name_part(client_name),
                _safe_name_part(str(site)),
                _safe_name_part(period_part),
                "data_upload",
            ]
        ) + ".xlsx"
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to build template: {e}")
    finally:
        if _tmp_template_file is not None:
            try:
                os.unlink(_tmp_template_file.name)
            except Exception:
                pass

    # Simple Content-Disposition header with quoted filename
    safe_filename = filename.replace('"', '\\"')
    headers = {
        "Content-Disposition": f'attachment; filename="{safe_filename}"'
    }
    print(f"DEBUG: Content-Disposition header: {headers['Content-Disposition']}")
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/jobs/{job_id}/excel-upload")
async def job_excel_upload(
    job_id: int,
    file: UploadFile = File(...),
    _user: dict[str, str] = Depends(_current_user),
):
    from api.parse_single_sheet_upload import is_single_sheet_format, parse_single_sheet_upload
    
    paths = _job_template_paths(int(job_id))
    filename = (file.filename or "").strip()
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty upload")

    errors: list[str] = []
    warnings: list[str] = []
    details: dict[str, object] = {
        "filename": filename,
        "size_bytes": int(len(raw)),
        "job_id": int(job_id),
        "job_excel_template_path": paths.get("excel_template_path"),
    }

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid xlsx file: {e}")

    details["sheets"] = list(wb.sheetnames)

    # Year-mismatch check: compare filename years against the job's reporting period.
    # The standard template filename contains the period dates, e.g.:
    #   J000002_ClientName_SiteName_2021-01-01-to-2021-12-31_data_upload.xlsx
    # Multi-year reporting periods are allowed, so we accept either boundary year.
    try:
        import re as _re
        with get_conn() as _con:
            _job_period = _con.execute(
                "SELECT reporting_period_start, reporting_period_end, reporting_year FROM jobs WHERE job_id=%s",
                [int(job_id)],
            ).fetchone()
        if _job_period:
            _job_start = _job_period[0]
            _job_end = _job_period[1]
            _job_year = int(_job_period[2]) if _job_period[2] else None
            _job_start_year = int(str(_job_start)[:4]) if _job_start else None
            _job_end_year = int(str(_job_end)[:4]) if _job_end else None
            allowed_years = {
                year
                for year in (_job_year, _job_start_year, _job_end_year)
                if year is not None
            }
            # Look for 4-digit years in the filename and allow any matching boundary year.
            _years_in_name = [int(y) for y in _re.findall(r'\b(20\d{2})\b', filename)]
            if _years_in_name and allowed_years:
                _file_year = _years_in_name[-1]
                if _file_year not in allowed_years:
                    period_label = (
                        f"{_job_start_year}-{_job_end_year}"
                        if _job_start_year and _job_end_year and _job_start_year != _job_end_year
                        else str(_job_year or _job_end_year or _job_start_year or "unknown")
                    )
                    errors.append(
                        f"Year mismatch: this file appears to be for {_file_year} "
                        f"but the job covers {period_label}. "
                        f"Please upload the correct file for the {period_label} reporting period."
                    )
    except Exception:
        pass  # Don't block upload if year check itself fails

    ds_map, auto_resolution, auto_ds_warnings = _resolve_scope_dataset_map(int(job_id))
    warnings.extend(auto_ds_warnings)
    details["datasets_by_scope"] = ds_map
    details["dataset_resolution_mode"] = "automatic" if auto_resolution else "legacy"
    if auto_resolution:
        details["dataset_resolution"] = {
            "country": auto_resolution.get("country"),
            "reporting_period_start": auto_resolution.get("reporting_period_start"),
            "reporting_period_end": auto_resolution.get("reporting_period_end"),
            "uses_legacy_fallback": bool(auto_resolution.get("uses_legacy_fallback")),
            "scope_summaries": auto_resolution.get("scope_summaries") or [],
            "datasets_for_report": auto_resolution.get("datasets_for_report") or [],
        }
    
    # Detect format and route to appropriate parser
    if is_single_sheet_format(wb):
        details["template_format"] = "single-sheet"
        
        # Parse single-sheet format
        try:
            rows_ready, parse_errors, parse_warnings, parse_details = parse_single_sheet_upload(
                raw, int(job_id), ds_map
            )
        except Exception as parse_exc:
            raise HTTPException(status_code=400, detail=f"Failed to parse workbook: {parse_exc}")
        
        errors.extend(parse_errors)
        warnings.extend(parse_warnings)
        details.update(parse_details)
        
        details["rows_ready_count"] = len(rows_ready)
        
        if rows_ready:
            details["rows_ready"] = rows_ready[:10]  # Preview first 10
        
        return {
            "ok": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
            "details": details,
            "rows_ready": rows_ready if len(errors) == 0 else []
        }
    
    # Otherwise, use legacy multi-sheet parser
    details["template_format"] = "multi-sheet"

    def _norm_scope(sheet_name: str) -> str | None:
        s = (sheet_name or "").strip().lower()
        if "scope 1" in s:
            return "Scope 1"
        if "scope 2" in s:
            return "Scope 2"
        if "scope 3" in s:
            return "Scope 3"
        return None

    scopes_present: set[str] = set()
    for name in wb.sheetnames:
        sn = _norm_scope(name)
        if sn:
            scopes_present.add(sn)

    for s in ["Scope 1", "Scope 2", "Scope 3"]:
        if s not in scopes_present:
            errors.append(f"Missing required scope sheets for: {s}")

    # Basic marker validation against any scope sheet
    for sheet_name in wb.sheetnames:
        scope_name = _norm_scope(sheet_name)
        if scope_name is None:
            continue
        ws = wb[sheet_name]
        a1 = (ws["A1"].value or "")
        a2 = (ws["A2"].value or "")
        if "Site Name" not in str(a1):
            warnings.append(f"{sheet_name}: A1 is not 'Site Name:' marker")
        if "Data Files" not in str(a2):
            warnings.append(f"{sheet_name}: A2 is not 'Data Files:' marker")

    # Optional: core sheet if present
    if "Core Data" not in wb.sheetnames:
        warnings.append("Core Data sheet not found (template may be older)")

    def _find_table_header_row(ws):
        for r in range(1, 60):
            values = [ws.cell(row=r, column=c).value for c in range(1, 80)]
            norm = [str(x).strip().lower() if x is not None else "" for x in values]
            if "id" in norm and "qty" in norm:
                idx = {}
                for name in ("id", "qty", "apply"):
                    if name in norm:
                        idx[name] = norm.index(name) + 1
                return r, idx
        return None, None

    def _to_tco2e(qty: float, factor: float, ghg_unit: str | None) -> float:
        ghg = (str(ghg_unit or "kgCO2e").replace(" ", "").lower())
        emissions = float(qty) * float(factor)
        if ghg.startswith("kg"):
            return emissions / 1000.0
        return emissions

    def _factor_lookup_by_original_ids(dataset_id: int, scope_name: str, original_ids: list[str]) -> pd.DataFrame:
        original_ids = [str(x).strip() for x in (original_ids or []) if x is not None and str(x).strip()]
        if not original_ids:
            return pd.DataFrame()

        with get_conn() as con:
            if db_backend() == "postgres":
                sql = """
                    SELECT db_id, original_id, level_1, level_2, level_3, level_4,
                           column_text, uom, ghg_unit, factor
                    FROM v_factor_lookup
                    WHERE dataset_id=%s AND scope=%s AND original_id = ANY(%s)
                """
                return con.execute(sql, [int(dataset_id), str(scope_name), original_ids]).df()

            ph = ",".join(["?"] * len(original_ids))
            sql = f"""
                SELECT db_id, original_id, level_1, level_2, level_3, level_4,
                       column_text, uom, ghg_unit, factor
                FROM v_factor_lookup
                WHERE dataset_id=? AND scope=? AND original_id IN ({ph})
            """
            return con.execute(sql, [int(dataset_id), str(scope_name)] + original_ids).df()

    parsed_rows: list[dict[str, object]] = []
    for ws in wb.worksheets:
        scope_name = _norm_scope(ws.title)
        if scope_name is None:
            continue

        header_row, idx = _find_table_header_row(ws)
        if header_row is None or idx is None:
            continue

        id_col = idx.get("id")
        qty_col = idx.get("qty")
        apply_col = idx.get("apply")
        if not id_col or not qty_col:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            oid = ws.cell(row=r, column=id_col).value
            if oid is None or str(oid).strip() == "":
                continue

            qv = ws.cell(row=r, column=qty_col).value
            av = ws.cell(row=r, column=apply_col).value if apply_col else None

            if apply_col and av is not None:
                try:
                    if float(av) != 1.0:
                        continue
                except Exception:
                    if str(av).strip() not in ("1", "true", "True", "YES", "Yes"):
                        continue

            try:
                qty_val = float(qv) if qv is not None and str(qv).strip() != "" else None
            except Exception:
                qty_val = None

            if qty_val is None:
                continue

            parsed_rows.append({"scope": scope_name, "original_id": str(oid).strip(), "qty": qty_val})

    details["parsed_row_count"] = int(len(parsed_rows))
    if not parsed_rows:
        errors.append("No rows found to import. Ensure you have filled 'Qty' and set 'Apply' to 1 where applicable.")

    rows_ready: list[dict[str, object]] = []
    missing_ids: dict[str, list[str]] = {}

    for scope_name in ["Scope 1", "Scope 2", "Scope 3"]:
        scope_rows = [r for r in parsed_rows if r.get("scope") == scope_name]
        if not scope_rows:
            continue

        dsid = ds_map.get(scope_name)
        if dsid is None:
            errors.append(f"{scope_name}: no dataset selected in Job Folder -> Data Collection.")
            continue

        ids = [str(r.get("original_id") or "").strip() for r in scope_rows]
        fdf = _factor_lookup_by_original_ids(int(dsid), scope_name, ids)
        if fdf is None or fdf.empty:
            errors.append(f"{scope_name}: none of the uploaded IDs matched factor_lookup for dataset {int(dsid)}.")
            continue

        factor_by_id: dict[str, dict[str, object]] = {}
        for _, fr in fdf.iterrows():
            oid = str(fr.get("original_id") or "").strip()
            if not oid:
                continue
            factor_by_id[oid] = {
                "db_id": int(fr.get("db_id")) if fr.get("db_id") is not None and str(fr.get("db_id")) != "nan" else None,
                "original_id": oid,
                "level_1": fr.get("level_1"),
                "level_2": fr.get("level_2"),
                "level_3": fr.get("level_3"),
                "level_4": fr.get("level_4"),
                "column_text": fr.get("column_text"),
                "uom": fr.get("uom"),
                "ghg_unit": fr.get("ghg_unit"),
                "factor": float(fr.get("factor")) if fr.get("factor") is not None and str(fr.get("factor")) != "nan" else None,
            }

        missing = [oid for oid in ids if oid not in factor_by_id]
        if missing:
            missing_ids[scope_name] = missing[:200]
            errors.append(f"{scope_name}: {len(missing)} IDs were not found in the selected dataset.")

        for r in scope_rows:
            oid = str(r.get("original_id") or "").strip()
            qty_val = r.get("qty")
            if oid not in factor_by_id:
                continue
            f = factor_by_id[oid]
            if f.get("factor") is None:
                errors.append(f"{scope_name}: factor missing for ID {oid}.")
                continue

            calc = _to_tco2e(float(qty_val), float(f.get("factor")), f.get("ghg_unit"))
            rows_ready.append(
                {
                    "scope": scope_name,
                    "dataset_id": int(dsid),
                    "db_id": f.get("db_id"),
                    "original_id": oid,
                    "qty": float(qty_val),
                    "uom": f.get("uom"),
                    "ghg_unit": f.get("ghg_unit"),
                    "factor": f.get("factor"),
                    "calc_tco2e": float(calc),
                    "level_1": f.get("level_1"),
                    "level_2": f.get("level_2"),
                    "level_3": f.get("level_3"),
                    "level_4": f.get("level_4"),
                    "column_text": f.get("column_text"),
                }
            )

    details["rows_ready_count"] = int(len(rows_ready))

    ok = len(errors) == 0
    return {
        "ok": ok,
        "errors": errors,
        "warnings": warnings,
        "details": details,
        "missing_ids": missing_ids,
        "rows_ready": rows_ready,
    }
