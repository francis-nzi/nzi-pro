"""
Parse single-sheet Excel upload format.
Handles the new simplified template with all scopes in one sheet.
"""

import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from core.database import get_conn, db_backend


def is_single_sheet_format(wb) -> bool:
    """
    Detect if uploaded workbook is single-sheet format.
    Single-sheet format has a sheet named "Data Upload" or has only 1 sheet with columns:
    Scope, Category, Report Label, ID, UOM, GHG Unit, Factor, Qty
    """
    if "Data Upload" in wb.sheetnames:
        return True
    
    # If only 1 sheet and has the right columns, assume single-sheet
    if len(wb.sheetnames) == 1:
        ws = wb[wb.sheetnames[0]]
        # Check for header row with expected columns
        for r in range(1, 10):
            row_values = [ws.cell(row=r, column=c).value for c in range(1, 25)]
            norm = [str(x).strip().lower() if x else "" for x in row_values]
            if "scope" in norm and "report label" in norm and "id" in norm and "qty" in norm:
                return True
    
    return False


def _to_float(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "-":
        return None
    try:
        out = float(value)
        if pd.isna(out):
            return None
        return out
    except (ValueError, TypeError):
        return None


def _merge_notes(values: list[str | None]) -> str | None:
    notes: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        notes.append(text)
        seen.add(text)
    return " | ".join(notes) if notes else None


def parse_single_sheet_upload(
    raw_bytes: bytes,
    job_id: int,
    datasets_by_scope: dict[str, int]
) -> tuple[list[dict], list[str], list[str], dict]:
    """
    Parse single-sheet Excel upload.
    
    Args:
        raw_bytes: Excel file bytes
        job_id: Job ID
        datasets_by_scope: Map of scope name to dataset_id
    
    Returns:
        (rows_ready, errors, warnings, details)
    """
    
    errors = []
    warnings = []
    details = {}
    rows_ready = []
    
    try:
        wb = load_workbook(BytesIO(raw_bytes), data_only=True)
    except Exception as e:
        errors.append(f"Failed to load Excel file: {e}")
        return rows_ready, errors, warnings, details
    
    # Find the data sheet
    ws = None
    if "Data Upload" in wb.sheetnames:
        ws = wb["Data Upload"]
    elif len(wb.sheetnames) == 1:
        ws = wb[wb.sheetnames[0]]
    else:
        errors.append("Could not find 'Data Upload' sheet in single-sheet template")
        return rows_ready, errors, warnings, details
    
    # Find header row
    header_row = None
    col_indices = {}
    
    for r in range(1, 20):
        row_values = [ws.cell(row=r, column=c).value for c in range(1, 30)]
        norm = [str(x).strip().lower() if x else "" for x in row_values]
        
        if "scope" in norm and "id" in norm and "qty" in norm:
            header_row = r
            for col_name in ["scope", "category", "report label", "id", "uom", "ghg unit", "factor", "qty", "apply", "data source", "notes"]:
                if col_name in norm:
                    col_indices[col_name] = norm.index(col_name) + 1
            break
    
    if header_row is None:
        errors.append("Could not find header row with columns: Scope, ID, Qty")
        return rows_ready, errors, warnings, details
    
    details["header_row"] = header_row
    details["columns_found"] = list(col_indices.keys())

    month_cols: list[int] = []
    uom_col = col_indices.get("uom")
    qty_col = col_indices.get("qty")
    if uom_col and qty_col and qty_col > uom_col:
        month_cols = list(range(uom_col + 1, qty_col))
    
    # Parse data rows.
    parsed_rows = []
    
    for r in range(header_row + 1, ws.max_row + 1):
        # Get ID and Qty
        id_col = col_indices.get("id")
        qty_col = col_indices.get("qty")
        scope_col = col_indices.get("scope")
        
        if not id_col or not qty_col or not scope_col:
            continue
        
        oid = ws.cell(row=r, column=id_col).value
        qty_val = ws.cell(row=r, column=qty_col).value
        scope_val = ws.cell(row=r, column=scope_col).value
        
        # Skip if ID is blank
        if not oid or str(oid).strip() == "":
            continue
        
        # Parse scope
        scope_name = None
        if scope_val:
            s = str(scope_val).strip()
            if "scope 1" in s.lower():
                scope_name = "Scope 1"
            elif "scope 2" in s.lower():
                scope_name = "Scope 2"
            elif "scope 3" in s.lower():
                scope_name = "Scope 3"
        
        if not scope_name:
            warnings.append(f"Row {r}: Could not determine scope from '{scope_val}'")
            continue
        
        # Parse monthly values from the workbook
        month_values: list[float | None] = []
        for month_col in month_cols[:12]:
            month_values.append(_to_float(ws.cell(row=r, column=month_col).value))
        while len(month_values) < 12:
            month_values.append(None)

        qty_float = _to_float(qty_val)
        has_months = any(value is not None for value in month_values)
        if has_months:
            qty_float = float(sum(float(value or 0.0) for value in month_values))
        if qty_float is None or qty_float == 0:
            continue

        # Parse Apply% if present
        apply_col = col_indices.get("apply")
        apply_pct = 100.0  # Default
        if apply_col:
            apply_val = ws.cell(row=r, column=apply_col).value
            if apply_val:
                try:
                    # Handle "100%" or 100 or 1.0
                    apply_str = str(apply_val).strip().replace("%", "")
                    apply_pct = float(apply_str)
                except (ValueError, TypeError):
                    apply_pct = 100.0
        
        # Get notes if present
        notes_col = col_indices.get("notes")
        notes = None
        if notes_col:
            notes_val = ws.cell(row=r, column=notes_col).value
            if notes_val:
                notes = str(notes_val).strip()

        data_source_col = col_indices.get("data source")
        data_source = None
        if data_source_col:
            data_source_val = ws.cell(row=r, column=data_source_col).value
            if data_source_val:
                data_source = str(data_source_val).strip()

        parsed_rows.append({
            "scope": scope_name,
            "original_id": str(oid).strip(),
            "qty": qty_float,
            "months": month_values,
            "apply_pct": apply_pct,
            "notes": notes,
            "data_source": data_source,
            "row_number": r
        })
    
    details["parsed_row_count"] = len(parsed_rows)
    
    if not parsed_rows:
        errors.append("No data rows found with non-blank Qty values")
        return rows_ready, errors, warnings, details
    
    # Merge duplicate IDs so copy/pasted rows become a single stored record.
    merged_rows: list[dict] = []
    merged_map: dict[tuple[str, str], dict] = {}
    for row in parsed_rows:
        key = (row["scope"], row["original_id"])
        merged = merged_map.get(key)
        if merged is None:
            merged = {
                "scope": row["scope"],
                "original_id": row["original_id"],
                "qty": 0.0,
                "months": [None] * 12,
                "apply_pct": row.get("apply_pct"),
                "notes": [],
                "data_source": [],
                "row_numbers": [],
                "row_count": 0,
            }
            merged_map[key] = merged

        merged["row_count"] += 1
        merged["row_numbers"].append(row.get("row_number"))
        merged["qty"] += float(row.get("qty") or 0.0)
        row_months = row.get("months") or []
        for idx in range(12):
            if idx >= len(row_months):
                continue
            month_value = row_months[idx]
            if month_value is None:
                continue
            merged["months"][idx] = (merged["months"][idx] or 0.0) + float(month_value)
        if row.get("apply_pct") is not None and merged.get("apply_pct") is None:
            merged["apply_pct"] = row.get("apply_pct")
        if row.get("notes"):
            merged["notes"].append(row.get("notes"))
        if row.get("data_source"):
            merged["data_source"].append(row.get("data_source"))

    for merged in merged_map.values():
        merged["notes"] = _merge_notes(merged.get("notes") or [])
        merged["data_source"] = _merge_notes(merged.get("data_source") or [])
        merged_rows.append(merged)

    merged_rows.sort(key=lambda item: min([int(v) for v in item.get("row_numbers") or [10**9] if v is not None] or [10**9]))
    details["duplicate_groups"] = sum(1 for row in merged_rows if int(row.get("row_count") or 0) > 1)
    details["merged_row_count"] = len(merged_rows)

    # Factor lookup for each scope
    missing_ids = {}
    
    for scope_name in ["Scope 1", "Scope 2", "Scope 3"]:
        scope_rows = [r for r in merged_rows if r["scope"] == scope_name]
        if not scope_rows:
            continue
        
        dataset_id = datasets_by_scope.get(scope_name)
        if not dataset_id:
            errors.append(f"{scope_name}: No dataset selected in job configuration")
            continue
        
        # Get all IDs for this scope
        ids = list(set([r["original_id"] for r in scope_rows]))
        
        # Lookup factors (check both factor_lookup and custom_conversion_factors)
        factor_map = _lookup_factors(dataset_id, scope_name, ids, job_id)
        
        # Check for missing IDs — report row numbers and the IDs themselves
        missing = [oid for oid in ids if oid not in factor_map]
        if missing:
            missing_ids[scope_name] = missing[:50]
            # Find which rows each missing ID came from
            for oid in missing[:10]:
                row_nums = [str(r["row_number"]) for r in scope_rows if r["original_id"] == oid]
                row_ref = f" (row{'s' if len(row_nums) > 1 else ''} {', '.join(row_nums)})" if row_nums else ""
                errors.append(f"{scope_name}: ID '{oid}' not found in factor database{row_ref}")
            if len(missing) > 10:
                errors.append(f"{scope_name}: …and {len(missing) - 10} more unrecognised IDs")
        
        # Build rows_ready
        for r in scope_rows:
            oid = r["original_id"]
            if oid not in factor_map:
                continue
            
            factor_info = factor_map[oid]

            # Calculate tCO2e
            qty = r["qty"]
            factor = factor_info["factor"]
            apply_pct = r["apply_pct"]
            ghg_unit = factor_info.get("ghg_unit") or "kgCO2e"

            # Convert to tCO2e
            emissions = qty * factor * (apply_pct / 100.0)
            if "kg" in str(ghg_unit).lower():
                emissions = emissions / 1000.0  # kg to tonnes
            
            rows_ready.append({
                "scope": scope_name,
                "dataset_id": dataset_id,
                "db_id": factor_info.get("db_id"),
                "original_id": oid,
                "qty": qty,
                "uom": factor_info.get("uom"),
                "ghg_unit": ghg_unit,
                "factor": factor,
                "calc_tco2e": round(emissions, 4),
                "level_1": factor_info.get("level_1"),
                "level_2": factor_info.get("level_2"),
                "level_3": factor_info.get("level_3"),
                "level_4": factor_info.get("level_4"),
                "column_text": factor_info.get("column_text"),
                "report_label": factor_info.get("report_label"),
                "apply_pct": apply_pct,
                "notes": r.get("notes"),
                "data_source": r.get("data_source") or "Business Travel Data",
                "data_confidence": "M",
                "month_1": r.get("months", [None] * 12)[0],
                "month_2": r.get("months", [None] * 12)[1],
                "month_3": r.get("months", [None] * 12)[2],
                "month_4": r.get("months", [None] * 12)[3],
                "month_5": r.get("months", [None] * 12)[4],
                "month_6": r.get("months", [None] * 12)[5],
                "month_7": r.get("months", [None] * 12)[6],
                "month_8": r.get("months", [None] * 12)[7],
                "month_9": r.get("months", [None] * 12)[8],
                "month_10": r.get("months", [None] * 12)[9],
                "month_11": r.get("months", [None] * 12)[10],
                "month_12": r.get("months", [None] * 12)[11],
            })
    
    details["missing_ids"] = missing_ids
    details["rows_ready_count"] = len(rows_ready)
    
    return rows_ready, errors, warnings, details


def _lookup_factors(dataset_id: int, scope: str, original_ids: list[str], job_id: int) -> dict[str, dict]:
    """
    Lookup factors from factor_lookup and custom_conversion_factors.
    Returns dict mapping original_id -> factor info
    """
    factor_map = {}
    
    with get_conn() as con:
        # Standard factors
        if db_backend() == "postgres":
            sql = """
                SELECT db_id, original_id, level_1, level_2, level_3, level_4,
                       column_text, uom, ghg_unit, factor, report_label
                FROM factor_lookup
                WHERE dataset_id=%s AND scope=%s AND original_id = ANY(%s)
            """
            df = con.execute(sql, [dataset_id, scope, original_ids]).df()
        else:
            ph = ",".join(["?"] * len(original_ids))
            sql = f"""
                SELECT db_id, original_id, level_1, level_2, level_3, level_4,
                       column_text, uom, ghg_unit, factor, report_label
                FROM factor_lookup
                WHERE dataset_id=? AND scope=? AND original_id IN ({ph})
            """
            df = con.execute(sql, [dataset_id, scope] + original_ids).df()
        
        if df is not None and not df.empty:
            for _, row in df.iterrows():
                oid = str(row["original_id"]).strip()
                factor_map[oid] = {
                    "db_id": int(row["db_id"]) if pd.notna(row["db_id"]) else None,
                    "original_id": oid,
                    "level_1": row.get("level_1"),
                    "level_2": row.get("level_2"),
                    "level_3": row.get("level_3"),
                    "level_4": row.get("level_4"),
                    "column_text": row.get("column_text"),
                    "uom": row.get("uom"),
                    "ghg_unit": row.get("ghg_unit"),
                    "factor": float(row["factor"]) if pd.notna(row["factor"]) else 0.0,
                    "report_label": row.get("report_label")
                }
        
        # Custom factors (check by custom_id matching original_ids)
        if db_backend() == "postgres":
            custom_sql = """
                SELECT custom_factor_id, custom_id, level_1, level_2, level_3, level_4,
                       uom, ghg_unit, factor, report_label, category
                FROM custom_conversion_factors
                WHERE scope=%s AND custom_id = ANY(%s)
                  AND (job_id=%s OR job_id IS NULL)
                  AND is_active=TRUE
            """
            custom_df = con.execute(custom_sql, [scope, original_ids, job_id]).df()
        else:
            custom_ph = ",".join(["?"] * len(original_ids))
            custom_sql = f"""
                SELECT custom_factor_id, custom_id, level_1, level_2, level_3, level_4,
                       uom, ghg_unit, factor, report_label, category
                FROM custom_conversion_factors
                WHERE scope=? AND custom_id IN ({custom_ph})
                  AND (job_id=? OR job_id IS NULL)
                  AND is_active=TRUE
            """
            custom_df = con.execute(custom_sql, [scope] + original_ids + [job_id]).df()
        
        if custom_df is not None and not custom_df.empty:
            for _, row in custom_df.iterrows():
                cid = str(row["custom_id"]).strip()
                # Custom factors override standard factors
                factor_map[cid] = {
                    "db_id": None,  # Custom factors don't have db_id in factor_lookup
                    "custom_factor_id": int(row["custom_factor_id"]),
                    "original_id": cid,
                    "level_1": row.get("level_1"),
                    "level_2": row.get("level_2"),
                    "level_3": row.get("level_3"),
                    "level_4": row.get("level_4"),
                    "column_text": row.get("category"),
                    "uom": row.get("uom"),
                    "ghg_unit": row.get("ghg_unit"),
                    "factor": float(row["factor"]) if pd.notna(row["factor"]) else 0.0,
                    "report_label": row.get("report_label")
                }
    
    return factor_map
