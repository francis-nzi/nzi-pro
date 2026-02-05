from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ID_PATTERNS: list[re.Pattern[str]] = [
    re.compile(r"\b\d+_\d+_\d+_\d+_\d+\b"),  # DESNZ-style numeric underscore IDs
    re.compile(r"\bSPEND-[A-Z0-9\-_.]+\b"),  # Spend IDs
]


def _iter_ids_in_text(text: str) -> list[str]:
    out: list[str] = []
    for pat in ID_PATTERNS:
        out.extend(pat.findall(text))
    return out


def _load_factor_ids(folder: Path) -> set[str]:
    ids: set[str] = set()
    for p in sorted(folder.glob("*.csv")):
        df = pd.read_csv(p)
        # We standardised ingestion on these headers
        # Year,ID,Method,Scope,Category,Level 2,Level 3,Level 4,UOM,GHG/Unit,GHG Conversion Factor ...,Valid From,Valid To,Source
        if "ID" not in df.columns:
            continue
        for v in df["ID"].dropna().astype(str).tolist():
            s = v.strip()
            if s and s.lower() != "nan":
                ids.add(s)
    return ids


def rebuild_template(*, template_in: Path, template_out: Path, factor_folder: Path) -> dict[str, object]:
    factor_ids = _load_factor_ids(factor_folder)

    wb = load_workbook(template_in)

    cleared: list[tuple[str, str, str]] = []  # (sheet, cell, id)
    kept: int = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                ids = _iter_ids_in_text(v)
                if not ids:
                    continue

                # If cell contains exactly one ID and nothing else, we can safely blank it if invalid.
                if len(ids) == 1 and v.strip() == ids[0]:
                    oid = ids[0]
                    if oid not in factor_ids:
                        cell.value = ""
                        cleared.append((ws.title, cell.coordinate, oid))
                    else:
                        kept += 1
                    continue

                # If the cell contains multiple IDs or other text, we leave it unchanged.
                # (This avoids corrupting formulas/notes.)

    template_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(template_out)

    return {
        "template_in": str(template_in),
        "template_out": str(template_out),
        "factor_folder": str(factor_folder),
        "kept_valid_id_cells": int(kept),
        "cleared_invalid_id_cells": int(len(cleared)),
        "cleared": cleared[:50],
        "cleared_truncated": (len(cleared) > 50),
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Rebuild/clean an Excel upload template by removing any prefilled factor IDs that do not exist in the current factor CSVs."
    )
    parser.add_argument(
        "--template-in",
        default=str(Path("templates") / "NZI Data Upload Template - Basic UK.xlsx"),
    )
    parser.add_argument(
        "--template-out",
        default=str(Path("templates") / "NZI Data Upload Template - Basic UK.xlsx"),
        help="Output path (defaults to overwrite input).",
    )
    parser.add_argument(
        "--factors",
        default=str(Path("assets") / "conversion_factors"),
        help="Folder containing the conversion factor CSVs (source of truth).",
    )

    args = parser.parse_args()

    info = rebuild_template(
        template_in=Path(args.template_in),
        template_out=Path(args.template_out),
        factor_folder=Path(args.factors),
    )

    print("Rebuild complete")
    for k, v in info.items():
        print(f"{k}: {v}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
