"""
Generate single-sheet Excel template from factor_lookup + custom_conversion_factors.
This replaces the multi-sheet template approach with a simplified, filterable single sheet.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from core.database import get_conn
from services.dataset_selector import get_applicable_datasets, get_monthly_headers, get_reporting_period_display
from nzi_pages.fetch_existing_data import fetch_existing_scope_entries


def generate_single_sheet_template(
    job_id: int,
    client_name: str = "",
    site_name: str = "",
    job_number: str = "",
    report_from: str = "",
    report_to: str = "",
    include_custom_factors: bool = True
) -> tuple[bytes, str]:
    """
    Generate a single-sheet Excel template with all conversion factors.
    
    Returns:
        (excel_bytes, filename)
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Upload"
    
    # Header rows (metadata)
    ws['A1'] = 'Client Name:'
    ws['B1'] = client_name
    ws['C1'] = 'Job Number:'
    ws['D1'] = job_number
    
    ws['A2'] = 'Site Name:'
    ws['B2'] = site_name
    ws['C2'] = 'Reporting Period:'
    reporting_period = get_reporting_period_display(job_id)
    ws['D2'] = reporting_period
    
    ws['A3'] = 'Data Files:'
    ws['B3'] = 'Standard UK Conversion Factors'
    
    # Style header rows
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    for row in range(1, 4):
        for col in ['A', 'C', 'E']:
            cell = ws[f'{col}{row}']
            cell.font = Font(bold=True)
            cell.fill = header_fill
    
    # Get monthly headers based on reporting period
    monthly_headers = get_monthly_headers(job_id)
    
    # Column headers (row 4)
    headers = [
        'Scope', 'Category', 'Report Label', 'ID', 'UOM', 'GHG Unit', 'Factor',
        'Qty'
    ] + monthly_headers + [
        'tCO2e', 'Apply%', 'Notes'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Fetch conversion factors from database
    factors = _fetch_conversion_factors(job_id, include_custom_factors)
    
    # Fetch existing scope entry data to pre-fill
    existing_data = fetch_existing_scope_entries(job_id)
    
    # Populate data rows (starting at row 5)
    current_row = 5
    for factor in factors:
        ws.cell(row=current_row, column=1, value=factor['scope'])
        ws.cell(row=current_row, column=2, value=factor['category'])
        ws.cell(row=current_row, column=3, value=factor['report_label'])
        ws.cell(row=current_row, column=4, value=factor['id'])
        ws.cell(row=current_row, column=5, value=factor['uom'])
        ws.cell(row=current_row, column=6, value=factor['ghg_unit'])
        ws.cell(row=current_row, column=7, value=factor['factor'])
        
        # Check if we have existing data for this factor ID
        factor_id = factor['id']
        existing = existing_data.get(factor_id, {})
        
        # Qty column - pre-fill if exists
        ws.cell(row=current_row, column=8, value=existing.get('qty'))
        
        # Monthly columns - pre-fill if exists (columns 9-20)
        monthly_values = existing.get('monthly', [])
        for month_idx, month_col in enumerate(range(9, 21)):
            value = monthly_values[month_idx] if month_idx < len(monthly_values) else None
            ws.cell(row=current_row, column=month_col, value=value)
        
        # tCO2e formula: =H{row}*G{row}*(V{row}/100)/1000
        # (Qty * Factor * Apply% / 100) / 1000 to convert kg to tonnes
        ws.cell(row=current_row, column=21, value=f"=IF(H{current_row}=\"\",\"\",H{current_row}*G{current_row}*(V{current_row}/100)/1000)")
        
        # Apply% - pre-fill if exists, default 100%
        apply_pct = existing.get('apply_pct', 100)
        ws.cell(row=current_row, column=22, value=f"{apply_pct}%")
        
        # Total tCO2e (same formula, for display)
        ws.cell(row=current_row, column=23, value=f"=U{current_row}")
        
        # Notes column - pre-fill if exists
        ws.cell(row=current_row, column=24, value=existing.get('notes'))
        
        current_row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 10  # Scope
    ws.column_dimensions['B'].width = 30  # Category
    ws.column_dimensions['C'].width = 50  # Report Label
    ws.column_dimensions['D'].width = 20  # ID (can be hidden)
    ws.column_dimensions['E'].width = 12  # UOM
    ws.column_dimensions['F'].width = 12  # GHG Unit
    ws.column_dimensions['G'].width = 12  # Factor
    ws.column_dimensions['H'].width = 12  # Qty
    
    for col_idx in range(9, 21):  # Monthly columns
        ws.column_dimensions[get_column_letter(col_idx)].width = 10
    
    ws.column_dimensions['U'].width = 12  # tCO2e
    ws.column_dimensions['V'].width = 10  # Apply
    ws.column_dimensions['W'].width = 12  # tCO2e Total
    ws.column_dimensions['X'].width = 30  # Notes
    
    # Optionally hide ID column (column D)
    ws.column_dimensions['D'].hidden = True
    
    # Freeze panes (freeze header rows)
    ws.freeze_panes = 'A5'
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"job_{job_id}_template_single_sheet.xlsx"
    
    return buffer.getvalue(), filename


def _fetch_conversion_factors(job_id: int, include_custom: bool = True) -> list[dict]:
    """
    Fetch conversion factors from factor_lookup and optionally custom_conversion_factors.
    Returns a list of dicts with keys: scope, category, report_label, id, uom, ghg_unit, factor
    """
    
    factors = []
    
    # Get applicable datasets based on reporting period (Activity + Spend + Custom)
    applicable_datasets = get_applicable_datasets(job_id)
    all_dataset_ids = list(set(
        applicable_datasets.get('Scope 1', []) +
        applicable_datasets.get('Scope 2', []) +
        applicable_datasets.get('Scope 3', [])
    ))
    
    # Remove custom factors dataset ID (999) for standard factors query
    standard_dataset_ids = [d for d in all_dataset_ids if d < 900]
    
    with get_conn() as con:
        # Fetch standard factors from all applicable datasets
        if standard_dataset_ids:
            placeholders = ','.join(['?'] * len(standard_dataset_ids))
            factor_query = f"""
                SELECT 
                    scope,
                    level_1 as category,
                    report_label,
                    original_id as id,
                    uom,
                    ghg_unit,
                    factor
                FROM factor_lookup
                WHERE dataset_id IN ({placeholders})
                ORDER BY 
                    CASE scope 
                        WHEN 'Scope 1' THEN 1 
                        WHEN 'Scope 2' THEN 2 
                        WHEN 'Scope 3' THEN 3 
                        ELSE 4 
                    END,
                    level_1,
                    report_label
            """
            
            result = con.execute(factor_query, standard_dataset_ids)
            while True:
                row = result.fetchone()
                if not row:
                    break
                factors.append({
                    'scope': row[0],
                    'category': row[1] or '',
                    'report_label': row[2] or '',
                    'id': row[3],
                    'uom': row[4] or '',
                    'ghg_unit': row[5] or '',
                    'factor': row[6] or 0
                })
        
        # Fetch custom factors if requested
        if include_custom:
            custom_query = """
                SELECT 
                    scope,
                    category,
                    report_label,
                    custom_id as id,
                    uom,
                    ghg_unit,
                    factor
                FROM custom_conversion_factors
                WHERE (job_id = ? OR job_id IS NULL)
                  AND is_active = TRUE
                ORDER BY scope, category, report_label
            """
            
            result = con.execute(custom_query, [job_id])
            while True:
                row = result.fetchone()
                if not row:
                    break
                factors.append({
                    'scope': row[0],
                    'category': row[1] or '',
                    'report_label': row[2] or '',
                    'id': row[3],
                    'uom': row[4] or '',
                    'ghg_unit': row[5] or '',
                    'factor': row[6] or 0
                })
    
    return factors
