import io

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from core.database import db_backend, get_conn
from services.sites import list_sites


def _replace_sheet(wb, name: str):
    if name in wb.sheetnames:
        ws_old = wb[name]
        wb.remove(ws_old)
    return wb.create_sheet(title=name)


def _append_kv(ws, rows: list[tuple[str, object]]):
    for k, v in rows:
        ws.append([k, v])


def _job_core_data(job_id: int):
    try:
        with get_conn() as con:
            hdr = con.execute(
                """
                SELECT j.job_id, j.job_number, j.title, j.job_type, j.reporting_year, j.status,
                       j.start_date, j.due_date,
                       j.client_db_id, c.client_name
                FROM jobs j
                JOIN clients c ON c.db_id = j.client_db_id
                WHERE j.job_id=%s
                """,
                [int(job_id)],
            ).fetchone()

            crp = con.execute(
                """
                SELECT reporting_period_from, reporting_period_to,
                       client_order_number, client_contact_name, client_contact_email,
                       report_signee_name, report_signee_position,
                       num_employees, turnover_gbp, premises_size_m2,
                       vehicles_owned, vehicles_leased, premises_owned, premises_leased
                FROM crp_job_details
                WHERE job_id=%s
                """,
                [int(job_id)],
            ).fetchone()

            plan = con.execute(
                """
                SELECT data_collection_due, first_draft_due, final_report_due
                FROM job_plan
                WHERE job_id=%s
                """,
                [int(job_id)],
            ).fetchone()

        return hdr, crp, plan
    except Exception:
        return None, None, None


def _to_tco2e(qty: float, factor: float, ghg_unit: str | None) -> float:
    ghg = (str(ghg_unit or "kgCO2e").replace(" ", "").lower())
    emissions = float(qty) * float(factor)
    if ghg.startswith("kg"):
        return emissions / 1000.0
    return emissions


def _norm_scope(sheet_name: str) -> str | None:
    s = (sheet_name or "").strip().lower()
    if "scope 1" in s:
        return "Scope 1"
    if "scope 2" in s:
        return "Scope 2"
    if "scope 3" in s:
        return "Scope 3"
    return None


def _find_table_header_row(ws):
    for r in range(1, 60):
        values = [ws.cell(row=r, column=c).value for c in range(1, 80)]
        norm = [str(x).strip().lower() if x is not None else "" for x in values]
        if "id" in norm and "qty" in norm:
            idx = {}
            for name in ("id", "qty", "apply"):
                if name in norm:
                    idx[name] = norm.index(name) + 1
            return r, idx
    return None, None


def _factor_lookup_by_original_ids(dataset_id: int, scope_name: str, original_ids: list[str]) -> pd.DataFrame:
    original_ids = [str(x).strip() for x in (original_ids or []) if x is not None and str(x).strip()]
    if not original_ids:
        return pd.DataFrame()

    with get_conn() as con:
        if db_backend() == "postgres":
            sql = """
                SELECT db_id, original_id, level_1, level_2, level_3, level_4,
                       column_text, uom, ghg_unit, factor
                FROM factor_lookup
                WHERE dataset_id=%s AND scope=%s AND original_id = ANY(%s)
            """
            return con.execute(sql, [int(dataset_id), str(scope_name), original_ids]).df()

        ph = ",".join(["?"] * len(original_ids))
        sql = f"""
            SELECT db_id, original_id, level_1, level_2, level_3, level_4,
                   column_text, uom, ghg_unit, factor
            FROM factor_lookup
            WHERE dataset_id=? AND scope=? AND original_id IN ({ph})
        """
        return con.execute(sql, [int(dataset_id), str(scope_name)] + original_ids).df()


def _upsert_job_scope_rows(job_id: int, scope_name: str, dataset_id: int, rows_df: pd.DataFrame) -> tuple[int, int]:
    if rows_df is None or rows_df.empty:
        return 0, 0

    inserted = 0
    updated = 0

    with get_conn() as con:
        for _, r in rows_df.iterrows():
            oid = str(r.get("original_id") or "").strip()
            if not oid:
                continue

            exists = con.execute(
                "SELECT row_id FROM job_scope_rows WHERE job_id=%s AND scope=%s AND original_id=%s LIMIT 1",
                [int(job_id), str(scope_name), oid],
            ).fetchone()

            qty = r.get("qty")
            calc = r.get("calc_tco2e")
            params_common = [
                float(qty) if qty is not None and str(qty) != "nan" else None,
                (r.get("uom") or "").strip() or None,
                float(r.get("factor")) if r.get("factor") is not None and str(r.get("factor")) != "nan" else None,
                (r.get("ghg_unit") or "").strip() or None,
                float(calc) if calc is not None and str(calc) != "nan" else None,
                (r.get("level_1") or "").strip() or None,
                (r.get("level_2") or "").strip() or None,
                (r.get("level_3") or "").strip() or None,
                (r.get("level_4") or "").strip() or None,
                (r.get("column_text") or "").strip() or None,
            ]

            if exists:
                con.execute(
                    """
                    UPDATE job_scope_rows
                    SET enabled=TRUE,
                        qty=%s,
                        uom=%s,
                        factor=%s,
                        ghg_unit=%s,
                        calc_tco2e=%s,
                        level_1=%s,
                        level_2=%s,
                        level_3=%s,
                        level_4=%s,
                        column_text=%s,
                        updated_at=NOW()
                    WHERE row_id=%s
                    """,
                    params_common + [int(exists[0])],
                )
                updated += 1
            else:
                con.execute(
                    """
                    INSERT INTO job_scope_rows
                      (job_id, scope, dataset_id, factor_db_id, original_id,
                       level_1, level_2, level_3, level_4, column_text,
                       report_label, notes, enabled,
                       qty, uom, factor, ghg_unit,
                       calc_tco2e, override_tco2e, override_reason,
                       created_at, updated_at)
                    VALUES
                      (%s, %s, %s, %s, %s,
                       %s, %s, %s, %s, %s,
                       NULL, NULL, TRUE,
                       %s, %s, %s, %s,
                       %s, NULL, NULL,
                       NOW(), NOW())
                    """,
                    [
                        int(job_id), str(scope_name), int(dataset_id), int(r.get("db_id")), oid,
                        (r.get("level_1") or "").strip() or None,
                        (r.get("level_2") or "").strip() or None,
                        (r.get("level_3") or "").strip() or None,
                        (r.get("level_4") or "").strip() or None,
                        (r.get("column_text") or "").strip() or None,
                    ]
                    + params_common[:5],
                )
                inserted += 1

    return inserted, updated


def render_excel_section(
    *,
    jid: int,
    job_number: str,
    client_db_id: int,
    client_name: str,
    reporting_year: int | None,
    rp_from,
    rp_to,
    scopes_df: pd.DataFrame,
):
    st.subheader("Excel upload template")

    sites_df = list_sites(int(client_db_id))
    site_options = []
    if not sites_df.empty and "site_name" in sites_df.columns:
        site_options = [str(x) for x in sites_df["site_name"].dropna().tolist()]

    selected_site = None
    if not site_options:
        st.info("Add at least one Client Site (Client Folder → Sites) to download a per-site template.")
    else:
        selected_site = st.selectbox("Site", site_options, index=0, key="excel_tpl_site")

    include_prev_year = st.checkbox(
        "Include previous year data (reference sheet)",
        value=True,
        key="excel_tpl_prev_year",
    )

    data_files_ref = ""
    try:
        with get_conn() as con:
            df_ds = con.execute(
                """
                SELECT jsc.scope, d.name, d.year, d.analysis_type, d.country
                FROM job_scope_config jsc
                LEFT JOIN datasets d ON d.dataset_id = jsc.dataset_id
                WHERE jsc.job_id=%s
                ORDER BY jsc.scope
                """,
                [int(jid)],
            ).df()

        parts = []
        if not df_ds.empty:
            for _, r in df_ds.iterrows():
                if r.get("name") is None:
                    continue
                label = str(r.get("name") or "").strip()
                y = r.get("year")
                if y is not None and str(y) != "nan":
                    label = f"{label} {int(y)}"
                parts.append(label)
        data_files_ref = ", ".join(parts)
    except Exception:
        data_files_ref = ""

    if selected_site and st.button("Generate template", type="primary"):
        template_path = "templates/NZI Data Upload Template - Basic UK.xlsx"
        wb = load_workbook(template_path)

        hdr, crp, plan = _job_core_data(int(jid))

        if hdr is not None:
            (
                _jid,
                _job_number,
                _title,
                _job_type,
                _year,
                _status,
                _start_date,
                _due_date,
                _client_db_id,
                _client_name,
            ) = hdr

            ws_core = _replace_sheet(wb, "Core Data")
            _append_kv(
                ws_core,
                [
                    ("Client", _client_name),
                    ("Job Number", _job_number),
                    ("Job Title", _title),
                    ("Job Type", _job_type),
                    ("Job Status", _status),
                    ("Reporting Year", _year),
                    ("Start Date", _start_date),
                    ("Due Date", _due_date),
                    ("Reporting Period From", rp_from),
                    ("Reporting Period To", rp_to),
                    ("Template Site", selected_site),
                ],
            )

            if crp is not None:
                (
                    crp_from,
                    crp_to,
                    client_order_number,
                    client_contact_name,
                    client_contact_email,
                    report_signee_name,
                    report_signee_position,
                    num_employees,
                    turnover_gbp,
                    premises_size_m2,
                    vehicles_owned,
                    vehicles_leased,
                    premises_owned,
                    premises_leased,
                ) = crp

                ws_core.append([])
                _append_kv(
                    ws_core,
                    [
                        ("Client Order Number", client_order_number),
                        ("Client Contact Name", client_contact_name),
                        ("Client Contact Email", client_contact_email),
                        ("Report Signee Name", report_signee_name),
                        ("Report Signee Position", report_signee_position),
                        ("Employees", num_employees),
                        ("Turnover GBP", turnover_gbp),
                        ("Premises Size (m2)", premises_size_m2),
                        ("Vehicles Owned", vehicles_owned),
                        ("Vehicles Leased", vehicles_leased),
                        ("Premises Owned", premises_owned),
                        ("Premises Leased", premises_leased),
                    ],
                )

                if crp_from or crp_to:
                    ws_core.append([])
                    _append_kv(
                        ws_core,
                        [
                            ("CRP Reporting Period From (stored)", crp_from),
                            ("CRP Reporting Period To (stored)", crp_to),
                        ],
                    )

            if plan is not None:
                (data_collection_due, first_draft_due, final_report_due) = plan
                ws_core.append([])
                _append_kv(
                    ws_core,
                    [
                        ("Milestone: Data collection due", data_collection_due),
                        ("Milestone: First draft due", first_draft_due),
                        ("Milestone: Final report due", final_report_due),
                    ],
                )

            try:
                ws_sites = _replace_sheet(wb, "Sites")
                if sites_df is not None and not sites_df.empty:
                    ws_sites.append(list(sites_df.columns))
                    for row_vals in sites_df.itertuples(index=False, name=None):
                        ws_sites.append(list(row_vals))
                else:
                    ws_sites.append(["No sites found for this client."])
            except Exception:
                pass

        for ws in wb.worksheets:
            if ws["A1"].value and str(ws["A1"].value).strip().startswith("Site Name:"):
                ws["B1"].value = selected_site
            if ws["C1"].value and str(ws["C1"].value).strip().startswith("Report From:"):
                ws["D1"].value = rp_from
            if ws["E1"].value and str(ws["E1"].value).strip().startswith("To:"):
                ws["F1"].value = rp_to
            if ws["A2"].value and str(ws["A2"].value).strip().startswith("Data Files:"):
                ws["B2"].value = data_files_ref or ws["B2"].value

            ws["C2"].value = "Client Name:"
            ws["D2"].value = client_name
            ws["E2"].value = "Job Number:"
            ws["F2"].value = job_number

        if include_prev_year:
            prev_year = int(reporting_year or 0) - 1 if reporting_year is not None else None
            prev_job_id = None
            if prev_year:
                try:
                    with get_conn() as con:
                        r = con.execute(
                            """
                            SELECT job_id
                            FROM jobs
                            WHERE client_db_id=%s AND reporting_year=%s
                            ORDER BY job_id DESC
                            LIMIT 1
                            """,
                            [int(client_db_id), int(prev_year)],
                        ).fetchone()
                    if r:
                        prev_job_id = int(r[0])
                except Exception:
                    prev_job_id = None

            if prev_job_id is not None:
                try:
                    with get_conn() as con:
                        prev_df = con.execute(
                            """
                            SELECT scope, category, subcategory, description, amount, unit, tco2e, method, notes, updated_at
                            FROM crp_scope_entries
                            WHERE job_id=%s AND is_archived=FALSE
                            ORDER BY scope, category, subcategory, entry_id
                            """,
                            [int(prev_job_id)],
                        ).df()

                    if not prev_df.empty:
                        name = f"Previous Year ({prev_year})"
                        if name in wb.sheetnames:
                            ws_prev = wb[name]
                            wb.remove(ws_prev)
                        ws_prev = wb.create_sheet(title=name)
                        ws_prev.append(["Client", client_name])
                        ws_prev.append(["Job", job_number])
                        ws_prev.append(["Site", selected_site])
                        ws_prev.append([])
                        ws_prev.append(list(prev_df.columns))
                        for row_vals in prev_df.itertuples(index=False, name=None):
                            ws_prev.append(list(row_vals))
                except Exception:
                    pass

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fn_site = "".join(ch if ch.isalnum() or ch in ("-", "_", " ") else "_" for ch in str(selected_site))
        filename = f"{job_number} - {fn_site} - NZI Data Upload Template.xlsx"

        st.download_button(
            "Download template",
            data=buf,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.markdown("---")
    st.subheader("Upload completed template")

    uploaded = st.file_uploader("Upload completed NZI template (.xlsx)", type=["xlsx"], key="excel_tpl_upload")
    if uploaded is None:
        return

    try:
        wb_up = load_workbook(uploaded, data_only=True)
    except Exception as e:
        st.error(f"Could not read Excel file: {e}")
        return

    parsed_rows = []
    for ws in wb_up.worksheets:
        scope_name = _norm_scope(ws.title)
        if scope_name is None:
            continue
        header_row, idx = _find_table_header_row(ws)
        if header_row is None:
            continue

        id_col = idx.get("id")
        qty_col = idx.get("qty")
        apply_col = idx.get("apply")
        if not id_col or not qty_col:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            oid = ws.cell(row=r, column=id_col).value
            if oid is None or str(oid).strip() == "":
                continue
            qv = ws.cell(row=r, column=qty_col).value
            av = ws.cell(row=r, column=apply_col).value if apply_col else None
            if apply_col and av is not None:
                try:
                    if float(av) != 1.0:
                        continue
                except Exception:
                    if str(av).strip() not in ("1", "true", "True", "YES", "Yes"):
                        continue

            try:
                qty_val = float(qv) if qv is not None and str(qv).strip() != "" else None
            except Exception:
                qty_val = None
            if qty_val is None:
                continue

            parsed_rows.append({"scope": scope_name, "original_id": str(oid).strip(), "qty": qty_val})

    if not parsed_rows:
        st.info("No rows found to import. Ensure you have filled 'Qty' and set 'Apply' to 1 where applicable.")
        return

    parsed_df = pd.DataFrame(parsed_rows)

    ds_map = {}
    try:
        if scopes_df is not None and (not scopes_df.empty):
            for _, rr in scopes_df.iterrows():
                ds_map[str(rr["scope"])] = rr["dataset_id"]
    except Exception:
        ds_map = {}

    enriched = []
    errors = []
    for scope_name in ["Scope 1", "Scope 2", "Scope 3"]:
        sdf = parsed_df[parsed_df["scope"] == scope_name].copy()
        if sdf.empty:
            continue

        dsid = ds_map.get(scope_name)
        if dsid is None or str(dsid) == "nan":
            errors.append(f"{scope_name}: no dataset selected in Job Folder → Data Collection.")
            continue

        fdf = _factor_lookup_by_original_ids(int(dsid), scope_name, sdf["original_id"].tolist())
        if fdf.empty:
            errors.append(f"{scope_name}: none of the uploaded IDs matched factor_lookup for dataset {int(dsid)}.")
            continue

        m = sdf.merge(fdf, on="original_id", how="left")
        missing = m[m["db_id"].isna()]
        if not missing.empty:
            errors.append(f"{scope_name}: {len(missing)} IDs were not found in the selected dataset.")
        m = m[m["db_id"].notna()].copy()
        m["dataset_id"] = int(dsid)
        m["calc_tco2e"] = m.apply(lambda r: _to_tco2e(float(r["qty"]), float(r["factor"]), r.get("ghg_unit")), axis=1)
        enriched.append(m)

    if errors:
        for e in errors:
            st.error(e)

    if not enriched:
        return

    final_df = pd.concat(enriched, ignore_index=True)
    st.caption(f"Rows ready to import: {len(final_df)}")
    st.dataframe(
        final_df[[
            "scope",
            "original_id",
            "qty",
            "uom",
            "ghg_unit",
            "factor",
            "calc_tco2e",
            "level_1",
            "level_2",
            "level_3",
            "level_4",
            "column_text",
        ]],
        use_container_width=True,
        hide_index=True,
    )

    if st.button("Import rows", type="primary", key="excel_tpl_import"):
        total_ins = 0
        total_upd = 0
        for scope_name in ["Scope 1", "Scope 2", "Scope 3"]:
            sub = final_df[final_df["scope"] == scope_name].copy()
            if sub.empty:
                continue
            dsid = int(sub.iloc[0]["dataset_id"])
            ins, upd = _upsert_job_scope_rows(int(jid), scope_name, dsid, sub)
            total_ins += int(ins)
            total_upd += int(upd)

        st.success(f"Import complete. Inserted {total_ins}, updated {total_upd}.")
        st.rerun()
