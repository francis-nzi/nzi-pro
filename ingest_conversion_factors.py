from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
from typing import Any
from dataclasses import dataclass

import pandas as pd
from openpyxl import load_workbook

from core.database import db_backend, get_conn


@dataclass
class DatasetReplacementBlocked(Exception):
    message: str
    dependency_summary: dict[str, int]

    def __str__(self) -> str:
        return self.message


@dataclass
class WorkbookImportSummary:
    dataset_id: int
    year: int
    sheet_name: str
    total_rows: int
    accepted_rows: int
    updated_rows: int
    inserted_rows: int
    deleted_rows: int
    blocked_rows: int
    rejected_rows: int
    rejected_details: list[dict[str, Any]]
    years_seen: list[int]


def _norm_col(c: str) -> str:
    s = str(c).replace("\ufeff", "").replace("\u200b", "").replace("\u200c", "").replace("\u200d", "")
    s = s.lower().strip().replace("_", " ")
    s = "".join(ch if (ch.isalnum() or ch.isspace()) else " " for ch in s)
    return " ".join(s.split())


def _pick(cols: dict[str, str], *names: str) -> str | None:
    for n in names:
        k = _norm_col(n)
        if k in cols:
            return cols[k]
        # allow year-suffixed variants like 'ghg conversion factor 2025'
        for ck, orig in cols.items():
            if ck.startswith(k):
                return orig
    return None


def _norm_scope(v: Any) -> str | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return None
    sl = s.replace(" ", "").replace("_", "").lower()
    if sl.startswith("scope1") or sl == "s1" or "scope1" in sl:
        return "Scope 1"
    if sl.startswith("scope2") or sl == "s2" or "scope2" in sl:
        return "Scope 2"
    if sl.startswith("scope3") or sl == "s3" or "scope3" in sl:
        return "Scope 3"
    if "scope 1" in s.lower():
        return "Scope 1"
    if "scope 2" in s.lower():
        return "Scope 2"
    if "scope 3" in s.lower():
        return "Scope 3"
    return s


def _norm_original_id(v: Any) -> str | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, float) and float(v).is_integer():
        return str(int(v))
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return None
    if s.endswith(".0"):
        head = s[:-2]
        if head.isdigit():
            return head
    return s


def _norm_ghg_unit(v: Any) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "kgCO2e"
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return "kgCO2e"
    s = s.replace(" ", "")
    if s.lower() in ("kgco2e", "kgco₂e"):
        return "kgCO2e"
    return s


@dataclass(frozen=True)
class ConversionFactorColumn:
    field: str
    label: str
    required: bool
    aliases: tuple[str, ...]


# Single source of truth for the conversion-factor upload column set, shared by
# the per-dataset CSV parser, the DESNZ/DEFRA workbook parser, and the blank
# upload template generator (api/admin_datasets_routes.py) so all three can
# never drift out of sync with each other again.
CONVERSION_FACTOR_COLUMNS: tuple[ConversionFactorColumn, ...] = (
    ConversionFactorColumn("original_id", "ID", True, ("ID", "Code", "Original ID", "original_id", "original id", "factor id")),
    ConversionFactorColumn("scope", "Scope", True, ("Scope", "scope", "emissions scope")),
    ConversionFactorColumn("category", "Category", False, ("Category", "category")),
    ConversionFactorColumn("level_1", "Level 1", False, ("Level 1", "level_1", "level 1")),
    ConversionFactorColumn("level_2", "Level 2", False, ("Level 2", "Subcategory", "level_2", "level 2")),
    ConversionFactorColumn("level_3", "Level 3", False, ("Level 3", "Detail", "level_3", "level 3")),
    ConversionFactorColumn("level_4", "Level 4", False, ("Level 4", "level_4", "level 4")),
    ConversionFactorColumn("column_text", "Column Text", False, ("Column Text", "Description", "Name", "Activity", "column_text", "column text", "activity name")),
    ConversionFactorColumn("report_label", "Report Label", False, ("Report Label", "report_label", "report label")),
    ConversionFactorColumn("uom", "UOM", False, ("UOM", "Unit", "Units", "uom", "unit of measure")),
    ConversionFactorColumn("ghg_unit", "GHG Unit", False, ("GHG Unit", "GHGUnit", "GHG/Unit", "GHG Unit per", "ghg_unit", "ghg unit", "co2 unit")),
    ConversionFactorColumn("factor", "Factor", True, ("Factor", "GHG Conversion Factor", "GHG Conversion Factor 2025", "GHG conversion factor", "kgCO2e per unit", "kgCO2e per gbp", "factor")),
    ConversionFactorColumn("year", "Year", False, ("Year", "year", "reporting year")),
    ConversionFactorColumn("source", "Source", False, ("Source", "source")),
    ConversionFactorColumn("region", "Region", False, ("Region", "region")),
    ConversionFactorColumn("method", "Method", False, ("Method", "method", "calculation method")),
    ConversionFactorColumn("valid_from", "Valid From", False, ("Valid From", "ValidFrom", "valid_from", "valid from")),
    ConversionFactorColumn("valid_to", "Valid To", False, ("Valid To", "ValidTo", "valid_to", "valid to")),
)


def _resolve_column_mapping(df_columns: list[str]) -> tuple[dict[str, str | None], list[dict[str, Any]], list[str]]:
    """Resolve which physical column header in an uploaded file maps to which
    canonical field, using the same fuzzy alias/prefix matching as `_pick`.

    Returns:
        field_columns: canonical field name -> physical column name (or None if unmatched)
        mapping_report: one entry per physical column, for display in an upload preview UI
        missing_required_labels: labels of any required field with no matching column
    """
    cols = {_norm_col(c): c for c in df_columns}
    field_columns: dict[str, str | None] = {}
    claimed: dict[str, str] = {}
    for spec in CONVERSION_FACTOR_COLUMNS:
        picked = _pick(cols, *spec.aliases)
        field_columns[spec.field] = picked
        if picked is not None:
            claimed[picked] = spec.field

    label_by_field = {spec.field: spec.label for spec in CONVERSION_FACTOR_COLUMNS}
    mapping_report = [
        {
            "source_header": c,
            "mapped_field": claimed.get(c),
            "mapped_label": label_by_field.get(claimed.get(c) or "", None),
        }
        for c in df_columns
    ]
    missing_required = [spec.label for spec in CONVERSION_FACTOR_COLUMNS if spec.required and field_columns[spec.field] is None]
    return field_columns, mapping_report, missing_required


def _norm_dupe_text(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower()
    return "" if s == "nan" else " ".join(s.split())


def _synth_column_text(r: pd.Series, c_text: str | None, c_l1: str | None, c_l2: str | None, c_l3: str | None, c_l4: str | None) -> str:
    if c_text is not None:
        v = r.get(c_text)
        if v is not None and not (isinstance(v, float) and pd.isna(v)):
            s = str(v).strip()
            if s and s.lower() != "nan":
                return s
    parts: list[str] = []
    for c in (c_l1, c_l2, c_l3, c_l4):
        if c is None:
            continue
        v = r.get(c)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip()
        if s and s.lower() != "nan":
            parts.append(s)
    return " - ".join(parts) if parts else ""


def _dataset_meta_from_filename(path: Path) -> tuple[str, str, str, int, str]:
    name = path.stem
    lower = name.lower()
    analysis_type = "Activity" if "activity" in lower else ("Spend" if "spend" in lower else "")
    year = 0
    try:
        year = int(name.split("-")[0])
    except Exception:
        year = 0
    # Keep source simple; you can edit in DB later
    source = "DESNZ" if "desnz" in lower else ("DEFRA" if "defra" in lower else "")
    country = "UK"
    version = str(year) if year else ""
    return name, source, analysis_type, year, country


def _ensure_dataset(*, name: str, source: str, analysis_type: str, country: str, year: int) -> int:
    with get_conn() as con:
        row = con.execute(
            """
            SELECT dataset_id
            FROM datasets
            WHERE coalesce(name,'')=? AND coalesce(source,'')=? AND coalesce(analysis_type,'')=? AND coalesce(country,'')=? AND year=?
            ORDER BY dataset_id DESC
            LIMIT 1
            """,
            [name, source, analysis_type, country, int(year)],
        ).fetchone()
        if row:
            return int(row[0])

        if db_backend() == "postgres":
            new_row = con.execute(
                """
                INSERT INTO datasets (name, source, analysis_type, country, region, currency, year, version, license, notes)
                VALUES (%s,%s,%s,%s,NULL,NULL,%s,%s,NULL,%s)
                RETURNING dataset_id
                """,
                [name, source, analysis_type, country, int(year), str(year), f"Ingested from {name}"],
            ).fetchone()
            return int(new_row[0])

        # DuckDB fallback
        con.execute(
            """
            INSERT INTO datasets (dataset_id, name, source, analysis_type, country, region, currency, year, version, license, notes)
            VALUES ((SELECT COALESCE(MAX(dataset_id),0)+1 FROM datasets), ?, ?, ?, ?, NULL, NULL, ?, ?, NULL, ?)
            """,
            [name, source, analysis_type, country, int(year), str(year), f"Ingested from {name}"],
        )
        row2 = con.execute(
            """
            SELECT dataset_id FROM datasets
            WHERE coalesce(name,'')=? AND coalesce(source,'')=? AND year=?
            ORDER BY dataset_id DESC
            LIMIT 1
            """,
            [name, source, int(year)],
        ).fetchone()
        return int(row2[0])


def _get_dataset_year(dataset_id: int) -> int | None:
    with get_conn() as con:
        row = con.execute(
            "SELECT year FROM datasets WHERE dataset_id = %s",
            [int(dataset_id)],
        ).fetchone()
        if not row or row[0] is None:
            return None
        try:
            return int(row[0])
        except Exception:
            return None


def _read_conversion_factor_csv(path: Path) -> pd.DataFrame:
    """Read uploaded factor CSVs with resilient encoding/delimiter handling."""
    attempts: list[tuple[str, str]] = []
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return pd.read_csv(path, sep=None, engine="python", encoding=encoding)
        except UnicodeDecodeError as e:
            attempts.append((encoding, str(e)))
            continue
    detail = "; ".join(f"{enc}: {msg}" for enc, msg in attempts) if attempts else "unknown decode failure"
    raise RuntimeError(f"{path.name}: could not decode CSV with supported encodings ({detail})")


def _table_exists(con, table_name: str) -> bool:
    row = con.execute(
        """
        SELECT 1
        FROM information_schema.tables
        WHERE table_schema = 'public' AND table_name = %s
        LIMIT 1
        """,
        [table_name],
    ).fetchone()
    return bool(row)


def _ensure_factor_lookup_schema() -> None:
    """Ensure workbook-friendly columns exist before importing."""
    with get_conn() as con:
        con.execute("ALTER TABLE factor_lookup ADD COLUMN IF NOT EXISTS category VARCHAR")
        con.execute("ALTER TABLE factor_lookup ADD COLUMN IF NOT EXISTS ghg_unit VARCHAR")
        con.execute("ALTER TABLE factor_lookup ADD COLUMN IF NOT EXISTS report_label VARCHAR")


def _normalize_dataset_name(name: str) -> str:
    return " ".join(str(name or "").strip().lower().split())


def _resolve_uk_dataset_id(con, year: int) -> int:
    rows = con.execute(
        """
        SELECT dataset_id, name, source, analysis_type, country, year, archived
        FROM datasets
        WHERE year = %s
          AND COALESCE(archived, FALSE) = FALSE
        ORDER BY
            CASE WHEN lower(coalesce(country, '')) IN ('united kingdom', 'uk') THEN 0 ELSE 1 END,
            CASE WHEN lower(coalesce(analysis_type, '')) LIKE '%%activity%%spend%%' THEN 0 ELSE 1 END,
            dataset_id
        """,
        [int(year)],
    ).fetchall()
    if not rows:
        raise RuntimeError(f"No dataset found for year {year}")

    preferred = {
        _normalize_dataset_name(f"UK Activity & Spend {year}"),
        _normalize_dataset_name(f"DESNZ Activity UK {year}"),
        _normalize_dataset_name(f"UK Activity {year}"),
    }
    for dataset_id, name, source, analysis_type, country, ds_year, archived in rows:
        if _normalize_dataset_name(name) in preferred:
            return int(dataset_id)
    return int(rows[0][0])


def _sheet_year_from_name(sheet_name: str) -> int | None:
    try:
        return int(str(sheet_name or "").strip().split()[0])
    except Exception:
        return None


def _read_workbook_sheets(path: Path) -> list[tuple[str, pd.DataFrame]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: list[tuple[str, pd.DataFrame]] = []
    for ws in workbook.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c).replace("\ufeff", "").replace("\u200b", "").replace("\u200c", "").replace("\u200d", "").strip() if c is not None else "" for c in rows[0]]
        data_rows = [list(r) for r in rows[1:] if not all(v is None for v in r)]
        if not data_rows:
            continue
        df = pd.DataFrame(data_rows, columns=headers)
        sheets.append((ws.title, df))
    return sheets


def _parse_conversion_factor_rows(
    df: pd.DataFrame,
    *,
    dataset_id: int,
    target_dataset_year: int | None,
    file_name: str,
    default_source: str | None = None,
) -> dict[str, Any]:
    """Pure parse+validate of an already-loaded factor upload DataFrame against a
    resolved dataset context. Issues no DB writes. Shared by the commit path
    (`_parse_conversion_factor_upload`) and the preview/dry-run path
    (`preview_conversion_factor_upload`).
    """
    field_columns, column_mapping, missing_required = _resolve_column_mapping(list(df.columns))
    if missing_required:
        return {
            "rows": [],
            "total_rows": int(len(df.index)),
            "accepted_rows": 0,
            "rejected_rows": 0,
            "rejected_details": [],
            "years_seen": [],
            "column_mapping": column_mapping,
            "missing_required_fields": missing_required,
            "content_duplicate_warnings": [],
        }

    c_year = field_columns["year"]
    c_id = field_columns["original_id"]
    c_scope = field_columns["scope"]
    c_category = field_columns["category"]
    c_l1 = field_columns["level_1"]
    c_l2 = field_columns["level_2"]
    c_l3 = field_columns["level_3"]
    c_l4 = field_columns["level_4"]
    c_text = field_columns["column_text"]
    c_report_label = field_columns["report_label"]
    c_uom = field_columns["uom"]
    c_ghg = field_columns["ghg_unit"]
    c_fac = field_columns["factor"]
    c_method = field_columns["method"]
    c_valid_from = field_columns["valid_from"]
    c_valid_to = field_columns["valid_to"]
    c_source = field_columns["source"]
    c_region = field_columns["region"]

    year = None
    if c_year and not df.empty:
        try:
            year = int(df.iloc[0][c_year])
        except Exception:
            pass

    total_rows = int(len(df.index))
    rows: list[list[Any]] = []
    rejected_rows: list[dict[str, Any]] = []
    years_seen: set[int] = set()
    duplicate_keys_seen: set[tuple[int, str, str]] = set()
    content_dupe_groups: dict[tuple[str, ...], dict[str, list[Any]]] = {}
    for _, r in df.iterrows():
        row_number = int(getattr(_, "item", lambda: _)()) + 2 if hasattr(_, "item") else int(_) + 2
        rejection_reasons: list[str] = []

        try:
            factor = float(r[c_fac])
        except Exception:
            factor = None
            rejection_reasons.append("Invalid factor value")

        oid = _norm_original_id(r[c_id])
        scope = _norm_scope(r[c_scope])
        if oid is None:
            rejection_reasons.append("Missing original ID")
        if scope is None:
            rejection_reasons.append("Missing scope")

        yr = int(year or 0)
        if c_year and r.get(c_year) is not None and not (isinstance(r.get(c_year), float) and pd.isna(r.get(c_year))):
            try:
                yr = int(r.get(c_year))
            except Exception:
                rejection_reasons.append("Invalid year value")

        if target_dataset_year is not None and int(yr) != int(target_dataset_year):
            rejection_reasons.append(
                f"Row year {int(yr)} does not match target dataset year {int(target_dataset_year)}"
            )

        if factor is not None and oid is not None and scope is not None:
            dedupe_key = (int(yr), str(scope), str(oid))
            if dedupe_key in duplicate_keys_seen:
                rejection_reasons.append("Duplicate year/scope/original_id in upload")
            duplicate_keys_seen.add(dedupe_key)

        if rejection_reasons:
            rejected_rows.append(
                {
                    "row_number": row_number,
                    "original_id": oid,
                    "scope": scope,
                    "reason": "; ".join(rejection_reasons),
                }
            )
            continue

        category = None
        if c_category:
            category = r.get(c_category)
            if category is not None and not (isinstance(category, float) and pd.isna(category)):
                category = str(category).strip()
                if not category or category.lower() == "nan":
                    category = None

        l1 = None
        if c_l1:
            l1 = r.get(c_l1)
            if l1 is not None and not (isinstance(l1, float) and pd.isna(l1)):
                l1 = str(l1).strip()
                if not l1 or l1.lower() == "nan":
                    l1 = None

        # Backward compatibility: if a file only has one of the two fields,
        # preserve that value in the missing field rather than dropping it.
        if category is None and l1 is not None:
            category = l1
        if l1 is None and category is not None:
            l1 = category

        l2 = None
        if c_l2:
            l2 = r.get(c_l2)
            if l2 is not None and not (isinstance(l2, float) and pd.isna(l2)):
                l2 = str(l2).strip()
                if not l2 or l2.lower() == "nan":
                    l2 = None
        l3 = None
        if c_l3:
            l3 = r.get(c_l3)
            if l3 is not None and not (isinstance(l3, float) and pd.isna(l3)):
                l3 = str(l3).strip()
                if not l3 or l3.lower() == "nan":
                    l3 = None
        l4 = None
        if c_l4:
            l4 = r.get(c_l4)
            if l4 is not None and not (isinstance(l4, float) and pd.isna(l4)):
                l4 = str(l4).strip()
                if not l4 or l4.lower() == "nan":
                    l4 = None

        report_label = None
        if c_report_label:
            report_label = r.get(c_report_label)
            if report_label is not None and not (isinstance(report_label, float) and pd.isna(report_label)):
                report_label = str(report_label).strip()
                if not report_label or report_label.lower() == "nan":
                    report_label = None
        col_text = _synth_column_text(r, c_text, c_category, c_l2, c_l3, c_l4)
        if not report_label:
            report_label = col_text or None
        if not report_label:
            report_label_parts = []
            for part in (category, l1, l2, l3, l4):
                if part and part.lower() != "nan":
                    report_label_parts.append(part)
            if report_label_parts:
                report_label = " ".join(report_label_parts)
        if not col_text:
            col_text = report_label or ""
        uom = r.get(c_uom) if c_uom else None
        ghg_unit = _norm_ghg_unit(r.get(c_ghg) if c_ghg else None)
        years_seen.add(int(yr))

        method = None
        if c_method:
            mv = r.get(c_method)
            if mv is not None and not (isinstance(mv, float) and pd.isna(mv)):
                ms = str(mv).strip()
                if ms and ms.lower() != "nan":
                    method = ms

        valid_from = None
        if c_valid_from:
            vf = r.get(c_valid_from)
            if vf is not None and not (isinstance(vf, float) and pd.isna(vf)):
                try:
                    dt = pd.to_datetime(vf, dayfirst=True, errors="coerce")
                    valid_from = None if pd.isna(dt) else dt.date()
                except Exception:
                    valid_from = None

        valid_to = None
        if c_valid_to:
            vt = r.get(c_valid_to)
            if vt is not None and not (isinstance(vt, float) and pd.isna(vt)):
                try:
                    dt = pd.to_datetime(vt, dayfirst=True, errors="coerce")
                    valid_to = None if pd.isna(dt) else dt.date()
                except Exception:
                    valid_to = None

        # Get source from CSV row if available, otherwise use dataset-level source
        row_source = None
        if c_source:
            src_val = r.get(c_source)
            if src_val is not None and not (isinstance(src_val, float) and pd.isna(src_val)):
                row_source = str(src_val).strip()
        if not row_source:
            row_source = default_source if default_source else ""

        region = None
        if c_region:
            region_val = r.get(c_region)
            if region_val is not None and not (isinstance(region_val, float) and pd.isna(region_val)):
                region = str(region_val).strip()
                if not region or region.lower() == "nan":
                    region = None

        rows.append(
            [
                int(dataset_id),
                file_name,
                int(yr),
                oid,
                scope,
                category,
                l1,
                l2,
                l3,
                l4,
                col_text,
                (str(uom).strip() if uom is not None else None),
                ghg_unit,
                float(factor),
                row_source,
                region or "",
                "",
                method,
                valid_from,
                valid_to,
                report_label,
            ]
        )

        dupe_key = (
            _norm_dupe_text(scope), _norm_dupe_text(category),
            _norm_dupe_text(l1), _norm_dupe_text(l2), _norm_dupe_text(l3), _norm_dupe_text(l4),
            _norm_dupe_text(report_label or col_text),
        )
        if any(dupe_key):
            group = content_dupe_groups.setdefault(dupe_key, {"original_ids": [], "row_numbers": []})
            if oid not in group["original_ids"]:
                group["original_ids"].append(oid)
            group["row_numbers"].append(row_number)

    content_duplicate_warnings = [
        {
            "original_ids": g["original_ids"],
            "row_numbers": g["row_numbers"],
            "shared_fields": {
                "scope": key[0], "category": key[1], "level_1": key[2], "level_2": key[3],
                "level_3": key[4], "level_4": key[5], "report_label_or_column_text": key[6],
            },
        }
        for key, g in content_dupe_groups.items()
        if len(g["original_ids"]) > 1
    ]

    return {
        "rows": rows,
        "total_rows": total_rows,
        "accepted_rows": len(rows),
        "rejected_rows": len(rejected_rows),
        "rejected_details": rejected_rows,
        "years_seen": sorted(years_seen),
        "column_mapping": column_mapping,
        "missing_required_fields": [],
        "content_duplicate_warnings": content_duplicate_warnings,
    }


def _parse_conversion_factor_upload(path: Path, *, dataset_id: int | None = None) -> dict[str, Any]:
    # Be forgiving with uploaded CSVs from different source systems by
    # autodetecting the separator and cleaning up header artifacts.
    df = _read_conversion_factor_csv(path)
    df.columns = [str(c).replace("\ufeff", "").replace("\u200b", "").replace("\u200c", "").replace("\u200d", "").strip() for c in df.columns]

    field_columns_probe, _mapping_probe, _missing_probe = _resolve_column_mapping(list(df.columns))
    c_year = field_columns_probe["year"]

    # Initialize year from CSV if available, otherwise use None
    year = None
    if c_year and not df.empty:
        try:
            year = int(df.iloc[0][c_year])
        except Exception:
            pass

    # If dataset_id is provided, use it; otherwise create/find dataset from filename
    if dataset_id is None:
        name, source, analysis_type, year_guess, country = _dataset_meta_from_filename(path)
        # Use year from CSV if available, otherwise use year from filename
        if year is None:
            year = year_guess

        dataset_id = _ensure_dataset(name=name, source=source, analysis_type=analysis_type, country=country, year=int(year or 0))
        target_dataset_year = int(year or 0) if year is not None else None
    else:
        # When uploading to existing dataset, source comes from CSV or is empty
        source = None
        target_dataset_year = _get_dataset_year(int(dataset_id))

    parsed = _parse_conversion_factor_rows(
        df,
        dataset_id=int(dataset_id),
        target_dataset_year=target_dataset_year,
        file_name=path.name,
        default_source=source,
    )

    if parsed["missing_required_fields"]:
        raise RuntimeError(f"{path.name}: missing required columns (need Scope, ID, Factor). Found: {list(df.columns)}")

    parsed["dataset_id"] = int(dataset_id)
    return parsed


def preview_conversion_factor_upload(path: Path, *, dataset_id: int) -> dict[str, Any]:
    """Parse + validate a CSV against an existing dataset_id and diff the accepted
    rows against factor_lookup, WITHOUT writing anything. Mirrors the replace=True
    semantics the commit route always uses (see ingest_csv_with_report)."""
    df = _read_conversion_factor_csv(path)
    df.columns = [str(c).replace("\ufeff", "").replace("\u200b", "").replace("\u200c", "").replace("\u200d", "").strip() for c in df.columns]
    target_dataset_year = _get_dataset_year(int(dataset_id))
    parsed = _parse_conversion_factor_rows(
        df,
        dataset_id=int(dataset_id),
        target_dataset_year=target_dataset_year,
        file_name=path.name,
        default_source=None,
    )

    result: dict[str, Any] = {
        "dataset_id": int(dataset_id),
        "target_dataset_year": target_dataset_year,
        "blocking_error": None,
        "column_mapping": parsed["column_mapping"],
        "missing_required_fields": parsed["missing_required_fields"],
        "rejected_details": parsed["rejected_details"],
        "content_duplicate_warnings": parsed["content_duplicate_warnings"],
        "years_seen": parsed["years_seen"],
    }

    if parsed["missing_required_fields"]:
        result["blocking_error"] = "Missing required column(s): " + ", ".join(parsed["missing_required_fields"])
        result["counts"] = {
            "total_rows": parsed["total_rows"],
            "accepted_rows": 0,
            "rejected_rows": 0,
            "would_insert": 0,
            "would_update": 0,
            "would_delete_stale": 0,
            "blocked_stale_rows": 0,
        }
        result["replacement_blocked"] = False
        result["dependency_summary"] = {}
        return result

    rows = parsed["rows"]
    incoming_keys = {(int(r[2]), str(r[4]), str(r[3])) for r in rows}
    with get_conn() as con:
        existing_rows = con.execute(
            "SELECT db_id, year, scope, original_id FROM factor_lookup WHERE dataset_id = %s",
            [int(dataset_id)],
        ).fetchall()
        existing_keys = {(int(yr or 0), str(scope or ""), str(oid or "")) for _, yr, scope, oid in existing_rows}
        stale_ids = [
            int(db_id)
            for db_id, yr, scope, oid in existing_rows
            if (int(yr or 0), str(scope or ""), str(oid or "")) not in incoming_keys
        ]
        deps = _replacement_dependency_summary(con, stale_ids) if stale_ids else {}

    would_update = sum(1 for r in rows if (int(r[2]), str(r[4]), str(r[3])) in existing_keys)
    would_insert = len(rows) - would_update
    replacement_blocked = bool(deps)

    result["counts"] = {
        "total_rows": parsed["total_rows"],
        "accepted_rows": parsed["accepted_rows"],
        "rejected_rows": parsed["rejected_rows"],
        "would_insert": would_insert,
        "would_update": would_update,
        "would_delete_stale": 0 if replacement_blocked else len(stale_ids),
        "blocked_stale_rows": len(stale_ids) if replacement_blocked else 0,
    }
    result["replacement_blocked"] = replacement_blocked
    result["dependency_summary"] = deps
    return result


def _parse_conversion_factor_workbook(path: Path) -> list[dict[str, Any]]:
    """Parse a year-split workbook into per-sheet factor rows."""
    workbook_sheets = _read_workbook_sheets(path)
    if not workbook_sheets:
        raise RuntimeError(f"{path.name}: no usable sheets found")

    summaries: list[dict[str, Any]] = []
    with get_conn() as con:
        for sheet_name, df in workbook_sheets:
            sheet_year = _sheet_year_from_name(sheet_name)
            if sheet_year is None:
                continue

            dataset_id = _resolve_uk_dataset_id(con, sheet_year)

            field_columns, _mapping, missing_required = _resolve_column_mapping(list(df.columns))
            c_id = field_columns["original_id"]
            c_scope = field_columns["scope"]
            c_category = field_columns["category"]
            c_l1 = field_columns["level_1"]
            c_l2 = field_columns["level_2"]
            c_l3 = field_columns["level_3"]
            c_l4 = field_columns["level_4"]
            c_text = field_columns["column_text"]
            c_report_label = field_columns["report_label"]
            c_uom = field_columns["uom"]
            c_ghg = field_columns["ghg_unit"]
            c_fac = field_columns["factor"]
            c_method = field_columns["method"]
            c_valid_from = field_columns["valid_from"]
            c_valid_to = field_columns["valid_to"]
            c_source = field_columns["source"]

            if missing_required:
                raise RuntimeError(
                    f"{path.name} / {sheet_name}: missing required columns (need Scope, ID, Factor). Found: {list(df.columns)}"
                )

            rows: list[list[Any]] = []
            rejected_rows: list[dict[str, Any]] = []
            years_seen: set[int] = set()
            duplicate_keys_seen: set[tuple[int, str]] = set()

            for idx, r in df.iterrows():
                row_number = int(idx) + 2
                rejection_reasons: list[str] = []

                try:
                    factor = float(r[c_fac])
                except Exception:
                    factor = None
                    rejection_reasons.append("Invalid factor value")

                oid = _norm_original_id(r[c_id])
                scope = _norm_scope(r[c_scope])
                if oid is None:
                    rejection_reasons.append("Missing original ID")
                if scope is None:
                    rejection_reasons.append("Missing scope")

                yr = int(sheet_year)
                years_seen.add(yr)

                if factor is not None and oid is not None and scope is not None:
                    dedupe_key = (yr, str(oid))
                    if dedupe_key in duplicate_keys_seen:
                        rejection_reasons.append("Duplicate year/original_id in upload")
                    duplicate_keys_seen.add(dedupe_key)

                if rejection_reasons:
                    rejected_rows.append(
                        {
                            "sheet_name": sheet_name,
                            "row_number": row_number,
                            "original_id": oid,
                            "scope": scope,
                            "reason": "; ".join(rejection_reasons),
                        }
                    )
                    continue

                category = None
                if c_category:
                    category = r.get(c_category)
                    if category is not None and not (isinstance(category, float) and pd.isna(category)):
                        category = str(category).strip()
                        if not category or category.lower() == "nan":
                            category = None

                l1 = None
                if c_l1:
                    l1 = r.get(c_l1)
                    if l1 is not None and not (isinstance(l1, float) and pd.isna(l1)):
                        l1 = str(l1).strip()
                        if not l1 or l1.lower() == "nan":
                            l1 = None

                l2 = None
                if c_l2:
                    l2 = r.get(c_l2)
                    if l2 is not None and not (isinstance(l2, float) and pd.isna(l2)):
                        l2 = str(l2).strip()
                        if not l2 or l2.lower() == "nan":
                            l2 = None
                l3 = None
                if c_l3:
                    l3 = r.get(c_l3)
                    if l3 is not None and not (isinstance(l3, float) and pd.isna(l3)):
                        l3 = str(l3).strip()
                        if not l3 or l3.lower() == "nan":
                            l3 = None
                l4 = None
                if c_l4:
                    l4 = r.get(c_l4)
                    if l4 is not None and not (isinstance(l4, float) and pd.isna(l4)):
                        l4 = str(l4).strip()
                        if not l4 or l4.lower() == "nan":
                            l4 = None

                report_label = None
                if c_report_label:
                    report_label = r.get(c_report_label)
                    if report_label is not None and not (isinstance(report_label, float) and pd.isna(report_label)):
                        report_label = str(report_label).strip()
                        if not report_label or report_label.lower() == "nan":
                            report_label = None

                col_text = _synth_column_text(r, c_text, c_category, c_l2, c_l3, c_l4)
                if not report_label:
                    report_label = col_text or None
                if not col_text:
                    col_text = report_label or ""

                uom = r.get(c_uom) if c_uom else None
                ghg_unit = _norm_ghg_unit(r.get(c_ghg) if c_ghg else None)

                method = None
                if c_method:
                    mv = r.get(c_method)
                    if mv is not None and not (isinstance(mv, float) and pd.isna(mv)):
                        ms = str(mv).strip()
                        if ms and ms.lower() != "nan":
                            method = ms

                valid_from = None
                if c_valid_from:
                    vf = r.get(c_valid_from)
                    if vf is not None and not (isinstance(vf, float) and pd.isna(vf)):
                        try:
                            dt = pd.to_datetime(vf, dayfirst=True, errors="coerce")
                            valid_from = None if pd.isna(dt) else dt.date()
                        except Exception:
                            valid_from = None

                valid_to = None
                if c_valid_to:
                    vt = r.get(c_valid_to)
                    if vt is not None and not (isinstance(vt, float) and pd.isna(vt)):
                        try:
                            dt = pd.to_datetime(vt, dayfirst=True, errors="coerce")
                            valid_to = None if pd.isna(dt) else dt.date()
                        except Exception:
                            valid_to = None

                row_source = None
                if c_source:
                    src_val = r.get(c_source)
                    if src_val is not None and not (isinstance(src_val, float) and pd.isna(src_val)):
                        row_source = str(src_val).strip()
                if not row_source:
                    row_source = "DESNZ & DEFRA"

                rows.append(
                    [
                        int(dataset_id),
                        path.name,
                        int(yr),
                        oid,
                        scope,
                        category,
                        l1,
                        l2,
                        l3,
                        l4,
                        col_text,
                        (str(uom).strip() if uom is not None else None),
                        ghg_unit,
                        float(factor),
                        row_source,
                        "UK",
                        "GBP",
                        method,
                        valid_from,
                        valid_to,
                        report_label,
                    ]
                )

            summaries.append(
                {
                    "sheet_name": sheet_name,
                    "year": int(sheet_year),
                    "dataset_id": int(dataset_id),
                    "rows": rows,
                    "total_rows": int(len(df.index)),
                    "accepted_rows": len(rows),
                    "rejected_rows": len(rejected_rows),
                    "rejected_details": rejected_rows,
                    "years_seen": sorted(years_seen),
                }
            )

    return summaries


def _replacement_dependency_summary(con, stale_factor_ids: list[int]) -> dict[str, Any]:
    if not stale_factor_ids:
        return {}

    placeholders = ",".join(["%s"] * len(stale_factor_ids))
    summary: dict[str, Any] = {}

    if _table_exists(con, "job_scope_rows"):
        # Exclude emissions-fallback rows (uom='tCO2e', factor≈1.0): those rows store
        # their value as pre-computed tCO2e and don't live-calculate from factor_db_id,
        # so deleting the referenced factor has no effect on their calculations.
        row = con.execute(
            f"""
            SELECT COUNT(*) FROM job_scope_rows
            WHERE factor_db_id IN ({placeholders})
              AND COALESCE(enabled, TRUE) = TRUE
              AND NOT (
                LOWER(COALESCE(uom, '')) = 'tco2e'
                AND factor IS NOT NULL
                AND ABS(factor - 1.0) < 0.0001
              )
            """,
            stale_factor_ids,
        ).fetchone()
        count = int(row[0] or 0)
        if count:
            summary["job_scope_rows"] = count
            try:
                jobs_rows = con.execute(
                    f"""
                    SELECT DISTINCT j.job_id, j.job_number, j.title
                    FROM job_scope_rows jsr
                    JOIN jobs j ON j.job_id = jsr.job_id
                    WHERE jsr.factor_db_id IN ({placeholders})
                      AND COALESCE(jsr.enabled, TRUE) = TRUE
                      AND NOT (
                        LOWER(COALESCE(jsr.uom, '')) = 'tco2e'
                        AND jsr.factor IS NOT NULL
                        AND ABS(jsr.factor - 1.0) < 0.0001
                      )
                    ORDER BY j.job_number
                    """,
                    stale_factor_ids,
                ).fetchall()
                summary["job_scope_rows_jobs"] = [
                    {"job_id": int(r[0]), "job_number": str(r[1] or ""), "title": str(r[2] or "")}
                    for r in jobs_rows
                ]
            except Exception:
                pass

    if _table_exists(con, "job_spend_entries"):
        row = con.execute(
            f"SELECT COUNT(*) FROM job_spend_entries WHERE factor_db_id IN ({placeholders}) AND COALESCE(is_deleted, FALSE) = FALSE",
            stale_factor_ids,
        ).fetchone()
        count = int(row[0] or 0)
        if count:
            summary["job_spend_entries"] = count
            try:
                jobs_rows = con.execute(
                    f"""
                    SELECT DISTINCT j.job_id, j.job_number, j.title
                    FROM job_spend_entries jse
                    JOIN jobs j ON j.job_id = jse.job_id
                    WHERE jse.factor_db_id IN ({placeholders})
                      AND COALESCE(jse.is_deleted, FALSE) = FALSE
                    ORDER BY j.job_number
                    """,
                    stale_factor_ids,
                ).fetchall()
                summary["job_spend_entries_jobs"] = [
                    {"job_id": int(r[0]), "job_number": str(r[1] or ""), "title": str(r[2] or "")}
                    for r in jobs_rows
                ]
            except Exception:
                pass

    if _table_exists(con, "lca_inventory_items"):
        row = con.execute(
            f"SELECT COUNT(*) FROM lca_inventory_items WHERE mapped_factor_db_id IN ({placeholders})",
            stale_factor_ids,
        ).fetchone()
        count = int(row[0] or 0)
        if count:
            summary["lca_inventory_items"] = count

    return {k: v for k, v in summary.items() if v}


def _sync_normalised_tables_pg(cur, rows: list[list[Any]]) -> None:
    """Sync emission_factor_* tables after factor_lookup writes (Postgres only).

    For each imported row:
    - If (dataset_id, original_id) has no alias yet → insert a new
      emission_factor_definitions row (using factor_lookup.db_id as factor_id,
      consistent with Phase 1 backfill seeding) and the matching alias.
    - Upsert emission_factor_year_values for all aliased rows.

    Description fields on existing definitions are intentionally NOT overwritten
    so that manual edits made via the admin UI are preserved.
    """
    if not rows:
        return

    # Row layout (matches _parse_conversion_factor_upload / _parse_conversion_factor_workbook):
    # [0]=dataset_id [1]=file_name [2]=year [3]=original_id [4]=scope
    # [5]=category   [6]=level_1   [7]=level_2 [8]=level_3 [9]=level_4
    # [10]=column_text [11]=uom [12]=ghg_unit [13]=factor [14]=source
    # [15]=region [16]=currency [17]=method [18]=valid_from [19]=valid_to [20]=report_label
    sync_rows = [
        (r[0], r[3], r[2], r[4], r[6], r[7], r[8], r[9],
         r[11], r[12], r[14], r[5], r[20], r[17], r[13], r[16], r[18], r[19], r[10])
        for r in rows
    ]

    cur.execute("""
        CREATE TEMP TABLE IF NOT EXISTS tmp_nf_sync (
            dataset_id   INTEGER,
            original_id  VARCHAR,
            year         INTEGER,
            scope        VARCHAR,
            level_1      VARCHAR,
            level_2      VARCHAR,
            level_3      VARCHAR,
            level_4      VARCHAR,
            uom          VARCHAR,
            ghg_unit     VARCHAR,
            source       VARCHAR,
            category     VARCHAR,
            report_label VARCHAR,
            method       VARCHAR,
            factor       NUMERIC,
            currency     VARCHAR,
            valid_from   DATE,
            valid_to     DATE,
            column_text  VARCHAR
        ) ON COMMIT DROP
    """)
    cur.execute("TRUNCATE tmp_nf_sync")
    cur.executemany(
        """
        INSERT INTO tmp_nf_sync
            (dataset_id, original_id, year, scope, level_1, level_2, level_3, level_4,
             uom, ghg_unit, source, category, report_label, method,
             factor, currency, valid_from, valid_to, column_text)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,
        sync_rows,
    )

    # Step 1+2: Insert new definitions and aliases for pairs not yet in emission_factor_aliases.
    # Use the lowest factor_lookup.db_id for this (dataset_id, original_id) as the factor_id —
    # consistent with Phase 1 backfill seeding.
    cur.execute("""
        WITH unaliased AS (
            SELECT DISTINCT ON (t.dataset_id, t.original_id)
                fl.db_id        AS factor_id,
                t.dataset_id,
                t.original_id,
                t.scope,
                t.level_1,
                t.level_2,
                t.level_3,
                t.level_4,
                t.uom,
                t.ghg_unit,
                COALESCE(NULLIF(TRIM(t.source), ''), '') AS source,
                t.category,
                t.report_label,
                t.method,
                t.column_text
            FROM tmp_nf_sync t
            JOIN factor_lookup fl
                ON  fl.dataset_id  = t.dataset_id
                AND TRIM(fl.original_id) = TRIM(t.original_id)
            WHERE NOT EXISTS (
                SELECT 1 FROM emission_factor_aliases efa
                WHERE efa.dataset_id = t.dataset_id
                  AND TRIM(efa.original_id) = TRIM(t.original_id)
            )
            ORDER BY t.dataset_id, t.original_id, fl.db_id ASC
        ),
        ins_defs AS (
            INSERT INTO emission_factor_definitions
                (factor_id, scope, level_1, level_2, level_3, level_4,
                 uom, ghg_unit, source, region, category, report_label, method, archived, column_text)
            SELECT
                factor_id, scope, level_1, level_2, level_3, level_4,
                uom, ghg_unit, source, '' AS region, category, report_label, method, FALSE, column_text
            FROM unaliased
            ON CONFLICT (factor_id) DO NOTHING
            RETURNING factor_id
        )
        INSERT INTO emission_factor_aliases (factor_id, dataset_id, original_id)
        SELECT u.factor_id, u.dataset_id, u.original_id
        FROM unaliased u
        WHERE EXISTS (SELECT 1 FROM ins_defs d WHERE d.factor_id = u.factor_id)
        ON CONFLICT (dataset_id, original_id) DO NOTHING
    """)

    # Step 3: Upsert year-values for every row that now has an alias.
    # ON CONFLICT uses the uq_efyv_dataset_original unique index (dataset_id, original_id).
    cur.execute("""
        INSERT INTO emission_factor_year_values
            (factor_id, dataset_id, original_id, year, factor, currency, valid_from, valid_to)
        SELECT
            efa.factor_id,
            t.dataset_id,
            t.original_id,
            t.year,
            t.factor,
            t.currency,
            t.valid_from,
            t.valid_to
        FROM tmp_nf_sync t
        JOIN emission_factor_aliases efa
            ON  efa.dataset_id  = t.dataset_id
            AND TRIM(efa.original_id) = TRIM(t.original_id)
        ON CONFLICT (dataset_id, original_id) DO UPDATE
            SET factor_id  = EXCLUDED.factor_id,
                year       = EXCLUDED.year,
                factor     = EXCLUDED.factor,
                currency   = EXCLUDED.currency,
                valid_from = EXCLUDED.valid_from,
                valid_to   = EXCLUDED.valid_to
    """)


def _apply_factor_rows(
    *,
    path: Path,
    dataset_id: int,
    rows: list[list[Any]],
    replace: bool,
    incoming_keys: set[tuple[int, str, str]],
) -> dict[str, Any]:
    inserted_rows = 0
    updated_rows = 0
    deleted_rows = 0
    blocked_rows = 0

    with get_conn(autocommit=False) as con:
        if db_backend() == "postgres":
            raw = con._conn  # type: ignore[attr-defined]
            with raw.cursor() as cur:
                cur.execute(
                    """
                    CREATE TEMP TABLE tmp_factor_import (
                        dataset_id INTEGER,
                        file_name VARCHAR,
                        year INTEGER,
                        original_id VARCHAR,
                        scope VARCHAR,
                        category VARCHAR,
                        level_1 VARCHAR,
                        level_2 VARCHAR,
                        level_3 VARCHAR,
                        level_4 VARCHAR,
                        column_text VARCHAR,
                        uom VARCHAR,
                        ghg_unit VARCHAR,
                        factor NUMERIC,
                        source VARCHAR,
                        region VARCHAR,
                        currency VARCHAR,
                        method VARCHAR,
                        valid_from DATE,
                        valid_to DATE,
                        report_label VARCHAR
                    ) ON COMMIT DROP
                    """
                )
                cur.executemany(
                    """
                    INSERT INTO tmp_factor_import (
                        dataset_id, file_name, year, original_id, scope,
                        category, level_1, level_2, level_3, level_4, column_text,
                        uom, ghg_unit, factor, source, region, currency,
                        method, valid_from, valid_to, report_label
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    rows,
                )
                if replace:
                    existing_rows = cur.execute(
                        """
                        SELECT db_id, year, scope, original_id
                        FROM factor_lookup
                        WHERE dataset_id = %s
                        """,
                        [dataset_id],
                    ).fetchall()
                    stale_ids = [
                        int(db_id)
                        for db_id, yr, scope, original_id in existing_rows
                        if (int(yr or 0), str(scope or ""), str(original_id or "")) not in incoming_keys
                    ]
                    if stale_ids:
                        referenced_ids: set[int] = set()
                        if _table_exists(con, "job_scope_rows"):
                            cur.execute(
                                "SELECT DISTINCT factor_db_id FROM job_scope_rows WHERE factor_db_id = ANY(%s)",
                                [stale_ids],
                            )
                            referenced_ids.update(int(r[0]) for r in cur.fetchall() if r and r[0] is not None)
                        if _table_exists(con, "job_spend_entries"):
                            cur.execute(
                                "SELECT DISTINCT factor_db_id FROM job_spend_entries WHERE factor_db_id = ANY(%s)",
                                [stale_ids],
                            )
                            referenced_ids.update(int(r[0]) for r in cur.fetchall() if r and r[0] is not None)
                        if _table_exists(con, "lca_inventory_items"):
                            cur.execute(
                                "SELECT DISTINCT mapped_factor_db_id FROM lca_inventory_items WHERE mapped_factor_db_id = ANY(%s)",
                                [stale_ids],
                            )
                            referenced_ids.update(int(r[0]) for r in cur.fetchall() if r and r[0] is not None)

                        delete_ids = [db_id for db_id in stale_ids if db_id not in referenced_ids]
                        if delete_ids:
                            placeholders = ",".join(["%s"] * len(delete_ids))
                            cur.execute(
                                f"DELETE FROM factor_lookup WHERE db_id IN ({placeholders})",
                                delete_ids,
                            )
                            deleted_rows = cur.rowcount or len(delete_ids)
                        blocked_rows = len(referenced_ids)

                    cur.execute(
                        """
                        UPDATE factor_lookup AS f
                        SET file_name = t.file_name,
                            category = t.category,
                            level_1 = t.level_1,
                            level_2 = t.level_2,
                            level_3 = t.level_3,
                            level_4 = t.level_4,
                            column_text = t.column_text,
                            uom = t.uom,
                            ghg_unit = t.ghg_unit,
                            factor = t.factor,
                            source = t.source,
                            region = t.region,
                            currency = t.currency,
                            method = t.method,
                            valid_from = t.valid_from,
                            valid_to = t.valid_to,
                            report_label = t.report_label
                        FROM tmp_factor_import AS t
                        WHERE f.dataset_id = t.dataset_id
                          AND f.year = t.year
                          AND f.scope = t.scope
                          AND f.original_id = t.original_id
                        """,
                    )
                    updated_rows = cur.rowcount or 0

                    cur.execute(
                        """
                        INSERT INTO factor_lookup (
                            dataset_id, file_name, year, original_id, scope,
                            category, level_1, level_2, level_3, level_4, column_text,
                            uom, ghg_unit, factor, source, region, currency,
                            method, valid_from, valid_to, report_label
                        )
                        SELECT
                            t.dataset_id, t.file_name, t.year, t.original_id, t.scope,
                            t.category, t.level_1, t.level_2, t.level_3, t.level_4, t.column_text,
                            t.uom, t.ghg_unit, t.factor, t.source, t.region, t.currency,
                            t.method, t.valid_from, t.valid_to, t.report_label
                        FROM tmp_factor_import AS t
                        LEFT JOIN factor_lookup AS f
                          ON f.dataset_id = t.dataset_id
                         AND f.year = t.year
                         AND f.scope = t.scope
                         AND f.original_id = t.original_id
                        WHERE f.db_id IS NULL
                        """,
                    )
                    inserted_rows = max(0, len(rows) - updated_rows)
                else:
                    # Non-replace mode still updates matching rows and inserts missing ones.
                    cur.execute(
                        """
                        UPDATE factor_lookup AS f
                        SET file_name = t.file_name,
                            category = t.category,
                            level_1 = t.level_1,
                            level_2 = t.level_2,
                            level_3 = t.level_3,
                            level_4 = t.level_4,
                            column_text = t.column_text,
                            uom = t.uom,
                            ghg_unit = t.ghg_unit,
                            factor = t.factor,
                            source = t.source,
                            region = t.region,
                            currency = t.currency,
                            method = t.method,
                            valid_from = t.valid_from,
                            valid_to = t.valid_to,
                            report_label = t.report_label
                        FROM tmp_factor_import AS t
                        WHERE f.dataset_id = t.dataset_id
                          AND f.year = t.year
                          AND f.scope = t.scope
                          AND f.original_id = t.original_id
                        """,
                    )
                    updated_rows = cur.rowcount or 0
                    cur.execute(
                        """
                        INSERT INTO factor_lookup (
                            dataset_id, file_name, year, original_id, scope,
                            category, level_1, level_2, level_3, level_4, column_text,
                            uom, ghg_unit, factor, source, region, currency,
                            method, valid_from, valid_to, report_label
                        )
                        SELECT
                            t.dataset_id, t.file_name, t.year, t.original_id, t.scope,
                            t.category, t.level_1, t.level_2, t.level_3, t.level_4, t.column_text,
                            t.uom, t.ghg_unit, t.factor, t.source, t.region, t.currency,
                            t.method, t.valid_from, t.valid_to, t.report_label
                        FROM tmp_factor_import AS t
                        LEFT JOIN factor_lookup AS f
                          ON f.dataset_id = t.dataset_id
                         AND f.year = t.year
                         AND f.scope = t.scope
                         AND f.original_id = t.original_id
                        WHERE f.db_id IS NULL
                        """,
                    )
                    inserted_rows = max(0, len(rows) - updated_rows)

                _sync_normalised_tables_pg(cur, rows)
        else:
            # DuckDB fallback keeps the older row-by-row path.
            for params in rows:
                (
                    p_dataset_id,
                    p_file_name,
                    p_year,
                    p_original_id,
                    p_scope,
                    p_category,
                    p_level_1,
                    p_level_2,
                    p_level_3,
                    p_level_4,
                    p_column_text,
                    p_uom,
                    p_ghg_unit,
                    p_factor,
                    p_source,
                    p_region,
                    p_currency,
                    p_method,
                    p_valid_from,
                    p_valid_to,
                    p_report_label,
                ) = params
                existing = con.execute(
                    """
                    SELECT db_id
                    FROM factor_lookup
                    WHERE dataset_id = ?
                      AND year = ?
                      AND scope = ?
                      AND original_id = ?
                    ORDER BY db_id
                    LIMIT 1
                    """,
                    [p_dataset_id, p_year, p_scope, p_original_id],
                ).fetchone()
                if replace and existing:
                    con.execute(
                        """
                        UPDATE factor_lookup
                        SET file_name = ?,
                            category = ?,
                            level_1 = ?,
                            level_2 = ?,
                            level_3 = ?,
                            level_4 = ?,
                            column_text = ?,
                            uom = ?,
                            ghg_unit = ?,
                            factor = ?,
                            source = ?,
                            region = ?,
                            currency = ?,
                            method = ?,
                            valid_from = ?,
                            valid_to = ?,
                            report_label = ?
                        WHERE db_id = ?
                        """,
                        [
                            p_file_name,
                            p_category,
                            p_level_1,
                            p_level_2,
                            p_level_3,
                            p_level_4,
                            p_column_text,
                            p_uom,
                            p_ghg_unit,
                            p_factor,
                            p_source,
                            p_region,
                            p_currency,
                            p_method,
                            p_valid_from,
                            p_valid_to,
                            p_report_label,
                            int(existing[0]),
                        ],
                    )
                    updated_rows += 1
                else:
                    con.execute(
                        """
                        INSERT INTO factor_lookup
                        (dataset_id, file_name, year, original_id, scope,
                         category, level_1, level_2, level_3, level_4, column_text,
                         uom, ghg_unit, factor, source, region, currency,
                         method, valid_from, valid_to, report_label)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """,
                        [
                            p_dataset_id,
                            p_file_name,
                            p_year,
                            p_original_id,
                            p_scope,
                            p_category,
                            p_level_1,
                            p_level_2,
                            p_level_3,
                            p_level_4,
                            p_column_text,
                            p_uom,
                            p_ghg_unit,
                            p_factor,
                            p_source,
                            p_region,
                            p_currency,
                            p_method,
                            p_valid_from,
                            p_valid_to,
                            p_report_label,
                        ],
                    )
                    inserted_rows += 1

    return {
        "dataset_id": int(dataset_id),
        "inserted_rows": int(inserted_rows),
        "updated_rows": int(updated_rows),
        "deleted_rows": int(deleted_rows),
        "blocked_rows": int(blocked_rows),
    }


def ingest_workbook_with_report(path: Path, *, replace: bool = True) -> list[dict[str, Any]]:
    _ensure_factor_lookup_schema()
    summaries = _parse_conversion_factor_workbook(path)
    output: list[dict[str, Any]] = []
    for summary in summaries:
        rows = summary["rows"]
        incoming_keys = {(int(r[2]), str(r[4]), str(r[3])) for r in rows}
        applied = _apply_factor_rows(
            path=path,
            dataset_id=int(summary["dataset_id"]),
            rows=rows,
            replace=replace,
            incoming_keys=incoming_keys,
        )
        summary.update(applied)
        summary["message"] = (
            f"{summary['sheet_name']}: {summary['accepted_rows']} accepted, "
            f"{summary['updated_rows']} updated, {summary['inserted_rows']} inserted, "
            f"{summary['deleted_rows']} deleted, {summary['blocked_rows']} blocked"
        )
        output.append(summary)
    return output


def ingest_csv_with_report(path: Path, *, replace: bool, dataset_id: int | None = None) -> dict[str, Any]:
    parsed = _parse_conversion_factor_upload(path, dataset_id=dataset_id)
    rows = parsed["rows"]
    dataset_id = int(parsed["dataset_id"])

    if not rows:
        parsed.update({"deleted_rows": 0, "message": "No valid factor rows found in upload"})
        return parsed

    incoming_keys = {(int(r[2]), str(r[4]), str(r[3])) for r in rows}

    with get_conn(autocommit=False) as con:
        deleted_rows = 0
        if replace:
            existing_rows = con.execute(
                """
                SELECT db_id, year, scope, original_id
                FROM factor_lookup
                WHERE dataset_id = %s
                """,
                [dataset_id],
            ).fetchall()
            stale_ids = [
                int(db_id)
                for db_id, yr, scope, original_id in existing_rows
                if (int(yr or 0), str(scope or ""), str(original_id or "")) not in incoming_keys
            ]
            if stale_ids:
                deps = _replacement_dependency_summary(con, stale_ids)
                if deps:
                    count_parts = []
                    job_refs: list[str] = []
                    for name, value in sorted(deps.items()):
                        if name.endswith("_jobs"):
                            continue
                        count_parts.append(f"{name}={value}")
                    for key in ("job_scope_rows_jobs", "job_spend_entries_jobs"):
                        for j in deps.get(key, []):
                            label = j.get("job_number") or f"job_id={j.get('job_id')}"
                            title = j.get("title") or ""
                            ref = f"{label} ({title})" if title else label
                            if ref not in job_refs:
                                job_refs.append(ref)
                    details = ", ".join(count_parts)
                    job_hint = f" Affected jobs: {'; '.join(job_refs)}." if job_refs else ""
                    raise DatasetReplacementBlocked(
                        "STOP: this upload would remove factor rows that are already referenced by active job rows. "
                        f"Delete those job rows first ({details}).{job_hint}",
                        deps,
                    )
                placeholders = ",".join(["%s"] * len(stale_ids))
                # Clear factor_db_id and dataset_id on any job_scope_rows that still
                # reference these stale factor IDs (emissions-fallback rows whose FK we
                # allowed through the 409 check). Must be NULLed before the DELETE or
                # Postgres will raise a foreign key violation.
                con.execute(
                    f"UPDATE job_scope_rows SET factor_db_id = NULL WHERE factor_db_id IN ({placeholders})",
                    stale_ids,
                )
                con.execute(
                    f"DELETE FROM factor_lookup WHERE db_id IN ({placeholders})",
                    stale_ids,
                )
                deleted_rows = len(stale_ids)

        for params in rows:
            (
                p_dataset_id,
                p_file_name,
                p_year,
                p_original_id,
                p_scope,
                p_category,
                p_level_1,
                p_level_2,
                p_level_3,
                p_level_4,
                p_column_text,
                p_uom,
                p_ghg_unit,
                p_factor,
                p_source,
                p_region,
                p_currency,
                p_method,
                p_valid_from,
                p_valid_to,
                p_report_label,
            ) = params

            if db_backend() == "postgres":
                existing = con.execute(
                    """
                    SELECT db_id
                    FROM factor_lookup
                    WHERE dataset_id = %s
                      AND year = %s
                      AND scope = %s
                      AND original_id = %s
                    ORDER BY db_id
                    LIMIT 1
                    """,
                    [p_dataset_id, p_year, p_scope, p_original_id],
                ).fetchone()
                if replace and existing:
                    con.execute(
                        """
                    UPDATE factor_lookup
                    SET file_name = %s,
                        category = %s,
                        level_1 = %s,
                        level_2 = %s,
                        level_3 = %s,
                        level_4 = %s,
                            column_text = %s,
                            uom = %s,
                            ghg_unit = %s,
                            factor = %s,
                            source = %s,
                            region = %s,
                            currency = %s,
                            method = %s,
                            valid_from = %s,
                            valid_to = %s,
                            report_label = %s
                        WHERE db_id = %s
                        """,
                        [
                            p_file_name,
                            p_category,
                            p_level_1,
                            p_level_2,
                            p_level_3,
                            p_level_4,
                            p_column_text,
                            p_uom,
                            p_ghg_unit,
                            p_factor,
                            p_source,
                            p_region,
                            p_currency,
                            p_method,
                            p_valid_from,
                            p_valid_to,
                            p_report_label,
                            int(existing[0]),
                        ],
                    )
                else:
                    con.execute(
                        """
                    INSERT INTO factor_lookup
                    (dataset_id, file_name, year, original_id, scope,
                     category, level_1, level_2, level_3, level_4, column_text,
                     uom, ghg_unit, factor, source, region, currency,
                     method, valid_from, valid_to, report_label)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                        [
                            p_dataset_id,
                            p_file_name,
                            p_year,
                            p_original_id,
                            p_scope,
                            p_category,
                            p_level_1,
                            p_level_2,
                            p_level_3,
                            p_level_4,
                            p_column_text,
                            p_uom,
                            p_ghg_unit,
                            p_factor,
                            p_source,
                            p_region,
                            p_currency,
                            p_method,
                            p_valid_from,
                            p_valid_to,
                            p_report_label,
                        ],
                    )
            else:
                existing = con.execute(
                    """
                    SELECT db_id
                    FROM factor_lookup
                    WHERE dataset_id = ?
                      AND year = ?
                      AND scope = ?
                      AND original_id = ?
                    ORDER BY db_id
                    LIMIT 1
                    """,
                    [p_dataset_id, p_year, p_scope, p_original_id],
                ).fetchone()
                if replace and existing:
                    con.execute(
                        """
                    UPDATE factor_lookup
                    SET file_name = ?,
                        category = ?,
                        level_1 = ?,
                        level_2 = ?,
                        level_3 = ?,
                            level_4 = ?,
                            column_text = ?,
                            uom = ?,
                            ghg_unit = ?,
                            factor = ?,
                            source = ?,
                            region = ?,
                            currency = ?,
                            method = ?,
                            valid_from = ?,
                            valid_to = ?,
                            report_label = ?
                        WHERE db_id = ?
                        """,
                        [
                            p_file_name,
                            p_category,
                            p_level_1,
                            p_level_2,
                            p_level_3,
                            p_level_4,
                            p_column_text,
                            p_uom,
                            p_ghg_unit,
                            p_factor,
                            p_source,
                            p_region,
                            p_currency,
                            p_method,
                            p_valid_from,
                            p_valid_to,
                            p_report_label,
                            int(existing[0]),
                        ],
                    )
                else:
                    con.execute(
                        """
                    INSERT INTO factor_lookup
                    (dataset_id, file_name, year, original_id, scope,
                     category, level_1, level_2, level_3, level_4, column_text,
                     uom, ghg_unit, factor, source, region, currency,
                     method, valid_from, valid_to, report_label)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                        [
                            p_dataset_id,
                            p_file_name,
                            p_year,
                            p_original_id,
                            p_scope,
                            p_category,
                            p_level_1,
                            p_level_2,
                            p_level_3,
                            p_level_4,
                            p_column_text,
                            p_uom,
                            p_ghg_unit,
                            p_factor,
                            p_source,
                            p_region,
                            p_currency,
                            p_method,
                            p_valid_from,
                            p_valid_to,
                            p_report_label,
                        ],
                    )

        if db_backend() == "postgres":
            raw = con._conn  # type: ignore[attr-defined]
            with raw.cursor() as cur:
                _sync_normalised_tables_pg(cur, rows)

    parsed.update(
        {
            "deleted_rows": deleted_rows,
            "message": f"Imported {len(rows)} factor rows",
        }
    )
    return parsed


def ingest_csv(path: Path, *, replace: bool, dataset_id: int | None = None) -> tuple[int, int]:
    report = ingest_csv_with_report(path, replace=replace, dataset_id=dataset_id)
    return int(report["dataset_id"]), int(report["accepted_rows"])


def main() -> int:
    parser = argparse.ArgumentParser(description="Ingest conversion factor CSVs or Excel workbooks into datasets + factor_lookup")
    parser.add_argument("--folder", default=str(Path(__file__).parent / "assets" / "conversion_factors"))
    parser.add_argument("--workbook", default="", help="Path to a year-split Excel workbook to import")
    parser.add_argument("--replace", action="store_true", help="Delete existing factor_lookup rows for matched datasets before inserting")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate only; do not write changes")
    args = parser.parse_args()

    print(f"DB backend: {db_backend()}")

    if args.workbook:
        workbook_path = Path(args.workbook)
        if not workbook_path.exists():
            raise SystemExit(f"Workbook not found: {workbook_path}")
        if workbook_path.suffix.lower() != ".xlsx":
            raise SystemExit("Workbook imports require an .xlsx file")

        if args.dry_run:
            summaries = _parse_conversion_factor_workbook(workbook_path)
            total_rows = 0
            for summary in summaries:
                total_rows += int(summary.get("accepted_rows") or 0)
                print(
                    f"{summary['sheet_name']}: dataset_id={summary['dataset_id']} "
                    f"accepted={summary['accepted_rows']} rejected={summary['rejected_rows']} "
                    f"year={summary['year']}"
                )
            print(f"Dry run complete. Workbook contains {total_rows} valid factor rows.")
            return 0

        summaries = ingest_workbook_with_report(workbook_path, replace=bool(args.replace))
        total_written = 0
        for summary in summaries:
            total_written += int(summary.get("inserted_rows") or 0) + int(summary.get("updated_rows") or 0)
            print(
                f"{summary['sheet_name']}: dataset_id={summary['dataset_id']} "
                f"accepted={summary['accepted_rows']} updated={summary['updated_rows']} "
                f"inserted={summary['inserted_rows']} deleted={summary['deleted_rows']} "
                f"blocked={summary['blocked_rows']}"
                )
            if summary.get("rejected_rows"):
                print(f"  rejected={summary['rejected_rows']}")
        print(f"Done. Applied {total_written} factor row changes from workbook.")
        return 0

    folder = Path(args.folder)
    if not folder.exists():
        raise SystemExit(f"Folder not found: {folder}")

    files = sorted([p for p in folder.iterdir() if p.is_file() and p.suffix.lower() == ".csv"])
    if not files:
        raise SystemExit(f"No .csv files found in: {folder}")

    total = 0
    for p in files:
        ds_id, n = ingest_csv(p, replace=bool(args.replace))
        total += n
        print(f"{p.name}: dataset_id={ds_id}, inserted={n}")

    print(f"Done. Inserted {total} factor rows.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
